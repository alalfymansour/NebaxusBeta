from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, session, send_from_directory, send_file, make_response, current_app
from uuid import uuid4
from sqlalchemy import extract, or_, and_, text, case, select
from sqlalchemy.orm import joinedload, selectinload, load_only
from datetime import datetime, timedelta, date
import io
from .models import Customer, Product, ProductVariant, Order, OrderItem, Issue, IssueItem, CustomerLog, Employee, SalaryTransaction, AttendanceRecord, EmployeeLoginLog, Supplier, Invoice, InvoiceItem, EmployeeActivityLog, Expense, ReplacementOrder, ReplacementOrderItem, DamagedProductLog, FollowUp, BundleItem, SupplierReturn, SupplierReturnItem, SupplierDebt, Party, Transaction, CapitalSnapshot, AppSettings, CapitalGrowthHistory, OrderEditLog, OrderStatusHistory, ReplacementOrderStatusHistory, StockTake, StockTakeItem, ReturnOrder, ReturnOrderItem, MonthlyGoal, ORDER_TYPE_DELIVERY, ORDER_TYPE_WALKIN
from . import db
from sqlalchemy.exc import IntegrityError
from functools import wraps
from sqlalchemy import func
from collections import defaultdict, Counter
import random
import string
import json
import os
import time
import re
from .permissions import build_permission_dict, require_permissions

GOV_TO_WASSALHA = {
    "القاهرة":       "CAIRO",
    "الجيزة":        "GIZA",
    "الإسكندرية":    "ALEXANDRIA",
    "البحيرة":       "BEHIRA",
    "القليوبية":     "QALIUBIA",
    "الغربية":       "GHARBIA",
    "المنوفية":      "MONOUFIA",
    "دمياط":         "DOMITTA",
    "الدقهلية":      "DAKAHLIA",
    "كفر الشيخ":     "KAFR EL SHEIKH",
    "مطروح":         "MARSA MATROUH",
    "الإسماعيلية":   "ISMAILIA",
    "السويس":        "SUEZ",
    "بورسعيد":       "PORT SAID",
    "الشرقية":       "SHARKIA",
    "الفيوم":        "FAYOUM",
    "بني سويف":      "BANI SWEIF",
    "المنيا":        "MENIA",
    "أسيوط":         "ASSIUT",
    "سوهاج":         "SOUHAGE",
    "قنا":           "QENA",
    "أسوان":         "ASWAN",
    "الأقصر":        "LOUXOR",
    "البحر الأحمر":  "RED SEA",
    "الوادي الجديد": "NEW VALLLEY",
    "شمال سيناء":    "NOURTH SINAI",
    "جنوب سيناء":    "SOUTH SINAI",
}

def _build_wassalha_description(items):
    parts = []
    for item in items:
        if getattr(item, 'is_damaged', None) is not None and item.state not in ('سليم', 'منتج جديد', 'جديد'):
            continue
        if not item.product:
            continue
        if item.product.is_bundle and item.product.bundle_items:
            bv_map = item.bundle_variants_map
            bundle_parts = []
            for idx, bi in enumerate(item.product.bundle_items):
                bp = bi.product.name if bi.product else 'منتج محذوف'
                bv = bv_map.get(idx, {})
                v_names = [f"{k}: {v.variant_name}" for k, v in bv.get('variants', {}).items()]
                if v_names:
                    bp += f" ({', '.join(v_names)})"
                bundle_parts.append(bp)
            part = f"{item.product.name} × {item.quantity}: {' ├ '.join(bundle_parts)}"
        else:
            part = f"{item.product.name} × {item.quantity}"
            if item.selected_variants:
                variant_names = [
                    v.variant_name
                    for v in item.selected_variants.values()
                    if getattr(v, 'variant_name', None)
                ]
                if variant_names:
                    part += f" - {' / '.join(variant_names)}"
        parts.append(part)
    return ' | '.join(parts)

def _generate_wassalha_xlsx(orders):
    import openpyxl, io, os
    merchant  = AppSettings.get_value('wassalha_merchant_name', '')
    warehouse = AppSettings.get_value('wassalha_warehouse_name', '')
    seller    = AppSettings.get_value('wassalha_seller_name', '')
    template_path = os.path.join(
        current_app.root_path, 'static', 'templates', 'wassalha_template.xlsx'
    )
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Sheet1']
    for i, order in enumerate(orders, start=2):
        description = _build_wassalha_description(order.items)
        gov       = order.customer_governorate or ''
        city_code = GOV_TO_WASSALHA.get(gov, gov.upper())
        if getattr(order, 'customer_refund_amount', None) is not None:
            cod_value = (order.total_amount or 0) - (order.amount_paid or 0) - (order.customer_refund_amount or 0)
        else:
            cod_value = getattr(order, 'remaining_amount', 0) or 0
        has_pod   = 'no'
        ws.cell(row=i, column=1).value  = None
        ws.cell(row=i, column=2).value  = description
        ws.cell(row=i, column=3).value  = order.weight or 0
        ws.cell(row=i, column=4).value  = order.package_volume
        ws.cell(row=i, column=5).value  = cod_value
        ws.cell(row=i, column=6).value  = order.delivery_notes
        ws.cell(row=i, column=7).value  = order.customer_name
        ws.cell(row=i, column=8).value  = order.customer_phone
        ws.cell(row=i, column=9).value  = order.customer_address_details
        ws.cell(row=i, column=10).value = city_code
        ws.cell(row=i, column=11).value = None
        ws.cell(row=i, column=12).value = merchant
        ws.cell(row=i, column=13).value = warehouse
        ws.cell(row=i, column=14).value = has_pod
        ws.cell(row=i, column=15).value = seller
        ws.cell(row=i, column=16).value = None
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _order_status_dt(order):
    return getattr(order, 'status_updated_at', None) or getattr(order, 'updated_at', None) or getattr(order, 'created_at', None) or getattr(order, 'date', None)

def generate_tracking_number():
    while True:
        tracking = ''.join(random.choices(string.digits, k=8))
        existing_order = Order.query.filter_by(tracking_number=tracking).first()
        if not existing_order:
            return tracking

def calculate_cod_fee_for_order(order):
    base_remaining = (order.total_amount or 0) + (order.delivery_fees or 0) - (order.amount_paid or 0)
    
    if base_remaining > 0:
        cod_fee = AppSettings.get_cod_fee()
        return cod_fee
    else:
        return 0.0

def calculate_cod_fee_for_replacement(replacement_order):
    base_remaining = (replacement_order.total_amount or 0) + (replacement_order.delivery_fees or 0) - (replacement_order.amount_paid or 0)
    
    if base_remaining > 0:
        cod_fee = AppSettings.get_cod_fee()
        return cod_fee
    else:
        return 0.0

def apply_cod_fee_to_order(order):
    order.cod_fee_applied = calculate_cod_fee_for_order(order)

def apply_inventory_deduction(order, product_required_qty, locked_products):
    if order.inventory_deducted:
        return
    for pid, req_qty in product_required_qty.items():
        p = locked_products.get(pid)
        if p and _block_if_deleted(p, 'الطلب', 'deduct'):
            raise ValueError(f"Product {p.id} is deleted — cannot deduct inventory")
    to_log = {}
    for pid, req_qty in product_required_qty.items():
        p = locked_products.get(pid)
        if not p:
            continue
        p.stock = (p.stock or 0) - req_qty
        to_log[pid] = -req_qty
    order.inventory_deducted = True
    _log_stock_delta("add_order deduct", to_log)

def apply_cod_fee_to_replacement(replacement_order):
    replacement_order.cod_fee_applied = calculate_cod_fee_for_replacement(replacement_order)

def log_activity(action, entity_type, entity_id=None, entity_name=None, details=None):
    if 'employee_id' in session:
        try:
            activity_log = EmployeeActivityLog(
                employee_id=session['employee_id'],
                action=action,
                entity_type=entity_type,
                entity_id=entity_id,
                entity_name=entity_name,
                details=details,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(activity_log)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.debug(f"Error logging activity: {e}")

def _log_stock_delta(context, deltas):
    try:
        if not deltas:
            return
        total = sum(deltas.values())
        current_app.logger.info(f"[StockDelta] {context} items={len(deltas)} total={total} details={deltas}")
    except Exception:
        pass

def _extract_invoice_product_indices(form_data, required_field):
    indices = set()
    pattern = re.compile(rf"^products\[(\d+)\]\[{re.escape(required_field)}\]$")
    for key in form_data.keys():
        match = pattern.match(key)
        if match:
            indices.add(int(match.group(1)))
    return sorted(indices)

def _format_invoice_merge_message(product_name, line_count, total_quantity):
    line_label = 'سطرين' if line_count == 2 else f'{line_count} سطور'
    return f"تم دمج منتج {product_name} ({line_label}) في سطر واحد بكمية إجمالية {total_quantity}"


def _get_locked_product(product_id):
    try:
        product_id = int(product_id)
    except (TypeError, ValueError):
        return None
    return Product.query.with_for_update().filter(Product.id == product_id).first()


def _get_locked_products(product_ids):
    locked_products = {}
    normalized_ids = []
    seen_ids = set()
    for product_id in product_ids or []:
        try:
            product_id = int(product_id)
        except (TypeError, ValueError):
            continue
        if product_id in seen_ids:
            continue
        seen_ids.add(product_id)
        normalized_ids.append(product_id)

    if not normalized_ids:
        return locked_products

    for product in Product.query.with_for_update().filter(Product.id.in_(normalized_ids)).all():
        locked_products[product.id] = product
    return locked_products


def _get_replacement_item_product_id(item):
    pid = getattr(item, 'product_id', None)
    if pid:
        return pid
    for attr in ('size_variant', 'color_variant', 'style_variant', 'variant'):
        variant_obj = getattr(item, attr, None)
        if variant_obj and getattr(variant_obj, 'product_id', None):
            return variant_obj.product_id
    return None


def _collect_replacement_required_quantities(items):
    required = defaultdict(int)
    for it in items:
        try:
            st = (it.state or '').strip()
        except Exception:
            st = ''
        if st not in ['منتج جديد', 'جديد']:
            continue

        pid = _get_replacement_item_product_id(it)
        if not pid:
            continue

        product = _get_locked_product(pid)
        if not product:
            continue

        qty = int(it.quantity or 0)
        if qty <= 0:
            continue

        if product.is_bundle:
            bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
            for bundle_item in bundle_items:
                required[bundle_item.product_id] += qty
        else:
            required[product.id] += qty
    return required


VALID_TRANSITIONS = {
    'جديد': {'خرج للتوصيل', 'إلغاء'},
    'خرج للتوصيل': {'وصل', 'رفض الاستلام', 'إلغاء'},
    'رفض الاستلام': {'خرج للتوصيل', 'إلغاء'},
    'وصل': set(),
    'إلغاء': set(),
}

SPECIAL_STATUSES = {'استبدال', 'مرتجع'}

ALL_STATUSES = sorted(
    set(VALID_TRANSITIONS.keys())
    | {s for v in VALID_TRANSITIONS.values() for s in v}
    | SPECIAL_STATUSES
)

class StatusService:
    @staticmethod
    def valid_next(status: str) -> list[str]:
        if status in SPECIAL_STATUSES:
            return [status]
        return sorted(set(VALID_TRANSITIONS.get(status, [])) | {status})

def is_replacement(obj) -> bool:
    return isinstance(obj, ReplacementOrder)

def validate_transition(old_status, new_status):
    if old_status == new_status:
        return True
    if old_status in SPECIAL_STATUSES or new_status in SPECIAL_STATUSES:
        raise ValueError(f"انتقال غير مسموح به: الحالة '{old_status}' -> '{new_status}' غير مسموح به")
    if new_status not in VALID_TRANSITIONS.get(old_status, set()):
        raise ValueError(f"انتقال غير مسموح به: {old_status} -> {new_status}")
    return True

def _order_status_counts(month, year):
    from sqlalchemy import func
    return dict(
        db.session.query(Order.status, func.count(Order.id))
        .filter(
            extract('month', Order.date) == month,
            extract('year', Order.date) == year
        )
        .group_by(Order.status)
        .all()
    )

def _annotate_status_age(orders, status_filter):
    try:
        now_dt = datetime.now()
        if status_filter == 'جديد':
            for o in orders:
                ts = _order_status_dt(o)
                if ts:
                    age_days = (now_dt - ts).days
                else:
                    age_days = 0
                o.status_age_days = age_days
                o.is_overdue_yellow = age_days >= 5
                o.is_overdue_red = age_days >= 10
                o.is_overdue = o.is_overdue_yellow
        else:
            for o in orders:
                o.status_age_days = 0
                o.is_overdue_yellow = False
                o.is_overdue_red = False
                o.is_overdue = False
    except Exception:
        for o in orders:
            o.status_age_days = 0
            o.is_overdue_yellow = False
            o.is_overdue_red = False
            o.is_overdue = False

def _collect_order_required_quantities(order):
    from collections import defaultdict
    req = defaultdict(int)
    for it in order.items:
        product = it.product
        if product:
            if product.is_bundle:
                bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                for bundle_item in bundle_items:
                    req[bundle_item.product_id] += (it.quantity or 0) * 1
            else:
                req[product.id] += (it.quantity or 0)
    return req


def _block_if_deleted(product, context_msg="هذا الطلب", direction="deduct"):
    if direction == "restock":
        return False
    if product and product.is_deleted:
        flash(f'{context_msg} يحتوي على منتج محذوف ("{product.name}") ولا يمكن إتمام العملية', 'error')
        return True
    return False


DELIVERY_CALL_NOTE_PREFIX = 'التحقق من التسليم: '
DELIVERY_CALL_LEGACY_NOTE_PREFIX = 'مكالمة تحقق من محاولة تسليم: '
DELIVERY_CALL_STATE_INFORMED = 'informed'
DELIVERY_CALL_STATE_CONTACTED = 'contacted'
DELIVERY_CALL_STATE_NOT_CONTACTED = 'not_contacted'
DELIVERY_CALL_STATE_NONE = 'none'


def _delivery_call_note_for_state(state):
    if state == DELIVERY_CALL_STATE_INFORMED:
        return f'{DELIVERY_CALL_NOTE_PREFIX}تم ابلاغ العميل بخروج الطلب للتوصيل'
    if state == DELIVERY_CALL_STATE_CONTACTED:
        return f'{DELIVERY_CALL_NOTE_PREFIX}تم التواصل مع العميل من جانب المندوب'
    if state == DELIVERY_CALL_STATE_NOT_CONTACTED:
        return f'{DELIVERY_CALL_NOTE_PREFIX}لم يتم التواصل مع العميل من جانب المندوب'
    return None


def _delivery_call_state_from_note(note):
    normalized = (note or '').strip()
    if not (
        normalized.startswith(DELIVERY_CALL_NOTE_PREFIX)
        or normalized.startswith(DELIVERY_CALL_LEGACY_NOTE_PREFIX)
    ):
        return DELIVERY_CALL_STATE_NONE
    if 'تم ابلاغ العميل بخروج الطلب للتوصيل' in normalized:
        return DELIVERY_CALL_STATE_INFORMED
    if 'لم يتم التواصل مع العميل من جانب المندوب' in normalized:
        return DELIVERY_CALL_STATE_NOT_CONTACTED
    if 'تم التواصل مع العميل من جانب المندوب' in normalized:
        return DELIVERY_CALL_STATE_CONTACTED
    return DELIVERY_CALL_STATE_NONE


def _delivery_call_review_description_from_note(note):
    normalized = (note or '').strip()
    if not normalized:
        return None
    if not (
        normalized.startswith(DELIVERY_CALL_NOTE_PREFIX)
        or normalized.startswith(DELIVERY_CALL_LEGACY_NOTE_PREFIX)
    ):
        return None
    if 'تم ابلاغ العميل بخروج الطلب للتوصيل' in normalized:
        return 'تم ابلاغ العميل بخروج الطلب للتوصيل'
    if 'لم يتم التواصل مع العميل من جانب المندوب' in normalized:
        return 'لم يتم التواصل مع العميل من جانب المندوب'
    if 'تم التواصل مع العميل من جانب المندوب' in normalized:
        return 'تم التواصل مع العميل من جانب المندوب'
    return None


def _order_delivery_call_state_map(order_ids):
    ids = [oid for oid in (order_ids or []) if oid]
    if not ids:
        return {}

    rows = (
        db.session.query(OrderStatusHistory.order_id, OrderStatusHistory.notes)
        .filter(
            OrderStatusHistory.order_id.in_(ids),
            OrderStatusHistory.notes.isnot(None),
            or_(
                OrderStatusHistory.notes.like(f"{DELIVERY_CALL_NOTE_PREFIX}%"),
                OrderStatusHistory.notes.like(f"{DELIVERY_CALL_LEGACY_NOTE_PREFIX}%")
            )
        )
        .order_by(OrderStatusHistory.timestamp.desc(), OrderStatusHistory.id.desc())
        .all()
    )

    state_map = {}
    for order_id, note in rows:
        if order_id in state_map:
            continue
        state_map[order_id] = _delivery_call_state_from_note(note)
    return state_map


def _replacement_delivery_call_state_map(order_ids):
    ids = [oid for oid in (order_ids or []) if oid]
    if not ids:
        return {}

    rows = (
        db.session.query(ReplacementOrderStatusHistory.replacement_order_id, ReplacementOrderStatusHistory.notes)
        .filter(
            ReplacementOrderStatusHistory.replacement_order_id.in_(ids),
            ReplacementOrderStatusHistory.notes.isnot(None),
            or_(
                ReplacementOrderStatusHistory.notes.like(f"{DELIVERY_CALL_NOTE_PREFIX}%"),
                ReplacementOrderStatusHistory.notes.like(f"{DELIVERY_CALL_LEGACY_NOTE_PREFIX}%")
            )
        )
        .order_by(ReplacementOrderStatusHistory.timestamp.desc(), ReplacementOrderStatusHistory.id.desc())
        .all()
    )

    state_map = {}
    for order_id, note in rows:
        if order_id in state_map:
            continue
        state_map[order_id] = _delivery_call_state_from_note(note)
    return state_map

def log_operation(action, entity_type):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            result = f(*args, **kwargs)

            entity_id = kwargs.get('customer_id') or kwargs.get('product_id') or kwargs.get(
                'order_id') or kwargs.get('employee_id') or kwargs.get('supplier_id') or kwargs.get('invoice_id')

            entity_name = None
            if entity_id:
                try:
                    if entity_type == 'customer':
                        entity = Customer.query.get(entity_id)
                        entity_name = entity.name if entity else None
                    elif entity_type == 'product':
                        entity = Product.query.get(entity_id)
                        entity_name = entity.name if entity else None
                    elif entity_type == 'order':
                        entity = Order.query.get(entity_id)
                        entity_name = f"طلب رقم {entity_id}" if entity else None
                    elif entity_type == 'employee':
                        entity = Employee.query.get(entity_id)
                        entity_name = entity.name if entity else None
                    elif entity_type == 'supplier':
                        entity = Supplier.query.get(entity_id)
                        entity_name = entity.name if entity else None
                    elif entity_type == 'invoice':
                        entity = Invoice.query.get(entity_id)
                        entity_name = f"فاتورة رقم {entity.invoice_number}" if entity else None
                except BaseException:
                    entity_name = None

            log_activity(action, entity_type, entity_id, entity_name)
            return result
        return decorated_function
    return decorator

main = Blueprint('main', __name__)

if os.environ.get('NEBAXUS_MODE') == 'trial':
    from .trial import before_request_check
    @main.before_request
    def _trial_limit():
        return before_request_check()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'employee_id' not in session:
            flash('يجب تسجيل الدخول أولاً')
            return redirect(url_for('main.login'))
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'employee_id' not in session:
                flash('يجب تسجيل الدخول أولاً')
                return redirect(url_for('main.login'))

            employee = Employee.query.get(session['employee_id'])
            if not employee or not employee.is_active:
                flash('حسابك غير نشط')
                session.clear()
                return redirect(url_for('main.login'))

            if not getattr(
                    employee,
                    permission,
                    False) and not employee.is_admin:
                flash('ليس لديك صلاحية للوصول لهذه الصفحة')
                return redirect(url_for('main.dashboard'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator


def _build_customers_query(search_query):
    query = Customer.query.options(
        load_only(Customer.id, Customer.name, Customer.phone, Customer.governorate, Customer.address_details)
    )
    if search_query:
        # Prefix matching is much cheaper than %term% and benefits from phone index.
        escaped_query = search_query.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
        query = query.filter(Customer.phone.like(f"{escaped_query}%", escape='\\'))
    return query


def _get_customers_total_count(base_query, search_query):
    if search_query:
        return base_query.with_entities(func.count(Customer.id)).scalar() or 0

    try:
        if db.session.bind and db.session.bind.dialect.name == 'postgresql':
            estimated = db.session.execute(
                text("SELECT COALESCE(reltuples, 0)::bigint FROM pg_class WHERE oid = 'customer'::regclass")
            ).scalar()
            if estimated is not None:
                return int(estimated)
    except Exception as exc:
        current_app.logger.debug(f"Failed to read customer estimated count: {exc}")

    return base_query.with_entities(func.count(Customer.id)).scalar() or 0

@main.route('/followups', methods=['GET'])
@login_required
@permission_required('can_view_followups')
def followups():
    status = request.args.get('status', 'قائمة')
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    phone_search = request.args.get('phone', '').strip()

    now = datetime.utcnow()
    if not month:
        month = now.month
    if not year:
        year = now.year

    query = FollowUp.query.join(Customer)
    if status and status != 'الكل':
        query = query.filter(FollowUp.status == status)
    if month:
        query = query.filter(extract('month', FollowUp.created_at) == month)
    if year:
        query = query.filter(extract('year', FollowUp.created_at) == year)
    if phone_search:
        query = query.filter(Customer.phone.ilike(f"%{phone_search}%"))

    followups = query.order_by(FollowUp.created_at.desc()).all()

    years = [r[0] for r in db.session.query(extract('year', FollowUp.created_at).label('y')).distinct().order_by('y').all()]
    if year not in years and years:
        years.append(year)
        years = sorted(set(years))

    return render_template('followups.html',
                           followups=followups,
                           selected_status=status,
                           selected_month=month,
                           selected_year=year,
                           phone_search=phone_search,
                           years=years,
                           now=datetime.utcnow())

@main.route('/followups/<int:followup_id>/toggle', methods=['POST'])
@login_required
@permission_required('can_edit_followups')
def toggle_followup_status(followup_id):
    fu = FollowUp.query.get_or_404(followup_id)
    fu.status = 'تم الحل' if fu.status != 'تم الحل' else 'قائمة'
    if fu.status == 'تم الحل':
        fu.next_contact_due = None
    db.session.commit()
    flash('تم تحديث حالة المشكلة')
    return redirect(url_for('main.followups', **request.args))

@main.route('/followups/<int:followup_id>/contact', methods=['POST'])
@login_required
@permission_required('can_edit_followups')
def followup_contact(followup_id):
    fu = FollowUp.query.get_or_404(followup_id)
    fu.last_contact_at = datetime.utcnow()
    if fu.status != 'تم الحل':
        fu.next_contact_due = fu.last_contact_at + timedelta(hours=24)
    try:
        contact_log = CustomerLog(
            customer_id=fu.customer_id,
            type='تواصل',
            employee_id=session.get('employee_id'),
            content=f'تم التواصل مع العميل بشأن المتابعة (رقم {fu.id}) في {fu.last_contact_at.strftime("%Y-%m-%d %H:%M")}'
        )
        db.session.add(contact_log)
    except Exception:
        pass
    db.session.commit()
    flash('تم تسجيل عملية التواصل وسيتم التذكير بعد 24 ساعة ما لم يتم الحل')
    return redirect(url_for('main.followups', **request.args))

@main.route('/followups/<int:followup_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_followups')
def delete_followup(followup_id):
    fu = FollowUp.query.get_or_404(followup_id)
    db.session.delete(fu)
    db.session.commit()
    flash('تم حذف المتابعة')
    return redirect(url_for('main.followups', **request.args))

@main.route('/followups/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_followups')
def add_followup():
    if request.method == 'POST':
        customer_id = request.form.get('customer_id', type=int)
        problem = (request.form.get('problem') or '').strip()
        if not customer_id or not problem:
            flash('يجب اختيار عميل وكتابة المشكلة')
            return redirect(url_for('main.add_followup'))
        customer = Customer.query.get(customer_id)
        if not customer:
            flash('العميل غير موجود')
            return redirect(url_for('main.add_followup'))
        fu = FollowUp(customer_id=customer_id, problem=problem, status='قائمة')
        db.session.add(fu)
        db.session.commit()
        flash('تم إضافة المشكلة إلى قائمة المتابعة')
        return redirect(url_for('main.followups'))

    return render_template('add_followup.html')

@main.route('/api/search_customers')
@login_required
@permission_required('can_view_customers')
def api_search_customers():
    term = (request.args.get('q') or '').strip()
    by_id = request.args.get('by') == 'id'
    q = Customer.query
    if term:
        if by_id:
            q = q.filter(Customer.id == int(term))
        else:
            like = f"%{term}%"
            q = q.filter(or_(Customer.phone.ilike(like), Customer.name.ilike(like)))
    customers = q.order_by(Customer.name.asc()).limit(25).all()
    results = []
    for c in customers:
        label = f"{c.name} - {c.phone}"
        results.append({
            'id': c.id,
            'text': label,
            'name': c.name,
            'phone': c.phone,
            'address': f"{c.governorate or ''} - {c.address_details or ''}"
        })
    return jsonify({'results': results})

@main.route('/api/orders/<int:order_id>/items')
@login_required
@permission_required('can_view_orders')
def api_get_order_items(order_id):
    order = Order.query.options(
        joinedload(Order.items).joinedload(OrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product)
    ).filter_by(id=order_id).first()
    if not order:
        return jsonify({'success': False, 'error': 'الطلب غير موجود'}), 404
    items = []
    for item in order.items:
        if item.product and item.product.is_bundle:
            bundle_items = BundleItem.query.filter_by(bundle_id=item.product.id).all()
            for bundle_item in bundle_items:
                if bundle_item.product:
                    variants = []
                    items.append({
                        'id': f"{item.id}_bundle_{bundle_item.id}",
                        'product_id': bundle_item.product_id,
                        'product_name': bundle_item.product.name,
                        'price': bundle_item.sale_price_in_bundle,
                        'purchase_price': bundle_item.product.purchase_price or 0,
                        'state': item.state,
                        'variants': variants,
                        'from_bundle': True,
                        'bundle_name': item.product.name
                    })
        else:
            variants = []
            if hasattr(item, 'selected_variants') and item.selected_variants:
                for label, variant in item.selected_variants.items():
                    variant_name = getattr(variant, 'variant_name', str(variant))
                    variants.append({'label': label, 'value': variant_name})
            items.append({
                'id': item.id,
                'product_id': item.product_id,
                'product_name': item.product.name if item.product else 'منتج محذوف',
                'price': item.price,
                'purchase_price': item.product.purchase_price if item.product else 0,
                'state': item.state,
                'variants': variants,
                'from_bundle': False
            })
    return jsonify({'success': True, 'items': items})

@main.route('/api/product_price/<int:product_id>')
@login_required
def get_product_price(product_id):
    product = Product.query.get_or_404(product_id)
    if product.is_bundle:
        current_price = product.bundle_total_price
    else:
        current_price = product.price or 0
    price_data = {
        'price': current_price,
        'wholesale_price': product.wholesale_price or 0,
        'is_bundle': product.is_bundle,
    }
    response = jsonify(price_data)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

def get_min_price(product):
    if product.is_bundle:
        total = 0
        for bi in product.bundle_items:
            if not bi.product:
                return float('inf')
            total += bi.product.purchase_price or 0
        return total
    return product.purchase_price or 0

@main.route('/api/validate_price/<int:product_id>')
@login_required
def validate_price(product_id):
    product = Product.query.get_or_404(product_id)
    price = request.args.get('price', type=float, default=0)
    min_price = get_min_price(product)
    return jsonify({'valid': price >= min_price})

@main.route('/api/validate_order_prices', methods=['POST'])
@login_required
def validate_order_prices():
    data = request.get_json() or []
    results = []
    all_valid = True
    for item in data:
        pid = item.get('product_id')
        if not pid:
            results.append({'product_id': None, 'valid': True})
            continue
        product = Product.query.get(int(pid))
        if not product:
            results.append({'product_id': pid, 'valid': True})
            continue
        price = float(item.get('price', 0))
        min_price = get_min_price(product)
        valid = price >= min_price
        if not valid:
            all_valid = False
        results.append({'product_id': pid, 'valid': valid})
    return jsonify({'all_valid': all_valid, 'results': results})

@main.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        fast_splash = os.environ.get('FAST_SPLASH') == '1'
        if fast_splash:
            return render_template('login.html')
        if request.args.get('from_splash') == '1':
            return render_template('login.html')
        session.clear()
        return render_template('splash.html', redirect_to_login=True)

    username = request.form['username']
    password = request.form['password']

    t0 = time.perf_counter()
    employee = Employee.query.filter_by(username=username, is_active=True).first()
    t1 = time.perf_counter()

    if employee and employee.check_password(password):
        t2 = time.perf_counter()
        session['employee_id'] = employee.id
        session['employee_name'] = employee.name
        session['employee_username'] = employee.username
        session['is_admin'] = employee.is_admin

        fast_login = os.environ.get('FAST_LOGIN') == '1'
        commit_duration = 0.0
        if not fast_login:
            login_log = EmployeeLoginLog(
                employee_id=employee.id,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(login_log)
            activity = EmployeeActivityLog(
                employee_id=employee.id,
                action='login',
                entity_type='session',
                entity_id=None,
                entity_name=employee.username,
                details='تسجيل دخول ناجح',
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(activity)
            db.session.commit()
            t3 = time.perf_counter()
            commit_duration = (t3 - t2) * 1000
        else:
            t3 = time.perf_counter()

        total_duration = (t3 - t0) * 1000
        query_duration = (t1 - t0) * 1000
        password_duration = (t2 - t1) * 1000
        current_app.logger.info(
            f"[LoginPerf] user={username} q={query_duration:.1f}ms pw={password_duration:.1f}ms commit={commit_duration:.1f}ms total={total_duration:.1f}ms fast={fast_login}")

        flash(f'مرحباً {employee.name}! تم تسجيل الدخول بنجاح')
        return redirect(url_for('main.dashboard'))
    else:
        fail_total = (time.perf_counter() - t0) * 1000
        current_app.logger.info(f"[LoginPerf] user={username} FAILED total={fail_total:.1f}ms")
        if employee:
            try:
                fail_activity = EmployeeActivityLog(
                    employee_id=employee.id,
                    action='login_failed',
                    entity_type='session',
                    entity_id=None,
                    entity_name=employee.username,
                    details='محاولة تسجيل دخول فاشلة (كلمة مرور خاطئة)',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
                db.session.add(fail_activity)
                db.session.commit()
            except Exception:
                db.session.rollback()
        flash('اسم المستخدم أو كلمة المرور غير صحيحة')

    return render_template('login.html')

@main.route('/logout')
def logout():
    session.clear()
    flash('تم تسجيل الخروج بنجاح')
    return redirect(url_for('main.splash'))

@main.route('/')
def splash():
    if 'employee_id' in session:
        return redirect(url_for('main.dashboard'))

    if session.get('splash_seen'):
        return redirect(url_for('main.login', from_splash=1))

    session['splash_seen'] = True
    return render_template('splash.html', redirect_to_login=True)

@main.route('/splash')
def splash_only():
    session.clear()
    response = make_response(render_template('splash.html'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@main.route('/health')
def health():
    return 'OK', 200

@main.route('/ping')
def ping():
    return 'pong', 200

@main.route('/test')
def test():
    return 'App is running! Database connection test...', 200

@main.route('/manifest.json')
def manifest():
    return send_from_directory(
        'static',
        'manifest.json',
        mimetype='application/json')

@main.route('/sw.js')
def service_worker():
    return send_from_directory(
        'static',
        'sw.js',
        mimetype='application/javascript')

@main.route('/dashboard')
@login_required
def dashboard():
    today = date.today()
    month_start = today.replace(day=1)

    recent_orders = (Order.query
                     .order_by(Order.id.desc())
                     .limit(10)
                     .all())

    orders_this_month = (db.session.query(func.count(Order.id))
                         .filter(Order.date >= month_start).scalar() or 0)
    target = MonthlyGoal.get_current_goal()

    out_stock_count = (db.session.query(func.count(OrderItem.id))
                       .join(Order)
                       .join(Product, OrderItem.product_id == Product.id)
                       .filter(Order.status == 'جديد', Product.stock <= 0)
                       .scalar() or 0)

    out_for_delivery = (db.session.query(func.count(Order.id))
                        .filter(Order.status == 'خرج للتوصيل', func.date(Order.date) == today)
                        .scalar() or 0)

    delivery_fees_total = (db.session.query(func.sum(Order.delivery_fees))
                           .scalar() or 0)

    new_customers_today = (db.session.query(func.count(Customer.id))
                           .filter(Customer.id.in_(
                               db.session.query(Order.customer_id)
                               .filter(func.date(Order.date) == today)
                           ))
                           .scalar() or 0)

    data = dict(
        total_orders=db.session.query(func.count(Order.id)).scalar() or 0,
        today_orders=(db.session.query(func.count(Order.id))
                      .filter(func.date(Order.date) == today).scalar() or 0),
        orders_this_month=orders_this_month,
        monthly_target=target,
        out_stock_count=out_stock_count,
        out_for_delivery=out_for_delivery,
        delivery_fees_total=delivery_fees_total,
        new_customers_today=new_customers_today,
        total_customers=db.session.query(func.count(Customer.id)).scalar() or 0,
        total_products=db.session.query(func.count(Product.id)).filter(Product.is_deleted == False).scalar() or 0,
        recent_orders=recent_orders,
    )
    return render_template('dashboard.html', **data)

@main.route('/transactions')
@login_required
@require_permissions('can_view_transactions')
def transactions():
    search_query = request.args.get('search', '').strip()
    page = 1
    per_page = 100

    base_query = Party.query
    if search_query:
        like_expr = f"%{search_query}%"
        base_query = base_query.filter(
            or_(
                Party.name.ilike(like_expr),
                Party.phone.ilike(like_expr)
            )
        )

    total_parties = base_query.count()
    parties_page = (base_query
                    .order_by(Party.id.desc())
                    .limit(per_page)
                    .offset((page - 1) * per_page)
                    .all())

    has_more = total_parties > per_page * page
    next_page = page + 1 if has_more else None

    return render_template(
        'transactions.html',
        parties=parties_page,
        search_query=search_query,
        total_parties=total_parties,
        has_more=has_more,
        next_page=next_page,
        per_page=per_page,
        page=page)

@main.route('/transactions/more')
@login_required
@require_permissions('can_view_transactions')
def transactions_more():
    try:
        page = int(request.args.get('page', '2'))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    per_page = 100
    search_query = request.args.get('search', '').strip()

    base_query = Party.query
    if search_query:
        like_expr = f"%{search_query}%"
        base_query = base_query.filter(
            or_(
                Party.name.ilike(like_expr),
                Party.phone.ilike(like_expr)
            )
        )

    total_parties = base_query.count()
    parties_page = (base_query
                    .order_by(Party.id.desc())
                    .limit(per_page)
                    .offset((page - 1) * per_page)
                    .all())

    has_more = total_parties > per_page * page
    next_page = page + 1 if has_more else None

    fragment_html = render_template('partials/_party_cards.html', parties=parties_page)
    return jsonify({
        'html': fragment_html,
        'has_more': has_more,
        'next_page': next_page,
        'loaded_count': len(parties_page),
        'total': total_parties
    })

@main.route('/transactions/add_party', methods=['GET', 'POST'])
@login_required
@require_permissions('can_add_transactions')
def add_party():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        notes = request.form.get('notes', '').strip()

        if not name:
            flash('يجب إدخال اسم المتعامل', 'danger')
            return redirect(url_for('main.add_party'))

        new_party = Party(
            name=name,
            phone=phone if phone else None,
            address=address if address else None,
            notes=notes if notes else None
        )
        db.session.add(new_party)
        db.session.commit()

        log_activity('add', 'party', new_party.id, new_party.name)
        flash(f'تم إضافة المتعامل {name} بنجاح', 'success')
        return redirect(url_for('main.transactions'))

    return render_template('add_party.html')

@main.route('/transactions/party/<int:party_id>/edit', methods=['GET', 'POST'])
@login_required
@require_permissions('can_edit_transactions')
def edit_party(party_id):
    party = Party.query.get_or_404(party_id)

    if request.method == 'POST':
        party.name = request.form.get('name', '').strip()
        party.phone = request.form.get('phone', '').strip() or None
        party.address = request.form.get('address', '').strip() or None
        party.notes = request.form.get('notes', '').strip() or None
        party.updated_at = datetime.utcnow()

        db.session.commit()
        log_activity('edit', 'party', party.id, party.name)
        flash('تم تحديث بيانات المتعامل بنجاح', 'success')
        return redirect(url_for('main.party_profile', party_id=party.id))

    return render_template('edit_party.html', party=party)

@main.route('/transactions/party/<int:party_id>/delete', methods=['POST'])
@login_required
@require_permissions('can_delete_transactions')
def delete_party(party_id):
    party = Party.query.get_or_404(party_id)
    party_name = party.name

    from .models import Transaction, TransactionPayment
    try:
        tx_ids = [t.id for t in db.session.query(Transaction.id).filter(Transaction.party_id == party.id).all()]
        db.session.query(TransactionPayment).filter(TransactionPayment.transaction_id.in_(tx_ids)).delete(synchronize_session=False)
        db.session.query(Transaction).filter(Transaction.party_id == party.id).delete(synchronize_session=False)
        db.session.delete(party)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('حدث خطأ أثناء حذف المتعامل أو سجلاته. تواصل مع الدعم.', 'danger')
        return redirect(url_for('main.party_profile', party_id=party.id))

    log_activity('delete', 'party', party_id, party_name)
    flash(f'تم حذف المتعامل {party_name} بنجاح', 'success')
    return redirect(url_for('main.transactions'))

@main.route('/transactions/party/<int:party_id>')
@login_required
@require_permissions('can_view_transactions')
def party_profile(party_id):
    party = Party.query.get_or_404(party_id)
    transactions = Transaction.query.filter_by(party_id=party_id).order_by(Transaction.transaction_date.desc()).all()
    from .models import TransactionPayment
    payment_logs = TransactionPayment.query.join(Transaction).filter(Transaction.party_id == party_id).order_by(TransactionPayment.created_at.desc()).all()

    # ponytail: running balance from all events sorted chronologically
    timeline = []
    for t in transactions:
        timeline.append((t.transaction_date, 'transaction', t))
    for p in payment_logs:
        timeline.append((p.created_at, 'payment', p))
    timeline.sort(key=lambda x: x[0])
    running = 0
    for date, typ, obj in timeline:
        if typ == 'transaction':
            running += obj.amount if obj.transaction_type == 'receivable' else -obj.amount
        else:
            running += -obj.amount if obj.transaction.transaction_type == 'receivable' else obj.amount
        obj._running_balance = running

    party.payment_logs = payment_logs

    return render_template('party_profile.html', party=party, transactions=transactions)

@main.route('/transactions/add', methods=['GET', 'POST'])
@login_required
@require_permissions('can_add_transactions')
def add_transaction_page():
    party_id_param = request.args.get('party_id', type=int)
    selected_party = None
    if party_id_param:
        selected_party = Party.query.get(party_id_param)
    
    if request.method == 'POST':
        party_id = request.form.get('party_id')
        party_name = request.form.get('party_name', '').strip()
        
        transaction_type = request.form.get('transaction_type')
        amount = request.form.get('amount', type=float)
        category = request.form.get('category', '').strip()
        description = request.form.get('description', '').strip()
        transaction_date = request.form.get('transaction_date')

        if not party_name:
            flash('يجب إدخال اسم المتعامل', 'danger')
            return redirect(url_for('main.add_transaction_page'))

        if not transaction_type or transaction_type not in ['receivable', 'payable']:
            flash('يجب اختيار نوع المعاملة', 'danger')
            return redirect(url_for('main.add_transaction_page'))

        if not amount or amount <= 0:
            flash('يجب إدخال مبلغ صحيح', 'danger')
            return redirect(url_for('main.add_transaction_page'))

        if party_id:
            party = Party.query.get(party_id)
            if not party:
                party = Party(name=party_name)
                db.session.add(party)
                db.session.flush()
        else:
            party = Party(name=party_name)
            db.session.add(party)
            db.session.flush()

        if transaction_date:
            try:
                trans_date = datetime.strptime(transaction_date, '%Y-%m-%d')
            except ValueError:
                trans_date = datetime.utcnow()
        else:
            trans_date = datetime.utcnow()

        new_transaction = Transaction(
            party_id=party.id,
            transaction_type=transaction_type,
            amount=amount,
            category=category if category else None,
            description=description if description else None,
            transaction_date=trans_date,
            created_by=session.get('employee_id')
        )
        db.session.add(new_transaction)
        db.session.commit()

        log_activity('add', 'transaction', new_transaction.id, f"{party.name} - {transaction_type}")
        flash('تم إضافة المعاملة بنجاح', 'success')
        return redirect(url_for('main.party_profile', party_id=party.id))

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('add_transaction.html', today=today, selected_party=selected_party)

@main.route('/api/search_parties')
@login_required
@permission_required('can_view_transactions')
def search_parties():
    query = request.args.get('q', '').strip()
    
    if len(query) < 2:
        return jsonify({'parties': []})
    
    parties = Party.query.filter(
        or_(
            Party.name.ilike(f'%{query}%'),
            Party.phone.ilike(f'%{query}%')
        )
    ).limit(10).all()
    
    results = []
    for party in parties:
        results.append({
            'id': party.id,
            'name': party.name,
            'phone': party.phone,
            'address': party.address,
            'balance': party.balance
        })
    
    return jsonify({'parties': results})

@main.route('/transactions/party/<int:party_id>/pay', methods=['POST'])
@login_required
@require_permissions('can_add_transactions')
def pay_party_transaction(party_id):
    party = Party.query.get_or_404(party_id)
    
    transaction_id = request.form.get('transaction_id', type=int)
    payment_amount = request.form.get('payment_amount', type=float)
    payment_notes = request.form.get('payment_notes', '').strip()
    
    if not transaction_id:
        flash('يجب اختيار المعاملة', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))
    
    if not payment_amount or payment_amount <= 0:
        flash('يجب إدخال مبلغ صحيح', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))
    
    transaction = Transaction.query.get_or_404(transaction_id)
    
    if transaction.party_id != party_id:
        flash('المعاملة لا تخص هذا المتعامل', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))
    
    remaining = transaction.remaining_amount
    if payment_amount > remaining:
        flash(f'المبلغ أكبر من المتبقي ({remaining:.2f} ج.م)', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))
    
    transaction.paid_amount = (transaction.paid_amount or 0) + payment_amount
    try:
        from .models import TransactionPayment
        payment_record = TransactionPayment(
            transaction_id=transaction.id,
            amount=payment_amount,
            notes=payment_notes if payment_notes else None,
            created_by=session.get('employee_id')
        )
        db.session.add(payment_record)
        db.session.commit()

        action_text = "استلام" if transaction.transaction_type == 'receivable' else "دفع"
        log_activity('payment', 'transaction', transaction.id, f"{action_text} {payment_amount} ج.م من {party.name}")
        flash(f'تم تسجيل {action_text} {payment_amount:.2f} ج.م بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error saving transaction payment: {e}")
        flash('تم تسجيل المبلغ ولكن حدث خطأ أثناء حفظ ملاحظة السداد', 'warning')

    return redirect(url_for('main.party_profile', party_id=party_id))

@main.route('/transactions/party/<int:party_id>/add_transaction', methods=['GET', 'POST'])
@login_required
@require_permissions('can_add_transactions')
def add_transaction(party_id):
    party = Party.query.get_or_404(party_id)

    if request.method == 'POST':
        transaction_type = request.form.get('transaction_type')
        amount = request.form.get('amount', type=float)
        description = request.form.get('description', '').strip()
        transaction_date = request.form.get('transaction_date')

        if not transaction_type or transaction_type not in ['receivable', 'payable']:
            flash('يجب اختيار نوع المعاملة', 'danger')
            return redirect(url_for('main.add_transaction', party_id=party_id))

        if not amount or amount <= 0:
            flash('يجب إدخال مبلغ صحيح', 'danger')
            return redirect(url_for('main.add_transaction', party_id=party_id))

        if transaction_date:
            try:
                trans_date = datetime.strptime(transaction_date, '%Y-%m-%d')
            except ValueError:
                trans_date = datetime.utcnow()
        else:
            trans_date = datetime.utcnow()

        new_transaction = Transaction(
            party_id=party_id,
            transaction_type=transaction_type,
            amount=amount,
            description=description if description else None,
            transaction_date=trans_date,
            created_by=session.get('employee_id')
        )
        db.session.add(new_transaction)
        db.session.commit()

        log_activity('add', 'transaction', new_transaction.id, f"{party.name} - {transaction_type}")
        flash('تم إضافة المعاملة بنجاح', 'success')
        return redirect(url_for('main.party_profile', party_id=party_id))

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('add_transaction.html', party=party, today=today)

@main.route('/transactions/transaction/<int:transaction_id>/delete', methods=['POST'])
@login_required
@require_permissions('can_delete_transactions')
def delete_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    party_id = transaction.party_id
    party_name = transaction.party.name

    if getattr(transaction, 'payments', None):
        if len(transaction.payments) > 0:
            flash('لا يمكن حذف المعاملة لأنها تحتوي على سجلات سداد. احذف سجلات السداد أولاً.', 'danger')
            return redirect(url_for('main.party_profile', party_id=party_id))

    try:
        db.session.delete(transaction)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('حدث خطأ أثناء حذف المعاملة. تأكد من عدم وجود سجلات مرتبطة أو تواصل مع الدعم.', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))

    log_activity('delete', 'transaction', transaction_id, f"{party_name} - {transaction.transaction_type}")
    flash('تم حذف المعاملة بنجاح', 'success')
    return redirect(url_for('main.party_profile', party_id=party_id))

@main.route('/transactions/payment/<int:payment_id>/delete', methods=['POST'])
@login_required
@require_permissions('can_delete_transactions')
def delete_payment(payment_id):
    from .models import TransactionPayment
    payment = TransactionPayment.query.get_or_404(payment_id)
    tx = payment.transaction
    party_id = tx.party_id if tx else None

    try:
        db.session.delete(payment)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('حدث خطأ أثناء حذف السداد. تواصل مع الدعم.', 'danger')
        return redirect(url_for('main.party_profile', party_id=party_id))

    log_activity('delete', 'transaction_payment', payment_id, f"payment for tx {payment.transaction_id}")
    flash('تم حذف السداد بنجاح', 'success')
    return redirect(url_for('main.party_profile', party_id=party_id))

@main.route('/customers')
@login_required
@permission_required('can_view_customers')
def customers():
    search_query = request.args.get('search', '').strip()
    page = 1
    per_page = 50

    base_query = _build_customers_query(search_query)
    rows = (base_query
            .order_by(Customer.id.desc())
            .limit(per_page + 1)
            .offset((page - 1) * per_page)
            .all())

    has_more = len(rows) > per_page
    customers_page = rows[:per_page]
    total_customers = _get_customers_total_count(base_query, search_query)
    next_page = page + 1 if has_more else None

    return render_template(
        'customers.html',
        customers=customers_page,
        search_query=search_query,
        total_customers=total_customers,
        has_more=has_more,
        next_page=next_page,
        per_page=per_page,
        page=page)

@main.route('/customers/more')
@login_required
@permission_required('can_view_customers')
def customers_more():
    try:
        page = int(request.args.get('page', '2'))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    per_page = 50
    search_query = request.args.get('search', '').strip()

    base_query = _build_customers_query(search_query)
    rows = (base_query
            .order_by(Customer.id.desc())
            .limit(per_page + 1)
            .offset((page - 1) * per_page)
            .all())

    has_more = len(rows) > per_page
    customers_page = rows[:per_page]
    next_page = page + 1 if has_more else None

    fragment_html = render_template('partials/_customer_cards.html', customers=customers_page)
    return jsonify({
        'html': fragment_html,
        'has_more': has_more,
        'next_page': next_page,
        'loaded_count': len(customers_page)
    })

@main.route('/customers/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_customers')
@log_operation('create', 'customer')
def add_customer():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        governorate = request.form['governorate']
        address_details = request.form['address_details']
        new_customer = Customer(
            name=name,
            phone=phone,
            governorate=governorate,
            address_details=address_details)
        db.session.add(new_customer)
        db.session.flush()

        log_activity('create', 'customer', new_customer.id,
                     name, f"إضافة عميل جديد: {name} - {phone}")

        db.session.commit()
        flash(f"تم إضافة العميل {name} بنجاح")
        return redirect(url_for('main.customers'))
    return render_template('add_customer.html')

@main.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_customers')
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    if request.method == 'POST':
        old_name = customer.name
        old_phone = customer.phone

        customer.name = request.form['name']
        customer.phone = request.form['phone']
        customer.governorate = request.form['governorate']
        customer.address_details = request.form['address_details']

        db.session.commit()

        details = f"تعديل بيانات العميل: {old_name} -> {customer.name}, {old_phone} -> {customer.phone}"
        log_activity('update', 'customer', customer_id, customer.name, details)

        flash("تم تعديل بيانات العميل بنجاح")
        return redirect(url_for('main.customers'))
    return render_template('edit_customer.html', customer=customer)

@main.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_customers')
def delete_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    customer_name = customer.name

    db.session.delete(customer)
    db.session.commit()

    log_activity(
        'delete',
        'customer',
        customer_id,
        customer_name,
        f"حذف العميل: {customer_name}")

    flash("تم حذف العميل بنجاح")
    return redirect(url_for('main.customers'))

@main.route('/products')
@login_required
@permission_required('can_view_products')
def products():
    search_query = request.args.get('search', '')
    view = request.args.get('view', 'active')

    q = Product.query.options(joinedload(Product.variants))
    if view == 'deleted':
        q = q.filter(Product.is_deleted == True)
    else:
        q = q.filter(Product.is_deleted == False)

    if search_query:
        q = q.filter(Product.name.ilike(f"%{search_query}%"))
    q = q.order_by(Product.stock.asc(), Product.name.asc())

    all_products = q.all()

    return render_template(
        'products.html',
        products=all_products,
        search_query=search_query,
        view=view)

@main.route('/products/inventory_pdf')
@login_required
@permission_required('can_view_products')
def inventory_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        flash('مكتبة إنشاء ملفات PDF غير مثبتة. الرجاء تثبيت reportlab', 'danger')
        return redirect(url_for('main.products'))

    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
    except Exception:
        flash('الـ python packages المطلوبة للغة العربية غير مثبتة. الرجاء تثبيت arabic-reshaper و python-bidi', 'danger')
        return redirect(url_for('main.products'))

    font_candidates = [
        '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
        '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
        '/usr/share/fonts/truetype/amiri/Amiri-Regular.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
    ]

    font_path = None
    for p in font_candidates:
        try:
            if os.path.exists(p):
                font_path = p
                break
        except Exception:
            continue

    if not font_path:
        flash('لم أجد خطًا يدعم العربية على السيرفر. ثبّت أحد الخطوط مثل: fonts-dejavu-core أو fonts-amiri أو google-noto-fonts-arabic ثم أعد المحاولة.', 'danger')
        return redirect(url_for('main.products'))

    try:
        pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
    except Exception:
        flash('فشل تسجيل الخط على السيرفر. الرجاء التأكد من أذونات الملف ومسار الخط.', 'danger')
        return redirect(url_for('main.products'))

    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    today_str = datetime.now().strftime('%d/%m/%Y')
    title = f'كشف جرد المنتجات بتاريخ {today_str}'
    try:
        reshaped_title = arabic_reshaper.reshape(title)
        display_title = get_display(reshaped_title)
    except Exception:
        display_title = title
    p.setFont('ArabicFont', 16)
    p.drawCentredString(width / 2.0, height - 2 * cm, display_title)

    y = height - 3 * cm
    left_x = 2 * cm
    right_x = width - 2 * cm
    mid_x = (left_x + right_x) / 2.0
    p.setFont('ArabicFont', 12)
    header_name = get_display(arabic_reshaper.reshape('اسم المنتج'))
    header_actual = get_display(arabic_reshaper.reshape('الكمية الفعلية'))
    header_qty = get_display(arabic_reshaper.reshape('الكمية'))
    p.drawString(left_x, y, header_name)
    p.drawCentredString(mid_x, y, header_actual)
    p.drawRightString(right_x, y, header_qty)
    y -= 0.8 * cm

    p.setFont('ArabicFont', 11)

    products = Product.query.filter(Product.is_deleted == False, Product.is_bundle == False).order_by(Product.name.asc()).all()

    for prod in products:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont('ArabicFont', 12)
            p.drawString(left_x, y, header_name)
            p.drawCentredString(mid_x, y, header_actual)
            p.drawRightString(right_x, y, header_qty)
            y -= 0.8 * cm
            p.setFont('ArabicFont', 11)

        name = (prod.name or '')
        qty = str(prod.stock or 0)

        try:
            reshaped = arabic_reshaper.reshape(name)
            display_name = get_display(reshaped)
        except Exception:
            display_name = name

        max_chars = 120
        if len(display_name) > max_chars:
            display_name = display_name[:max_chars - 3] + '...'

        p.drawString(left_x, y, display_name)
        p.drawCentredString(mid_x, y, '')
        p.drawRightString(right_x, y, qty)

        try:
            p.setStrokeColorRGB(0, 0, 0)
            p.setLineWidth(0.35)
            p.line(left_x, y - 0.18 * cm, right_x, y - 0.18 * cm)
        except Exception:
            pass

        y -= 0.6 * cm

    p.save()
    buf.seek(0)
    data = buf.read()

    resp = make_response(data)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename="inventory.pdf"'
    return resp

@main.route('/products/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_products')
def add_product():
    if request.method == 'POST':
        try:
            product_type = request.form.get('product_type', 'normal')
            
            if product_type == 'bundle':
                bundle_name = request.form.get('bundle_name', '').strip()
                if not bundle_name:
                    flash('يجب إدخال اسم البندل', 'danger')
                    return redirect(url_for('main.add_product'))
                
                existing_product = Product.query.filter(
                    Product.name.ilike(bundle_name),
                    Product.is_deleted == False
                ).first()
                if existing_product:
                    flash(f'المنتج "{bundle_name}" موجود بالفعل في النظام', 'warning')
                    return redirect(url_for('main.add_product'))
                
                bundle_product_ids = request.form.getlist('bundle_product_ids[]')
                bundle_sale_prices = request.form.getlist('bundle_sale_prices[]')
                
                if not bundle_product_ids or len(bundle_product_ids) == 0:
                    flash('يجب إضافة منتج واحد على الأقل إلى البندل', 'danger')
                    return redirect(url_for('main.add_product'))
                
                total_bundle_price = 0
                total_purchase_price = 0
                
                bundle_items_data = []
                for i in range(len(bundle_product_ids)):
                    product_id = int(bundle_product_ids[i])
                    sale_price = float(bundle_sale_prices[i])
                    
                    product = Product.query.get(product_id)
                    if not product:
                        continue
                    
                    bundle_items_data.append({
                        'product_id': product_id,
                        'sale_price': sale_price,
                        'purchase_price': product.purchase_price
                    })
                    
                    total_bundle_price += sale_price
                    total_purchase_price += product.purchase_price
                
                new_bundle = Product(
                    name=bundle_name,
                    price=total_bundle_price,
                    purchase_price=total_purchase_price,
                    stock=0,
                    is_bundle=True
                )
                db.session.add(new_bundle)
                db.session.flush()
                
                for item_data in bundle_items_data:
                    bundle_item = BundleItem(
                        bundle_id=new_bundle.id,
                        product_id=item_data['product_id'],
                        sale_price_in_bundle=item_data['sale_price']
                    )
                    db.session.add(bundle_item)
                
                db.session.commit()
                flash(f"تم إضافة البندل '{bundle_name}' بنجاح", "success")
                return redirect(url_for('main.products'))
            
            else:
                name = request.form['name'].strip()
                purchase_price = float(request.form['purchase_price'])
                price = float(request.form['price'])
                stock = int(request.form['stock'])

                existing_product = Product.query.filter(
                    Product.name.ilike(name),
                    Product.is_deleted == False
                ).first()
                if existing_product:
                    flash(f'المنتج "{name}" موجود بالفعل في النظام', 'warning')
                    return redirect(url_for('main.add_product'))

                new_product = Product(name=name, purchase_price=purchase_price, price=price, stock=stock, is_bundle=False)
                db.session.add(new_product)
                db.session.flush()

                variant_group_names = request.form.getlist('variant_group_name[]')
                variant_names = request.form.getlist('variant_name[]')

                if variant_group_names and variant_names:
                    count = min(len(variant_names), len(variant_group_names))
                    for i in range(count):
                        gname = (variant_group_names[i] or '').strip()
                        vname = (variant_names[i] or '').strip()
                        if not gname or not vname:
                            continue
                        db.session.add(ProductVariant(
                            product_id=new_product.id,
                            group_name=gname,
                            variant_name=vname,
                            price=0
                        ))

                db.session.commit()
                flash("تم إضافة المنتج بنجاح", "success")
                return redirect(url_for('main.products'))
                
        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ أثناء إضافة المنتج: {e}", "danger")
            return redirect(url_for('main.add_product'))

    return render_template('add_product.html')

@main.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_products')
def edit_product(product_id):
    product = Product.query.options(
        joinedload(Product.variants),
        joinedload(Product.bundle_items).joinedload(BundleItem.product)
    ).get_or_404(product_id)

    available_products = Product.query.filter_by(is_deleted=False, is_bundle=False).order_by(Product.name).all()
    available_products_json = json.dumps([
        {'id': p.id, 'name': p.name, 'purchase_price': float(p.purchase_price or 0)}
        for p in available_products
    ])

    if request.method == 'POST':
        try:
            old_name = product.name
            product.name = (request.form.get('name') or '').strip()
            
            if product.is_bundle:
                bundle_item_ids = request.form.getlist('bundle_item_ids[]')
                bundle_product_ids = request.form.getlist('bundle_product_ids[]')
                bundle_sale_prices = request.form.getlist('bundle_sale_prices[]')
                
                for old_item in list(product.bundle_items):
                    db.session.delete(old_item)
                db.session.flush()
                
                new_total_price = 0
                new_total_purchase = 0
                for idx in range(len(bundle_product_ids)):
                    if idx < len(bundle_sale_prices):
                        try:
                            product_id_in_bundle = int(bundle_product_ids[idx])
                            sale_price = float(bundle_sale_prices[idx])
                            
                            bp = Product.query.get(product_id_in_bundle)
                            new_bundle_item = BundleItem(
                                bundle_id=product.id,
                                product_id=product_id_in_bundle,
                                sale_price_in_bundle=sale_price
                            )
                            db.session.add(new_bundle_item)
                            new_total_price += sale_price
                            if bp:
                                new_total_purchase += (bp.purchase_price or 0)
                        except (ValueError, TypeError):
                            continue
                
                product.price = new_total_price
                product.purchase_price = new_total_purchase
                
                update_new_orders_with_bundle_changes(product.id)
                
            else:
                product.purchase_price = float(request.form.get('purchase_price') or 0)
                product.price = float(request.form.get('price') or 0)
                product.stock = int(request.form.get('stock') or 0)

            group_names = request.form.getlist('variant_group_name[]')
            var_names = request.form.getlist('variant_name[]')
            var_prices = request.form.getlist('variant_price[]')

            for v in list(product.variants):
                db.session.delete(v)

            count = min(len(group_names), len(var_names), len(var_prices))
            for i in range(count):
                g = (group_names[i] or '').strip()
                n = (var_names[i] or '').strip()
                if not g or not n:
                    continue
                try:
                    p = float(var_prices[i]) if var_prices[i] not in (None, '',) else 0.0
                except Exception:
                    p = 0.0
                db.session.add(ProductVariant(
                    product_id=product.id,
                    group_name=g,
                    variant_name=n,
                    price=p,
                    stock=0
                ))

            product.update_variant_types()

            db.session.commit()

            try:
                details = f"تعديل منتج: {old_name} -> {product.name}, سعر شراء {product.purchase_price}, سعر بيع {product.price}, كمية {product.stock}"
                log_activity('update', 'product', product.id, product.name, details)
            except Exception:
                pass

            flash('تم حفظ تعديلات المنتج بنجاح', 'success')
            return redirect(url_for('main.products'))
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تعديل المنتج: {e}', 'danger')
            return redirect(url_for('main.edit_product', product_id=product.id))

    grouped = {}
    for v in (product.variants or []):
        grouped.setdefault(v.group_name, []).append(v)

    return render_template('edit_product.html', 
                         product=product, 
                         grouped_variants=grouped,
                         available_products=available_products,
                         available_products_json=available_products_json)

def update_new_orders_with_bundle_changes(bundle_id):
    try:
        bundle = Product.query.options(
            joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
        ).get(bundle_id)
        
        if not bundle or not bundle.is_bundle:
            return
        
        new_orders = Order.query.filter(
            Order.status == 'جديد',
            Order.items.any(OrderItem.product_id == bundle_id)
        ).all()
        
        for order in new_orders:
            for item in order.items:
                if item.product_id == bundle_id:
                    bundle_variants = {}
                    for bundle_item in bundle.bundle_items:
                        if bundle_item.product:
                            product_variants = []
                            for variant in bundle_item.product.variants:
                                product_variants.append({
                                    'id': variant.id,
                                    'group_name': variant.group_name,
                                    'variant_name': variant.variant_name,
                                    'price': float(variant.price or 0)
                                })
                            bundle_variants[str(bundle_item.product.id)] = {
                                'product_id': bundle_item.product.id,
                                'product_name': bundle_item.product.name,
                                'product_variants': product_variants
                            }
                    
                    item.bundle_variants_json = json.dumps(bundle_variants) if bundle_variants else None
        
        db.session.commit()
        current_app.logger.info(f"Updated {len(new_orders)} new orders with bundle changes for bundle_id={bundle_id}")
    except Exception as e:
        current_app.logger.error(f"Error updating new orders with bundle changes: {e}", exc_info=True)
        db.session.rollback()

@main.route('/products/<int:product_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_products')
def delete_product(product_id):
    try:
        product = Product.query.get_or_404(product_id)
        product_name = product.name
        
        product.is_deleted = True
        db.session.commit()

        try:
            log_activity(
                'delete',
                'product',
                product_id,
                product_name,
                f"حذف منتج: {product_name}")
        except Exception as log_err:
            current_app.logger.warning(f"Failed to log activity: {log_err}")

        flash("تم حذف المنتج بنجاح", "success")
        current_app.logger.info(f"DELETE PRODUCT SUCCESS: id={product_id} name={product_name}")
        
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف المنتج: {str(e)}", "danger")
        current_app.logger.error(f"DELETE PRODUCT ERROR: {e}", exc_info=True)
    
    return redirect(url_for('main.products'))

@main.route('/products/delete_all', methods=['POST'])
@login_required
@permission_required('can_delete_all_products')
def delete_all_products():
    Product.query.update({Product.is_deleted: True})
    db.session.commit()
    flash("تم حذف كل المنتجات بنجاح")
    return redirect(url_for('main.products'))

@main.route('/orders')
@login_required
@permission_required('can_view_orders')
def orders():
    from datetime import date
    from sqlalchemy import func, extract

    month = request.args.get('month', type=int)
    status = request.args.get('status', type=str)
    phone = request.args.get('phone', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = min(max(per_page, 10), 200)

    today = date.today()
    if not month:
        month = today.month

    cache_key = f"orders:{today.year}:{month}:{status or 'ALL'}:{phone or 'NO'}:{page}:{per_page}"
    orders_cache = current_app.config.get('_ORDERS_CACHE', {})
    cached_entry = orders_cache.get(cache_key)
    if cached_entry and (datetime.utcnow() - cached_entry['ts']).total_seconds() < 20:
        payload = cached_entry['payload']
        return render_template('orders.html', **payload)

    query = Order.query.options(joinedload(Order.customer)) 
    query = query.filter(extract('month', Order.date) == month, extract('year', Order.date) == today.year)
    if status:
        query = query.filter(Order.status == status)
    if phone:
        query = query.join(Order.customer).filter(Customer.phone.ilike(f"%{phone}%"))

    status_counts_rows = (db.session.query(Order.status, func.count(Order.id))
                          .filter(extract('month', Order.date) == month, extract('year', Order.date) == today.year)
                          .group_by(Order.status)
                          .all())
    status_counts = {row[0]: row[1] for row in status_counts_rows}
    current_month_orders = sum(status_counts.values())

    try:
        from .models import MonthlyGoal
        monthly_goal = MonthlyGoal.get_current_goal()
    except Exception:
        monthly_goal = 100

    pagination = query.order_by(Order.is_urgent.desc(), Order.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    orders_page_items = pagination.items

    try:
        def _sort_key(o):
            is_urg = bool(getattr(o, 'is_urgent', False))
            dt = getattr(o, 'date', None)
            ts = dt.timestamp() if dt is not None else 0
            return (0 if is_urg else 1, -ts)
        orders_page_items = sorted(orders_page_items, key=_sort_key)
    except Exception:
        pass

    payload = dict(
        orders=orders_page_items,
        selected_month=month,
        selected_status=status,
        daily_status_counts=status_counts,
        monthly_goal=monthly_goal,
        current_month_orders=current_month_orders,
        page=page,
        per_page=per_page,
        total_pages=pagination.pages,
        total_orders=pagination.total,
        has_next=pagination.has_next,
        has_prev=pagination.has_prev
    )
    orders_cache[cache_key] = {'payload': payload, 'ts': datetime.utcnow()}
    current_app.config['_ORDERS_CACHE'] = orders_cache
    return render_template('orders.html', **payload)

@main.route('/orders/search-by-tracking', methods=['POST'])
@login_required
@permission_required('can_view_orders')
def search_orders_by_tracking():
    try:
        data = request.get_json(force=True, silent=True) or {}
        q = (data.get('tracking_number') or '').strip()
        if not q:
            return jsonify(success=False, error='empty_query')
        matches = (Order.query
                   .options(joinedload(Order.customer), joinedload(Order.employee))
                   .outerjoin(Customer, Order.customer_id == Customer.id)
                   .filter(db.or_(
                       Order.tracking_number.ilike(f"%{q}%"),
                       Order._customer_phone.ilike(f"%{q}%"),
                       Customer.phone.ilike(f"%{q}%")
                   ))
                   .order_by(Order.date.desc())
                   .limit(50)
                   .all())

        out = []
        for o in matches:
            out.append({
                'id': o.id,
                'tracking_number': o.tracking_number,
                'customer_name': o.customer.name if o.customer else None,
                'customer_phone': o.customer.phone if o.customer else (o.customer_phone or None),
                'employee_name': o.employee.username if o.employee else None,
                'date': o.date.strftime('%Y-%m-%d %H:%M') if getattr(o, 'date', None) else None,
                'status': o.status,
                'total_amount': float(o.total_amount or 0),
                'delivery_fees': float(o.delivery_fees or 0),
                'amount_paid': float(o.amount_paid or 0),
                'remaining_amount': float(o.remaining_amount or 0),
            })

        return jsonify(success=True, orders=out)
    except Exception as e:
        current_app.logger.exception('search_orders_by_tracking failed')
        return jsonify(success=False, error=str(e))

@main.route('/replacement-orders/search-by-tracking', methods=['POST'])
@login_required
@permission_required('can_view_replacements')
def search_replacement_orders_by_tracking():
    try:
        data = request.get_json(force=True, silent=True) or {}
        q = (data.get('tracking_number') or '').strip()
        if not q:
            return jsonify(success=False, error='empty_query')

        matches = (ReplacementOrder.query
                   .options(joinedload(ReplacementOrder.customer), joinedload(ReplacementOrder.registered_by))
                   .filter(db.or_(
                       ReplacementOrder.tracking_number.ilike(f"%{q}%"),
                       ReplacementOrder.alternative_phone.ilike(f"%{q}%"),
                       Customer.phone.ilike(f"%{q}%")
                   ))
                   .join(Customer, ReplacementOrder.customer_id == Customer.id, isouter=True)
                   .order_by(ReplacementOrder.date.desc())
                   .limit(50)
                   .all())

        out = []
        for o in matches:
            total_amount = float(getattr(o, 'total_amount', 0) or 0)
            amount_paid = float(getattr(o, 'amount_paid', 0) or 0)
            customer_refund_amount = float(getattr(o, 'customer_refund_amount', 0) or 0)
            remaining_amount = total_amount - amount_paid - customer_refund_amount
            out.append({
                'id': o.id,
                'tracking_number': o.tracking_number,
                'customer_name': o.customer_name,
                'customer_phone': o.customer_phone,
                'employee_name': o.registered_by.username if o.registered_by else None,
                'date': o.date.strftime('%Y-%m-%d %H:%M') if getattr(o, 'date', None) else None,
                'status': o.status,
                'total_amount': total_amount,
                'amount_paid': amount_paid,
                'remaining_amount': remaining_amount,
            })

        return jsonify(success=True, orders=out)
    except Exception as e:
        current_app.logger.exception('search_replacement_orders_by_tracking failed')
        return jsonify(success=False, error=str(e))

@main.route('/orders/update_goal', methods=['POST'])
@login_required
@permission_required('can_view_orders')
def update_monthly_goal():
    try:
        goal = request.form.get('target', 0)
        
        if not goal:
            return jsonify({'success': False, 'error': 'لم يتم إرسال قيمة الهدف'})
        
        goal = int(goal)
        if goal < 0:
            return jsonify({'success': False, 'error': 'الهدف يجب أن يكون رقم موجب'})
        
        from .models import MonthlyGoal
        MonthlyGoal.set_current_goal(goal)
        
        session['monthly_goal'] = goal
        
        current_app.logger.info(f"✅ Monthly goal updated to: {goal}")
        
        return jsonify({
            'success': True, 
            'message': f'تم تحديث الهدف إلى {goal}',
            'goal': goal
        })
        
    except Exception as e:
        current_app.logger.error(f"❌ Error updating goal: {e}")
        return jsonify({'success': False, 'error': 'خطأ في تحديث الهدف'})

@main.route('/orders/monthly_total')
@login_required
@permission_required('can_view_orders')
def get_monthly_orders_total():
    try:
        from datetime import date
        from sqlalchemy import extract, func
        
        today = date.today()
        month = today.month
        year = today.year
        
        status_counts = dict(
            db.session.query(Order.status, func.count(Order.id))
            .filter(extract('month', Order.date) == month)
            .filter(extract('year', Order.date) == year)
            .group_by(Order.status)
            .all()
        )
        
        total_orders = sum(status_counts.values()) if status_counts else 0
        
        current_app.logger.debug(f"status_counts={status_counts} total_orders={total_orders}")
        
        try:
            total_amount = db.session.query(func.sum(Order.total_amount))\
                .filter(extract('month', Order.date) == month)\
                .filter(extract('year', Order.date) == year)\
                .scalar() or 0
        except Exception as e:
            current_app.logger.warning(f"Error calculating total amounts: {e}")
            total_amount = 0
        
        current_app.logger.debug(f"Monthly totals m={month} y={year} orders={total_orders} amount={total_amount}")
        
        from .models import MonthlyGoal
        monthly_goal = MonthlyGoal.get_current_goal()
        
        return jsonify({
            'success': True,
            'total': total_orders,
            'total_amount': total_amount,
            'month': month,
            'year': year,
            'monthly_goal': monthly_goal,
            'remaining': monthly_goal - total_orders
        })
        
    except Exception as e:
        current_app.logger.error(f"Error calculating monthly orders total: {e}")
        return jsonify({'success': False, 'error': 'خطأ في حساب إجمالي طلبات الشهر'})

@main.route('/orders/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_orders')
def add_order():
    customers = Customer.query.all()
    employees = Employee.query.filter(
        Employee.is_active.is_(True),
        Employee.sales_commission_percentage > 0
    ).all()

    referenced_product_ids = set()

    base_products = Product.query.options(
        joinedload(Product.variants),
        joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
    ).filter_by(is_deleted=False).all()
    base_ids = {p.id for p in base_products}
    extra_products = []
    extra_ids = referenced_product_ids - base_ids
    if extra_ids:
        extra_products = Product.query.options(
            joinedload(Product.variants),
            joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
        ).filter(Product.id.in_(list(extra_ids))).all()
    products = base_products + [p for p in extra_products if p.id not in base_ids]

    def serialize_product(product):
        result = {
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'stock': product.stock,
            'wholesale_price': product.wholesale_price,
            'is_bundle': product.is_bundle if hasattr(product, 'is_bundle') else False,
            'variants': [
                {
                    'id': v.id,
                    'group_name': v.group_name,
                    'variant_name': v.variant_name,
                    'price': v.price
                }
                for v in product.variants
            ]
        }
        
        if result['is_bundle'] and hasattr(product, 'bundle_items'):
            result['bundle_items'] = []
            for bundle_item in product.bundle_items:
                if bundle_item.product:
                    result['bundle_items'].append({
                        'product_id': bundle_item.product.id,
                        'product_name': bundle_item.product.name,
                        'sale_price_in_bundle': bundle_item.sale_price_in_bundle,
                        'product_variants': [
                            {
                                'id': v.id,
                                'group_name': v.group_name,
                                'variant_name': v.variant_name,
                                'price': v.price
                            }
                            for v in bundle_item.product.variants
                        ] if bundle_item.product.variants else []
                    })
        
        return result

    products_data = [serialize_product(p) for p in products]

    if request.method == 'POST':
        current_app.logger.debug(f"add_order POST form_keys={list(request.form.keys())}")
        
        print("\n=== DEBUG: request.form keys (first 10) ===")
        form_items = list(request.form.items())[:10]
        for key, value in form_items:
            print(f"  {key} = {value}")
        print("=== END DEBUG ===\n")
        
        try:
            submission_id = request.form.get('submission_id')
            if submission_id:
                processed = session.get('processed_submissions', {})
                if submission_id in processed:
                    flash('تم استلام هذا الطلب مسبقًا. تم منع التكرار.', 'info')
                    return redirect(url_for('main.orders'))
            order_type = request.form.get('order_type', ORDER_TYPE_DELIVERY)
            if order_type not in (ORDER_TYPE_DELIVERY, ORDER_TYPE_WALKIN):
                flash('نوع الطلب غير صالح')
                return redirect(url_for('main.add_order'))

            employee_id = request.form.get('employee_id')
            if not employee_id or not employee_id.strip():
                flash('لا يمكن حفظ الطلب بدون تحديد الموظف المسؤول', 'danger')
                return redirect(url_for('main.add_order'))
            employee_id = None if employee_id == '0' else int(employee_id)

            customer = None

            if order_type == ORDER_TYPE_WALKIN:
                # ------------------------------------------------------------
                # Walk-in (floor customer) path
                # No customer data needed. order is 'وصل' at creation with
                # instant inventory deduction and full payment.
                # NOTE: 'وصل' here means "order completed at creation", not
                # "physically delivered" as in the delivery flow. This is a
                # semantic override of the same status field — do NOT build
                # SLA/fulfillment-time metrics on this field without accounting
                # for the mixed semantics.
                # ------------------------------------------------------------
                status = "وصل"
                delivery_fees = 0
                notes = request.form.get('notes', "")
                amount_paid = 0  # will be set to total after calculation
                tracking_number = None
                is_nearest_post_branch = False

                order = Order(
                    customer_id=None,
                    employee_id=employee_id,
                    status=status,
                    order_type=ORDER_TYPE_WALKIN,
                    delivery_fees=delivery_fees,
                    notes=notes,
                    amount_paid=amount_paid,
                    tracking_number=tracking_number,
                    is_nearest_post_branch=is_nearest_post_branch,
                    customer_name='زبون أرضية',
                    customer_phone=None,
                    customer_governorate=None,
                    customer_address_details=None,
                    weight=request.form.get('weight', type=float),
                    package_volume=request.form.get('package_volume') or None,
                    delivery_notes=request.form.get('delivery_notes') or None,
                )

            else:
                # ------------------------------------------------------------
                # Delivery path — existing logic unchanged
                # ------------------------------------------------------------
                customer_name = request.form.get('customer_name')
                customer_governorate = request.form.get('customer_governorate')
                customer_address_details = request.form.get(
                    'customer_address_details')
                customer_phone = request.form.get('customer_phone')
                
                current_app.logger.debug(f"Customer new order name={customer_name} phone={customer_phone}")

                if not (
                        customer_name and customer_governorate and customer_address_details and customer_phone):
                    flash('يجب إدخال جميع بيانات العميل')
                    return redirect(url_for('main.add_order'))

                customer = Customer.query.filter_by(phone=customer_phone).first()
                if not customer:
                    customer = Customer(
                        name=customer_name,
                        phone=customer_phone,
                        governorate=customer_governorate,
                        address_details=customer_address_details)
                    db.session.add(customer)
                    db.session.flush()
                else:
                    if customer_name and customer_name != customer.name:
                        customer.name = customer_name
                    if customer_governorate and customer_governorate != customer.governorate:
                        customer.governorate = customer_governorate
                    if customer_address_details and customer_address_details != customer.address_details:
                        customer.address_details = customer_address_details
                    db.session.flush()

                status = "جديد"
                delivery_fees = float(request.form.get('delivery_fees', 0))
                notes = request.form.get('notes', "")
                amount_paid = float(request.form.get('amount_paid', 0))
                tracking_number = request.form.get('tracking_number', "").strip() or None

                original_status = status
                if tracking_number:
                    status = "خرج للتوصيل"

                is_nearest_post_branch = request.form.get('is_nearest_post_branch') == '1'

                order = Order(
                    customer_id=customer.id,
                    employee_id=employee_id,
                    status=status,
                    order_type=ORDER_TYPE_DELIVERY,
                    delivery_fees=delivery_fees,
                    notes=notes,
                    amount_paid=amount_paid,
                    tracking_number=tracking_number,
                    is_nearest_post_branch=is_nearest_post_branch,
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    customer_governorate=customer_governorate,
                    customer_address_details=customer_address_details,
                    weight=request.form.get('weight', type=float),
                    package_volume=request.form.get('package_volume') or None,
                    delivery_notes=request.form.get('delivery_notes') or None,
                )
            db.session.add(order)
            db.session.flush()
            import time
            t0 = time.perf_counter()
            total = 0

            # قراءة بيانات المنتجات من order_items_data JSON
            items_data_json = request.form.get('order_items_data')
            items_data = []
            if items_data_json:
                try:
                    items_data = json.loads(items_data_json)
                except Exception as e:
                    current_app.logger.error(f"Failed to parse order_items_data: {e}")
                    items_data = []

            # جمع معرفات المنتجات والـ variants
            product_ids = []
            variant_ids = set()
            
            for row in items_data:
                product_id = row.get('product_id')
                if product_id:
                    try:
                        product_ids.append(int(product_id))
                    except (TypeError, ValueError):
                        pass
                
                for key in ('size_variant_id', 'color_variant_id', 'style_variant_id', 'variant_id'):
                    vid = row.get(key)
                    if vid:
                        try:
                            variant_ids.add(int(vid))
                        except (TypeError, ValueError):
                            pass

            products_map = {}
            if product_ids:
                products_map = _get_locked_products(product_ids)

            variants_map = {}
            if variant_ids:
                for v in ProductVariant.query.filter(ProductVariant.id.in_(list(variant_ids))).all():
                    variants_map[v.id] = v

            order_items_to_add = []
            product_required_qty = defaultdict(int)

            for row in items_data:
                product_id = row.get('product_id')
                if not product_id:
                    continue
                
                try:
                    product_id = int(product_id)
                except (TypeError, ValueError):
                    continue

                product = products_map.get(product_id)
                if not product:
                    continue

                try:
                    quantity = int(row.get('quantity') or 1)
                except (TypeError, ValueError):
                    quantity = 1

                try:
                    price = float(row.get('price') or 0)
                except (TypeError, ValueError):
                    price = 0

                # التحقق من أن السعر لا يقل عن سعر الشراء
                min_price = get_min_price(product)
                if price < min_price:
                    flash(f'سعر المنتج {product.name} أقل من سعر الشراء', 'danger')
                    db.session.rollback()
                    return redirect(url_for('main.add_order'))

                # حساب الكميات المطلوبة للمخزون
                if product.is_bundle:
                    bundle_items = BundleItem.query.filter_by(bundle_id=product_id).all()
                    for bundle_item in bundle_items:
                        if bundle_item.product:
                            product_required_qty[bundle_item.product_id] += quantity
                else:
                    product_required_qty[product_id] += quantity

                # إنشاء OrderItem
                size_variant_id = row.get('size_variant_id')
                color_variant_id = row.get('color_variant_id')
                style_variant_id = row.get('style_variant_id')
                variant_id = row.get('variant_id')

                try:
                    size_variant_id = int(size_variant_id) if size_variant_id else None
                except (TypeError, ValueError):
                    size_variant_id = None
                try:
                    color_variant_id = int(color_variant_id) if color_variant_id else None
                except (TypeError, ValueError):
                    color_variant_id = None
                try:
                    style_variant_id = int(style_variant_id) if style_variant_id else None
                except (TypeError, ValueError):
                    style_variant_id = None
                try:
                    variant_id = int(variant_id) if variant_id else None
                except (TypeError, ValueError):
                    variant_id = None

                order_item = OrderItem(
                    order_id=order.id,
                    product_id=product_id,
                    variant_id=variant_id,
                    size_variant_id=size_variant_id,
                    color_variant_id=color_variant_id,
                    style_variant_id=style_variant_id,
                    quantity=quantity,
                    price=price,
                    purchase_price_snapshot=float(product.purchase_price or 0)
                )

                # حفظ bundle_variants إذا كانت موجودة
                bundle_variants = row.get('bundle_variants')
                if bundle_variants:
                    try:
                        order_item.bundle_variants_json = json.dumps(bundle_variants)
                    except Exception:
                        order_item.bundle_variants_json = None

                total += price * quantity
                order_items_to_add.append(order_item)

            db.session.add_all(order_items_to_add)

            locked_products = _get_locked_products(product_required_qty.keys())

            insufficient_products = []
            for pid, req_qty in product_required_qty.items():
                p = locked_products.get(pid)
                if not p:
                    continue
                if (p.stock or 0) < req_qty:
                    insufficient_products.append((p.name, req_qty, p.stock or 0))

            if insufficient_products and order_type == ORDER_TYPE_WALKIN:
                msg_lines = [f"{name}: المطلوب {req} المتوفر {avail}" for name, req, avail in insufficient_products]
                flash('لا يمكن إنشاء طلب أرضية لعدم كفاية المخزون:\n' + '\n'.join(msg_lines), 'danger')
                db.session.rollback()
                return redirect(url_for('main.add_order'))
            if insufficient_products and status == 'خرج للتوصيل':
                msg_lines = [f"{name}: المطلوب {req} المتوفر {avail}" for name, req, avail in insufficient_products]
                flash('لا يمكن إنشاء الطلب كـ "خرج للتوصيل" لعدم كفاية المخزون:\n' + '\n'.join(msg_lines), 'danger')
                db.session.rollback()
                return redirect(url_for('main.add_order'))
            elif insufficient_products and status == 'جديد':
                if len(insufficient_products) == 1:
                    name, req, avail = insufficient_products[0]
                    flash(f'تنبيه: المنتج {name} غير متوفر حالياً (المطلوب {req} / المتوفر {avail}). تم حفظ الطلب كـ "جديد" ويمكن تعديله لاحقاً.', 'warning')
                else:
                    msg_lines = ['المنتجات التالية غير متوفرة حالياً:'] + [f'• {name} (المطلوب {req} / المتوفر {avail})' for name, req, avail in insufficient_products]
                    flash('\n'.join(msg_lines) + '\n\nتم حفظ الطلب كـ "جديد" ولن يُخصم المخزون حتى يتوفر.', 'warning')
                current_app.logger.info(f"Order saved with insufficient stock (deferred). items={insufficient_products}")

            # Delivery path deduction — unchanged trigger
            deducted_delivery = False
            if status == 'خرج للتوصيل':
                try:
                    apply_inventory_deduction(order, product_required_qty, locked_products)
                    deducted_delivery = True
                except ValueError:
                    db.session.rollback()
                    return redirect(url_for('main.add_order'))

            # Walk-in path deduction — fully independent trigger.
            # Partial-return for a single item while keeping the rest of the
            # order intact is out of scope. Delete or edit quantities are the
            # supported workaround for "order created by mistake."
            deducted_walkin = False
            if order_type == ORDER_TYPE_WALKIN:
                apply_inventory_deduction(order, product_required_qty, locked_products)
                deducted_walkin = True

            if current_app.logger.isEnabledFor(10):
                current_app.logger.debug(f"[Perf] add_order items={len(order_items_to_add)} total={total} ms={(time.perf_counter()-t0)*1000:.2f}")

            if total == 0:
                current_app.logger.debug("No products found - total is 0")
                flash('يجب إضافة منتج واحد على الأقل للطلب')
                return redirect(url_for('main.add_order'))

            order.total_amount = total

            if order_type == ORDER_TYPE_WALKIN:
                # Walk-in: full cash sale at creation
                order.amount_paid = total  # remaining_amount naturally computes to 0
            else:
                apply_cod_fee_to_order(order)

            creation_msg = f"تم إنشاء الطلب بالحالة: {status}"
            if deducted_delivery or deducted_walkin:
                creation_msg += " - تم خصم المخزون"

            if order_type == ORDER_TYPE_DELIVERY and customer is not None:
                db.session.add(CustomerLog(
                    customer_id=customer.id,
                    order_id=order.id,
                    employee_id=session.get('employee_id'),
                    type="تلقائي",
                    content=creation_msg
                ))

            try:
                order.add_status_history(order.status, employee_id=session.get('employee_id'))
            except Exception as _e:
                current_app.logger.error(f"Failed to add status history on order creation: {_e}")

            db.session.commit()

            customer_label = order.customer_name or 'زبون أرضية'
            details = f"إضافة طلب جديد للعميل: {customer_label} - إجمالي المبلغ: {total} ج.م"
            log_activity(
                'create',
                'order',
                order.id,
                f"طلب رقم {order.id}",
                details)

            if submission_id:
                processed[submission_id] = order.id
                session['processed_submissions'] = processed
                session.modified = True

            current_app.logger.info(f"Order created id={order.id} type={order_type} total={total} customer_id={customer.id if customer else None}")
            
            flash('تم حفظ الطلب بنجاح - رقم الطلب: ' + str(order.id))
            
            return redirect(url_for('main.orders'))
        except Exception as e:
            current_app.logger.exception(f"ERROR in add_order: {e}")
            
            flash('حدث خطأ أثناء إضافة الطلب: {}'.format(e))
            return redirect(url_for('main.add_order'))
    from .models import GovernorateFee
    nearest_post_fee = GovernorateFee.query.filter_by(name="لأقرب فرع بريد").first()
    nearest_post_fee_value = nearest_post_fee.fee if nearest_post_fee else 40
    all_fees = GovernorateFee.query.all()
    fees_dict = {fee.name: fee.fee for fee in all_fees}
    submission_id = str(uuid4())
    return render_template('add_order.html', products=products_data, employees=employees, nearest_post_branch_fee=nearest_post_fee_value, governorate_fees=fees_dict, submission_id=submission_id)

@main.route('/orders/<int:order_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_orders')
def edit_order(order_id):
    order = Order.query.options(
        joinedload(Order.items).joinedload(OrderItem.variant),
        joinedload(Order.items).joinedload(OrderItem.size_variant),
        joinedload(Order.items).joinedload(OrderItem.color_variant),
        joinedload(Order.items).joinedload(OrderItem.style_variant),
        joinedload(Order.items).joinedload(OrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(Order.customer),
        joinedload(Order.registered_by),
        joinedload(Order.verified_by),
        joinedload(Order.called_by)
    ).get_or_404(order_id)
    customers = Customer.query.all()
    products = Product.query.options(
        joinedload(Product.variants),
        joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
    ).filter_by(is_deleted=False).all()
    old_status = order.status

    def serialize_product(product):
        result = {
            'id': product.id,
            'name': (product.name + ' (محذوف)') if getattr(product, 'is_deleted', False) else product.name,
            'price': product.price,
            'stock': product.stock,
            'wholesale_price': getattr(product, 'wholesale_price', 0),
            'is_bundle': getattr(product, 'is_bundle', False),
            'variants': [
                {
                    'id': v.id,
                    'group_name': v.group_name,
                    'variant_name': v.variant_name,
                    'price': v.price
                }
                for v in product.variants
            ]
        }
        
        if result['is_bundle'] and hasattr(product, 'bundle_items'):
            result['bundle_items'] = []
            for bundle_item in product.bundle_items:
                if bundle_item.product:
                    result['bundle_items'].append({
                        'product_id': bundle_item.product.id,
                        'product_name': bundle_item.product.name,
                        'sale_price_in_bundle': bundle_item.sale_price_in_bundle,
                        'product_variants': [
                            {
                                'id': v.id,
                                'group_name': v.group_name,
                                'variant_name': v.variant_name,
                                'price': v.price
                            }
                            for v in bundle_item.product.variants
                        ] if bundle_item.product.variants else []
                    })
        
        return result
    products_data = [serialize_product(p) for p in products]

    order_items_json = []
    for item in order.items:
        prod_id = item.product_id or None
        if not prod_id:
            if item.size_variant is not None:
                prod_id = item.size_variant.product_id
            elif item.color_variant is not None:
                prod_id = item.color_variant.product_id
            elif item.style_variant is not None:
                prod_id = item.style_variant.product_id
            elif item.variant is not None:
                prod_id = item.variant.product_id

        product_name = None
        if prod_id:
            try:
                prod_obj = Product.query.filter_by(id=prod_id).first()
                if prod_obj:
                    product_name = (prod_obj.name + ' (محذوف)') if getattr(prod_obj, 'is_deleted', False) else prod_obj.name
            except Exception:
                product_name = None

        try:
            bundle_variants = json.loads(item.bundle_variants_json) if getattr(item, 'bundle_variants_json', None) else None
        except Exception:
            bundle_variants = None

        payload = {
            'id': item.id,
            'quantity': item.quantity,
            'price': item.price,
            'product_id': prod_id,
            'product_name': product_name,
            'variant_id': item.variant_id,
            'size_variant_id': item.size_variant_id,
            'color_variant_id': item.color_variant_id,
            'style_variant_id': item.style_variant_id,
            'bundle_variants': bundle_variants,
        }
        order_items_json.append(payload)

    order_products_total = sum(
        item.price *
        item.quantity for item in order.items if item.variant)

    if request.method == 'POST':

        # order_type is immutable after creation — explicitly ignore any submitted value
        if 'order_type' in request.form:
            current_app.logger.debug(
                f"edit_order: ignoring submitted order_type='{request.form['order_type']}' "
                f"(immutable, current={order.order_type})"
            )

        # لقطة الحالة القديمة قبل التعديل
        _old_status_snap = order.status
        _old_notes_snap = order.notes or ''
        _old_amount_paid_snap = order.amount_paid
        _old_delivery_fees_snap = order.delivery_fees
        _old_tracking_snap = order.tracking_number or ''
        _old_employee_id_snap = order.employee_id
        _old_items_snap = []
        for _it in list(order.items):
            _pname = None
            try:
                _pid = _it.product_id
                if not _pid:
                    for _attr in ('size_variant', 'color_variant', 'style_variant', 'variant'):
                        _v = getattr(_it, _attr, None)
                        if _v:
                            _pid = getattr(_v, 'product_id', None)
                            break
                if _pid:
                    _p = Product.query.get(_pid)
                    if _p:
                        _pname = _p.name
            except Exception:
                pass
            _old_items_snap.append((_pname or f'منتج#{_it.product_id}', _it.quantity, _it.price))

        employee_id = request.form.get("employee_id")
        if employee_id and employee_id.strip():
            order.employee_id = None if employee_id == '0' else int(employee_id)

        customer_updated = False
        if order.status == 'جديد':
            customer_name = request.form.get("customer_name")
            customer_phone = request.form.get("customer_phone")
            customer_governorate = request.form.get("customer_governorate")
            customer_address_details = request.form.get("customer_address_details")
            
            changes = []
            if customer_name and customer_name != order.customer_name:
                changes.append(f"الاسم من '{order.customer_name}' إلى '{customer_name}'")
                order.customer_name = customer_name
                customer_updated = True
            if customer_phone and customer_phone != order.customer_phone:
                changes.append(f"الهاتف من '{order.customer_phone}' إلى '{customer_phone}'")
                order.customer_phone = customer_phone
                customer_updated = True
            if customer_governorate and customer_governorate != order.customer_governorate:
                changes.append(f"المحافظة من '{order.customer_governorate}' إلى '{customer_governorate}'")
                order.customer_governorate = customer_governorate
                customer_updated = True
            if customer_address_details and customer_address_details != order.customer_address_details:
                changes.append(f"العنوان من '{order.customer_address_details}' إلى '{customer_address_details}'")
                order.customer_address_details = customer_address_details
                customer_updated = True
            
            # تحديث بيانات العميل في ملفه الشخصي إذا اختار المستخدم ذلك
            update_customer_profile = request.form.get('update_customer_profile', '0')
            if customer_updated and update_customer_profile == 'order_and_customer' and order.customer:
                if customer_name and customer_name != order.customer.name:
                    order.customer.name = customer_name
                if customer_phone and customer_phone != order.customer.phone:
                    order.customer.phone = customer_phone
                if customer_governorate and customer_governorate != order.customer.governorate:
                    order.customer.governorate = customer_governorate
                if customer_address_details and customer_address_details != order.customer.address_details:
                    order.customer.address_details = customer_address_details

            if changes and order.customer:
                update_scope = "في الطلب وملف العميل" if update_customer_profile == 'order_and_customer' else "في الطلب فقط"
                log_content = f"تم تحديث بيانات العميل {update_scope}: " + " | ".join(changes)
                db.session.add(
                    CustomerLog(
                        customer_id=order.customer.id,
                        order_id=order.id,
                        employee_id=session.get('employee_id'),
                        type="تلقائي",
                        content=log_content
                    )
                )

        old_status = order.status
        new_status = request.form.get("status")
        try:
            validate_transition(old_status, new_status)
        except ValueError as e:
            flash(str(e), 'error')
            db.session.rollback()
            return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))
        order.status = new_status
        
        if old_status != new_status:
            order.status_updated_at = datetime.utcnow()
        
        delivery_fees_value = request.form.get("delivery_fees")
        try:
            current_app.logger.debug(f"edit_order: incoming delivery_fees raw: {delivery_fees_value!r}")
            current_app.logger.debug(f"edit_order: existing order.delivery_fees before: {order.delivery_fees!r}")
        except Exception:
            pass

        if 'delivery_fees' in request.form:
            try:
                order.delivery_fees = float(request.form.get('delivery_fees') or 0)
                try:
                    current_app.logger.debug(f"edit_order: updated order.delivery_fees to: {order.delivery_fees!r}")
                except Exception:
                    pass
            except Exception as e:
                current_app.logger.error(f"edit_order: failed to parse delivery_fees: {e}")
        order.notes = request.form.get("notes", "")
        order.amount_paid = float(request.form.get("amount_paid", 0))
        order.tracking_number = request.form.get("tracking_number")
        order.weight         = request.form.get('weight', type=float)
        order.package_volume = request.form.get('package_volume') or None
        order.delivery_notes = request.form.get('delivery_notes') or None
        
        try:
            order.is_urgent = bool(request.form.get('is_urgent'))
        except Exception:
            order.is_urgent = False

        from collections import defaultdict

        def _resolve_product_id(item):
            try:
                if item.product_id:
                    return int(item.product_id)
                if getattr(item, 'size_variant', None):
                    return int(item.size_variant.product_id)
                if getattr(item, 'color_variant', None):
                    return int(item.color_variant.product_id)
                if getattr(item, 'style_variant', None):
                    return int(item.style_variant.product_id)
                if getattr(item, 'variant', None):
                    return int(item.variant.product_id)
            except Exception:
                return None
            return None

        before_qty = defaultdict(int)
        try:
            for it in list(order.items):
                pid = _resolve_product_id(it)
                if pid:
                    product = _get_locked_product(pid)
                    if product and product.is_bundle:
                        bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                        for bundle_item in bundle_items:
                            before_qty[bundle_item.product_id] += int(it.quantity or 0) * 1
                    else:
                        before_qty[pid] += int(it.quantity or 0)
        except Exception:
            before_qty = defaultdict(int)

        items_data_json = request.form.get('order_items_data')
        if items_data_json:
            try:
                items_data = json.loads(items_data_json)
            except Exception:
                items_data = []
            existing_items = {it.id: it for it in order.items}
            seen_ids = set()

            product_ids = {
                int(row.get('product_id'))
                for row in items_data
                if row.get('product_id')
            }
            products_map = _get_locked_products(product_ids)

            for row in items_data:
                row_id_raw = row.get('id')
                row_id = int(row_id_raw) if row_id_raw is not None else None
                product_id = row.get('product_id')
                variant_id = row.get('variant_id')
                size_variant_id = row.get('size_variant_id')
                color_variant_id = row.get('color_variant_id')
                style_variant_id = row.get('style_variant_id')
                bundle_variants = row.get('bundle_variants')
                quantity = int(row.get('quantity') or 1)
                price = float(row.get('price') or 0)

                product = products_map.get(product_id)
                if product:
                    min_price = get_min_price(product)
                    if price < min_price:
                        flash(f'سعر المنتج {product.name} أقل من سعر الشراء', 'danger')
                        db.session.rollback()
                        return redirect(url_for('main.edit_order', order_id=order.id))

                if row_id and row_id in existing_items:
                    it = existing_items[row_id]
                    old_product_id = it.product_id
                    it.variant_id = None
                    it.size_variant_id = None
                    it.color_variant_id = None
                    it.style_variant_id = None
                    it.product_id = product_id
                    if variant_id:
                        it.variant_id = variant_id
                    if size_variant_id:
                        it.size_variant_id = size_variant_id
                    if color_variant_id:
                        it.color_variant_id = color_variant_id
                    if style_variant_id:
                        it.style_variant_id = style_variant_id
                    try:
                        it.bundle_variants_json = json.dumps(bundle_variants) if bundle_variants else None
                    except Exception:
                        it.bundle_variants_json = None
                    it.quantity = quantity
                    it.price = price
                    snapshot_product = products_map.get(product_id) if product_id else None
                    if snapshot_product and (it.purchase_price_snapshot is None or old_product_id != product_id):
                        it.purchase_price_snapshot = float(snapshot_product.purchase_price or 0)
                    seen_ids.add(row_id)
                else:
                    snapshot_product = products_map.get(product_id) if product_id else None
                    it = OrderItem(
                        order_id=order.id,
                        product_id=product_id,
                        variant_id=variant_id,
                        size_variant_id=size_variant_id,
                        color_variant_id=color_variant_id,
                        style_variant_id=style_variant_id,
                        quantity=quantity,
                        price=price,
                        purchase_price_snapshot=float(snapshot_product.purchase_price or 0) if snapshot_product else 0.0,
                        state='لم يتم التجربة بعد'
                    )
                    try:
                        it.bundle_variants_json = json.dumps(bundle_variants) if bundle_variants else None
                    except Exception:
                        it.bundle_variants_json = None
                    db.session.add(it)
                    db.session.flush()
                    seen_ids.add(it.id)

            to_delete = [it for it in order.items if it.id not in seen_ids]
            for it in to_delete:
                db.session.delete(it)

            if to_delete and not items_data:
                flash('خطأ في البيانات: لم يتم استلام أي منتجات. لم يتم حفظ التعديلات.', 'danger')
                db.session.rollback()
                return redirect(url_for('main.edit_order', order_id=order.id))

            try:
                if 'items_data' in locals() and isinstance(items_data, list) and len(items_data) > 0:
                    new_products_total = 0
                    for r in items_data:
                        try:
                            qty = int(r.get('quantity') or 1)
                        except Exception:
                            qty = 1
                        try:
                            pr = float(r.get('price') or 0)
                        except Exception:
                            pr = 0
                        new_products_total += pr * qty
                    order.total_amount = new_products_total
                else:
                    order.total_amount = sum((it.price or 0) * (it.quantity or 1) for it in order.items)
            except Exception:
                try:
                    order.total_amount = sum((it.price or 0) * (it.quantity or 1) for it in order.items)
                except Exception:
                    pass

            after_qty = defaultdict(int)
            try:
                variant_ids = set()
                for r in items_data:
                    for key in ('size_variant_id', 'color_variant_id', 'style_variant_id', 'variant_id'):
                        vid = r.get(key)
                        if vid:
                            try:
                                variant_ids.add(int(vid))
                            except Exception:
                                pass
                variant_product_map = {}
                if variant_ids:
                    for v_id, p_id in db.session.query(ProductVariant.id, ProductVariant.product_id).filter(ProductVariant.id.in_(list(variant_ids))).all():
                        variant_product_map[int(v_id)] = int(p_id)

                def _resolve_pid_from_row(r):
                    pid = r.get('product_id')
                    if pid:
                        try:
                            return int(pid)
                        except Exception:
                            pass
                    for key in ('size_variant_id', 'color_variant_id', 'style_variant_id', 'variant_id'):
                        vid = r.get(key)
                        if vid:
                            try:
                                vid_int = int(vid)
                            except Exception:
                                continue
                            mapped = variant_product_map.get(vid_int)
                            if mapped:
                                return int(mapped)
                    return None

                for r in items_data:
                    pid = _resolve_pid_from_row(r)
                    if not pid:
                        continue
                    try:
                        qty = int(r.get('quantity') or 0)
                    except Exception:
                        qty = 0
                    
                    product = _get_locked_product(pid)
                    if product and product.is_bundle:
                        bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                        for bundle_item in bundle_items:
                            after_qty[bundle_item.product_id] += qty * 1
                    else:
                        after_qty[pid] += qty
            except Exception:
                after_qty = defaultdict(int)

            increase = {}
            decrease = {}
            product_ids_union = set(before_qty.keys()) | set(after_qty.keys())
            for pid in product_ids_union:
                before = int(before_qty.get(pid, 0))
                after = int(after_qty.get(pid, 0))
                delta = after - before
                if delta > 0:
                    increase[pid] = delta
                elif delta < 0:
                    decrease[pid] = -delta

            if order.inventory_deducted:
                if increase:
                    pmap = _get_locked_products(increase.keys())
                    for pid, inc in increase.items():
                        p = pmap.get(pid)
                        if p and _block_if_deleted(p, 'الطلب', 'deduct'):
                            db.session.rollback()
                            return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))
                    for pid, inc in increase.items():
                        p = pmap.get(pid)
                        if not p:
                            continue
                        if (p.stock or 0) < inc:
                            flash(f"المخزون غير كافي للمنتج {p.name}. الزيادة المطلوبة: {inc} - المتوفر: {p.stock}")
                            try:
                                db.session.rollback()
                            except Exception:
                                pass
                            return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))

                    inc_log = {}
                    for pid, inc in increase.items():
                        p = pmap.get(pid)
                        if p:
                            p.stock = (p.stock or 0) - inc
                            inc_log[pid] = -inc
                    _log_stock_delta("edit_order increase deduct", inc_log)

                if decrease:
                    pmap2 = _get_locked_products(decrease.keys())
                    dec_log = {}
                    for pid, dec in decrease.items():
                        p = pmap2.get(pid)
                        if p:
                            p.stock = (p.stock or 0) + dec
                            dec_log[pid] = dec
                    _log_stock_delta("edit_order decrease restock", dec_log)

        if order.status == "استبدال":
            call_dates = request.form.getlist("call_logs[][datetime]")
            call_notes = request.form.getlist("call_logs[][note]")
            for d, note in zip(call_dates, call_notes):
                if d.strip():
                    call_time = datetime.strptime(d.strip(), "%Y-%m-%dT%H:%M")
                    db.session.add(
                        Issue(
                            order_id=order.id,
                            timestamp=call_time,
                            note=note))

        if old_status != order.status:
            if order.status == "خرج للتوصيل" and not order.inventory_deducted:
                from collections import defaultdict
                req = defaultdict(int)
                for it in order.items:
                    product = it.product
                    if product:
                        req[product.id] += (it.quantity or 0)
                if req:
                    pmap = _get_locked_products(req.keys())
                    for pid, q in req.items():
                        p = pmap.get(pid)
                        if not p:
                            continue
                        
                        if p.is_bundle:
                            bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                            for bundle_item in bundle_items:
                                bundle_product = _get_locked_product(bundle_item.product_id)
                                if bundle_product:
                                    required_qty = q * 1
                                    if (bundle_product.stock or 0) < required_qty:
                                        flash(f"المخزون غير كافي للمنتج {bundle_product.name} (داخل {p.name}). المطلوب: {required_qty} - المتوفر: {bundle_product.stock}")
                                        try:
                                            db.session.rollback()
                                        except Exception:
                                            pass
                                        return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))
                        else:
                            if (p.stock or 0) < q:
                                flash(f"المخزون غير كافي للمنتج {p.name}. المطلوب: {q} - المتوفر: {p.stock}")
                                try:
                                    db.session.rollback()
                                except Exception:
                                    pass
                                return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))
                    blocked = False
                    for pid, q in req.items():
                        p = pmap.get(pid)
                        if p:
                            if p.is_bundle:
                                bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                                for bundle_item in bundle_items:
                                    bundle_product = _get_locked_product(bundle_item.product_id)
                                    if bundle_product and _block_if_deleted(bundle_product, 'الطلب', 'deduct'):
                                        blocked = True
                                        break
                                if blocked:
                                    break
                            else:
                                if _block_if_deleted(p, 'الطلب', 'deduct'):
                                    blocked = True
                                    break
                    if blocked:
                        try:
                            db.session.rollback()
                        except Exception:
                            pass
                        return redirect(url_for('main.edit_order', order_id=order.id, next=request.args.get('next')))
                    deduct_log = {}
                    for pid, q in req.items():
                        p = pmap.get(pid)
                        if p:
                            if p.is_bundle:
                                bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                                for bundle_item in bundle_items:
                                    bundle_product = _get_locked_product(bundle_item.product_id)
                                    if bundle_product:
                                        bundle_product.stock = (bundle_product.stock or 0) - (q * 1)
                                        deduct_log[bundle_product.id] = deduct_log.get(bundle_product.id, 0) - (q * 1)
                            else:
                                p.stock = (p.stock or 0) - q
                                deduct_log[p.id] = deduct_log.get(p.id, 0) - q
                order.inventory_deducted = True
                _log_stock_delta("edit_order status->delivering deduct", deduct_log)

            if order.status in ["إلغاء", "رفض الاستلام"] and order.inventory_deducted:
                restock_log = {}
                for item in order.items:
                    product = item.product
                    if product:
                        if product.is_bundle:
                            bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                            for bundle_item in bundle_items:
                                bundle_product = _get_locked_product(bundle_item.product_id)
                                if bundle_product:
                                    bundle_product.stock = (bundle_product.stock or 0) + (item.quantity * 1)
                                    restock_log[bundle_product.id] = restock_log.get(bundle_product.id, 0) + (item.quantity * 1)
                        else:
                            product.stock = (product.stock or 0) + (item.quantity or 0)
                            restock_log[product.id] = restock_log.get(product.id, 0) + (item.quantity or 0)
                _log_stock_delta(f"edit_order status->{order.status} restock", restock_log)
                order.inventory_deducted = False

            if order.customer:
                db.session.add(
                    CustomerLog(
                        customer_id=order.customer.id,
                        order_id=order.id,
                        employee_id=session.get('employee_id'),
                        type="تلقائي",
                        content=f"تم تغيير حالة الطلب من {old_status} إلى {order.status}"
                    )
                )

        try:
            current_app.logger.debug(f"edit_order: about to commit - order_id={order.id}, total_amount={order.total_amount!r}, amount_paid={order.amount_paid!r}, delivery_fees={order.delivery_fees!r}")
        except Exception:
            pass
        
        apply_cod_fee_to_order(order)
        
        db.session.commit()
        
        if old_status != order.status:
            order.add_status_history(order.status, employee_id=session.get('employee_id'))
            db.session.commit()
        
        try:
            refreshed = Order.query.get(order.id)
            current_app.logger.debug(f"edit_order: after commit - order_id={order.id}, refreshed.total_amount={refreshed.total_amount!r}, refreshed.amount_paid={refreshed.amount_paid!r}, refreshed.delivery_fees={refreshed.delivery_fees!r}, refreshed.remaining_amount={refreshed.remaining_amount!r}")
        except Exception as e:
            current_app.logger.error(f"edit_order: post-commit check failed: {e}")

        success_msg = "تم تعديل الطلب بنجاح"
        if customer_updated:
            update_customer_profile = request.form.get('update_customer_profile', '0')
            if update_customer_profile == 'order_and_customer':
                success_msg += " وتم تحديث بيانات العميل في الطلب وملف العميل"
            else:
                success_msg += " وتم تحديث بيانات العميل في الطلب فقط"

        # تسجيل التعديل في سجل التعديلات مع تفاصيل التغييرات
        try:
            _edit_details = []
            if _old_status_snap != order.status:
                _edit_details.append(f"الحالة: {_old_status_snap} ← {order.status}")
            if customer_updated and changes:
                _edit_details.append("بيانات العميل: " + " | ".join(changes))
            if _old_notes_snap != (order.notes or ''):
                _old_n = (_old_notes_snap[:30] + '...') if len(_old_notes_snap) > 30 else _old_notes_snap
                _new_n = ((order.notes or '')[:30] + '...') if len(order.notes or '') > 30 else (order.notes or '')
                _edit_details.append(f"الملاحظات: '{_old_n}' ← '{_new_n}'" if (_old_n or _new_n) else "الملاحظات: تم التعديل")
            if _old_amount_paid_snap != order.amount_paid:
                _edit_details.append(f"المبلغ المدفوع: {_old_amount_paid_snap} ← {order.amount_paid}")
            if _old_delivery_fees_snap != order.delivery_fees:
                _edit_details.append(f"رسوم التوصيل: {_old_delivery_fees_snap} ← {order.delivery_fees}")
            if _old_tracking_snap != (order.tracking_number or ''):
                _new_t = order.tracking_number or 'محذوف'
                _edit_details.append(f"رقم التتبع: '{_old_tracking_snap}' ← '{_new_t}'" if _old_tracking_snap else f"رقم التتبع: {_new_t}")
            if _old_employee_id_snap != order.employee_id:
                try:
                    _new_emp = Employee.query.get(order.employee_id)
                    _edit_details.append(f"المسؤول: تم التغيير إلى {_new_emp.name if _new_emp else order.employee_id}")
                except Exception:
                    _edit_details.append("المسؤول: تم التغيير")
            # مقارنة المنتجات
            try:
                _new_items_snap = []
                for _it in order.items:
                    _pname_new = None
                    try:
                        _pid_new = _it.product_id
                        if not _pid_new:
                            for _attr in ('size_variant', 'color_variant', 'style_variant', 'variant'):
                                _v = getattr(_it, _attr, None)
                                if _v:
                                    _pid_new = getattr(_v, 'product_id', None)
                                    break
                        if _pid_new:
                            _p_new = Product.query.get(_pid_new)
                            if _p_new:
                                _pname_new = _p_new.name
                    except Exception:
                        pass
                    _new_items_snap.append((_pname_new or f'منتج#{_it.product_id}', _it.quantity, _it.price))
                if sorted(_old_items_snap) != sorted(_new_items_snap):
                    _items_str = "، ".join(f"{n} x{q} ({p}ج)" for n, q, p in _new_items_snap)
                    _edit_details.append(f"المنتجات: [{_items_str}]")
            except Exception:
                pass
            _desc = " | ".join(_edit_details) if _edit_details else "تم تعديل الطلب"
            db.session.add(OrderEditLog(
                order_id=order.id,
                employee_id=session.get('employee_id'),
                description=_desc
            ))
            db.session.commit()
        except Exception:
            pass

        flash(success_msg)
        
        next_url = request.args.get('next')
        if next_url:
            return redirect(next_url)
        else:
            if order.status == 'جديد':
                return redirect(url_for('main.orders', status='جديد'))
            return redirect(url_for('main.orders'))

    next_url = request.args.get('next')
    employees = Employee.query.filter(
        Employee.is_active.is_(True),
        or_(
            Employee.sales_commission_percentage > 0,
            Employee.id == order.employee_id
        )
    ).all()
    
    from .models import GovernorateFee
    all_fees = GovernorateFee.query.all()
    fees_dict = {f.name: f.fee for f in all_fees}
    
    returned_product_ids = {
        r.product_id for r in ReturnOrderItem.query.join(ReturnOrder).filter(
            ReturnOrder.original_order_id == order_id, ReturnOrderItem.return_order_id == ReturnOrder.id
        ).with_entities(ReturnOrderItem.product_id).distinct().all()
    }

    return render_template(
        'edit_order.html',
        order=order,
        customers=customers,
        employees=employees,
        products=products_data,
        order_items_json=order_items_json,
        order_products_total=order_products_total,
        governorate_fees=fees_dict,
        next_url=next_url,
        all_statuses=ALL_STATUSES,
        valid_next_statuses=StatusService.valid_next(order.status),
        returned_product_ids=returned_product_ids)

@main.route('/orders/<int:order_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_orders')
def delete_order(order_id):
    import time
    t0 = time.perf_counter()
    order = (Order.query.options(selectinload(Order.items).load_only(
        OrderItem.id,
        OrderItem.product_id,
        OrderItem.variant_id,
        OrderItem.size_variant_id,
        OrderItem.color_variant_id,
        OrderItem.style_variant_id,
        OrderItem.quantity))
        .filter_by(id=order_id)
        .first_or_404())

    if db.session.query(ReplacementOrder.id).filter_by(original_order_id=order.id).first():
        flash(f'لا يمكن حذف الطلب رقم {order.id} لأنه مرتبط بطلب استبدال. استخدم "حذف مع الطلبات المرتبطة" بدلاً من ذلك.', 'warning')
        return redirect(url_for('main.orders'))

    stock_returns = {}
    if order.inventory_deducted and order.status != 'إلغاء':
        variant_ids = set()
        for it in order.items:
            for vid in (it.variant_id, it.size_variant_id, it.color_variant_id, it.style_variant_id):
                if vid:
                    variant_ids.add(vid)

        variant_product_map = {}
        if variant_ids:
            for vid, pid in db.session.query(ProductVariant.id, ProductVariant.product_id).filter(ProductVariant.id.in_(variant_ids)).all():
                variant_product_map[vid] = pid

        from collections import defaultdict
        stock_returns = defaultdict(int)
        for it in order.items:
            pid = it.product_id
            if not pid:
                pid = variant_product_map.get(it.size_variant_id) or \
                      variant_product_map.get(it.color_variant_id) or \
                      variant_product_map.get(it.style_variant_id) or \
                      variant_product_map.get(it.variant_id)
            if pid:
                product = _get_locked_product(pid)
                if product and product.is_bundle:
                    bundle_items = BundleItem.query.filter_by(bundle_id=pid).all()
                    for bundle_item in bundle_items:
                        stock_returns[bundle_item.product_id] += (it.quantity or 0) * 1
                else:
                    stock_returns[pid] += (it.quantity or 0)

        if stock_returns:
            for pid, qty in stock_returns.items():
                product = _get_locked_product(pid)
                if product:
                    product.stock = (product.stock or 0) + qty
            _log_stock_delta("delete_order restock", dict(stock_returns))

        if order.inventory_deducted and order.status != 'إلغاء':
            order.inventory_deducted = False

    if 'employee_id' in session:
        try:
            activity_log = EmployeeActivityLog(
                employee_id=session['employee_id'],
                action='delete',
                entity_type='order',
                entity_id=order.id,
                entity_name=f"طلب رقم {order.id}",
                details=("حذف الطلب واسترجاع مخزون %d منتج" % len(stock_returns)) if stock_returns else "حذف الطلب (لم يتم خصم المخزون مسبقاً)",
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(activity_log)
        except Exception as e:
            current_app.logger.warning(f"delete_order: failed to log activity inline: {e}")

    try:
        db.session.delete(order)
        db.session.commit()
        elapsed_ms = (time.perf_counter() - t0) * 1000
        if current_app.logger.isEnabledFor(10):
            current_app.logger.debug(f"[Perf] delete_order id={order.id} items={len(order.items)} returns={len(stock_returns)} ms={elapsed_ms:.2f}")
        flash("تم حذف الطلب بنجاح", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف الطلب: {e}", 'error')
        return redirect(url_for('main.orders'))

    next_url = request.args.get('next')
    return redirect(next_url or url_for('main.orders'))

@main.route('/orders/status/<status>')
@login_required
@permission_required('can_view_orders_by_status')
def orders_by_status(status):
    from datetime import date
    today = date.today()
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    day = request.args.get('day', type=int)
    phone = request.args.get('phone', '').strip()
    tracking_number = request.args.get('tracking_number', '').strip()
    page = request.args.get('page', 1, type=int) or 1
    if page < 1:
        page = 1
    per_page = 50
    if not month:
        month = today.month
    if not year:
        year = today.year
    query = Order.query.filter(Order.status == status)
    if year:
        query = query.filter(extract('year', Order.date) == year)
    if month:
        query = query.filter(extract('month', Order.date) == month)
    if day:
        query = query.filter(extract('day', Order.date) == day)
    if phone:
        query = query.join(Order.customer).filter(Customer.phone.ilike(f"%{phone}%"))
    if tracking_number:
        query = query.filter(Order.tracking_number.ilike(f"%{tracking_number}%"))

    total_matching_orders = query.with_entities(func.count(Order.id)).scalar() or 0

    options_q = query.options(
        selectinload(Order.items).joinedload(OrderItem.variant),
        selectinload(Order.items).joinedload(OrderItem.size_variant),
        selectinload(Order.items).joinedload(OrderItem.color_variant),
        selectinload(Order.items).joinedload(OrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(Order.customer)
    )

    orders = (
        options_q
        .order_by(func.coalesce(Order.status_updated_at, Order.date).desc(), Order.id.desc())
        .limit(per_page)
        .offset((page - 1) * per_page)
        .all()
    )

    delivery_call_state_map = {}
    if status == 'خرج للتوصيل':
        delivery_call_state_map = _order_delivery_call_state_map([o.id for o in orders])

    if status == 'جديد':
        try:
            for _o in orders:
                has_oos = False
                for it in _o.items:
                    try:
                        prod = it.product
                        if not prod:
                            continue
                        if (prod.stock or 0) == 0 and it.quantity > 0:
                            has_oos = True
                            break
                    except Exception:
                        continue
                _o.has_out_of_stock_products = has_oos
        except Exception:
            for _o in orders:
                _o.has_out_of_stock_products = False
    else:
        for _o in orders:
            _o.has_out_of_stock_products = False
    
    status_counts = _order_status_counts(month, year)
    _annotate_status_age(orders, status)
    
    try:
        from .models import GovernorateFee
        _np = GovernorateFee.query.filter_by(name="لأقرب فرع بريد").first()
        nearest_post_branch_fee = _np.fee if _np else None
    except Exception:
        nearest_post_branch_fee = None

    has_next_page = (page * per_page) < total_matching_orders

    for o in orders:
        o.valid_next_statuses = StatusService.valid_next(o.status)

    return render_template(
        'orders_by_status.html',
        orders=orders,
        delivery_call_state_map=delivery_call_state_map,
        selected_status=status,
        selected_month=month,
        selected_year=year,
        selected_day=day,
        phone=phone,
        tracking_number=tracking_number,
        status_counts=status_counts,
        nearest_post_branch_fee=nearest_post_branch_fee,
        total_matching_orders=total_matching_orders,
        page=page,
        per_page=per_page,
        has_next_page=has_next_page,
        all_statuses=ALL_STATUSES,
    )


@main.route('/orders/new_products_summary')
@login_required
@permission_required('can_view_sold_products_by_quantity')
def new_products_summary():
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        from collections import Counter
        now_date = datetime.utcnow().date()
        if not month:
            month = now_date.month
        if not year:
            year = now_date.year

        orders_q = Order.query.filter(
            Order.status == 'جديد',
            extract('month', Order.date) == month,
            extract('year', Order.date) == year
        ).options(
            joinedload(Order.items).joinedload(OrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product)
        )

        orders = orders_q.all()

        counts = Counter()
        profits = Counter()

        for o in orders:
            for item in o.items:
                try:
                    prod = item.product
                    if not prod:
                        continue
                    if not prod.is_bundle:
                        q = (item.quantity or 0)
                        counts[prod.id] += q
                        try:
                            unit_price = item.price or prod.price or 0
                            profit_per_unit = (unit_price - (prod.purchase_price or 0))
                            profits[prod.id] += profit_per_unit * q
                        except Exception:
                            pass
                    else:
                        # bundle: always collect internal products both from JSON and from BundleItem rows
                        q = (item.quantity or 0)
                        seen = set()
                        if item.bundle_variants_json:
                            try:
                                bv = json.loads(item.bundle_variants_json)
                                for k, v in bv.items():
                                    pid = v.get('product_id')
                                    inner_q = v.get('quantity') or v.get('qty') or v.get('count') or 1
                                    if pid:
                                        total_add = q * int(inner_q)
                                        counts[int(pid)] += total_add
                                        seen.add(int(pid))
                                        # profit from JSON if available
                                        sale_price = v.get('sale_price') or v.get('price')
                                        if sale_price is not None:
                                            try:
                                                prod_obj = Product.query.get(int(pid))
                                                profits[int(pid)] += (float(sale_price) - (prod_obj.purchase_price or 0)) * total_add
                                            except Exception:
                                                pass
                            except Exception:
                                pass  # ignore JSON errors
                        # include any bundle_items not covered by JSON
                        for bi in prod.bundle_items:
                            if bi.product:
                                pid = bi.product_id
                                if pid in seen:
                                    continue
                                counts[pid] += q
                                try:
                                    profits[pid] += (bi.sale_price_in_bundle - (bi.product.purchase_price or 0)) * q
                                except Exception:
                                    pass
                except Exception:
                    continue

        # fetch product details
        results = []
        if counts:
            prod_rows = Product.query.filter(
                Product.id.in_(list(counts.keys())),
                Product.is_deleted == False
            ).all()
            prod_map = {p.id: p for p in prod_rows}
            for pid, qty in counts.items():
                p = prod_map.get(pid)
                if not p:
                    continue
                total_profit = float(profits.get(pid, 0.0) or 0.0)
                results.append({
                    'product_id': p.id,
                    'name': p.name,
                    'total_quantity': int(qty),
                    'total_profit': round(total_profit, 2),
                    'stock': int(p.stock or 0)
                })

        results = sorted(results, key=lambda x: x['total_quantity'], reverse=True)
        return jsonify(success=True, data=results)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@main.route('/orders/new_products_summary_page')
@login_required
@permission_required('can_view_sold_products_by_quantity')
def new_products_summary_page():
    from datetime import date
    today = date.today()
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    if not month:
        month = today.month
    if not year:
        year = today.year
    return render_template('new_products_summary.html', selected_month=month, selected_year=year)

@main.route('/orders/overdue_new_count')
@login_required
@permission_required('can_view_orders')
def overdue_new_count():
    try:
        now_dt = datetime.now()
        cutoff = now_dt - timedelta(days=7)
        current_month = now_dt.month
        current_year = now_dt.year
        new_orders = (
            Order.query
            .filter(
                Order.status == 'جديد',
                extract('month', Order.date) == current_month,
                extract('year', Order.date) == current_year
            )
            .all()
        )
        count = 0
        for o in new_orders:
            ts = _order_status_dt(o)
            if ts and ts <= cutoff:
                count += 1
        return jsonify(success=True, count=int(count))
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@main.route('/replacement-orders/overdue_new_count')
@login_required
@permission_required('can_view_replacements')
def replacement_overdue_new_count():
    try:
        now_dt = datetime.now()
        cutoff = now_dt - timedelta(days=7)
        current_month = now_dt.month
        current_year = now_dt.year
        new_repl = (
            ReplacementOrder.query
            .filter(
                ReplacementOrder.status == 'جديد',
                extract('month', ReplacementOrder.date) == current_month,
                extract('year', ReplacementOrder.date) == current_year
            )
            .all()
        )
        count = 0
        for o in new_repl:
            ts = _order_status_dt(o)
            if ts and ts <= cutoff:
                count += 1
        return jsonify(success=True, count=int(count))
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@main.route('/api/overdue_all')
@login_required
def overdue_all():
    try:
        now_dt = datetime.now()
        cutoff = now_dt - timedelta(days=7)
        orders = Order.query.filter(Order.status == 'جديد').all()
        o_count = sum(1 for o in orders if _order_status_dt(o) and _order_status_dt(o) <= cutoff)
        repl = ReplacementOrder.query.filter(ReplacementOrder.status == 'جديد').all()
        r_count = sum(1 for o in repl if _order_status_dt(o) and _order_status_dt(o) <= cutoff)
        return jsonify(success=True, orders=int(o_count), replacements=int(r_count))
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500
@login_required
@permission_required('can_view_orders')
def orders_all():
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    phone = request.args.get('phone', '').strip()
    tracking_number = request.args.get('tracking_number', '').strip()
    query = Order.query
    if year:
        query = query.filter(extract('year', Order.date) == year)
    if month:
        query = query.filter(extract('month', Order.date) == month)
    if phone:
        query = query.join(
            Order.customer).filter(
            Customer.phone.ilike(
                f"%{phone}%"))
    if tracking_number:
        query = query.filter(
            Order.tracking_number.ilike(
                f"%{tracking_number}%"))
    orders = query.options(
        joinedload(Order.items).joinedload(OrderItem.variant),
        joinedload(Order.items).joinedload(OrderItem.size_variant),
        joinedload(Order.items).joinedload(OrderItem.color_variant),
        joinedload(Order.customer)
    ).order_by(Order.date.desc()).all()
    
    from datetime import date
    today = date.today()
    
    if not month:
        month = today.month
    if not year:
        year = today.year
    
    status_counts = _order_status_counts(month, year)
    
    for o in orders:
        o.valid_next_statuses = StatusService.valid_next(o.status)

    return render_template(
        'orders_by_status.html',
        orders=orders,
        selected_status='كل الطلبات',
        selected_month=month,
        selected_year=year,
        status_counts=status_counts,
        all_statuses=ALL_STATUSES)

@main.route('/orders/returns/customer-orders', methods=['POST'])
@login_required
def return_customer_orders():
    try:
        data = request.get_json(force=True)
        customer_id = int(data['customer_id'])
        source_type = data['source_type']

        if source_type == 'order':
            orders = Order.query.filter(Order.customer_id == customer_id, Order.status == 'وصل').order_by(Order.date.desc()).limit(20).all()
            result = [{'id': o.id, 'date': o.date.strftime('%Y-%m-%d') if o.date else '', 'total_amount': o.total_amount or 0} for o in orders]
        else:
            orders = ReplacementOrder.query.filter(ReplacementOrder.customer_id == customer_id, ReplacementOrder.status == 'وصل', ReplacementOrder.is_draft == False).order_by(ReplacementOrder.date.desc()).limit(20).all()
            result = [{'id': o.id, 'date': o.date.strftime('%Y-%m-%d') if o.date else '', 'total_amount': o.total_amount or 0} for o in orders]

        return jsonify({'success': True, 'orders': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/orders/returns/order-items', methods=['POST'])
@login_required
def return_order_items():
    try:
        data = request.get_json(force=True)
        source_id = int(data['source_id'])
        source_type = data['source_type']

        if source_type == 'order':
            items = OrderItem.query.filter_by(order_id=source_id).all()
            result = [{
                'product_id': item.product_id,
                'product_name': item.product.name if item.product else 'منتج محذوف',
                'quantity': item.quantity or 1,
                'price': item.price or 0
            } for item in items]
        else:
            items = ReplacementOrderItem.query.filter_by(replacement_order_id=source_id).all()
            result = [{
                'product_id': item.product_id,
                'product_name': item.product.name if item.product else 'منتج محذوف',
                'quantity': item.quantity or 1,
                'price': item.price or 0
            } for item in items]

        return jsonify({'success': True, 'items': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/orders/returns/popup-data', methods=['POST'])
@login_required
@permission_required('can_view_returns')
def returns_popup_data():
    try:
        data = request.get_json(force=True)
        month = int(data.get('month', datetime.now().month))
        year = int(data.get('year', datetime.now().year))

        returns = ReturnOrder.query.filter(
            extract('month', ReturnOrder.created_at) == month,
            extract('year', ReturnOrder.created_at) == year
        ).options(
            joinedload(ReturnOrder.customer),
            joinedload(ReturnOrder.received_by)
        ).order_by(ReturnOrder.created_at.desc()).all()

        result = []
        all_items = []
        for r in returns:
            items_summary = []
            reasons = []
            for item in r.items:
                pname = item.product.name if item.product else f'منتج #{item.product_id}'
                items_summary.append(f'{pname} x{item.quantity} ({item.item_condition})')
                reasons.append(item.return_reason)
                all_items.append({
                    'return_order_id': r.id,
                    'created_at': r.created_at.strftime('%Y-%m-%d') if r.created_at else '',
                    'customer_name': r.customer.name if r.customer else '-',
                    'product_name': pname,
                    'quantity': item.quantity,
                    'condition': item.item_condition,
                    'reason': item.return_reason,
                    'refund': (item.unit_sale_price_snapshot or 0) * item.quantity
                })

            result.append({
                'id': r.id,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                'customer_name': r.customer.name if r.customer else '-',
                'items_summary': ' | '.join(items_summary) if items_summary else '-',
                'return_reason': ' | '.join(reasons) if reasons else '-',
                'customer_refund_amount': r.customer_refund_amount or 0,
                'received_by_name': r.received_by.name if r.received_by else '-'
            })

        total_amount = sum(r.customer_refund_amount or 0 for r in returns)
        total_items_count = len(all_items)

        return jsonify({
            'success': True,
            'returns': result,
            'all_items': all_items,
            'total_amount': total_amount,
            'total_items_count': total_items_count,
            'count': len(result)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/orders/returns/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_returns')
def add_return_order():
    if request.method == 'GET':
        return render_template('add_return_order.html')

    try:
        data = request.get_json(force=True)
        customer_id = data.get('customer_id')
        source_type = data.get('source_type')
        source_id = data.get('source_id')
        items = data.get('items', [])

        if not customer_id or not source_type or not source_id or not items:
            return jsonify({'error': 'الرجاء إدخال جميع البيانات المطلوبة'}), 400

        customer_id = int(customer_id)
        source_id = int(source_id)

        if source_type not in ('order', 'replacement_order'):
            return jsonify({'error': 'نوع المصدر غير صالح'}), 400

        product_ids = [int(it['product_id']) for it in items]
        for it in items:
            if int(it['quantity']) < 1:
                return jsonify({'error': 'الكمية يجب أن تكون 1 على الأقل'}), 400
            if it.get('item_condition') not in ('سليم', 'تالف'):
                return jsonify({'error': 'حالة المنتج غير صالحة'}), 400
            if not it.get('return_reason', '').strip():
                return jsonify({'error': 'سبب الإرجاع مطلوب'}), 400

        for it in items:
            product_id = int(it['product_id'])
            new_qty = int(it['quantity'])

            already_returned = db.session.query(
                func.coalesce(func.sum(ReturnOrderItem.quantity), 0)
            ).join(ReturnOrder, ReturnOrder.id == ReturnOrderItem.return_order_id).filter(
                ReturnOrderItem.product_id == product_id
            )

            if source_type == 'order':
                already_returned = already_returned.filter(
                    ReturnOrder.original_order_id == source_id
                )
                original_qty = db.session.query(
                    func.coalesce(func.sum(OrderItem.quantity), 0)
                ).filter(
                    OrderItem.order_id == source_id,
                    OrderItem.product_id == product_id
                ).scalar() or 0
            else:
                already_returned = already_returned.filter(
                    ReturnOrder.original_replacement_order_id == source_id
                )
                original_qty = db.session.query(
                    func.coalesce(func.sum(ReplacementOrderItem.quantity), 0)
                ).filter(
                    ReplacementOrderItem.replacement_order_id == source_id,
                    ReplacementOrderItem.product_id == product_id
                ).scalar() or 0

            already_returned = already_returned.scalar() or 0

            if already_returned + new_qty > original_qty:
                return jsonify({'error': 'الكمية المرتجعة تتجاوز الكمية الأصلية للطلب'}), 400

        locked_products = _get_locked_products(product_ids)

        return_order = ReturnOrder(
            customer_id=customer_id,
            received_by_id=session['employee_id'],
            customer_refund_amount=0.0
        )
        if source_type == 'order':
            return_order.original_order_id = source_id
        else:
            return_order.original_replacement_order_id = source_id

        db.session.add(return_order)
        db.session.flush()

        customer_refund_amount = 0.0

        for it in items:
            product_id = int(it['product_id'])
            quantity = int(it['quantity'])
            item_condition = it['item_condition']
            return_reason = it['return_reason'].strip()

            product = locked_products.get(product_id)
            unit_sale_price = 0.0
            unit_purchase_price = 0.0

            if source_type == 'order':
                order_item = OrderItem.query.filter_by(
                    order_id=source_id, product_id=product_id
                ).first()
                if order_item:
                    unit_sale_price = order_item.price or 0.0
                    unit_purchase_price = order_item.purchase_price_snapshot or 0.0
            else:
                rep_item = ReplacementOrderItem.query.filter_by(
                    replacement_order_id=source_id, product_id=product_id
                ).first()
                if rep_item:
                    unit_sale_price = rep_item.price or 0.0
                    unit_purchase_price = rep_item.purchase_price or 0.0

            return_order_item = ReturnOrderItem(
                return_order_id=return_order.id,
                product_id=product_id,
                quantity=quantity,
                item_condition=item_condition,
                return_reason=return_reason,
                unit_sale_price_snapshot=unit_sale_price,
                unit_purchase_price_snapshot=unit_purchase_price,
                inspected_by_id=session['employee_id']
            )
            db.session.add(return_order_item)
            db.session.flush()

            customer_refund_amount += unit_sale_price * quantity

            if item_condition == 'سليم':
                if product:
                    product.stock = (product.stock or 0) + quantity
            elif item_condition == 'تالف':
                total_loss = unit_purchase_price * quantity
                damaged_log = DamagedProductLog(
                    product_id=product_id,
                    quantity=quantity,
                    purchase_price_snapshot=unit_purchase_price,
                    total_loss=total_loss,
                    notes=f'إرجاع - طلب استرجاع رقم {return_order.id}',
                    created_by=session['employee_id'],
                    return_order_item_id=return_order_item.id
                )
                db.session.add(damaged_log)

        return_order.customer_refund_amount = customer_refund_amount

        customer_log = CustomerLog(
            customer_id=customer_id,
            return_order_id=return_order.id,
            employee_id=session['employee_id'],
            type='تلقائي',
            content=f'تم تسجيل طلب استرجاع رقم {return_order.id} بقيمة {customer_refund_amount}'
        )
        db.session.add(customer_log)

        db.session.flush()

        if source_type == 'order':
            status_history_entry = OrderStatusHistory(
                order_id=source_id,
                status='مرتجع',
                changed_by_employee_id=session['employee_id'],
                timestamp=datetime.utcnow()
            )
        else:
            status_history_entry = ReplacementOrderStatusHistory(
                replacement_order_id=source_id,
                status='مرتجع',
                changed_by_employee_id=session['employee_id'],
                timestamp=datetime.utcnow()
            )
        db.session.add(status_history_entry)

        if source_type == 'order':
            order_obj = db.session.get(Order, source_id)
            if order_obj:
                order_obj.status = 'مرتجع'
                order_obj.status_updated_at = datetime.utcnow()
        else:
            rep_obj = db.session.get(ReplacementOrder, source_id)
            if rep_obj:
                rep_obj.status = 'مرتجع'
                rep_obj.status_updated_at = datetime.utcnow()

        db.session.commit()
        flash(f'تم تسجيل طلب الإرجاع رقم {return_order.id} بنجاح', 'success')
        return jsonify({'success': True, 'return_order_id': return_order.id, 'redirect': url_for('main.replacement_orders')})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error in add_return_order: {e}')
        return jsonify({'error': 'حدث خطأ أثناء تسجيل طلب الإرجاع'}), 500


@main.route('/orders/replacements')
@login_required
@permission_required('can_view_replacements')
def replacement_orders():
    current_year = datetime.now().year
    current_month = datetime.now().month

    # Get non-draft new orders only
    new_orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'جديد',
        ReplacementOrder.is_draft == False,
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).options(
        joinedload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.is_urgent.desc(), ReplacementOrder.date.desc()).limit(5).all()
    new_count = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'جديد',
        ReplacementOrder.is_draft == False,
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).count()
    
    # Get drafts count for badge
    drafts_count = ReplacementOrder.query.filter(
        ReplacementOrder.is_draft == True
    ).count()

    shipped_orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'خرج للتوصيل',
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).options(
        joinedload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.date.desc()).limit(5).all()
    shipped_count = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'خرج للتوصيل',
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).count()

    delivered_orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'وصل',
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).options(
        joinedload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.date.desc()).limit(5).all()
    delivered_count = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'وصل',
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).count()
    
    total_delivery_fees_loss_all = db.session.query(
        func.coalesce(func.sum(ReplacementOrder.delivery_fees), 0)
    ).filter(
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).scalar() or 0
    total_damaged_products_cost_all = db.session.query(
        func.coalesce(func.sum(ReplacementOrder.damaged_products_loss), 0)
    ).filter(
        extract('year', ReplacementOrder.date) == current_year,
        extract('month', ReplacementOrder.date) == current_month
    ).scalar() or 0

    return render_template(
        'replacements.html',
        drafts_count=drafts_count,
        new_count=new_count,
        shipped_count=shipped_count,
        delivered_count=delivered_count,
        total_delivery_fees_loss=total_delivery_fees_loss_all,
        total_damaged_products_cost=total_damaged_products_cost_all)

@main.route('/orders/replacements/drafts')
@login_required
@permission_required('can_view_replacements')
def replacement_drafts():
    """Display all replacement order drafts"""
    draft_orders = ReplacementOrder.query.filter(
        ReplacementOrder.is_draft == True
    ).options(
        joinedload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.date.desc()).all()
    
    return render_template(
        'replacement_drafts.html',
        draft_orders=draft_orders
    )

@main.route('/orders/replacements/status/<status>')
@login_required
@permission_required('can_view_replacements_by_state')

def replacement_orders_by_status(status):
    phone = request.args.get('phone', '').strip()
    tracking_number = request.args.get('tracking_number', '').strip()
    selected_status = status
    page = request.args.get('page', 1, type=int) or 1
    if page < 1:
        page = 1
    per_page = 50
    
    selected_year = request.args.get('year', datetime.now().year, type=int)
    selected_month = request.args.get('month', datetime.now().month, type=int)
    selected_day = request.args.get('day', type=int)
    
    # Exclude drafts from status-based queries
    query = ReplacementOrder.query.filter_by(status=selected_status, is_draft=False)
    if selected_year:
        query = query.filter(extract('year', ReplacementOrder.date) == selected_year)
    if selected_month:
        query = query.filter(extract('month', ReplacementOrder.date) == selected_month)
    if selected_day:
        query = query.filter(extract('day', ReplacementOrder.date) == selected_day)
    if phone:
        query = query.join(ReplacementOrder.customer).filter(Customer.phone.ilike(f"%{phone}%"))
    if tracking_number:
        query = query.filter(ReplacementOrder.tracking_number.ilike(f"%{tracking_number}%"))
    total_matching_orders = query.with_entities(func.count(ReplacementOrder.id)).scalar() or 0
    query = query.order_by(func.coalesce(ReplacementOrder.status_updated_at, ReplacementOrder.date).desc(), ReplacementOrder.id.desc())
    
    orders = query.options(
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.size_variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.color_variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(ReplacementOrder.customer)
    ).limit(per_page).offset((page - 1) * per_page).all()

    delivery_call_state_map = {}
    if selected_status == 'خرج للتوصيل':
        delivery_call_state_map = _replacement_delivery_call_state_map([o.id for o in orders])

    _annotate_status_age(orders, selected_status)
    
    try:
        from .models import GovernorateFee
        _np = GovernorateFee.query.filter_by(name="لأقرب فرع بريد").first()
        nearest_post_branch_fee = _np.fee if _np else None
    except Exception:
        nearest_post_branch_fee = None

    has_next_page = (page * per_page) < total_matching_orders

    for o in orders:
        o.valid_next_statuses = StatusService.valid_next(o.status)

    return render_template(
        'replacements_by_state.html',
        orders=orders,
        delivery_call_state_map=delivery_call_state_map,
        phone=phone,
    tracking_number=tracking_number,
        selected_status=selected_status,
        selected_year=selected_year,
    selected_month=selected_month,
    selected_day=selected_day,
    nearest_post_branch_fee=nearest_post_branch_fee,
    total_matching_orders=total_matching_orders,
    page=page,
    per_page=per_page,
    has_next_page=has_next_page,
    all_statuses=ALL_STATUSES,
    )

@main.route('/replacement-orders/<int:order_id>/update-status', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def update_replacement_order_status(order_id):
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'success': False, 'message': 'الحالة مطلوبة'})
        
        replacement_order = ReplacementOrder.query.get(order_id)
        if not replacement_order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'})
        
        old_status = replacement_order.status
        try:
            validate_transition(old_status, new_status)
        except ValueError as e:
            return jsonify({'success': False, 'message': str(e)}), 400
        if new_status == 'خرج للتوصيل' and not getattr(replacement_order, 'inventory_deducted', False):
            required = _collect_replacement_required_quantities(replacement_order.items)
            if required:
                pmap = _get_locked_products(required.keys())
                for p in pmap.values():
                    if _block_if_deleted(p, 'طلب الاستبدال', 'deduct'):
                        return jsonify({'success': False, 'message': f'لا يمكن تغيير الحالة: طلب الاستبدال يحتوي على منتج محذوف ("{p.name}")'}), 400
                insufficient = []
                for pid, q in required.items():
                    p = pmap.get(pid)
                    if not p:
                        continue
                    if (p.stock or 0) < q:
                        insufficient.append(f"{p.name} (المطلوب {q} المتوفر {p.stock})")
                if insufficient:
                    return jsonify({'success': False, 'message': f'لا يمكن تغيير الحالة إلى {new_status} لعدم كفاية المخزون: ' + ' ؛ '.join(insufficient)}), 400
                rep_deduct_log = {}
                for pid, q in required.items():
                    p = pmap.get(pid)
                    if p:
                        p.stock = (p.stock or 0) - q
                        rep_deduct_log[pid] = rep_deduct_log.get(pid, 0) - q
                replacement_order.inventory_deducted = True
                _log_stock_delta("replacement_order status->delivering deduct", rep_deduct_log)

        if new_status in ['إلغاء', 'رفض الاستلام'] and getattr(replacement_order, 'inventory_deducted', False):
            try:
                returns = _collect_replacement_required_quantities(replacement_order.items)
                if returns:
                    for pid, q in returns.items():
                        product = _get_locked_product(pid)
                        if product:
                            product.stock = (product.stock or 0) + q
                    _log_stock_delta(f"replacement_order status->{new_status} restock", dict(returns))
            except Exception:
                pass
            replacement_order.inventory_deducted = False

        old_status = replacement_order.status
        replacement_order.status = new_status
        replacement_order.status_updated_at = datetime.utcnow()
        db.session.commit()
        
        if old_status != new_status:
            try:
                replacement_order.add_status_history(
                    new_status,
                    employee_id=session.get('employee_id'),
                    notes=f'تم تغيير الحالة من {old_status}'
                )
                db.session.commit()
            except Exception as e:
                current_app.logger.error(f"Failed to add replacement status history: {e}")
        
        log_activity(
            action='update',
            entity_type='replacement_order',
            entity_id=replacement_order.id,
            entity_name=f'طلب استبدال رقم {replacement_order.id}',
            details=f'تم تحديث الحالة إلى: {new_status}'
        )
        
        return jsonify({'success': True, 'message': 'تم تحديث الحالة بنجاح'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/replacement-orders/<int:order_id>/update-tracking', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def update_replacement_order_tracking(order_id):
    try:
        data = request.get_json() or {}
        tracking_number = (data.get('tracking_number') or '').strip()

        replacement_order = ReplacementOrder.query.get(order_id)
        if not replacement_order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'})

        previous_status = replacement_order.status
        replacement_order.tracking_number = tracking_number or None
        if tracking_number and replacement_order.status == 'جديد':
            try:
                validate_transition(previous_status, 'خرج للتوصيل')
            except ValueError as e:
                return jsonify({'success': False, 'message': str(e)}), 400
            if not getattr(replacement_order, 'inventory_deducted', False):
                required = _collect_replacement_required_quantities(replacement_order.items)
                if required:
                    pmap = _get_locked_products(required.keys())
                    for p in pmap.values():
                        if _block_if_deleted(p, 'طلب الاستبدال', 'deduct'):
                            return jsonify({'success': False, 'message': f'لا يمكن تعيين رقم تتبعي: طلب الاستبدال يحتوي على منتج محذوف ("{p.name}")'}), 400
                    insufficient = []
                    for pid, q in required.items():
                        p = pmap.get(pid)
                        if (p.stock or 0) < q:
                            insufficient.append(f"{p.name} (المطلوب {q} المتوفر {p.stock})")
                    if insufficient:
                        return jsonify({'success': False, 'message': 'لا يمكن تعيين رقم تتبعي / التحويل إلى خرج للتوصيل بسبب نقص المخزون: ' + ' ؛ '.join(insufficient)}), 400
                    for pid, q in required.items():
                        p = pmap.get(pid)
                        if p:
                            p.stock = (p.stock or 0) - q
                    replacement_order.inventory_deducted = True
            replacement_order.status = 'خرج للتوصيل'
            replacement_order.status_updated_at = datetime.utcnow()

        db.session.commit()
        
        if previous_status != replacement_order.status:
            try:
                replacement_order.add_status_history(
                    replacement_order.status,
                    employee_id=session.get('employee_id'),
                    notes='تم تغيير الحالة تلقائيًا عند إضافة رقم تتبع'
                )
                db.session.commit()
            except Exception as e:
                current_app.logger.error(f"Failed to add replacement status history: {e}")

        details_parts = [f"تحديث الرقم التتبعي إلى: {tracking_number or 'فارغ'}"]
        if previous_status != replacement_order.status:
            details_parts.append(f"وتغيير الحالة من {previous_status} إلى {replacement_order.status}")
        log_activity(
            action='update',
            entity_type='replacement_order',
            entity_id=replacement_order.id,
            entity_name=f'طلب استبدال رقم {replacement_order.id}',
            details='، '.join(details_parts)
        )

        return jsonify({'success': True, 'new_status': replacement_order.status or ''})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/replacement-orders/<int:order_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_replacements')
def delete_replacement_order(order_id):
    try:
        replacement_order = ReplacementOrder.query.get(order_id)
        if not replacement_order:
            flash('طلب الاستبدال غير موجود', 'error')
            return redirect(url_for('main.replacement_orders'))
        
        log_activity(
            action='delete',
            entity_type='replacement_order',
            entity_id=replacement_order.id,
            entity_name=f'طلب استبدال رقم {replacement_order.id}',
            details=f'تم حذف طلب الاستبدال للعميل: {replacement_order.customer.name}'
        )
        
        try:
            if getattr(replacement_order, 'inventory_deducted', False) and replacement_order.status not in ['إلغاء', 'رفض الاستلام']:
                returns = _collect_replacement_required_quantities(replacement_order.items)
                if returns:
                    for pid, q in returns.items():
                        product = _get_locked_product(pid)
                        if product:
                            product.stock = (product.stock or 0) + q
        except Exception:
            pass

        db.session.delete(replacement_order)
        db.session.commit()
        
        flash('تم حذف طلب الاستبدال بنجاح', 'success')
        return redirect(url_for('main.replacement_orders'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الحذف: {str(e)}', 'error')
        return redirect(url_for('main.replacement_orders'))


@main.route('/orders/<int:order_id>/delivery-call-verification', methods=['POST'])
@login_required
@permission_required('can_edit_orders')
def order_delivery_call_verification(order_id):
    try:
        data = request.get_json() or {}
        outcome = str(data.get('outcome', '')).strip().lower()
        if outcome not in (DELIVERY_CALL_STATE_INFORMED, DELIVERY_CALL_STATE_CONTACTED, DELIVERY_CALL_STATE_NOT_CONTACTED):
            return jsonify({'success': False, 'message': 'حالة المكالمة غير صحيحة'}), 400

        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'}), 404
        if order.status != 'خرج للتوصيل':
            return jsonify({'success': False, 'message': 'هذا الإجراء متاح فقط لطلبات خرج للتوصيل'}), 400

        employee_id = session.get('employee_id')
        note = _delivery_call_note_for_state(outcome)

        db.session.add(OrderStatusHistory(
            order_id=order.id,
            status=order.status,
            changed_by_employee_id=employee_id,
            notes=note
        ))
        if order.is_delivery:
            db.session.add(CustomerLog(
                customer_id=order.customer_id,
                order_id=order.id,
                employee_id=employee_id,
                type='تلقائي',
                content=f'{note} (طلب رقم {order.id})'
            ))
        db.session.commit()

        log_activity(
            action='update',
            entity_type='order',
            entity_id=order.id,
            entity_name=f'طلب رقم {order.id}',
            details=note
        )

        return jsonify({'success': True, 'state': outcome, 'message': 'تم حفظ نتيجة التحقق من التسليم'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})


@main.route('/replacement-orders/<int:order_id>/delivery-call-verification', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def replacement_delivery_call_verification(order_id):
    try:
        data = request.get_json() or {}
        outcome = str(data.get('outcome', '')).strip().lower()
        if outcome not in (DELIVERY_CALL_STATE_INFORMED, DELIVERY_CALL_STATE_CONTACTED, DELIVERY_CALL_STATE_NOT_CONTACTED):
            return jsonify({'success': False, 'message': 'حالة المكالمة غير صحيحة'}), 400

        order = ReplacementOrder.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'}), 404
        if order.status != 'خرج للتوصيل':
            return jsonify({'success': False, 'message': 'هذا الإجراء متاح فقط لطلبات خرج للتوصيل'}), 400

        employee_id = session.get('employee_id')
        note = _delivery_call_note_for_state(outcome)

        db.session.add(ReplacementOrderStatusHistory(
            replacement_order_id=order.id,
            status=order.status,
            changed_by_employee_id=employee_id,
            notes=note
        ))
        db.session.add(CustomerLog(
            customer_id=order.customer_id,
            employee_id=employee_id,
            type='تلقائي',
            content=f'{note} (طلب استبدال رقم {order.id})'
        ))
        db.session.commit()

        log_activity(
            action='update',
            entity_type='replacement_order',
            entity_id=order.id,
            entity_name=f'طلب استبدال رقم {order.id}',
            details=note
        )

        return jsonify({'success': True, 'state': outcome, 'message': 'تم حفظ نتيجة التحقق من التسليم'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/orders/<int:order_id>/toggle-called', methods=['POST'])
@login_required
@permission_required('can_edit_orders')
def toggle_order_called(order_id):
    try:
        data = request.get_json() or {}
        called = str(data.get('called', False)).strip().lower() in ('true', '1', 'yes', 'on')
        
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'})
        
        was_called = bool(order.customer_called)
        order.customer_called = called
        if called and not was_called:
            order.customer_called_at = datetime.utcnow()
            order.customer_called_by_id = session.get('employee_id')
        elif not called and was_called:
            order.customer_called_at = None
            order.customer_called_by_id = None
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/orders/<int:order_id>/toggle-verified', methods=['POST'])
@login_required
@permission_required('can_edit_orders')
def toggle_order_verified(order_id):
    try:
        data = request.get_json() or {}
        verified = str(data.get('verified', False)).strip().lower() in ('true', '1', 'yes', 'on')
        
        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'})
        
        was_verified = bool(order.customer_verified)
        order.customer_verified = verified
        if verified and not was_verified:
            order.customer_verified_at = datetime.utcnow()
            order.customer_verified_by_id = session.get('employee_id')
        elif not verified and was_verified:
            order.customer_verified_at = None
            order.customer_verified_by_id = None
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/replacement-orders/<int:order_id>/toggle-called', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def toggle_replacement_order_called(order_id):
    try:
        data = request.get_json() or {}
        called = str(data.get('called', False)).strip().lower() in ('true', '1', 'yes', 'on')
        
        order = ReplacementOrder.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'})
        
        was_called = bool(order.customer_called)
        order.customer_called = called
        if called and not was_called:
            order.customer_called_at = datetime.utcnow()
            order.customer_called_by_id = session.get('employee_id')
        elif not called and was_called:
            order.customer_called_at = None
            order.customer_called_by_id = None
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/replacement-orders/<int:order_id>/toggle-verified', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def toggle_replacement_order_verified(order_id):
    try:
        data = request.get_json() or {}
        verified = str(data.get('verified', False)).strip().lower() in ('true', '1', 'yes', 'on')
        
        order = ReplacementOrder.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'})
        
        was_verified = bool(order.customer_verified)
        order.customer_verified = verified
        if verified and not was_verified:
            order.customer_verified_at = datetime.utcnow()
            order.customer_verified_by_id = session.get('employee_id')
        elif not verified and was_verified:
            order.customer_verified_at = None
            order.customer_verified_by_id = None
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/orders/<int:order_id>/toggle-registered', methods=['POST'])
@login_required
@permission_required('can_edit_orders')
def toggle_order_registered(order_id):
    try:
        data = request.get_json() or {}
        registered = str(data.get('registered', False)).strip().lower() in ('true', '1', 'yes', 'on')

        order = Order.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'الطلب غير موجود'})

        was_registered = bool(order.registered)
        order.registered = registered
        if registered and not was_registered:
            order.registered_at = datetime.utcnow()
            order.registered_by_id = session.get('employee_id')
        elif not registered and was_registered:
            order.registered_at = None
            order.registered_by_id = None
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/replacement-orders/<int:order_id>/toggle-registered', methods=['POST'])
@login_required
@permission_required('can_edit_replacements')
def toggle_replacement_order_registered(order_id):
    try:
        data = request.get_json() or {}
        registered = str(data.get('registered', False)).strip().lower() in ('true', '1', 'yes', 'on')

        order = ReplacementOrder.query.get(order_id)
        if not order:
            return jsonify({'success': False, 'message': 'طلب الاستبدال غير موجود'})

        was_registered = bool(order.registered)
        order.registered = registered
        if registered and not was_registered:
            order.registered_at = datetime.utcnow()
            order.registered_by_id = session.get('employee_id')
        elif not registered and was_registered:
            order.registered_at = None
            order.registered_by_id = None
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'})

@main.route('/api/search-replacement-orders')
@login_required
@permission_required('can_view_replacements_by_state')
def api_search_replacement_orders():
    phone = request.args.get('phone', '').strip()
    if not phone:
        return jsonify({'success': False, 'message': 'رقم التليفون مطلوب'})
    
    orders = Order.query.filter_by(status='استبدال').join(
        Order.customer).filter(
        Customer.phone.ilike(f"%{phone}%")
    ).options(
        joinedload(Order.items).joinedload(OrderItem.variant),
        joinedload(Order.items).joinedload(OrderItem.size_variant),
        joinedload(Order.items).joinedload(OrderItem.color_variant),
        joinedload(Order.customer)
    ).order_by(Order.date.desc()).all()
    
    orders_data = []
    for order in orders:
        order_data = {
            'id': order.id,
            'date': order.date.isoformat(),
            'customer': {
                'id': order.customer.id if order.customer else None,
                'name': order.customer_name,
                'phone': order.customer_phone,
                'governorate': order.customer_governorate,
                'address_details': order.customer_address_details
            },

            'items': []
        }
        
        for item in order.items:
            if item.is_valid_product:
                item_data = {
                    'id': item.id,
                    'product': {
                        'id': item.product.id,
                        'name': item.product.name
                    },
                    'quantity': item.quantity,
                    'price': float(item.price),
                    'state': item.state,
                    'selected_variants': {}
                }
                
                if item.variant:
                    item_data['selected_variants']['النوع'] = {
                        'variant_name': item.variant.variant_name
                    }
                if item.size_variant:
                    item_data['selected_variants']['المقاس'] = {
                        'variant_name': item.size_variant.variant_name
                    }
                if item.color_variant:
                    item_data['selected_variants']['اللون'] = {
                        'variant_name': item.color_variant.variant_name
                    }
                
                order_data['items'].append(item_data)
        
        orders_data.append(order_data)
    
    return jsonify({'success': True, 'orders': orders_data})

@main.route('/replacements/shipped')
@login_required
@permission_required('can_view_replacements_by_state')
def replacements_shipped():
    phone = request.args.get('phone', '').strip()
    
    selected_year = request.args.get('year', datetime.now().year, type=int)
    selected_month = request.args.get('month', datetime.now().month, type=int)
    
    query = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'خرج للتوصيل'
    )
    
    if selected_year:
        query = query.filter(extract('year', ReplacementOrder.date) == selected_year)
    
    if selected_month:
        query = query.filter(extract('month', ReplacementOrder.date) == selected_month)
    
    if phone:
        query = query.join(
            Order.customer).filter(
            Customer.phone.ilike(
                f"%{phone}%"))
    
    orders = query.options(
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.variant),
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.size_variant),
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.color_variant),
        joinedload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.date.desc()).all()
    
    return render_template(
        'replacements_shipped.html',
        orders=orders,
        phone=phone,
        selected_year=selected_year,
        selected_month=selected_month)

@main.route('/customer/<int:customer_id>/orders')
@login_required
@permission_required('can_view_customer_orders')
def view_customer_orders(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    selected_status = request.args.get('status')
    selected_month = request.args.get('month', None, type=int)
    selected_year = request.args.get('year', None, type=int)
    
    orders_query = Order.query.filter(
        and_(
            Order.customer_id == customer.id,
            Order.status != 'مسودة استبدال'
        )
    )
    if selected_status:
        orders_query = orders_query.filter_by(status=selected_status)
    if selected_year:
        orders_query = orders_query.filter(extract('year', Order.date) == selected_year)
    if selected_month:
        orders_query = orders_query.filter(extract('month', Order.date) == selected_month)

    orders = orders_query.options(
        joinedload(Order.items).joinedload(OrderItem.variant),
        joinedload(Order.items).joinedload(OrderItem.size_variant),
        joinedload(Order.items).joinedload(OrderItem.color_variant),
        joinedload(Order.items).joinedload(OrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(Order.customer)
    ).all()

    replacements_query = (ReplacementOrder.query
        .outerjoin(Order, ReplacementOrder.original_order_id == Order.id)
        .filter(or_(
            ReplacementOrder.customer_id == customer.id,
            Order.customer_id == customer.id
        )))
    if selected_status in ('جديد', 'خرج للتوصيل', 'وصل'):
        replacements_query = replacements_query.filter(ReplacementOrder.status == selected_status)
    if selected_year:
        replacements_query = replacements_query.filter(extract('year', ReplacementOrder.date) == selected_year)
    if selected_month:
        replacements_query = replacements_query.filter(extract('month', ReplacementOrder.date) == selected_month)

    replacement_orders = replacements_query.options(
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.variant),
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.size_variant),
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem.color_variant),
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(ReplacementOrder.customer)
    ).distinct(ReplacementOrder.id).all()

    combined = orders + replacement_orders
    combined.sort(key=lambda o: o.date or datetime.min, reverse=True)
    
    for o in combined:
        o.valid_next_statuses = StatusService.valid_next(o.status)
        o.is_replacement_order = is_replacement(o)

    order_ids = [o.id for o in combined]
    returned_rows = db.session.query(
        ReturnOrder.original_order_id, ReturnOrder.original_replacement_order_id, ReturnOrderItem.product_id
    ).join(ReturnOrderItem).filter(
        or_(ReturnOrder.original_order_id.in_(order_ids), ReturnOrder.original_replacement_order_id.in_(order_ids)),
        ReturnOrderItem.return_order_id == ReturnOrder.id
    ).distinct().all()
    returned_by_order = {}
    for orig_order_id, orig_rep_id, pid in returned_rows:
        oid = orig_order_id or orig_rep_id
        if oid not in returned_by_order:
            returned_by_order[oid] = set()
        returned_by_order[oid].add(pid)

    return render_template(
        'customer_orders.html',
        customer=customer,
        orders=combined,
        selected_status=selected_status,
        selected_month=selected_month,
        selected_year=selected_year,
        all_statuses=ALL_STATUSES,
        returned_by_order=returned_by_order)

@main.route('/customers/<int:customer_id>/history')
@login_required
@permission_required('can_view_customer_logs')
def customer_history(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    logs = CustomerLog.query.options(
        joinedload(CustomerLog.employee),
        joinedload(CustomerLog.order)
    ).filter_by(
        customer_id=customer.id).order_by(
        CustomerLog.timestamp.desc()).all()
    return render_template(
        'customer_history.html',
        customer=customer,
        logs=logs)

@main.route('/customers/<int:customer_id>/add_log', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_customer_logs')
def add_customer_log(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    replacement_orders = ReplacementOrder.query.filter_by(
        customer_id=customer.id,
        is_draft=False
    ).order_by(ReplacementOrder.date.desc()).all()
    if request.method == 'POST':
        log_type = request.form.get('type', 'يدوي')
        if log_type == 'تحويل رسوم توصيل':
            replacement_order_id = request.form.get('replacement_order_id')
            transferred_fees = request.form.get('transferred_delivery_fees', 0, type=float)
            transfer_note = request.form.get('transfer_note', '').strip()

            if not replacement_order_id:
                flash('يجب اختيار طلب الاستبدال')
                return redirect(
                    url_for(
                        'main.add_customer_log',
                        customer_id=customer.id))

            if not transferred_fees or transferred_fees <= 0:
                flash('يجب إدخال قيمة رسوم محوّلة أكبر من صفر')
                return redirect(
                    url_for(
                        'main.add_customer_log',
                        customer_id=customer.id))

            replacement_order = ReplacementOrder.query.filter_by(
                id=int(replacement_order_id),
                customer_id=customer.id,
                is_draft=False
            ).first()

            if not replacement_order:
                flash('طلب الاستبدال غير موجود')
                return redirect(
                    url_for(
                        'main.add_customer_log',
                        customer_id=customer.id))

            replacement_order.delivery_fees = (replacement_order.delivery_fees or 0) + transferred_fees
            try:
                replacement_order.calculate_losses()
            except Exception:
                pass

            content = f"تحويل رسوم توصيل لطلب استبدال رقم {replacement_order.id} بقيمة {transferred_fees:.2f} ج.م"
            if transfer_note:
                content = f"{content} | ملاحظة: {transfer_note}"

            log = CustomerLog(
                customer_id=customer.id,
                employee_id=session.get('employee_id'),
                type=log_type,
                content=content)

            db.session.add(log)
            db.session.commit()
            flash('تمت إضافة السجل بنجاح')
            return redirect(
                url_for(
                    'main.customer_history',
                    customer_id=customer.id))

        content = request.form.get('content', '').strip()

        if not content:
            flash('يجب إدخال محتوى السجل')
            return redirect(
                url_for(
                    'main.add_customer_log',
                    customer_id=customer.id))

        log = CustomerLog(
            customer_id=customer.id,
            employee_id=session.get('employee_id'),
            type=log_type,
            content=content)

        if log_type == 'تنبيه بالمتابعة':
            reminder_duration = request.form.get('reminder_duration')
            reminder_duration_type = request.form.get('reminder_duration_type')

            if not reminder_duration or not reminder_duration_type:
                flash('يجب تحديد مدة التنبيه')
                return redirect(
                    url_for(
                        'main.add_customer_log',
                        customer_id=customer.id))

            try:
                duration_value = int(reminder_duration)
                if duration_value <= 0:
                    flash('يجب أن تكون مدة التنبيه أكبر من صفر')
                    return redirect(
                        url_for(
                            'main.add_customer_log',
                            customer_id=customer.id))

                from datetime import timedelta
                current_time = datetime.now()
                
                if reminder_duration_type == 'دقائق':
                    reminder_time = current_time + timedelta(minutes=duration_value)
                elif reminder_duration_type == 'ساعات':
                    reminder_time = current_time + timedelta(hours=duration_value)
                elif reminder_duration_type == 'أيام':
                    reminder_time = current_time + timedelta(days=duration_value)
                else:
                    flash('نوع المدة غير صحيح')
                    return redirect(
                        url_for(
                            'main.add_customer_log',
                            customer_id=customer.id))

                log.reminder_time = reminder_time
                log.reminder_duration = duration_value
                log.reminder_duration_type = reminder_duration_type
                log.follow_up_reason = content

            except ValueError:
                flash('مدة التنبيه غير صحيحة')
                return redirect(
                    url_for(
                        'main.add_customer_log',
                        customer_id=customer.id))

        db.session.add(log)

        db.session.commit()
        flash('تمت إضافة السجل بنجاح')

        return redirect(
            url_for(
                'main.customer_history',
                customer_id=customer.id))
    return render_template(
        'add_customer_log.html',
        customer=customer,
        replacement_orders=replacement_orders
    )

@main.route('/customer/<int:customer_id>/profile')
@login_required
@permission_required('can_view_customer_profile')
def customer_profile(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    return render_template('customer_profile.html', customer=customer)

@main.route('/api/customer_by_phone')
@login_required
@permission_required('can_view_customers')
def api_customer_by_phone():
    phone = request.args.get('phone', '').strip()
    customer = Customer.query.filter(Customer.phone == phone).first()
    if customer:
        return jsonify({
            'found': True,
            'name': customer.name,
            'governorate': customer.governorate,
            'address_details': customer.address_details
        })
    else:
        return jsonify({'found': False})

@main.route('/api/supplier_by_phone')
@login_required
@permission_required('can_view_suppliers')
def api_supplier_by_phone():
    phone = request.args.get('phone', '').strip()
    if not phone:
        return jsonify({'success': False, 'message': 'رقم التليفون مطلوب'})
    
    supplier = Supplier.query.filter_by(phone=phone).first()
    if supplier:
        return jsonify({
            'success': True,
            'supplier': {
                'id': supplier.id,
                'name': supplier.name,
                'phone': supplier.phone,
                'address': supplier.address,
                'email': supplier.email
            }
        })
    else:
        return jsonify({'success': False, 'message': 'لم يتم العثور على المورد'})

@main.route('/api/search_suppliers')
@login_required
@permission_required('can_view_suppliers')
def api_search_suppliers():
    term = request.args.get('q', '').strip()
    query = Supplier.query
    if term:
        like = f"%{term}%"
        query = query.filter(Supplier.name.ilike(like))
    suppliers = query.order_by(Supplier.name.asc()).all()
    return jsonify({
        'results': [
            {
                'id': s.id,
                'name': s.name,
                'phone': getattr(s, 'phone', None),
                'address': getattr(s, 'address', None)
            } for s in suppliers
        ]
    })

@main.route('/api/search_products')
@login_required
@permission_required('can_view_products')
def api_search_products():
    term = request.args.get('term', '').strip()
    if not term:
        return jsonify({'success': False, 'message': 'مصطلح البحث مطلوب'})
    
    products = Product.query.filter(
        Product.name.ilike(f'%{term}%'),
        Product.is_deleted == False,
        Product.is_bundle == False
    ).limit(10).all()
    
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'purchase_price': product.purchase_price,
            'wholesale_price': product.wholesale_price,
            'stock': product.stock
        })
    
    return jsonify({
        'success': True,
        'products': products_data
    })

@main.route('/api/products')
@login_required
@permission_required('can_view_products')
def api_get_products():
    products = Product.query.filter(
        Product.is_deleted == False,
        Product.is_bundle == False
    ).all()
    
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'purchase_price': product.purchase_price,
            'wholesale_price': product.wholesale_price,
            'stock': product.stock,
            'is_bundle': product.is_bundle,
            'is_deleted': product.is_deleted
        })
    
    return jsonify(products_data)

@main.route('/api/products/<int:product_id>/orders')
@login_required
@permission_required('can_view_orders')
def api_product_orders(product_id):
    product = Product.query.get_or_404(product_id)
    variant_ids = [v.id for v in product.variants]

    filters = [OrderItem.product_id == product_id]
    if variant_ids:
        filters += [
            OrderItem.variant_id.in_(variant_ids),
            OrderItem.size_variant_id.in_(variant_ids),
            OrderItem.color_variant_id.in_(variant_ids),
            OrderItem.style_variant_id.in_(variant_ids),
        ]

    items = OrderItem.query.options(
        joinedload(OrderItem.order).joinedload(Order.customer)
    ).filter(or_(*filters)).order_by(OrderItem.id.desc()).limit(200).all()

    return jsonify({
        'success': True,
        'product_name': product.name,
        'orders': [{
            'order_id': item.order_id,
            'customer_name': item.order.customer.name if item.order.customer else 'عميل أرضية',
            'status': item.order.status,
            'date': item.order.date.strftime('%Y-%m-%d %H:%M'),
            'quantity': item.quantity,
            'price': item.price,
        } for item in items]
    })

@main.route('/orders/replacement', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_replacements')
def add_replacement_order():
    if request.method == 'GET':
        customer_phone = request.args.get('customer_phone', '')
        selected_order_id = request.args.get('selected_order')
        selected_replacement_order_id = request.args.get('selected_replacement_order')
        draft_id = request.args.get('draft_id')

        customer = None
        customer_orders = []
        customer_replacement_orders = []
        selected_order = None
        selected_replacement_order = None
        draft = None
        draft_data = None

        # Check if resuming a draft
        if draft_id:
            try:
                draft = ReplacementOrder.query.filter_by(id=int(draft_id), is_draft=True).first()
                if draft:
                    customer = draft.customer
                    if customer:
                        customer_phone = customer.phone
                        customer_orders = Order.query.filter_by(customer_id=customer.id).order_by(Order.date.desc()).all()
                        customer_replacement_orders = ReplacementOrder.query.filter_by(
                            customer_id=customer.id, is_draft=False
                        ).order_by(ReplacementOrder.date.desc()).all()
                    
                    # Parse draft_data
                    if draft.draft_data:
                        try:
                            draft_data = json.loads(draft.draft_data)
                            if draft_data.get('orderId'):
                                selected_order = Order.query.get(int(draft_data['orderId']))
                        except:
                            pass
            except Exception as e:
                current_app.logger.error(f'Error loading draft: {e}')
                flash('حدث خطأ أثناء تحميل المسودة', 'error')

        products = Product.query.options(
            joinedload(Product.variants),
            joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
        ).filter_by(is_deleted=False).all()
        def serialize_product(product):
            result = {
                'id': product.id,
                'name': product.name,
                'price': product.price,
                'purchase_price': product.purchase_price,
                'wholesale_price': product.wholesale_price,
                'stock': product.stock,
                'has_size': product.has_size,
                'has_color': product.has_color,
                'has_style': product.has_style,
                'is_bundle': getattr(product, 'is_bundle', False),
                'variants': [
                    {
                        'id': v.id,
                        'group_name': v.group_name,
                        'variant_name': v.variant_name,
                        'price': v.price
                    }
                    for v in product.variants
                ]
            }
            
            if result['is_bundle'] and hasattr(product, 'bundle_items'):
                result['bundle_items'] = []
                for bundle_item in product.bundle_items:
                    if bundle_item.product:
                        result['bundle_items'].append({
                            'product_id': bundle_item.product.id,
                            'product_name': bundle_item.product.name,
                            'sale_price_in_bundle': bundle_item.sale_price_in_bundle,
                            'product_variants': [
                                {
                                    'id': v.id,
                                    'group_name': v.group_name,
                                    'variant_name': v.variant_name,
                                    'price': v.price
                                }
                                for v in bundle_item.product.variants
                            ] if bundle_item.product.variants else []
                        })
            
            return result
        products_data = [serialize_product(p) for p in products]

        if customer_phone and not draft_id:
            customer = Customer.query.filter_by(phone=customer_phone).first()

            if customer:
                customer_orders = Order.query.filter_by(customer_id=customer.id).order_by(Order.date.desc()).all()
                customer_replacement_orders = ReplacementOrder.query.filter_by(
                    customer_id=customer.id, is_draft=False
                ).order_by(ReplacementOrder.date.desc()).all()

                if selected_order_id:
                    selected_order = Order.query.get(selected_order_id)

                if selected_replacement_order_id:
                    selected_replacement_order = ReplacementOrder.query.get(selected_replacement_order_id)

        return render_template('add_replacement_order.html',
                             customer=customer,
                             customer_orders=customer_orders,
                             customer_replacement_orders=customer_replacement_orders,
                             selected_order=selected_order,
                             selected_replacement_order=selected_replacement_order,
                             customer_phone=customer_phone,
                             products=products_data,
                             draft=draft,
                             draft_data=draft_data)

    customer_id = request.form.get('customer_id')
    customer_name = request.form.get('customer_name')
    customer_phone = request.form.get('customer_phone')
    customer_address = request.form.get('customer_address')
    update_customer_data = request.form.get('update_customer_data', '0')
    customer_governorate = None
    customer_address_details = None
    if customer_address:
        parts = customer_address.split('-', 1)
        customer_governorate = parts[0].strip() if len(parts) > 0 else ''
        customer_address_details = parts[1].strip() if len(parts) > 1 else ''
    original_order_id = request.form.get('original_order_id')
    original_replacement_order_id = request.form.get('original_replacement_order_id')
    delivery_fees_me = request.form.get('delivery_fees_me', 0, type=float)
    delivery_fees_customer = request.form.get('delivery_fees_customer', 0, type=float)
    customer_refund = request.form.get('customer_refund', 0, type=float)
    amount_paid = request.form.get('amount_paid', 0, type=float)
    order_total = request.form.get('order_total', 0, type=float)
    order_remaining = request.form.get('order_remaining', 0, type=float)
    tracking_number = request.form.get('tracking_number', '')
    notes = request.form.get('notes', '')

    if not customer_id:
        flash('يجب اختيار العميل', 'error')
        return redirect(request.url)

    try:
        replacement_order = ReplacementOrder(
            customer_id=int(customer_id),
            original_order_id=int(original_order_id) if original_order_id else None,
            original_replacement_order_id=int(original_replacement_order_id) if original_replacement_order_id else None,
            date=datetime.now(),
            status='جديد',
            delivery_fees=delivery_fees_me,
            delivery_fees_customer=delivery_fees_customer,
            total_amount=order_total,
            amount_paid=amount_paid,
            customer_refund_amount=customer_refund,
            tracking_number=tracking_number or None,
            notes=notes,
            alternative_name=customer_name if update_customer_data == '1' else None,
            alternative_phone=customer_phone if update_customer_data == '1' else None,
            alternative_governorate=customer_governorate if update_customer_data == '1' else None,
            alternative_address_details=customer_address_details if update_customer_data == '1' else None,
            weight=request.form.get('weight', type=float),
            package_volume=request.form.get('package_volume') or None,
            delivery_notes=request.form.get('delivery_notes') or None,
        )
        db.session.add(replacement_order)
        db.session.flush()

        if replacement_order.original_order_id:
            original_order = Order.query.get(replacement_order.original_order_id)
            if original_order and original_order.status != 'استبدال':
                old_status = original_order.status
                original_order.status = 'استبدال'
                try:
                    db.session.flush()
                    original_order.add_status_history(
                        'استبدال',
                        employee_id=session.get('employee_id'),
                        notes=f'تم تعيين الطلب كـ استبدال بسبب إنشاء طلب استبدال رقم {replacement_order.id}'
                    )
                except Exception as e:
                    current_app.logger.error(f"Failed to add status history: {e}")
                try:
                    log_activity(
                        action='update',
                        entity_type='order',
                        entity_id=original_order.id,
                        entity_name=f'طلب رقم {original_order.id}',
                        details=f'تم تعيين الطلب كـ استبدال بسبب إنشاء طلب استبدال رقم {replacement_order.id}'
                    )
                except Exception:
                    pass

        # Handle original replacement order - change its status to 'استبدال'
        if replacement_order.original_replacement_order_id:
            orig_replacement = ReplacementOrder.query.get(replacement_order.original_replacement_order_id)
            if orig_replacement and orig_replacement.status != 'استبدال':
                orig_replacement.status = 'استبدال'
                try:
                    db.session.flush()
                    orig_replacement.add_status_history(
                        'استبدال',
                        employee_id=session.get('employee_id'),
                        notes=f'تم تعيين الطلب كـ استبدال بسبب إنشاء طلب استبدال رقم {replacement_order.id}'
                    )
                except Exception as e:
                    current_app.logger.error(f"Failed to add replacement status history: {e}")
                try:
                    log_activity(
                        action='update',
                        entity_type='replacement_order',
                        entity_id=orig_replacement.id,
                        entity_name=f'طلب استبدال رقم {orig_replacement.id}',
                        details=f'تم تعيين الطلب كـ استبدال بسبب إنشاء طلب استبدال رقم {replacement_order.id}'
                    )
                except Exception:
                    pass

        final_products_data = request.form.get('final_products_data')
        if final_products_data:
            try:
                final_products = json.loads(final_products_data)
                for prod in final_products:
                    state_val = prod.get('state', 'سليم')
                    is_damaged = prod.get('is_damaged', False)
                    is_returned = prod.get('is_returned', False)
                    
                    product = _get_locked_product(prod.get('product_id'))
                    if product and product.is_bundle:
                        bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                        for bundle_item in bundle_items:
                            if bundle_item.product:
                                item_purchase_price = bundle_item.product.purchase_price or 0
                                
                                item_damage_loss = (item_purchase_price * prod.get('quantity', 1)) if is_damaged else 0
                                
                                replacement_item = ReplacementOrderItem(
                                    replacement_order_id=replacement_order.id,
                                    product_id=bundle_item.product_id,
                                    variant_id=None,
                                    size_variant_id=None,
                                    color_variant_id=None,
                                    style_variant_id=None,
                                    quantity=prod.get('quantity', 1),
                                    price=bundle_item.sale_price_in_bundle,
                                    state=state_val,
                                    is_damaged=is_damaged,
                                    is_returned=is_returned,
                                    purchase_price=item_purchase_price,
                                    damage_loss=item_damage_loss
                                )
                                db.session.add(replacement_item)
                    else:
                        purchase_price = prod.get('purchase_price', 0)
                        if (is_damaged or is_returned) and not purchase_price:
                            try:
                                if product:
                                    purchase_price = product.purchase_price or 0
                            except Exception:
                                purchase_price = 0
                        
                        damage_loss = (purchase_price * prod.get('quantity', 1)) if is_damaged else 0
                        
                        replacement_item = ReplacementOrderItem(
                            replacement_order_id=replacement_order.id,
                            product_id=prod.get('product_id'),
                            variant_id=prod.get('variant_id'),
                            size_variant_id=prod.get('size_variant_id'),
                            color_variant_id=prod.get('color_variant_id'),
                            style_variant_id=prod.get('style_variant_id'),
                            quantity=prod.get('quantity', 1),
                            price=prod.get('price', 0),
                            state=state_val,
                            is_damaged=is_damaged,
                            is_returned=is_returned,
                            purchase_price=purchase_price,
                            damage_loss=damage_loss
                        )
                        db.session.add(replacement_item)
            except Exception as e:
                flash(f'خطأ في معالجة المنتجات النهائية: {str(e)}', 'warning')

        # Flush items to DB and expire the cached (possibly empty) items collection
        # so that replacement_order.items is reloaded from DB on next access
        db.session.flush()
        db.session.expire(replacement_order, ['items'])

        try:
            for it in replacement_order.items:
                is_dmg = bool(it.is_damaged) or (it.state == 'تالف')
                it.is_damaged = is_dmg
                if is_dmg:
                    try:
                        it.damage_loss = (it.purchase_price or 0) * (it.quantity or 1)
                    except Exception:
                        pass
                else:
                    it.damage_loss = 0

            try:
                new_products_total = 0.0
                for it in replacement_order.items:
                    if it.state in ['منتج جديد', 'جديد']:
                        new_products_total += (it.price or 0.0) * (it.quantity or 1)
                replacement_order.total_amount = float(new_products_total) + float(replacement_order.delivery_fees_customer or 0.0)
            except Exception:
                pass

            try:
                replacement_order.calculate_losses()
            except Exception:
                pass
        except Exception:
            pass

        apply_cod_fee_to_replacement(replacement_order)

        try:
            replacement_order.add_status_history('جديد', employee_id=session.get('employee_id'))
        except Exception as _e:
            current_app.logger.error(f"Failed to add status history on replacement order creation: {_e}")

        # زيادة المخزون للمنتجات المرتجعة عند حفظ طلب الاستبدال
        try:
            ret_log = {}
            for it in replacement_order.items:
                if it.is_returned and it.product_id:
                    quantity_to_return = it.quantity or 0
                    if quantity_to_return > 0:
                        product = _get_locked_product(it.product_id)
                        if product:
                            if product.is_bundle:
                                bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                                for bundle_item in bundle_items:
                                    bundle_product = _get_locked_product(bundle_item.product_id)
                                    if bundle_product:
                                        bundle_product.stock = (bundle_product.stock or 0) + (quantity_to_return * 1)
                                        ret_log[bundle_product.id] = ret_log.get(bundle_product.id, 0) + (quantity_to_return * 1)
                            else:
                                product.stock = (product.stock or 0) + quantity_to_return
                                ret_log[product.id] = ret_log.get(product.id, 0) + quantity_to_return
            if ret_log:
                _log_stock_delta("replacement_order created returned-items restock", ret_log)
        except Exception as e:
            current_app.logger.error(f'Error returning stock for returned products on creation: {str(e)}')

        db.session.commit()

        log_activity(
            action='create',
            entity_type='replacement_order',
            entity_id=replacement_order.id,
            entity_name=f'طلب استبدال رقم {replacement_order.id}',
            details=f'تم إنشاء طلب استبدال جديد للعميل: {replacement_order.customer.name}'
        )

        flash('تم إنشاء طلب الاستبدال بنجاح', 'success')
        return redirect(url_for('main.replacement_orders'))

    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء إنشاء طلب الاستبدال: {str(e)}', 'error')
        return redirect(request.url)

@main.route('/replacement-order/save-draft', methods=['POST'])
@login_required
@permission_required('can_add_replacements')
def save_replacement_draft():
    """Save or update a replacement order draft"""
    try:
        draft_id = request.form.get('draft_id')
        customer_id = request.form.get('customer_id')
        draft_step = request.form.get('draft_step')  # 'product_condition', 'new_products', 'review'
        draft_data_json = request.form.get('draft_data')
        
        if not customer_id or not draft_step or not draft_data_json:
            return jsonify({'success': False, 'message': 'بيانات غير مكتملة'}), 400
        
        # Parse draft data
        try:
            draft_data = json.loads(draft_data_json)
        except:
            return jsonify({'success': False, 'message': 'خطأ في تنسيق البيانات'}), 400
        
        if draft_id:
            # Update existing draft
            draft = ReplacementOrder.query.filter_by(id=draft_id, is_draft=True).first()
            if not draft:
                return jsonify({'success': False, 'message': 'المسودة غير موجودة'}), 404
            
            draft.draft_step = draft_step
            draft.draft_data = draft_data_json
            draft.date = datetime.now()
        else:
            # Create new draft
            draft = ReplacementOrder(
                customer_id=int(customer_id),
                date=datetime.now(),
                status='جديد',
                is_draft=True,
                draft_step=draft_step,
                draft_data=draft_data_json,
                delivery_fees=0,
                total_amount=0,
                amount_paid=0
            )
            db.session.add(draft)
        
        db.session.commit()
        
        log_activity(
            action='create' if not draft_id else 'update',
            entity_type='replacement_draft',
            entity_id=draft.id,
            entity_name=f'مسودة طلب استبدال رقم {draft.id}',
            details=f'تم حفظ المسودة في خطوة: {draft_step}'
        )
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ المسودة بنجاح',
            'draft_id': draft.id
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error saving draft: {str(e)}')
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

@main.route('/replacement-order/draft/<int:draft_id>')
@login_required
@permission_required('can_add_replacements')
def get_replacement_draft(draft_id):
    """Retrieve a draft to resume editing"""
    draft = ReplacementOrder.query.filter_by(id=draft_id, is_draft=True).first_or_404()
    
    try:
        draft_data = json.loads(draft.draft_data) if draft.draft_data else {}
    except:
        draft_data = {}
    
    return jsonify({
        'id': draft.id,
        'customer_id': draft.customer_id,
        'draft_step': draft.draft_step,
        'draft_data': draft_data,
        'date': draft.date.isoformat() if draft.date else None
    })

@main.route('/replacement-order/draft/<int:draft_id>/delete', methods=['POST'])
@login_required
@permission_required('can_add_replacements')
def delete_replacement_draft(draft_id):
    """Delete a draft"""
    try:
        draft = ReplacementOrder.query.filter_by(id=draft_id, is_draft=True).first_or_404()
        
        log_activity(
            action='delete',
            entity_type='replacement_draft',
            entity_id=draft.id,
            entity_name=f'مسودة طلب استبدال رقم {draft.id}',
            details='تم حذف المسودة'
        )
        
        db.session.delete(draft)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم حذف المسودة بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

@main.route('/replacement-order/draft/<int:draft_id>/convert', methods=['POST'])
@login_required
@permission_required('can_add_replacements')
def convert_draft_to_order(draft_id):
    """Convert a draft to a real order"""
    try:
        draft = ReplacementOrder.query.filter_by(id=draft_id, is_draft=True).first_or_404()
        
        # Mark as non-draft
        draft.is_draft = False
        draft.draft_step = None
        draft.draft_data = None
        draft.date = datetime.now()
        
        db.session.commit()
        
        log_activity(
            action='convert',
            entity_type='replacement_order',
            entity_id=draft.id,
            entity_name=f'طلب استبدال رقم {draft.id}',
            details='تم تحويل المسودة إلى طلب فعلي'
        )
        
        return jsonify({'success': True, 'message': 'تم إنشاء الطلب بنجاح', 'order_id': draft.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

@main.route('/orders/replacement/<int:order_id>', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_replacements')
def edit_replacement_order(order_id):
    next_url = request.args.get('next')

    replacement_order = ReplacementOrder.query.options(
        joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product),
        joinedload(ReplacementOrder.customer),
        joinedload(ReplacementOrder.registered_by),
        joinedload(ReplacementOrder.verified_by),
        joinedload(ReplacementOrder.called_by)
    ).get_or_404(order_id)

    for item in replacement_order.items:
        if item.size_variant_id:
            item.size_variant
        if item.color_variant_id:
            item.color_variant
        if item.variant_id:
            item.variant

    products_data = []
    items_data = []

    if request.method == 'GET':
        # ===== serialize products and items for JS =====
        products = Product.query.options(
            joinedload(Product.variants),
            joinedload(Product.bundle_items).joinedload(BundleItem.product).joinedload(Product.variants)
        ).filter_by(is_deleted=False).all()

        def serialize_product(product):
            result = {
                'id': product.id,
                'name': (product.name + ' (محذوف)') if getattr(product, 'is_deleted', False) else product.name,
                'price': product.price,
                'stock': product.stock,
                'wholesale_price': getattr(product, 'wholesale_price', 0),
                'is_bundle': getattr(product, 'is_bundle', False),
                'variants': [
                    {'id': v.id, 'group_name': v.group_name, 'variant_name': v.variant_name, 'price': v.price}
                    for v in product.variants
                ]
            }
            return result
        products_data = [serialize_product(p) for p in products]

        items_data = []
        for item in replacement_order.items:
            if item.state not in ['سليم', 'منتج جديد', 'جديد']:
                continue
            prod_id = item.product_id or None
            if not prod_id:
                if item.size_variant:
                    prod_id = item.size_variant.product_id
                elif item.color_variant:
                    prod_id = item.color_variant.product_id
                elif item.style_variant:
                    prod_id = item.style_variant.product_id
                elif item.variant:
                    prod_id = item.variant.product_id

            product_name = None
            if prod_id:
                try:
                    prod_obj = Product.query.filter_by(id=prod_id).first()
                    if prod_obj:
                        product_name = (prod_obj.name + ' (محذوف)') if getattr(prod_obj, 'is_deleted', False) else prod_obj.name
                except Exception:
                    pass

            items_data.append({
                'id': item.id,
                'product_id': prod_id,
                'product_name': product_name,
                'variant_id': item.variant_id,
                'size_variant_id': item.size_variant_id,
                'color_variant_id': item.color_variant_id,
                'style_variant_id': item.style_variant_id,
                'quantity': item.quantity,
                'price': item.price,
                'state': item.state,
            })

    if request.method == 'POST':
        # لقطة الحالة القديمة قبل التعديل
        _r_old_status_snap = replacement_order.status
        _r_old_notes_snap = replacement_order.notes or ''
        _r_old_amount_paid_snap = replacement_order.amount_paid
        _r_old_delivery_fees_snap = replacement_order.delivery_fees
        _r_old_tracking_snap = replacement_order.tracking_number or ''
        _r_old_items_snap = []
        for _it in list(replacement_order.items):
            _pname_r = getattr(_it.product, 'name', None) if _it.product else None
            _r_old_items_snap.append((_pname_r or f'منتج#{_it.product_id}', _it.quantity, _it.state or ''))

        delivery_fees = request.form.get('delivery_fees', 0, type=float)
        amount_paid = request.form.get('amount_paid', 0, type=float)
        status = request.form.get('status', replacement_order.status)
        tracking_number = request.form.get('tracking_number', '')
        delivery_fees_customer = request.form.get('delivery_fees_customer', 0, type=float)
        notes = request.form.get('notes', '')

        # customer fields
        customer_name = request.form.get('customer_name', '').strip()
        customer_phone = request.form.get('customer_phone', '').strip()
        customer_governorate = request.form.get('customer_governorate', '').strip()
        customer_address_details = request.form.get('customer_address_details', '').strip()
        update_customer_profile = request.form.get('update_customer_profile', '0')

        replacement_order.alternative_name = customer_name or None
        replacement_order.alternative_phone = customer_phone or None
        replacement_order.alternative_governorate = customer_governorate or None
        replacement_order.alternative_address_details = customer_address_details or None

        if update_customer_profile == 'order_and_customer' and replacement_order.customer:
            cust = replacement_order.customer
            if customer_name and customer_name != cust.name:
                cust.name = customer_name
            if customer_phone and customer_phone != cust.phone:
                cust.phone = customer_phone
            if customer_governorate and customer_governorate != cust.governorate:
                cust.governorate = customer_governorate
            if customer_address_details and customer_address_details != cust.address_details:
                cust.address_details = customer_address_details

        replacement_order.delivery_fees = delivery_fees
        replacement_order.amount_paid = amount_paid
        replacement_order.delivery_fees_customer = delivery_fees_customer
        old_status = replacement_order.status
        if status:
            try:
                validate_transition(old_status, status)
            except ValueError as e:
                flash(str(e), 'error')
                db.session.rollback()
                return redirect(request.url)
            if status == 'خرج للتوصيل' and not getattr(replacement_order, 'inventory_deducted', False):
                required = _collect_replacement_required_quantities(replacement_order.items)
                if required:
                    pmap = _get_locked_products(required.keys())
                    for p in pmap.values():
                        if _block_if_deleted(p, 'طلب الاستبدال', 'deduct'):
                            db.session.rollback()
                            flash(f'لا يمكن تغيير الحالة: طلب الاستبدال يحتوي على منتج محذوف ("{p.name}")', 'error')
                            return redirect(request.url)
                    insufficient = []
                    for pid, q in required.items():
                        p = pmap.get(pid)
                        if (p.stock or 0) < q:
                            insufficient.append(f"{p.name} (المطلوب {q} المتوفر {p.stock})")
                    if insufficient:
                        flash(f'لا يمكن التحويل إلى {status} لعدم كفاية المخزون: ' + ' ؛ '.join(insufficient), 'error')
                        return redirect(request.url)
                    for pid, q in required.items():
                        p = pmap.get(pid)
                        if p:
                            p.stock = (p.stock or 0) - q
                    replacement_order.inventory_deducted = True
            if status in ['إلغاء', 'رفض الاستلام'] and getattr(replacement_order, 'inventory_deducted', False):
                try:
                    returns = _collect_replacement_required_quantities(replacement_order.items)
                    if returns:
                        for pid, q in returns.items():
                            product = _get_locked_product(pid)
                            if product:
                                product.stock = (product.stock or 0) + q
                except Exception:
                    pass
                replacement_order.inventory_deducted = False
            old_status_for_history = replacement_order.status
            replacement_order.status = status
            replacement_order.status_updated_at = datetime.utcnow()
            
            status_changed_to = status
            status_changed_from = old_status_for_history
        else:
            status_changed_to = None
            status_changed_from = None

        if (replacement_order.status or '') == 'خرج للتوصيل':
            replacement_order.tracking_number = tracking_number or None
        else:
            replacement_order.tracking_number = None
        replacement_order.notes = notes
        replacement_order.weight         = request.form.get('weight', type=float)
        replacement_order.package_volume = request.form.get('package_volume') or None
        replacement_order.delivery_notes = request.form.get('delivery_notes') or None
        try:
            replacement_order.is_urgent = bool(request.form.get('is_urgent'))
        except Exception:
            replacement_order.is_urgent = False

        # معالجة المنتجات من الحقل المخفي order_items_data
        items_data_json = request.form.get('order_items_data')
        if items_data_json:
            try:
                items_data = json.loads(items_data_json)
            except Exception:
                items_data = []

            existing_items = {it.id: it for it in replacement_order.items}
            seen_ids = set()

            product_ids = {
                int(row['product_id'])
                for row in items_data
                if row.get('product_id')
            }
            products_map = _get_locked_products(product_ids) if product_ids else {}

            for row in items_data:
                row_id_raw = row.get('id')
                row_id = int(row_id_raw) if row_id_raw is not None else None
                product_id = row.get('product_id')
                variant_id = row.get('variant_id')
                size_variant_id = row.get('size_variant_id')
                color_variant_id = row.get('color_variant_id')
                style_variant_id = row.get('style_variant_id')
                state = row.get('state', 'منتج جديد')

                if row_id and row_id in existing_items:
                    it = existing_items[row_id]
                    it.variant_id = None
                    it.size_variant_id = None
                    it.color_variant_id = None
                    it.style_variant_id = None
                    if it.state == 'سليم':
                        # سليم: ONLY variants — quantity, price, product_id NEVER from request
                        if variant_id:
                            it.variant_id = variant_id
                        if size_variant_id:
                            it.size_variant_id = size_variant_id
                        if color_variant_id:
                            it.color_variant_id = color_variant_id
                        if style_variant_id:
                            it.style_variant_id = style_variant_id
                    else:
                        # منتج جديد/جديد: full update
                        quantity = int(row.get('quantity') or 1)
                        price = float(row.get('price') or 0)
                        it.product_id = product_id
                        if variant_id:
                            it.variant_id = variant_id
                        if size_variant_id:
                            it.size_variant_id = size_variant_id
                        if color_variant_id:
                            it.color_variant_id = color_variant_id
                        if style_variant_id:
                            it.style_variant_id = style_variant_id
                        it.quantity = quantity
                        it.price = price
                    seen_ids.add(row_id)
                else:
                    # New item
                    quantity = int(row.get('quantity') or 1)
                    price = float(row.get('price') or 0)
                    purchase_price = 0.0
                    if product_id and products_map.get(product_id):
                        purchase_price = products_map[product_id].purchase_price or 0
                    it = ReplacementOrderItem(
                        replacement_order_id=replacement_order.id,
                        product_id=product_id,
                        variant_id=variant_id,
                        size_variant_id=size_variant_id,
                        color_variant_id=color_variant_id,
                        style_variant_id=style_variant_id,
                        quantity=quantity,
                        price=price,
                        state=state,
                        is_damaged=False,
                        is_returned=False,
                        purchase_price=purchase_price,
                        damage_loss=0.0
                    )
                    db.session.add(it)
                    db.session.flush()
                    seen_ids.add(it.id)

            # Delete removed items (only non-سليم items)
            to_delete = [it for it in replacement_order.items if it.id not in seen_ids and it.state != 'سليم']
            for it in to_delete:
                db.session.delete(it)

            # Recalculate total_amount
            try:
                new_products_total = 0.0
                for it in replacement_order.items:
                    if it.state in ['منتج جديد', 'جديد']:
                        new_products_total += (it.price or 0.0) * (it.quantity or 1)
                replacement_order.total_amount = float(new_products_total) + float(delivery_fees_customer)
            except Exception:
                pass

            # Stock delta when inventory_deducted
            try:
                if getattr(replacement_order, 'inventory_deducted', False):
                    # after_qty from submitted data
                    after_qty = defaultdict(int)
                    for row in items_data:
                        _st = (row.get('state') or '').strip()
                        if _st not in ['منتج جديد', 'جديد']:
                            continue
                        _pid = row.get('product_id')
                        if not _pid:
                            continue
                        _p = _get_locked_product(_pid)
                        if _p and _p.is_bundle:
                            _bis = BundleItem.query.filter_by(bundle_id=_pid).all()
                            for _bi in _bis:
                                after_qty[_bi.product_id] += int(row.get('quantity') or 0)
                        else:
                            after_qty[_pid] += int(row.get('quantity') or 0)

                    # compute from currently saved items (already updated)
                    current_qty = defaultdict(int)
                    for _it in replacement_order.items:
                        _st = (_it.state or '').strip()
                        if _st not in ['منتج جديد', 'جديد']:
                            continue
                        _pid = _get_replacement_item_product_id(_it)
                        if not _pid:
                            continue
                        _p = _get_locked_product(_pid)
                        if _p and _p.is_bundle:
                            _bis = BundleItem.query.filter_by(bundle_id=_pid).all()
                            for _bi in _bis:
                                current_qty[_bi.product_id] += int(_it.quantity or 0)
                        else:
                            current_qty[_pid] += int(_it.quantity or 0)

                    all_pids = set(current_qty.keys()) | set(after_qty.keys())
                    for _pid in all_pids:
                        _cur = int(current_qty.get(_pid, 0))
                        _new = int(after_qty.get(_pid, 0))
                        _delta = _new - _cur
                        if _delta > 0:
                            _pp = _get_locked_product(_pid)
                            if _pp:
                                if (_pp.stock or 0) < _delta:
                                    flash(f'المخزون غير كافٍ للمنتج {_pp.name}', 'warning')
                                _pp.stock = (_pp.stock or 0) - _delta
                        elif _delta < 0:
                            _pp = _get_locked_product(_pid)
                            if _pp:
                                _pp.stock = (_pp.stock or 0) + (-_delta)
            except Exception as e:
                current_app.logger.error(f"Replacement stock delta error: {e}")
        
        try:
            try:
                replacement_order.calculate_losses()
            except Exception:
                pass

            apply_cod_fee_to_replacement(replacement_order)

            # CustomerLog for customer data changes
            try:
                _cust_changes = []
                if replacement_order.alternative_name:
                    _cust_changes.append(f"الاسم: {replacement_order.alternative_name}")
                if replacement_order.alternative_phone:
                    _cust_changes.append(f"الهاتف: {replacement_order.alternative_phone}")
                if replacement_order.alternative_governorate:
                    _cust_changes.append(f"المحافظة: {replacement_order.alternative_governorate}")
                if replacement_order.alternative_address_details:
                    _cust_changes.append(f"العنوان: {replacement_order.alternative_address_details}")
                if _cust_changes and replacement_order.customer:
                    _scope = "في الطلب وملف العميل" if update_customer_profile == 'order_and_customer' else "في الطلب فقط"
                    db.session.add(CustomerLog(
                        customer_id=replacement_order.customer.id,
                        employee_id=session.get('employee_id'),
                        type="تلقائي",
                        content=f"تم تحديث بيانات العميل {_scope}: " + " | ".join(_cust_changes)
                    ))
            except Exception:
                pass

            db.session.commit()
            
            if status_changed_to and status_changed_from != status_changed_to:
                try:
                    replacement_order.add_status_history(
                        status_changed_to,
                        employee_id=session.get('employee_id'),
                        notes=f'تم تغيير الحالة من {status_changed_from}'
                    )
                    db.session.commit()
                except Exception as e:
                    current_app.logger.error(f"Failed to add replacement status history: {e}")
            
            log_activity(
                action='update',
                entity_type='replacement_order',
                entity_id=replacement_order.id,
                entity_name=f'طلب استبدال رقم {replacement_order.id}',
                details=f'تم تحديث طلب الاستبدال للعميل: {replacement_order.customer.name}'
            )
            
            # تسجيل التعديل في سجل التعديلات مع تفاصيل التغييرات
            try:
                _r_edit_details = []
                if status_changed_to and status_changed_from != status_changed_to:
                    _r_edit_details.append(f"الحالة: {status_changed_from} ← {status_changed_to}")
                if _r_old_notes_snap != (replacement_order.notes or ''):
                    _r_old_n = (_r_old_notes_snap[:30] + '...') if len(_r_old_notes_snap) > 30 else _r_old_notes_snap
                    _r_new_n = ((replacement_order.notes or '')[:30] + '...') if len(replacement_order.notes or '') > 30 else (replacement_order.notes or '')
                    _r_edit_details.append(f"الملاحظات: '{_r_old_n}' ← '{_r_new_n}'" if (_r_old_n or _r_new_n) else "الملاحظات: تم التعديل")
                if _r_old_amount_paid_snap != replacement_order.amount_paid:
                    _r_edit_details.append(f"المبلغ المدفوع: {_r_old_amount_paid_snap} ← {replacement_order.amount_paid}")
                if _r_old_delivery_fees_snap != replacement_order.delivery_fees:
                    _r_edit_details.append(f"رسوم التوصيل (المورد): {_r_old_delivery_fees_snap} ← {replacement_order.delivery_fees}")
                if _r_old_tracking_snap != (replacement_order.tracking_number or ''):
                    _r_new_t = replacement_order.tracking_number or 'محذوف'
                    _r_edit_details.append(f"رقم التتبع: '{_r_old_tracking_snap}' ← '{_r_new_t}'" if _r_old_tracking_snap else f"رقم التتبع: {_r_new_t}")
                # مقارنة المنتجات
                try:
                    _r_new_items_snap = []
                    for _it in replacement_order.items:
                        _pname_rn = getattr(_it.product, 'name', None) if _it.product else None
                        _r_new_items_snap.append((_pname_rn or f'منتج#{_it.product_id}', _it.quantity, _it.state or ''))
                    if sorted(_r_old_items_snap) != sorted(_r_new_items_snap):
                        _r_items_str = "، ".join(f"{n} x{q} ({s})" for n, q, s in _r_new_items_snap)
                        _r_edit_details.append(f"المنتجات: [{_r_items_str}]")
                except Exception:
                    pass
                _r_desc = " | ".join(_r_edit_details) if _r_edit_details else "تم تعديل طلب الاستبدال"
                db.session.add(OrderEditLog(
                    replacement_order_id=replacement_order.id,
                    employee_id=session.get('employee_id'),
                    description=_r_desc
                ))
                db.session.commit()
            except Exception:
                pass

            flash('تم تحديث طلب الاستبدال بنجاح', 'success')
            if replacement_order.status == 'جديد':
                return redirect(url_for('main.replacement_orders_by_status', status='جديد'))
            return redirect(url_for('main.replacement_orders'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء التحديث: {str(e)}', 'error')
            return redirect(request.url)
    
    returned_product_ids = {
        r.product_id for r in ReturnOrderItem.query.join(ReturnOrder).filter(
            ReturnOrder.original_replacement_order_id == order_id, ReturnOrderItem.return_order_id == ReturnOrder.id
        ).with_entities(ReturnOrderItem.product_id).distinct().all()
    }

    return render_template('edit_replacement_order.html', order=replacement_order, next_url=next_url,
        all_statuses=ALL_STATUSES,
        valid_next_statuses=StatusService.valid_next(replacement_order.status),
        returned_product_ids=returned_product_ids,
        products_data=products_data,
        items_data=items_data)

@main.route('/admin/replacements/recalculate', methods=['GET', 'POST'])
@login_required
@permission_required('is_admin')
def recalc_replacement_losses():
    updated = 0
    try:
        orders = ReplacementOrder.query.options(
            joinedload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product).joinedload(Product.bundle_items).joinedload(BundleItem.product)
        ).all()
        for order in orders:
            for it in order.items:
                is_dmg = bool(it.is_damaged) or (it.state == 'تالف')
                it.is_damaged = is_dmg
                it.damage_loss = (it.purchase_price or 0) * (it.quantity or 1) if is_dmg else 0

            try:
                new_products_total = 0.0
                for it in order.items:
                    if it.state in ['منتج جديد', 'جديد']:
                        new_products_total += (it.price or 0.0) * (it.quantity or 1)
                order.total_amount = float(new_products_total) + float(order.delivery_fees_customer or 0.0)
            except Exception:
                pass

            try:
                order.calculate_losses()
            except Exception:
                pass
            updated += 1

        db.session.commit()
        flash(f'تمت إعادة حساب الخسائر لعدد {updated} من طلبات الاستبدال', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'فشل تحديث الخسائر: {str(e)}', 'error')
    return redirect(url_for('main.replacement_orders'))

@main.route('/customer_log/<int:log_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_customer_logs')
def delete_customer_log(log_id):
    log = CustomerLog.query.get_or_404(log_id)
    customer_id = log.customer_id
    db.session.delete(log)
    db.session.commit()
    flash('تم حذف السجل بنجاح')
    return redirect(url_for('main.customer_history', customer_id=customer_id))

@main.route('/api/customer/<int:customer_id>/reminders')
@login_required
@permission_required('can_manage_reminders')
def get_customer_reminders(customer_id):
    try:
        current_time = datetime.now()
        reminders = CustomerLog.query.filter(
            CustomerLog.customer_id == customer_id,
            CustomerLog.type == 'تنبيه بالمتابعة',
            CustomerLog.is_dismissed == False
        ).all()

        reminder_data = []
        for reminder in reminders:
            reminder_data.append({
                'id': reminder.id,
                'customer_name': reminder.customer.name,
                'customer_phone': reminder.customer.phone,
                'content': reminder.content,
                'follow_up_reason': reminder.follow_up_reason,
                'reminder_time': reminder.reminder_time.strftime('%Y-%m-%d %I:%M %p'),
                'timestamp': reminder.timestamp.strftime('%Y-%m-%d %I:%M %p'),
                'is_due': reminder.reminder_time <= current_time
            })

        return jsonify({'reminders': reminder_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/reminder/<int:reminder_id>/dismiss', methods=['POST'])
@login_required
@permission_required('can_manage_reminders')
def dismiss_reminder(reminder_id):
    try:
        reminder = CustomerLog.query.get_or_404(reminder_id)
        reminder.is_dismissed = True
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/reminders/all')
@login_required
@permission_required('can_manage_reminders')
def get_all_reminders():
    try:
        current_time = datetime.now()

        reminders = CustomerLog.query.filter(
            CustomerLog.type == 'تنبيه بالمتابعة',
            CustomerLog.is_dismissed == False
        ).join(Customer).all()

        reminder_data = []
        for reminder in reminders:
            reminder_data.append({
                'id': reminder.id,
                'customer_name': reminder.customer.name,
                'customer_id': reminder.customer_id,
                'customer_phone': reminder.customer.phone,
                'content': reminder.content,
                'follow_up_reason': reminder.follow_up_reason,
                'reminder_time': reminder.reminder_time.strftime('%Y-%m-%d %I:%M %p'),
                'reminder_duration': reminder.reminder_duration,
                'reminder_duration_type': reminder.reminder_duration_type,
                'timestamp': reminder.timestamp.strftime('%Y-%m-%d %I:%M %p'),
                'is_due': reminder.reminder_time <= current_time
            })

        return jsonify({'reminders': reminder_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/api/order/<int:order_id>/update_tracking', methods=['POST'])
@login_required
@permission_required('can_update_tracking')
def update_tracking_number(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.get_json()
    new_tracking_number = data.get('tracking_number', '').strip()

    log_content = f"تم تحديث الرقم التتبعي إلى: {new_tracking_number}"
    if not new_tracking_number:
        log_content = "تمت إزالة الرقم التتبعي"

    if order.is_delivery:
        log = CustomerLog(
            customer_id=order.customer_id,
            order_id=order.id,
            employee_id=session.get('employee_id'),
            type="تلقائي",
            content=log_content
        )
        db.session.add(log)

    order.tracking_number = new_tracking_number

    new_status = None
    if order.status == 'جديد' and new_tracking_number:
        from collections import defaultdict
        req = defaultdict(int)
        from .models import Product
        for it in order.items:
            product = it.product
            if product:
                if product.is_bundle:
                    bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                    for bundle_item in bundle_items:
                        req[bundle_item.product_id] += (it.quantity or 0) * 1
                else:
                    req[product.id] += (it.quantity or 0)
        insufficient = []
        if req:
            pmap = _get_locked_products(req.keys())
            for pid, q in req.items():
                p = pmap.get(pid)
                if not p:
                    continue
                if (p.stock or 0) < q:
                    insufficient.append(f"{p.name} (المطلوب {q} المتوفر {p.stock})")
        if insufficient:
            return jsonify({'success': False, 'message': 'لا يمكن تعيين رقم تتبعي / التحويل إلى خرج للتوصيل بسبب نقص المخزون: ' + ' ؛ '.join(insufficient)}), 400
        old_status = order.status
        try:
            validate_transition(old_status, 'خرج للتوصيل')
        except ValueError as e:
            return jsonify({'success': False, 'message': str(e)}), 400
        order.status = 'خرج للتوصيل'
        if not order.inventory_deducted:
            for pid, q in req.items():
                p = pmap.get(pid)
                if p and _block_if_deleted(p, 'الطلب', 'deduct'):
                    return jsonify({'success': False, 'message': f'لا يمكن تعيين رقم تتبعي: الطلب يحتوي على منتج محذوف ("{p.name}")'}), 400
            for pid, q in req.items():
                p = pmap.get(pid)
                if p:
                    p.stock = (p.stock or 0) - q
            order.inventory_deducted = True
        new_status = order.status
        
        if old_status != new_status:
            try:
                order.add_status_history(
                    new_status,
                    employee_id=session.get('employee_id'),
                    notes='تم تغيير الحالة تلقائيًا عند إضافة رقم تتبع'
                )
            except Exception as e:
                current_app.logger.error(f"Failed to add status history: {e}")
        
        if order.is_delivery:
            status_log = CustomerLog(
                customer_id=order.customer_id,
                order_id=order.id,
                employee_id=session.get('employee_id'),
                type="تلقائي",
                content=f"تم تغيير الحالة تلقائيًا إلى: {order.status}"
            )
            db.session.add(status_log)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Tracking number updated successfully',
        'new_status': new_status
    })

@main.route('/api/order/<int:order_id>/update_status', methods=['POST'])
@login_required
@permission_required('can_update_order_status')
def update_status(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.get_json()
    new_status = data.get('status')

    if not new_status:
        return jsonify(
            {'success': False, 'message': 'Status cannot be empty'}), 400

    old_status = order.status
    try:
        validate_transition(old_status, new_status)
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    if old_status != new_status:
        if new_status == "خرج للتوصيل" and not order.inventory_deducted:
            req = _collect_order_required_quantities(order)
            if req:
                pmap = _get_locked_products(req.keys())
                insufficient = []
                for pid, q in req.items():
                    p = pmap.get(pid)
                    if not p:
                        continue
                    if (p.stock or 0) < q:
                        insufficient.append(f"{p.name} (المطلوب {q} المتوفر {p.stock})")
                if insufficient:
                    return jsonify({'success': False, 'message': 'لا يمكن تغيير الحالة إلى خرج للتوصيل لعدم كفاية المخزون: ' + ' ؛ '.join(insufficient)}), 400

        if new_status in ["إلغاء", "رفض الاستلام"] and order.inventory_deducted:
            restock_log = {}
            for item in order.items:
                product = item.product
                if product:
                    if product.is_bundle:
                        bundle_items = BundleItem.query.filter_by(bundle_id=product.id).all()
                        for bundle_item in bundle_items:
                            bundle_product = _get_locked_product(bundle_item.product_id)
                            if bundle_product:
                                bundle_product.stock = (bundle_product.stock or 0) + (item.quantity * 1)
                                restock_log[bundle_product.id] = restock_log.get(bundle_product.id, 0) + (item.quantity * 1)
                    else:
                        product.stock = (product.stock or 0) + (item.quantity or 0)
                        restock_log[product.id] = restock_log.get(product.id, 0) + (item.quantity or 0)
            if restock_log:
                _log_stock_delta(f"update_status status->{new_status} restock", restock_log)
                order.inventory_deducted = False

        if order.is_delivery:
            log = CustomerLog(
                customer_id=order.customer_id,
                order_id=order.id,
                employee_id=session.get('employee_id'),
                type="تلقائي",
                content=f"تم تغيير الحالة من '{old_status}' إلى '{new_status}'"
            )
            db.session.add(log)

        if new_status == "خرج للتوصيل" and not order.inventory_deducted:
            req = _collect_order_required_quantities(order)
            if req:
                pmap = _get_locked_products(req.keys())
                for p in pmap.values():
                    if _block_if_deleted(p, 'الطلب', 'deduct'):
                        return jsonify({'success': False, 'message': f'لا يمكن تغيير الحالة: الطلب يحتوي على منتج محذوف ("{p.name}")'}), 400
                for pid, q in req.items():
                    p = pmap.get(pid)
                    if p:
                        p.stock = (p.stock or 0) - q
            order.inventory_deducted = True

        order.status = new_status
        db.session.commit()
        
        try:
            order.add_status_history(
                new_status,
                employee_id=session.get('employee_id'),
                notes=f'تم تغيير الحالة من {old_status}'
            )
            db.session.commit()
        except Exception as e:
            current_app.logger.error(f"Failed to add status history in update_status: {e}")

    return jsonify({'success': True, 'message': 'Status updated successfully'})

@main.route('/governorate-fees', methods=['GET', 'POST'])
@login_required
@permission_required('can_view_fees')
def governorate_fees():
    from .models import GovernorateFee
    if request.method == 'POST':
        name = request.form.get('name')
        fee = request.form.get('fee')
        if name and fee is not None:
            fee_obj = GovernorateFee.query.filter_by(name=name).first()
            if fee_obj:
                fee_obj.fee = float(fee)
            else:
                fee_obj = GovernorateFee(name=name, fee=float(fee))
                db.session.add(fee_obj)
            db.session.commit()
            flash('تم حفظ الرسوم بنجاح')
        return redirect(url_for('main.governorate_fees'))
    fees = GovernorateFee.query.order_by(GovernorateFee.name).all()
    
    cod_fee = AppSettings.get_cod_fee()
    
    return render_template('governorate_fees.html', fees=fees, cod_fee=cod_fee)

@main.route('/governorate-fees/delete/<int:fee_id>', methods=['POST'])
@login_required
@permission_required('can_delete_fees')
def delete_governorate_fee(fee_id):
    from .models import GovernorateFee
    fee = GovernorateFee.query.get_or_404(fee_id)
    db.session.delete(fee)
    db.session.commit()
    flash('تم حذف الرسوم بنجاح')
    return redirect(url_for('main.governorate_fees'))

@main.route('/api/governorate-fee')
@login_required
@permission_required('can_view_fees')
def api_governorate_fee():
    from .models import GovernorateFee
    name = request.args.get('name')
    fee = 0
    if name:
        fee_obj = GovernorateFee.query.filter_by(name=name).first()
        if fee_obj:
            fee = fee_obj.fee
    return jsonify({'fee': fee})

@main.route('/api/cod-fee', methods=['GET'])
@login_required
@permission_required('can_view_fees')
def get_cod_fee():
    cod_fee = AppSettings.get_cod_fee()
    return jsonify({'success': True, 'cod_fee': cod_fee})

@main.route('/api/cod-fee', methods=['POST'])
@login_required
@permission_required('can_edit_fees')
def update_cod_fee():
    try:
        data = request.get_json()
        cod_fee = float(data.get('cod_fee', 0))
        
        if cod_fee < 0:
            return jsonify({'success': False, 'message': 'رسوم التحصيل يجب أن تكون قيمة موجبة أو صفر'}), 400
        
        AppSettings.set_value('cod_fee', cod_fee, 'رسوم التحصيل (COD Fee)')
        db.session.commit()
        
        log_activity(
            action='update',
            entity_type='settings',
            entity_id=0,
            entity_name='رسوم التحصيل',
            details=f'تم تحديث رسوم التحصيل إلى: {cod_fee} ج.م'
        )
        
        return jsonify({'success': True, 'message': 'تم تحديث رسوم التحصيل بنجاح', 'cod_fee': cod_fee})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

@main.route('/employees')
@login_required
@permission_required('can_view_employees')
def employees():
    search_query = request.args.get('search', '')
    if search_query:
        all_employees = Employee.query.filter(
            or_(Employee.name.ilike(f"%{search_query}%"),
                Employee.phone.ilike(f"%{search_query}%"),
                Employee.position.ilike(f"%{search_query}%"))
        ).all()
    else:
        all_employees = Employee.query.all()

    current_employee = Employee.query.get(session['employee_id'])
    is_admin = current_employee.is_admin if current_employee else False

    today = date.today()
    attended_ids = {r.employee_id for r in AttendanceRecord.query.filter_by(date=today).all()}

    return render_template(
        'employees.html',
        employees=all_employees,
        search_query=search_query,
        is_admin=is_admin,
        attended_ids=attended_ids)

@main.route('/employees/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_employees')
def add_employee_basic():
    if request.method == 'POST':
        field_names = {
            'name': 'اسم الموظف',
            'phone': 'رقم الهاتف',
            'position': 'الوظيفة',
            'hire_date': 'تاريخ التعيين',
            'salary': 'الراتب',
            'sales_commission_percentage': 'نسبة المبيعات',
            'username': 'اسم المستخدم',
            'password': 'كلمة المرور'
        }

        for field in [
            'name',
            'phone',
            'position',
            'hire_date',
            'salary',
            'username',
                'password']:
            if not request.form.get(field):
                flash(f'حقل {field_names[field]} مطلوب')
                return render_template('add_employee_basic.html')

        if Employee.query.filter_by(username=request.form['username']).first():
            flash('اسم المستخدم موجود بالفعل، يرجى اختيار اسم مستخدم آخر')
            return render_template('add_employee_basic.html')

        if Employee.query.filter_by(phone=request.form['phone']).first():
            flash('رقم الهاتف موجود بالفعل، يرجى إدخال رقم هاتف آخر')
            return render_template('add_employee_basic.html')

        try:
            salary = float(request.form['salary'])
            if salary < 0:
                flash('الراتب يجب أن يكون رقم موجب')
                return render_template('add_employee_basic.html')
        except ValueError:
            flash('الراتب يجب أن يكون رقم صحيح')
            return render_template('add_employee_basic.html')

        try:
            hire_date = datetime.strptime(
                request.form['hire_date'], '%Y-%m-%d').date()
            if hire_date > date.today():
                flash('تاريخ التعيين لا يمكن أن يكون في المستقبل')
                return render_template('add_employee_basic.html')
        except ValueError:
            flash('تاريخ التعيين غير صحيح')
            return render_template('add_employee_basic.html')

        session['new_employee_data'] = {
            'name': request.form['name'],
            'phone': request.form['phone'],
            'email': request.form.get('email', ''),
            'position': request.form['position'],
            'hire_date': request.form['hire_date'],
            'salary': request.form['salary'],
            'sales_commission_percentage': request.form.get('sales_commission_percentage', '0'),
            'requires_attendance': 'requires_attendance' in request.form,
            'username': request.form['username'],
            'password': request.form['password']
        }
        return redirect(url_for('main.add_employee_permissions'))

    return render_template('add_employee_basic.html')

@main.route('/employees/add/permissions', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_employees')
def add_employee_permissions():
    if 'new_employee_data' not in session:
        flash('يرجى إدخال البيانات الأساسية أولاً')
        return redirect(url_for('main.add_employee_basic'))

    if request.method == 'POST':
        employee_data = session['new_employee_data']

        can_view_customers = 'can_view_customers' in request.form
        can_add_customers = 'can_add_customers' in request.form
        can_edit_customers = 'can_edit_customers' in request.form
        can_delete_customers = 'can_delete_customers' in request.form
        can_view_customer_orders = 'can_view_customer_orders' in request.form
        can_view_customer_logs = 'can_view_customer_logs' in request.form
        can_view_customer_profile = 'can_view_customer_profile' in request.form
        can_add_customer_logs = 'can_add_customer_logs' in request.form
        can_delete_customer_logs = 'can_delete_customer_logs' in request.form

        can_view_products = 'can_view_products' in request.form
        can_add_products = 'can_add_products' in request.form
        can_edit_products = 'can_edit_products' in request.form
        can_delete_products = 'can_delete_products' in request.form
        can_restore_products = 'can_restore_products' in request.form
        can_delete_all_products = 'can_delete_all_products' in request.form
        can_view_purchase_price = 'can_view_purchase_price' in request.form
        can_view_sold_products_by_quantity = 'can_view_sold_products_by_quantity' in request.form
        can_manage_stocktake = 'can_manage_stocktake' in request.form

        can_view_orders = 'can_view_orders' in request.form
        can_add_orders = 'can_add_orders' in request.form
        can_edit_orders = 'can_edit_orders' in request.form
        can_delete_orders = 'can_delete_orders' in request.form
        can_update_order_status = 'can_update_order_status' in request.form
        can_update_tracking = 'can_update_tracking' in request.form
        can_view_orders_by_status = 'can_view_orders_by_status' in request.form

        can_view_replacements = 'can_view_replacements' in request.form
        can_add_replacements = 'can_add_replacements' in request.form
        can_edit_replacements = 'can_edit_replacements' in request.form
        can_delete_replacements = 'can_delete_replacements' in request.form
        can_view_replacements_by_state = 'can_view_replacements_by_state' in request.form

        can_view_returns = 'can_view_returns' in request.form
        can_add_returns = 'can_add_returns' in request.form
        can_view_returns_by_state = 'can_view_returns_by_state' in request.form
        can_delete_returns = 'can_delete_returns' in request.form

        can_view_fees = 'can_view_fees' in request.form
        can_edit_fees = 'can_edit_fees' in request.form
        can_delete_fees = 'can_delete_fees' in request.form

        can_view_employees = 'can_view_employees' in request.form
        can_add_employees = 'can_add_employees' in request.form
        can_edit_employees = 'can_edit_employees' in request.form
        can_delete_employees = 'can_delete_employees' in request.form
        can_view_employee_logs = 'can_view_employee_logs' in request.form
        can_manage_employee_salary = 'can_manage_employee_salary' in request.form
        can_delete_salary_transactions = 'can_delete_salary_transactions' in request.form
        can_view_employee_activity = 'can_view_employee_activity' in request.form
        can_manage_attendance = 'can_manage_attendance' in request.form

        can_view_statistics = 'can_view_statistics' in request.form
        can_view_stats_stock = 'can_view_stats_stock' in request.form
        can_view_stats_fixed_assets = 'can_view_stats_fixed_assets' in request.form
        can_view_stats_total_debt = 'can_view_stats_total_debt' in request.form
        can_view_stats_capital_growth = 'can_view_stats_capital_growth' in request.form
        can_view_stats_daily = 'can_view_stats_daily' in request.form
        can_view_stats_pending_orders = 'can_view_stats_pending_orders' in request.form
        can_view_stats_net_profit = 'can_view_stats_net_profit' in request.form
        can_view_stats_losses = 'can_view_stats_losses' in request.form
        can_view_stats_delivery_rate = 'can_view_stats_delivery_rate' in request.form
        can_view_stats_sales = 'can_view_stats_sales' in request.form
        can_view_stats_amount_paid = 'can_view_stats_amount_paid' in request.form
        can_view_stats_monthly_delivered = 'can_view_stats_monthly_delivered' in request.form
        can_view_stats_monthly_invoices = 'can_view_stats_monthly_invoices' in request.form
        can_view_stats_fixed_assets_expenses = 'can_view_stats_fixed_assets_expenses' in request.form
        can_view_stats_employee_salaries = 'can_view_stats_employee_salaries' in request.form
        can_view_stats_employee_debt = 'can_view_stats_employee_debt' in request.form
        can_view_stats_operational_expenses = 'can_view_stats_operational_expenses' in request.form

        can_view_suppliers = 'can_view_suppliers' in request.form
        can_add_suppliers = 'can_add_suppliers' in request.form
        can_edit_suppliers = 'can_edit_suppliers' in request.form
        can_delete_suppliers = 'can_delete_suppliers' in request.form
        can_view_supplier_invoices = 'can_view_supplier_invoices' in request.form
        can_view_supplier_history = 'can_view_supplier_history' in request.form
        can_pay_supplier_debt = 'can_pay_supplier_debt' in request.form
        can_add_supplier_debt = 'can_add_supplier_debt' in request.form

        can_view_invoices = 'can_view_invoices' in request.form
        can_add_invoices = 'can_add_invoices' in request.form
        can_edit_invoices = 'can_edit_invoices' in request.form
        can_delete_invoices = 'can_delete_invoices' in request.form
        can_view_invoice_details = 'can_view_invoice_details' in request.form

        can_view_expenses = 'can_view_expenses' in request.form
        can_add_expenses = 'can_add_expenses' in request.form
        can_edit_expenses = 'can_edit_expenses' in request.form
        can_delete_expenses = 'can_delete_expenses' in request.form
        can_view_operational_expenses = 'can_view_operational_expenses' in request.form
        can_view_fixed_assets = 'can_view_fixed_assets' in request.form

        can_view_activity_log = 'can_view_activity_log' in request.form
        can_manage_reminders = 'can_manage_reminders' in request.form
        is_admin = 'is_admin' in request.form

        can_view_followups = 'can_view_followups' in request.form
        can_add_followups = 'can_add_followups' in request.form
        can_edit_followups = 'can_edit_followups' in request.form
        can_delete_followups = 'can_delete_followups' in request.form

        can_view_damaged_products = 'can_view_damaged_products' in request.form
        can_add_damaged_products = 'can_add_damaged_products' in request.form
        can_delete_damaged_products = 'can_delete_damaged_products' in request.form

        can_view_transactions = 'can_view_transactions' in request.form
        can_add_transactions = 'can_add_transactions' in request.form
        can_edit_transactions = 'can_edit_transactions' in request.form
        can_delete_transactions = 'can_delete_transactions' in request.form

        new_employee = Employee(
            name=employee_data['name'],
            phone=employee_data['phone'],
            email=employee_data['email'],
            address='',
            position=employee_data['position'],
            hire_date=datetime.strptime(
                employee_data['hire_date'], '%Y-%m-%d').date(),
            salary=float(employee_data['salary']),
            username=employee_data['username'],
            can_view_customers=can_view_customers,
            can_add_customers=can_add_customers,
            can_edit_customers=can_edit_customers,
            can_delete_customers=can_delete_customers,
            can_view_customer_orders=can_view_customer_orders,
            can_view_customer_logs=can_view_customer_logs,
            can_view_customer_profile=can_view_customer_profile,
            can_add_customer_logs=can_add_customer_logs,
            can_delete_customer_logs=can_delete_customer_logs,
            can_view_products=can_view_products,
            can_add_products=can_add_products,
            can_edit_products=can_edit_products,
            can_delete_products=can_delete_products,
            can_restore_products=can_restore_products,
            can_delete_all_products=can_delete_all_products,
            can_view_purchase_price=can_view_purchase_price,
            can_view_sold_products_by_quantity=can_view_sold_products_by_quantity,
            can_manage_stocktake=can_manage_stocktake,
            can_view_orders=can_view_orders,
            can_add_orders=can_add_orders,
            can_edit_orders=can_edit_orders,
            can_delete_orders=can_delete_orders,
            can_update_order_status=can_update_order_status,
            can_update_tracking=can_update_tracking,
            can_view_orders_by_status=can_view_orders_by_status,
            can_view_replacements=can_view_replacements,
            can_add_replacements=can_add_replacements,
            can_edit_replacements=can_edit_replacements,
            can_delete_replacements=can_delete_replacements,
            can_view_replacements_by_state=can_view_replacements_by_state,
            can_view_returns=can_view_returns,
            can_add_returns=can_add_returns,
            can_view_returns_by_state=can_view_returns_by_state,
            can_delete_returns=can_delete_returns,
            can_view_fees=can_view_fees,
            can_edit_fees=can_edit_fees,
            can_delete_fees=can_delete_fees,
            can_view_employees=can_view_employees,
            can_add_employees=can_add_employees,
            can_edit_employees=can_edit_employees,
            can_delete_employees=can_delete_employees,
            can_view_employee_logs=can_view_employee_logs,
            can_manage_employee_salary=can_manage_employee_salary,
            can_delete_salary_transactions=can_delete_salary_transactions,
            can_view_employee_activity=can_view_employee_activity,
            requires_attendance=new_employee_data.get('requires_attendance', False),
            sales_commission_percentage=float(new_employee_data.get('sales_commission_percentage', '0')),
            can_manage_attendance=can_manage_attendance,
            can_view_statistics=can_view_statistics,
            can_view_stats_stock=can_view_stats_stock,
            can_view_stats_fixed_assets=can_view_stats_fixed_assets,
            can_view_stats_total_debt=can_view_stats_total_debt,
            can_view_stats_capital_growth=can_view_stats_capital_growth,
            can_view_stats_daily=can_view_stats_daily,
            can_view_stats_pending_orders=can_view_stats_pending_orders,
            can_view_stats_net_profit=can_view_stats_net_profit,
            can_view_stats_losses=can_view_stats_losses,
            can_view_stats_delivery_rate=can_view_stats_delivery_rate,
            can_view_stats_sales=can_view_stats_sales,
            can_view_stats_amount_paid=can_view_stats_amount_paid,
            can_view_stats_monthly_delivered=can_view_stats_monthly_delivered,
            can_view_stats_monthly_invoices=can_view_stats_monthly_invoices,
            can_view_stats_fixed_assets_expenses=can_view_stats_fixed_assets_expenses,
            can_view_stats_employee_salaries=can_view_stats_employee_salaries,
            can_view_stats_employee_debt=can_view_stats_employee_debt,
            can_view_stats_operational_expenses=can_view_stats_operational_expenses,
            can_view_suppliers=can_view_suppliers,
            can_add_suppliers=can_add_suppliers,
            can_edit_suppliers=can_edit_suppliers,
            can_delete_suppliers=can_delete_suppliers,
            can_view_supplier_invoices=can_view_supplier_invoices,
            can_view_supplier_history=can_view_supplier_history,
            can_pay_supplier_debt=can_pay_supplier_debt,
            can_add_supplier_debt=can_add_supplier_debt,
            can_view_invoices=can_view_invoices,
            can_add_invoices=can_add_invoices,
            can_edit_invoices=can_edit_invoices,
            can_delete_invoices=can_delete_invoices,
            can_view_invoice_details=can_view_invoice_details,
            can_view_expenses=can_view_expenses,
            can_add_expenses=can_add_expenses,
            can_edit_expenses=can_edit_expenses,
            can_delete_expenses=can_delete_expenses,
            can_view_operational_expenses=can_view_operational_expenses,
            can_view_fixed_assets=can_view_fixed_assets,
            can_view_activity_log=can_view_activity_log,
            can_manage_reminders=can_manage_reminders,
            is_admin=is_admin,
            can_view_followups=can_view_followups,
            can_add_followups=can_add_followups,
            can_edit_followups=can_edit_followups,
            can_delete_followups=can_delete_followups,
            can_view_damaged_products=can_view_damaged_products,
            can_add_damaged_products=can_add_damaged_products,
            can_delete_damaged_products=can_delete_damaged_products,
            can_view_transactions=can_view_transactions,
            can_add_transactions=can_add_transactions,
            can_edit_transactions=can_edit_transactions,
            can_delete_transactions=can_delete_transactions
        )

        new_employee.set_password(employee_data['password'])

        db.session.add(new_employee)
        db.session.commit()

        session.pop('new_employee_data', None)

        flash(f'تم إضافة الموظف {employee_data["name"]} بنجاح!')
        return redirect(url_for('main.employees'))

    return render_template(
        'add_employee_permissions.html',
        employee_data=session['new_employee_data'])

@main.route('/employees/add/old', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_employees')
def add_employee():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form.get('email', '')
        position = request.form['position']
        hire_date = datetime.strptime(
            request.form['hire_date'], '%Y-%m-%d').date()
        salary = float(request.form['salary'])
        username = request.form['username']
        password = request.form['password']

        can_view_orders = 'can_view_orders' in request.form
        can_edit_orders = 'can_edit_orders' in request.form
        can_delete_orders = 'can_delete_orders' in request.form
        can_view_customers = 'can_view_customers' in request.form
        can_edit_customers = 'can_edit_customers' in request.form
        can_delete_customers = 'can_delete_customers' in request.form
        can_view_products = 'can_view_products' in request.form
        can_edit_products = 'can_edit_products' in request.form
        can_delete_products = 'can_delete_products' in request.form
        can_view_employees = 'can_view_employees' in request.form
        can_edit_employees = 'can_edit_employees' in request.form
        can_delete_employees = 'can_delete_employees' in request.form
        is_admin = 'is_admin' in request.form
        requires_attendance = 'requires_attendance' in request.form

        can_view_replacements = 'can_view_replacements' in request.form
        can_edit_replacements = 'can_edit_replacements' in request.form

        can_view_returns = 'can_view_returns' in request.form
        can_add_returns = 'can_add_returns' in request.form
        can_view_returns_by_state = 'can_view_returns_by_state' in request.form
        can_delete_returns = 'can_delete_returns' in request.form

        can_view_fees = 'can_view_fees' in request.form
        can_edit_fees = 'can_edit_fees' in request.form

        can_view_suppliers = 'can_view_suppliers' in request.form
        can_edit_suppliers = 'can_edit_suppliers' in request.form

        can_view_invoices = 'can_view_invoices' in request.form
        can_edit_invoices = 'can_edit_invoices' in request.form

        can_view_statistics = 'can_view_statistics' in request.form

        if Employee.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود بالفعل')
            return redirect(url_for('main.add_employee'))

        if Employee.query.filter_by(phone=phone).first():
            flash('رقم الهاتف موجود بالفعل')
            return redirect(url_for('main.add_employee'))

        new_employee = Employee(
            name=name,
            phone=phone,
            email=email,
            address='',
            position=position,
            hire_date=hire_date,
            salary=salary,
            username=username,
            can_view_orders=can_view_orders,
            can_edit_orders=can_edit_orders,
            can_delete_orders=can_delete_orders,
            can_view_customers=can_view_customers,
            can_edit_customers=can_edit_customers,
            can_delete_customers=can_delete_customers,
            can_view_products=can_view_products,
            can_edit_products=can_edit_products,
            can_delete_products=can_delete_products,
            can_view_employees=can_view_employees,
            can_edit_employees=can_edit_employees,
            can_delete_employees=can_delete_employees,
            is_admin=is_admin,
            requires_attendance=requires_attendance,
            sales_commission_percentage=float(request.form.get('sales_commission_percentage', '0')),
            can_view_replacements=can_view_replacements,
            can_edit_replacements=can_edit_replacements,
            can_view_returns=can_view_returns,
            can_add_returns=can_add_returns,
            can_view_returns_by_state=can_view_returns_by_state,
            can_delete_returns=can_delete_returns,
            can_view_fees=can_view_fees,
            can_edit_fees=can_edit_fees,
            can_view_suppliers=can_view_suppliers,
            can_edit_suppliers=can_edit_suppliers,
            can_view_invoices=can_view_invoices,
            can_edit_invoices=can_edit_invoices,
            can_view_statistics=can_view_statistics)
        new_employee.set_password(password)

        db.session.add(new_employee)
        db.session.commit()
        flash('تم إضافة الموظف بنجاح')
        return redirect(url_for('main.employees'))

    return render_template('add_employee.html')

@main.route('/employees/<int:employee_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_employees')
def edit_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    if request.method == 'POST':

        employee.name = request.form.get('name', '').strip()
        employee.phone = request.form.get('phone', '').strip()
        employee.position = request.form.get('position', '').strip()
        
        hire_date_str = request.form.get('hire_date', '')
        if hire_date_str:
            try:
                employee.hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('تاريخ التعيين غير صحيح')
                return redirect(url_for('main.edit_employee', employee_id=employee_id))
        
        salary_str = request.form.get('salary', '0')
        try:
            employee.salary = float(salary_str)
        except ValueError:
            flash('المرتب غير صحيح')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))
        
        sales_commission_str = request.form.get('sales_commission_percentage', '0')
        try:
            employee.sales_commission_percentage = float(sales_commission_str)
        except ValueError:
            flash('نسبة المبيعات غير صحيحة')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))
        
        employee.requires_attendance = 'requires_attendance' in request.form

        new_username = request.form.get('username', '').strip()
        if not new_username:
            flash('اسم المستخدم مطلوب')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))
        
        existing_employee = Employee.query.filter_by(username=new_username).first()
        if existing_employee and existing_employee.id != employee.id:
            flash('اسم المستخدم موجود بالفعل')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))
        employee.username = new_username

        if not employee.phone:
            flash('رقم الهاتف مطلوب')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))
            
        existing_employee = Employee.query.filter_by(
            phone=employee.phone).first()
        if existing_employee and existing_employee.id != employee.id:
            flash('رقم الهاتف موجود بالفعل')
            return redirect(
                url_for(
                    'main.edit_employee',
                    employee_id=employee_id))

        new_password = request.form.get('new_password', '').strip()
        if new_password:
            if len(new_password) < 6:
                flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل')
                return redirect(url_for('main.edit_employee', employee_id=employee_id))
            employee.set_password(new_password)

        try:
            db.session.commit()
            flash('تم تعديل بيانات الموظف بنجاح')
            return redirect(url_for('main.employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حفظ البيانات: {str(e)}')
            return redirect(url_for('main.edit_employee', employee_id=employee_id))

    return render_template('edit_employee.html', employee=employee)

@main.route('/employees/<int:employee_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_employees')
def delete_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    db.session.delete(employee)
    db.session.commit()
    flash('تم حذف الموظف بنجاح')
    return redirect(url_for('main.employees'))

@main.route('/employees/<int:employee_id>/profile')
@login_required
@permission_required('can_view_employees')
def employee_profile(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    return render_template('employee_profile.html', employee=employee)

@main.route('/employees/<int:employee_id>/salary', methods=['GET', 'POST'])
@login_required
@permission_required('can_manage_employee_salary')
def employee_salary(employee_id):
    employee = Employee.query.get_or_404(employee_id)

    current_month = request.args.get('month', datetime.now().month, type=int)
    current_year = request.args.get('year', datetime.now().year, type=int)

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_transaction':
            transaction_type = request.form['transaction_type']
            amount = float(request.form['amount'])
            description = request.form.get('description', '')
            transaction_date = datetime.strptime(
                request.form.get(
                    'transaction_date',
                    date.today().isoformat()),
                '%Y-%m-%d').date()

            transaction = SalaryTransaction(
                employee_id=employee.id,
                transaction_type=transaction_type,
                amount=amount,
                description=description,
                transaction_date=transaction_date,
                month=transaction_date.month,
                year=transaction_date.year,
                created_by=session.get('employee_id', 1)
            )

            db.session.add(transaction)
            db.session.commit()

            details = f"إضافة معاملة مرتب: {transaction_type} - {amount} ج.م"
            log_activity(
                'create',
                'salary_transaction',
                transaction.id,
                employee.name,
                details)

            flash('تم إضافة المعاملة بنجاح')
            return redirect(
                url_for(
                    'main.employee_salary',
                    employee_id=employee_id,
                    month=current_month,
                    year=current_year))

        elif action == 'delete_transaction':
            transaction_id = request.form.get('transaction_id', type=int)
            transaction = SalaryTransaction.query.get_or_404(transaction_id)

            from app.permissions import has_permission
            if not has_permission('can_delete_salary_transactions'):
                flash('ليس لديك صلاحية حذف معاملات الراتب', 'error')
                return redirect(url_for('main.employee_salary', employee_id=employee_id,
                                        month=current_month, year=current_year))

            if transaction.employee_id == employee.id:
                transaction_type = transaction.transaction_type
                amount = transaction.amount
                db.session.delete(transaction)
                db.session.commit()

                details = f"حذف معاملة مرتب: {transaction_type} - {amount} ج.م"
                log_activity(
                    'delete',
                    'salary_transaction',
                    transaction_id,
                    employee.name,
                    details)

                flash('تم حذف المعاملة بنجاح')
            else:
                flash('غير مسموح بحذف هذه المعاملة', 'error')

            return redirect(
                url_for(
                    'main.employee_salary',
                    employee_id=employee_id,
                    month=current_month,
                    year=current_year))

    salary_summary = employee.get_salary_summary(current_month, current_year)

    monthly_transactions = employee.get_monthly_transactions(
        current_month, current_year)

    start_date = datetime(current_year, current_month, 1)
    end_date = datetime(current_year + 1, 1, 1) if current_month == 12 else datetime(current_year, current_month + 1, 1)
    days_in_month = (end_date - start_date).days
    if employee.requires_attendance:
        accrued_salary = Employee._calc_accrued_salary_attendance(
            salary_summary['net_salary'] or 0, employee.id, current_month, current_year, days_in_month)
        month_start_date = start_date.date()
        month_end_date = end_date.date()
        month_records = AttendanceRecord.query.filter(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.date >= month_start_date,
            AttendanceRecord.date < month_end_date
        ).order_by(AttendanceRecord.date).all()
        month_records_map = {r.date.isoformat(): r.status for r in month_records}
    else:
        accrued_salary = Employee._calc_accrued_salary(salary_summary['net_salary'] or 0, current_month, current_year, days_in_month)
        month_records_map = {}
        month_start_date = None
        month_end_date = None

    recent_orders = Order.query.filter(
        Order.employee_id == employee.id,
        Order.date >= start_date,
        Order.date < end_date
    ).order_by(Order.date.desc()).all()

    total_orders = len(recent_orders)

    # Use history-based query: orders that reached 'وصل' this month,
    # even if their current status is 'استبدال'.
    delivered_this_month = employee._query_period_orders(start_date, end_date)
    delivered_orders_count = salary_summary['delivered_orders_count']
    total_sales = salary_summary['delivered_sales']
    total_commission = salary_summary['sales_commission']
    return render_template('employee_salary.html',
                           employee=employee,
                           transactions=monthly_transactions,
                           base_salary=salary_summary['base_salary'],
                           sales_commission=salary_summary['sales_commission'],
                           total_salary=salary_summary['total_salary'],
                           total_deductions=salary_summary['total_deductions'],
                           total_bonuses=salary_summary['total_bonuses'],
                           total_advances=salary_summary['total_advances'],
                           net_salary=salary_summary['net_salary'],
                           accrued_salary=accrued_salary,
                           current_month=current_month,
                           current_year=current_year,
                           recent_orders=recent_orders,
                           delivered_orders_this_month=delivered_this_month,
                           total_orders=total_orders,
                           delivered_orders_count=delivered_orders_count,
                           total_sales=total_sales,
                            total_commission=total_commission,
                             today=date.today().isoformat(),
                             running_balance=employee.get_running_balance(),
                             previous_months_dues=employee.get_previous_months_dues(),
                            month_records_map=month_records_map,
                            month_start=month_start_date,
                            month_end=month_end_date,
                            today_date=date.today())

@main.route('/employees/<int:employee_id>/permissions',
            methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_employees')
def employee_permissions(employee_id):
    employee = Employee.query.get_or_404(employee_id)

    if request.method == 'POST':
        employee.can_view_customers = 'can_view_customers' in request.form
        employee.can_add_customers = 'can_add_customers' in request.form
        employee.can_edit_customers = 'can_edit_customers' in request.form
        employee.can_delete_customers = 'can_delete_customers' in request.form
        employee.can_view_customer_orders = 'can_view_customer_orders' in request.form
        employee.can_view_customer_logs = 'can_view_customer_logs' in request.form
        employee.can_view_customer_profile = 'can_view_customer_profile' in request.form
        employee.can_add_customer_logs = 'can_add_customer_logs' in request.form
        employee.can_delete_customer_logs = 'can_delete_customer_logs' in request.form

        employee.can_view_products = 'can_view_products' in request.form
        employee.can_add_products = 'can_add_products' in request.form
        employee.can_edit_products = 'can_edit_products' in request.form
        employee.can_delete_products = 'can_delete_products' in request.form
        employee.can_restore_products = 'can_restore_products' in request.form
        employee.can_delete_all_products = 'can_delete_all_products' in request.form
        employee.can_view_purchase_price = 'can_view_purchase_price' in request.form
        employee.can_view_sold_products_by_quantity = 'can_view_sold_products_by_quantity' in request.form
        employee.can_manage_stocktake = 'can_manage_stocktake' in request.form

        employee.can_view_orders = 'can_view_orders' in request.form
        employee.can_add_orders = 'can_add_orders' in request.form
        employee.can_edit_orders = 'can_edit_orders' in request.form
        employee.can_delete_orders = 'can_delete_orders' in request.form
        employee.can_update_order_status = 'can_update_order_status' in request.form
        employee.can_update_tracking = 'can_update_tracking' in request.form
        employee.can_view_orders_by_status = 'can_view_orders_by_status' in request.form

        employee.can_view_replacements = 'can_view_replacements' in request.form
        employee.can_add_replacements = 'can_add_replacements' in request.form
        employee.can_edit_replacements = 'can_edit_replacements' in request.form
        employee.can_delete_replacements = 'can_delete_replacements' in request.form
        employee.can_view_replacements_by_state = 'can_view_replacements_by_state' in request.form

        employee.can_view_returns = 'can_view_returns' in request.form
        employee.can_add_returns = 'can_add_returns' in request.form
        employee.can_view_returns_by_state = 'can_view_returns_by_state' in request.form
        employee.can_delete_returns = 'can_delete_returns' in request.form

        employee.can_view_fees = 'can_view_fees' in request.form
        employee.can_edit_fees = 'can_edit_fees' in request.form
        employee.can_delete_fees = 'can_delete_fees' in request.form

        employee.can_view_employees = 'can_view_employees' in request.form
        employee.can_add_employees = 'can_add_employees' in request.form
        employee.can_edit_employees = 'can_edit_employees' in request.form
        employee.can_delete_employees = 'can_delete_employees' in request.form
        employee.can_view_employee_logs = 'can_view_employee_logs' in request.form
        employee.can_manage_employee_salary = 'can_manage_employee_salary' in request.form
        employee.can_delete_salary_transactions = 'can_delete_salary_transactions' in request.form
        employee.can_view_employee_activity = 'can_view_employee_activity' in request.form
        employee.can_manage_attendance = 'can_manage_attendance' in request.form

        employee.can_view_statistics = 'can_view_statistics' in request.form
        employee.can_view_stats_stock = 'can_view_stats_stock' in request.form
        employee.can_view_stats_fixed_assets = 'can_view_stats_fixed_assets' in request.form
        employee.can_view_stats_total_debt = 'can_view_stats_total_debt' in request.form
        employee.can_view_stats_capital_growth = 'can_view_stats_capital_growth' in request.form
        employee.can_view_stats_daily = 'can_view_stats_daily' in request.form
        employee.can_view_stats_pending_orders = 'can_view_stats_pending_orders' in request.form
        employee.can_view_stats_net_profit = 'can_view_stats_net_profit' in request.form
        employee.can_view_stats_losses = 'can_view_stats_losses' in request.form
        employee.can_view_stats_delivery_rate = 'can_view_stats_delivery_rate' in request.form
        employee.can_view_stats_sales = 'can_view_stats_sales' in request.form
        employee.can_view_stats_amount_paid = 'can_view_stats_amount_paid' in request.form
        employee.can_view_stats_monthly_delivered = 'can_view_stats_monthly_delivered' in request.form
        employee.can_view_stats_monthly_invoices = 'can_view_stats_monthly_invoices' in request.form
        employee.can_view_stats_fixed_assets_expenses = 'can_view_stats_fixed_assets_expenses' in request.form
        employee.can_view_stats_employee_salaries = 'can_view_stats_employee_salaries' in request.form
        employee.can_view_stats_employee_debt = 'can_view_stats_employee_debt' in request.form
        employee.can_view_stats_operational_expenses = 'can_view_stats_operational_expenses' in request.form

        employee.can_view_suppliers = 'can_view_suppliers' in request.form
        employee.can_add_suppliers = 'can_add_suppliers' in request.form
        employee.can_edit_suppliers = 'can_edit_suppliers' in request.form
        employee.can_delete_suppliers = 'can_delete_suppliers' in request.form
        employee.can_view_supplier_invoices = 'can_view_supplier_invoices' in request.form
        employee.can_view_supplier_history = 'can_view_supplier_history' in request.form
        employee.can_pay_supplier_debt = 'can_pay_supplier_debt' in request.form
        employee.can_add_supplier_debt = 'can_add_supplier_debt' in request.form

        employee.can_view_invoices = 'can_view_invoices' in request.form
        employee.can_add_invoices = 'can_add_invoices' in request.form
        employee.can_edit_invoices = 'can_edit_invoices' in request.form
        employee.can_delete_invoices = 'can_delete_invoices' in request.form
        employee.can_view_invoice_details = 'can_view_invoice_details' in request.form

        employee.can_view_expenses = 'can_view_expenses' in request.form
        employee.can_add_expenses = 'can_add_expenses' in request.form
        employee.can_edit_expenses = 'can_edit_expenses' in request.form
        employee.can_delete_expenses = 'can_delete_expenses' in request.form
        employee.can_view_operational_expenses = 'can_view_operational_expenses' in request.form
        employee.can_view_fixed_assets = 'can_view_fixed_assets' in request.form

        employee.can_view_activity_log = 'can_view_activity_log' in request.form
        employee.can_manage_reminders = 'can_manage_reminders' in request.form
        employee.is_admin = 'is_admin' in request.form

        employee.can_view_followups = 'can_view_followups' in request.form
        employee.can_add_followups = 'can_add_followups' in request.form
        employee.can_edit_followups = 'can_edit_followups' in request.form
        employee.can_delete_followups = 'can_delete_followups' in request.form

        employee.can_view_damaged_products = 'can_view_damaged_products' in request.form
        employee.can_add_damaged_products = 'can_add_damaged_products' in request.form
        employee.can_delete_damaged_products = 'can_delete_damaged_products' in request.form

        employee.can_view_transactions = 'can_view_transactions' in request.form
        employee.can_add_transactions = 'can_add_transactions' in request.form
        employee.can_edit_transactions = 'can_edit_transactions' in request.form
        employee.can_delete_transactions = 'can_delete_transactions' in request.form

        try:
            db.session.commit()
            flash('تم تحديث صلاحيات الموظف بنجاح')
            if session.get('employee_id') == employee_id:
                session['employee_permissions'] = {
                    'can_view_customers': employee.can_view_customers,
                    'can_add_customers': employee.can_add_customers,
                    'can_edit_customers': employee.can_edit_customers,
                    'can_delete_customers': employee.can_delete_customers,
                    'can_view_customer_orders': employee.can_view_customer_orders,
                    'can_view_customer_logs': employee.can_view_customer_logs,
                    'can_view_customer_profile': employee.can_view_customer_profile,
                    'can_add_customer_logs': employee.can_add_customer_logs,
                    'can_delete_customer_logs': employee.can_delete_customer_logs,
                    'can_view_products': employee.can_view_products,
                    'can_add_products': employee.can_add_products,
                    'can_edit_products': employee.can_edit_products,
                    'can_delete_products': employee.can_delete_products,
                    'can_restore_products': employee.can_restore_products,
                    'can_delete_all_products': employee.can_delete_all_products,
                    'can_view_purchase_price': employee.can_view_purchase_price,
                    'can_view_sold_products_by_quantity': employee.can_view_sold_products_by_quantity,
                    'can_manage_stocktake': employee.can_manage_stocktake,
                    'can_view_orders': employee.can_view_orders,
                    'can_add_orders': employee.can_add_orders,
                    'can_edit_orders': employee.can_edit_orders,
                    'can_delete_orders': employee.can_delete_orders,
                    'can_update_order_status': employee.can_update_order_status,
                    'can_update_tracking': employee.can_update_tracking,
                    'can_view_orders_by_status': employee.can_view_orders_by_status,
                    'can_view_replacements': employee.can_view_replacements,
                    'can_add_replacements': employee.can_add_replacements,
                    'can_edit_replacements': employee.can_edit_replacements,
                    'can_delete_replacements': employee.can_delete_replacements,
                    'can_view_replacements_by_state': employee.can_view_replacements_by_state,
                    'can_view_returns': employee.can_view_returns,
                    'can_add_returns': employee.can_add_returns,
                    'can_view_returns_by_state': employee.can_view_returns_by_state,
                    'can_delete_returns': employee.can_delete_returns,
                    'can_view_fees': employee.can_view_fees,
                    'can_edit_fees': employee.can_edit_fees,
                    'can_delete_fees': employee.can_delete_fees,
                    'can_view_employees': employee.can_view_employees,
                    'can_add_employees': employee.can_add_employees,
                    'can_edit_employees': employee.can_edit_employees,
                    'can_delete_employees': employee.can_delete_employees,
                    'can_view_employee_logs': employee.can_view_employee_logs,
                    'can_manage_employee_salary': employee.can_manage_employee_salary,
                    'can_delete_salary_transactions': employee.can_delete_salary_transactions,
                    'can_view_employee_activity': employee.can_view_employee_activity,
                    'can_manage_attendance': employee.can_manage_attendance,
                    'can_view_statistics': employee.can_view_statistics,
                    'can_view_stats_stock': employee.can_view_stats_stock,
                    'can_view_stats_fixed_assets': employee.can_view_stats_fixed_assets,
                    'can_view_stats_total_debt': employee.can_view_stats_total_debt,
                    'can_view_stats_capital_growth': employee.can_view_stats_capital_growth,
                    'can_view_stats_daily': employee.can_view_stats_daily,
                    'can_view_stats_pending_orders': employee.can_view_stats_pending_orders,
                    'can_view_stats_net_profit': employee.can_view_stats_net_profit,
                    'can_view_stats_losses': employee.can_view_stats_losses,
                    'can_view_stats_delivery_rate': employee.can_view_stats_delivery_rate,
                    'can_view_stats_sales': employee.can_view_stats_sales,
                    'can_view_stats_amount_paid': employee.can_view_stats_amount_paid,
                    'can_view_stats_monthly_delivered': employee.can_view_stats_monthly_delivered,
                    'can_view_stats_monthly_invoices': employee.can_view_stats_monthly_invoices,
                    'can_view_stats_fixed_assets_expenses': employee.can_view_stats_fixed_assets_expenses,
                    'can_view_stats_employee_salaries': employee.can_view_stats_employee_salaries,
                    'can_view_stats_employee_debt': employee.can_view_stats_employee_debt,
                    'can_view_stats_operational_expenses': employee.can_view_stats_operational_expenses,
                    'can_view_suppliers': employee.can_view_suppliers,
                    'can_add_suppliers': employee.can_add_suppliers,
                    'can_edit_suppliers': employee.can_edit_suppliers,
                    'can_delete_suppliers': employee.can_delete_suppliers,
                    'can_view_supplier_invoices': employee.can_view_supplier_invoices,
                    'can_view_supplier_history': employee.can_view_supplier_history,
                    'can_pay_supplier_debt': employee.can_pay_supplier_debt,
                    'can_add_supplier_debt': employee.can_add_supplier_debt,
                    'can_view_invoices': employee.can_view_invoices,
                    'can_add_invoices': employee.can_add_invoices,
                    'can_edit_invoices': employee.can_edit_invoices,
                    'can_delete_invoices': employee.can_delete_invoices,
                    'can_view_invoice_details': employee.can_view_invoice_details,
                    'can_view_expenses': employee.can_view_expenses,
                    'can_add_expenses': employee.can_add_expenses,
                    'can_edit_expenses': employee.can_edit_expenses,
                    'can_delete_expenses': employee.can_delete_expenses,
                    'can_view_operational_expenses': employee.can_view_operational_expenses,
                    'can_view_fixed_assets': employee.can_view_fixed_assets,
                    'can_view_activity_log': employee.can_view_activity_log,
                    'can_manage_reminders': employee.can_manage_reminders,
                    'is_admin': employee.is_admin,
                    'is_active': employee.is_active
                    ,
                    'can_view_followups': employee.can_view_followups,
                    'can_add_followups': employee.can_add_followups,
                    'can_edit_followups': employee.can_edit_followups,
                    'can_delete_followups': employee.can_delete_followups,
                    'can_view_damaged_products': employee.can_view_damaged_products,
                    'can_add_damaged_products': employee.can_add_damaged_products,
                    'can_delete_damaged_products': employee.can_delete_damaged_products,
                    'can_view_transactions': employee.can_view_transactions,
                    'can_add_transactions': employee.can_add_transactions,
                    'can_edit_transactions': employee.can_edit_transactions,
                    'can_delete_transactions': employee.can_delete_transactions
                }
            return redirect(
                url_for(
                    'main.employee_profile',
                    employee_id=employee_id))
        except Exception as e:
            db.session.rollback()
            flash('حدث خطأ في تحديث الصلاحيات')
            return redirect(
                url_for(
                    'main.employee_permissions',
                    employee_id=employee_id))

    return render_template('employee_permissions.html', employee=employee)

@main.route('/employees/<int:employee_id>/attendance', methods=['GET', 'POST'])
@login_required
@permission_required('can_manage_attendance')
def attendance(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    today = date.today()
    selected_date = today
    date_str = request.args.get('date') or request.form.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            selected_date = today

    existing = AttendanceRecord.query.filter_by(
        employee_id=employee.id, date=selected_date).first()

    if request.method == 'POST':
        status = request.form.get('status')
        if status in ('present', 'half_day', 'absent'):
            if existing:
                existing.status = status
            else:
                db.session.add(AttendanceRecord(employee_id=employee.id, date=selected_date, status=status))
            try:
                db.session.commit()
                flash('تم تسجيل الحضور')
            except Exception as e:
                db.session.rollback()
                flash('حدث خطأ في تسجيل الحضور')
        return redirect(url_for('main.attendance', employee_id=employee.id, date=selected_date.isoformat()))

    month_start = selected_date.replace(day=1)
    if selected_date.month == 12:
        month_end = selected_date.replace(year=selected_date.year + 1, month=1, day=1)
    else:
        month_end = selected_date.replace(month=selected_date.month + 1, day=1)
    month_records = AttendanceRecord.query.filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.date >= month_start,
        AttendanceRecord.date < month_end
    ).order_by(AttendanceRecord.date).all()
    month_records_map = {r.date.isoformat(): r.status for r in month_records}

    return render_template('attendance_popup.html',
                           employee=employee,
                           selected_date=selected_date,
                           today=today,
                           existing=existing,
                           month_records_map=month_records_map,
                           month_start=month_start,
                            month_end=month_end)

@main.route('/attendance-today')
@login_required
@permission_required('can_manage_attendance')
def attendance_today():
    today = date.today()
    selected_date = today
    date_str = request.args.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            selected_date = today
    employees_list = Employee.query.filter_by(is_active=True, requires_attendance=True).order_by(Employee.name).all()
    records = AttendanceRecord.query.filter_by(date=selected_date).all()
    today_map = {r.employee_id: r.status for r in records}
    return render_template('attendance_today.html',
                           employees=employees_list,
                           today=today,
                           selected_date=selected_date,
                           today_map=today_map)

def _item_product_name(item):
    item_product = getattr(item, 'product', None)
    if item_product and getattr(item_product, 'name', None):
        return item_product.name
    product_id = getattr(item, 'product_id', None)
    return f'منتج محذوف #{product_id}' if product_id else 'منتج محذوف'


def _aggregate_products_from_orders(orders, only_new_replacement_products=False):
    aggregated = {}
    for order in orders:
        for item in (order.items or []):
            if only_new_replacement_products:
                item_state = (item.state or '').strip()
                if item_state not in ['منتج جديد', 'جديد']:
                    continue

            quantity = int(item.quantity or 0)
            if quantity <= 0:
                quantity = 1

            product_name = _item_product_name(item)
            key = product_name

            if key not in aggregated:
                aggregated[key] = {
                    'product_name': product_name,
                    'total_quantity': 0
                }

            aggregated[key]['total_quantity'] += quantity

    rows = []
    for grouped in aggregated.values():
        rows.append({
            'product_name': grouped['product_name'],
            'total_quantity': grouped['total_quantity']
        })

    rows.sort(key=lambda row: (-row['total_quantity'], row['product_name']))
    return rows


def _build_stats_context(selected_month, selected_year):
    month_start = datetime(selected_year, selected_month, 1)
    if selected_month == 12:
        month_end = datetime(selected_year + 1, 1, 1)
    else:
        month_end = datetime(selected_year, selected_month + 1, 1)

    days_in_month = (month_end - month_start).days
    today = datetime.now()
    _sel = (selected_year, selected_month)
    _cur = (today.year, today.month)
    if _sel == _cur:
        accrued_days = today.day
    elif _sel < _cur:
        accrued_days = days_in_month
    else:
        accrued_days = 0
    
    total_products = Product.query.filter_by(is_deleted=False).count()

    total_stock_value = db.session.query(
        func.coalesce(func.sum(func.coalesce(Product.purchase_price, 0) * func.coalesce(Product.stock, 0)), 0)
    ).filter(Product.is_deleted == False).scalar() or 0

    delivery_orders = Order.query.filter(
        Order.status == 'خرج للتوصيل',
        Order.status_updated_at >= month_start,
        Order.status_updated_at < month_end
    ).options(
        selectinload(Order.items).joinedload(OrderItem._product),
        selectinload(Order.items).joinedload(OrderItem.variant),
        selectinload(Order.items).joinedload(OrderItem.size_variant),
        selectinload(Order.items).joinedload(OrderItem.color_variant),
        selectinload(Order.items).joinedload(OrderItem.style_variant)
    ).all()

    replacement_delivery_orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'خرج للتوصيل',
        ReplacementOrder.status_updated_at >= month_start,
        ReplacementOrder.status_updated_at < month_end
    ).options(
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.size_variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.color_variant),
        selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem.style_variant)
    ).all()

    delivery_orders_remaining = sum((o.remaining_amount or 0) for o in delivery_orders)

    # المتبقي على الاستبدالات = (قيمة المنتجات الجديدة + رسوم العميل) - المدفوع - المسترد
    # نستخدم delivery_fees_customer (رسوم على العميل) وليس delivery_fees (رسوم على الشركة)
    replacement_delivery_remaining = sum(
        (o.total_amount or 0)
        + (o.delivery_fees_customer or 0)
        - (o.amount_paid or 0)
        - (o.customer_refund_amount or 0)
        for o in replacement_delivery_orders
    )

    delivery_orders_pending_value = delivery_orders_remaining or 0
    replacement_delivery_orders_pending_value = replacement_delivery_remaining or 0
    total_pending_orders_value = delivery_orders_pending_value + replacement_delivery_orders_pending_value

    pending_delivery_products_rows = _aggregate_products_from_orders(delivery_orders)
    pending_replacement_new_products_rows = _aggregate_products_from_orders(
        replacement_delivery_orders,
        only_new_replacement_products=True
    )

    total_employees = Employee.query.filter_by(is_active=True).count()
    total_salary = db.session.query(func.coalesce(func.sum(Employee.salary), 0)).filter(Employee.is_active == True).scalar() or 0

    monthly_tx_rows = db.session.query(
        SalaryTransaction.transaction_type,
        func.coalesce(func.sum(SalaryTransaction.amount), 0).label('total_amount')
    ).filter(
        SalaryTransaction.month == selected_month,
        SalaryTransaction.year == selected_year
    ).group_by(SalaryTransaction.transaction_type).all()

    monthly_tx_map = {row.transaction_type: row.total_amount for row in monthly_tx_rows}
    monthly_bonuses = monthly_tx_map.get('مكافأة', 0) or 0
    monthly_deductions = monthly_tx_map.get('خصم من المرتب', 0) or 0
    monthly_advances = monthly_tx_map.get('سلفه', 0) or 0

    employees_salary_details = []
    active_employees = Employee.query.filter_by(is_active=True).order_by(Employee.id.asc()).all()

    for emp in active_employees:
        summary = emp.get_salary_summary(selected_month, selected_year)
        if emp.requires_attendance:
            accrued_salary = Employee._calc_accrued_salary_attendance(
                summary['net_salary'] or 0, emp.id, selected_month, selected_year, days_in_month)
        else:
            accrued_salary = Employee._calc_accrued_salary(summary['net_salary'] or 0, selected_month, selected_year, days_in_month)
        employees_salary_details.append({
            'id': emp.id,
            'name': emp.name,
            'base_salary': summary['base_salary'] or 0,
            'bonuses': summary['total_bonuses'] or 0,
            'deductions': summary['total_deductions'] or 0,
            'advances': summary['total_advances'] or 0,
            'net_salary': summary['net_salary'] or 0,
            'accrued_salary': accrued_salary,
            'sales_commission': summary['sales_commission'] or 0,
            'delivered_sales': summary['delivered_sales'] or 0,
            'delivered_orders_count': summary['delivered_orders_count'] or 0
        })

    monthly_employee_salaries = sum(e['accrued_salary'] for e in employees_salary_details)

    employee_debt_list = []
    for emp in active_employees:
        running_balance = emp.get_running_balance()
        employee_debt_list.append({'id': emp.id, 'name': emp.name, 'balance': running_balance})
    employee_debt_list.sort(key=lambda x: x['balance'])

    total_supplier_debt = 0
    supplier_debts_details = []
    supplier_invoice_rows = db.session.query(
        Invoice.supplier_id,
        func.coalesce(func.sum(func.coalesce(Invoice.total_amount, 0) - func.coalesce(Invoice.paid_amount, 0)), 0).label('total_debt')
    ).group_by(Invoice.supplier_id).all()
    supplier_manual_rows = db.session.query(
        SupplierDebt.supplier_id,
        func.coalesce(func.sum(func.coalesce(SupplierDebt.amount, 0) - func.coalesce(SupplierDebt.paid_amount, 0)), 0).label('total_debt')
    ).filter(
        SupplierDebt.is_payment == False
    ).group_by(SupplierDebt.supplier_id).all()
    supplier_return_rows = db.session.query(
        Invoice.supplier_id,
        func.coalesce(func.sum(func.coalesce(SupplierReturn.total_amount, 0)), 0).label('total_returns')
    ).join(
        Invoice,
        SupplierReturn.invoice_id == Invoice.id
    ).group_by(Invoice.supplier_id).all()

    supplier_debt_map = defaultdict(float)
    for row in supplier_invoice_rows:
        supplier_debt_map[row.supplier_id] += row.total_debt or 0
    for row in supplier_manual_rows:
        supplier_debt_map[row.supplier_id] += row.total_debt or 0
    for row in supplier_return_rows:
        supplier_debt_map[row.supplier_id] -= row.total_returns or 0

    positive_supplier_ids = [sid for sid, debt_val in supplier_debt_map.items() if debt_val > 0]
    if positive_supplier_ids:
        suppliers_map = {
            s.id: s.name
            for s in Supplier.query.filter(Supplier.id.in_(positive_supplier_ids)).all()
        }
        for sid in positive_supplier_ids:
            supplier_debt = supplier_debt_map[sid]
            supplier_debts_details.append({
                'name': suppliers_map.get(sid, f'مورد {sid}'),
                'debt': supplier_debt,
                'type': 'مورد'
            })
            total_supplier_debt += supplier_debt

    party_debts_details = []
    total_party_debt = 0
    party_balance_rows = db.session.query(
        Party.id,
        Party.name,
        func.coalesce(func.sum(case(
            (Transaction.transaction_type == 'payable', func.coalesce(Transaction.amount, 0) - func.coalesce(Transaction.paid_amount, 0)),
            else_=0
        )), 0).label('total_payable'),
        func.coalesce(func.sum(case(
            (Transaction.transaction_type == 'receivable', func.coalesce(Transaction.amount, 0) - func.coalesce(Transaction.paid_amount, 0)),
            else_=0
        )), 0).label('total_receivable')
    ).outerjoin(
        Transaction,
        Transaction.party_id == Party.id
    ).group_by(Party.id, Party.name).all()

    for row in party_balance_rows:
        party_debt = (row.total_payable or 0) - (row.total_receivable or 0)
        if party_debt > 0:
            party_debts_details.append({
                'name': row.name,
                'debt': party_debt,
                'type': 'متعامل'
            })
            total_party_debt += party_debt
    
    total_debt = total_supplier_debt + total_party_debt
    all_debts_details = supplier_debts_details + party_debts_details

    for emp in active_employees:
        bal = emp.get_running_balance()
        if bal > 0:
            all_debts_details.append({
                'name': emp.name,
                'debt': bal,
                'type': 'موظف'
            })
            total_debt += bal
        elif bal < 0:
            all_debts_details.append({
                'name': emp.name,
                'debt': -bal,
                'type': 'موظف'
            })
            total_debt += -bal

    all_debts_details.sort(key=lambda x: x['debt'], reverse=True)

    total_expenses = db.session.query(
        func.coalesce(func.sum(Expense.amount), 0)
    ).filter(
        extract('month', Expense.date) == selected_month,
        extract('year', Expense.date) == selected_year
    ).scalar() or 0

    monthly_fixed_assets_list = Expense.query.filter(
        Expense.category == 'أصول ثابتة',
        extract('month', Expense.date) == selected_month,
        extract('year', Expense.date) == selected_year
    ).order_by(Expense.date.desc()).all()
    monthly_fixed_assets_expenses = sum(expense.amount for expense in monthly_fixed_assets_list)
    
    monthly_operational_expenses_list = Expense.query.filter(
        Expense.category != 'أصول ثابتة',
        extract('month', Expense.date) == selected_month,
        extract('year', Expense.date) == selected_year
    ).order_by(Expense.date.desc()).all()
    monthly_operational_expenses = sum(expense.amount for expense in monthly_operational_expenses_list)

    # ===== الطلبات العادية اللي وصلت في الشهر =====
    delivered_order_ids = db.session.scalars(
        select(OrderStatusHistory.order_id)
        .where(
            OrderStatusHistory.status == 'وصل',
            extract('month', OrderStatusHistory.timestamp) == selected_month,
            extract('year', OrderStatusHistory.timestamp) == selected_year
        )
        .distinct()
    ).all()

    delivered_orders_list = Order.query.filter(
        Order.id.in_(delivered_order_ids)
    ).options(
        selectinload(Order.items).joinedload(OrderItem._product),
        selectinload(Order.items).joinedload(OrderItem.variant),
        selectinload(Order.items).joinedload(OrderItem.size_variant),
        selectinload(Order.items).joinedload(OrderItem.color_variant),
        selectinload(Order.items).joinedload(OrderItem.style_variant)
    ).all()

    # ===== الاستبدالات اللي وصلت في الشهر =====
    delivered_replacement_ids = db.session.scalars(
        select(ReplacementOrderStatusHistory.replacement_order_id)
        .where(
            ReplacementOrderStatusHistory.status == 'وصل',
            extract('month', ReplacementOrderStatusHistory.timestamp) == selected_month,
            extract('year', ReplacementOrderStatusHistory.timestamp) == selected_year
        )
        .distinct()
    ).all()

    delivered_replacements_list = ReplacementOrder.query.filter(
        ReplacementOrder.id.in_(delivered_replacement_ids)
    ).options(
        selectinload(ReplacementOrder.items)
        .joinedload(ReplacementOrderItem._product)
    ).all()

    # ===== الإيراد =====
    total_sales_orders = sum(o.total_amount or 0 for o in delivered_orders_list)
    total_sales_replacements = sum(
        (r.total_amount or 0) - (r.delivery_fees_customer or 0) - (r.customer_refund_amount or 0)
        for r in delivered_replacements_list
    )
    total_sales_value = total_sales_orders + total_sales_replacements

    # ===== COGS =====
    total_cost_of_goods_sold = 0

    # الطلبات العادية — مؤقتاً من product.purchase_price
    for order in delivered_orders_list:
        for item in order.items:
            snapshot_price = item.purchase_price_snapshot
            if snapshot_price is None:
                snapshot_price = (item.product.purchase_price or 0) if item.product else 0
            total_cost_of_goods_sold += (snapshot_price or 0) * (item.quantity or 1)

    # الاستبدالات — من item.purchase_price مباشرة (المنتجات الجديدة فقط)
    for rep_order in delivered_replacements_list:
        for item in rep_order.items:
            if not item.is_returned:
                total_cost_of_goods_sold += (item.purchase_price or 0) * (item.quantity or 1)

    damaged_stock_loss = 0
    if delivered_replacement_ids:
        damaged_stock_loss = db.session.query(
            func.coalesce(func.sum(ReplacementOrder.damaged_products_loss), 0)
        ).filter(
            ReplacementOrder.id.in_(delivered_replacement_ids)
        ).scalar() or 0

    damaged_logs_sum = db.session.query(func.coalesce(func.sum(DamagedProductLog.total_loss), 0)).filter(
        extract('month', DamagedProductLog.created_at) == selected_month,
        extract('year', DamagedProductLog.created_at) == selected_year
    ).scalar() or 0
    damaged_stock_loss += damaged_logs_sum

    delivery_fees_loss = 0
    if delivered_replacement_ids:
        delivery_fees_loss = db.session.query(
            func.coalesce(func.sum(ReplacementOrder.delivery_fees_loss), 0)
        ).filter(
            ReplacementOrder.id.in_(delivered_replacement_ids)
        ).scalar() or 0

    total_losses = (damaged_stock_loss or 0) + (delivery_fees_loss or 0)

    # ReturnOrder statistics
    return_orders_count = db.session.query(func.count(ReturnOrder.id)).filter(
        extract('month', ReturnOrder.created_at) == selected_month,
        extract('year', ReturnOrder.created_at) == selected_year
    ).scalar() or 0

    return_order_refund_total = db.session.query(
        func.coalesce(func.sum(ReturnOrder.customer_refund_amount), 0)
    ).filter(
        extract('month', ReturnOrder.created_at) == selected_month,
        extract('year', ReturnOrder.created_at) == selected_year
    ).scalar() or 0

    # ===== صافي الربح =====
    total_monthly_expenses = monthly_fixed_assets_expenses + monthly_operational_expenses
    total_purchase_cost = total_cost_of_goods_sold

    net_profit = total_sales_value - total_monthly_expenses - (monthly_employee_salaries or 0) - total_purchase_cost - total_losses

    total_fixed_assets_expenses = db.session.query(
        func.coalesce(func.sum(Expense.amount), 0)
    ).filter(Expense.category == 'أصول ثابتة').scalar() or 0
    
    total_fixed_assets_value = total_fixed_assets_expenses

    fixed_assets_rows = db.session.query(
        Expense.title,
        func.coalesce(func.sum(Expense.amount), 0).label('total_value')
    ).filter(
        Expense.category == 'أصول ثابتة'
    ).group_by(Expense.title).order_by(func.coalesce(func.sum(Expense.amount), 0).desc()).all()

    all_fixed_assets_list = [
        {'name': row.title, 'total_value': row.total_value}
        for row in fixed_assets_rows
    ]
    
    monthly_invoices_list = Invoice.query.filter(
        extract('month', Invoice.date) == selected_month,
        extract('year', Invoice.date) == selected_year
    ).order_by(Invoice.date.desc()).all()
    monthly_inventory_purchases = sum(invoice.total_amount for invoice in monthly_invoices_list)

    # ─── نسبة التوصيل ───────────────────────────────────────────────────────
    # المقام: الطلبات التي تم إنشاؤها خلال الشهر.
    # البسط: من هذه الطلبات، كم طلباً وصل (بغض النظر عن تاريخ الوصول).
    total_orders_count = db.session.query(func.count(Order.id)).filter(
        Order.date >= month_start,
        Order.date < month_end
    ).scalar() or 0
    total_replacement_count = db.session.query(func.count(ReplacementOrder.id)).filter(
        ReplacementOrder.date >= month_start,
        ReplacementOrder.date < month_end
    ).scalar() or 0
    total_orders_count = total_orders_count + total_replacement_count

    delivered_orders_count = db.session.query(func.count(func.distinct(Order.id))).join(
        OrderStatusHistory, OrderStatusHistory.order_id == Order.id
    ).filter(
        Order.date >= month_start,
        Order.date < month_end,
        OrderStatusHistory.status == 'وصل'
    ).scalar() or 0
    delivered_replacement_count = db.session.query(
        func.count(func.distinct(ReplacementOrder.id))
    ).join(
        ReplacementOrderStatusHistory,
        ReplacementOrderStatusHistory.replacement_order_id == ReplacementOrder.id
    ).filter(
        ReplacementOrder.date >= month_start,
        ReplacementOrder.date < month_end,
        ReplacementOrderStatusHistory.status == 'وصل'
    ).scalar() or 0
    delivered_orders_count = delivered_orders_count + delivered_replacement_count

    try:
        delivered_percentage = round((delivered_orders_count / total_orders_count) * 100, 2) if total_orders_count else 0.0
    except Exception:
        delivered_percentage = 0.0

    order_status_rows = db.session.query(
        Order.status,
        func.count(Order.id)
    ).filter(
        Order.date >= month_start,
        Order.date < month_end
    ).group_by(Order.status).all()
    order_status_counts = {status: count for status, count in order_status_rows}
    monthly_orders_created_count = sum(order_status_counts.values())
    monthly_orders_new_count = order_status_counts.get('جديد', 0)
    monthly_orders_delivery_count = order_status_counts.get('خرج للتوصيل', 0)
    monthly_orders_delivered_count = order_status_counts.get('وصل', 0)
    monthly_orders_other_count = max(
        monthly_orders_created_count - monthly_orders_new_count - monthly_orders_delivery_count - monthly_orders_delivered_count,
        0
    )

    replacement_status_rows = db.session.query(
        ReplacementOrder.status,
        func.count(ReplacementOrder.id)
    ).filter(
        ReplacementOrder.date >= month_start,
        ReplacementOrder.date < month_end
    ).group_by(ReplacementOrder.status).all()
    replacement_status_counts = {status: count for status, count in replacement_status_rows}
    monthly_replacements_created_count = sum(replacement_status_counts.values())
    monthly_replacements_new_count = replacement_status_counts.get('جديد', 0)
    monthly_replacements_delivery_count = replacement_status_counts.get('خرج للتوصيل', 0)
    monthly_replacements_delivered_count = replacement_status_counts.get('وصل', 0)
    monthly_replacements_other_count = max(
        monthly_replacements_created_count - monthly_replacements_new_count - monthly_replacements_delivery_count - monthly_replacements_delivered_count,
        0
    )

    monthly_orders_other_list = Order.query.filter(
        Order.date >= month_start,
        Order.date < month_end,
        ~Order.status.in_(['جديد', 'خرج للتوصيل', 'وصل'])
    ).options(
        selectinload(Order.customer)
    ).order_by(Order.date.desc(), Order.id.desc()).all()

    monthly_replacements_other_list = ReplacementOrder.query.filter(
        ReplacementOrder.date >= month_start,
        ReplacementOrder.date < month_end,
        ~ReplacementOrder.status.in_(['جديد', 'خرج للتوصيل', 'وصل'])
    ).options(
        selectinload(ReplacementOrder.customer)
    ).order_by(ReplacementOrder.date.desc(), ReplacementOrder.id.desc()).all()

    delivered_rate_orders_list = Order.query.join(
        OrderStatusHistory, OrderStatusHistory.order_id == Order.id
    ).filter(
        Order.date >= month_start,
        Order.date < month_end,
        OrderStatusHistory.status == 'وصل'
    ).options(
        selectinload(Order.customer)
    ).distinct().order_by(Order.id.desc()).all()

    delivered_rate_replacements_list = ReplacementOrder.query.join(
        ReplacementOrderStatusHistory,
        ReplacementOrderStatusHistory.replacement_order_id == ReplacementOrder.id
    ).filter(
        ReplacementOrder.date >= month_start,
        ReplacementOrder.date < month_end,
        ReplacementOrderStatusHistory.status == 'وصل'
    ).options(
        selectinload(ReplacementOrder.customer)
    ).distinct().order_by(ReplacementOrder.id.desc()).all()
    
    delivered_in_month_ids = db.session.query(OrderStatusHistory.order_id).filter(
        OrderStatusHistory.status == 'وصل',
        OrderStatusHistory.timestamp >= month_start,
        OrderStatusHistory.timestamp < month_end
    ).distinct().all()
    
    delivered_in_month_ids = [row[0] for row in delivered_in_month_ids]
    
    delivered_in_month_orders = []
    delivered_at_map = {}
    monthly_delivered_value = 0
    if delivered_in_month_ids:
        delivered_in_month_orders = Order.query.filter(
            Order.id.in_(delivered_in_month_ids)
        ).order_by(Order.id.desc()).all()
        # قيمة التسليم = مجموع قيمة المنتجات فقط (total_amount)
        # رسوم التوصيل و COD لا تُعدّ إيراداً للمبيعات، لذا لا تُضاف هنا
        monthly_delivered_value = sum(
            (order.total_amount or 0)
            for order in delivered_in_month_orders
        )
        delivered_at_rows = db.session.query(
            OrderStatusHistory.order_id,
            func.max(OrderStatusHistory.timestamp).label('delivered_at')
        ).filter(
            OrderStatusHistory.status == 'وصل',
            OrderStatusHistory.timestamp >= month_start,
            OrderStatusHistory.timestamp < month_end,
            OrderStatusHistory.order_id.in_(delivered_in_month_ids)
        ).group_by(OrderStatusHistory.order_id).all()
        delivered_at_map = {row.order_id: row.delivered_at for row in delivered_at_rows}
    
    longest_status_info = None
    if delivered_in_month_ids:
        try:
            status_durations = []
            timeline_rows = OrderStatusHistory.query.filter(
                OrderStatusHistory.order_id.in_(delivered_in_month_ids)
            ).order_by(OrderStatusHistory.order_id.asc(), OrderStatusHistory.timestamp.asc()).all()

            timeline_by_order = defaultdict(list)
            for rec in timeline_rows:
                timeline_by_order[rec.order_id].append(rec)

            for order_id, timeline in timeline_by_order.items():
                if len(timeline) < 2:
                    continue
                
                for i in range(len(timeline) - 1):
                    current_record = timeline[i]
                    next_record = timeline[i + 1]
                    
                    if next_record.status == 'وصل':
                        duration_seconds = (next_record.timestamp - current_record.timestamp).total_seconds()
                        duration_hours = duration_seconds / 3600
                        
                        status_durations.append({
                            'order_id': order_id,
                            'from_status': current_record.status,
                            'duration_hours': duration_hours,
                            'start_time': current_record.timestamp,
                            'end_time': next_record.timestamp
                        })
            
            if status_durations:
                longest = max(status_durations, key=lambda x: x['duration_hours'])
                longest_status_info = {
                    'order_id': longest['order_id'],
                    'status': longest['from_status'],
                    'duration_hours': round(longest['duration_hours'], 2),
                    'duration_days': round(longest['duration_hours'] / 24, 2)
                }
        except Exception as e:
            current_app.logger.error(f"Error calculating longest status duration: {e}")
            longest_status_info = None

    total_amount_paid_orders = db.session.query(
        func.coalesce(func.sum(Order.amount_paid), 0)
    ).filter(
        extract('month', Order.date) == selected_month,
        extract('year', Order.date) == selected_year
    ).scalar() or 0
    total_amount_paid_replacements = db.session.query(
        func.coalesce(func.sum(ReplacementOrder.amount_paid), 0)
    ).filter(
        extract('month', ReplacementOrder.date) == selected_month,
        extract('year', ReplacementOrder.date) == selected_year
    ).scalar() or 0
    total_amount_paid = (total_amount_paid_orders or 0) + (total_amount_paid_replacements or 0)

    orders_with_payments = Order.query.filter(
        extract('month', Order.date) == selected_month,
        extract('year', Order.date) == selected_year,
        Order.amount_paid > 0
    ).order_by(Order.date.desc()).all()
    replacement_orders_with_payments = ReplacementOrder.query.filter(
        extract('month', ReplacementOrder.date) == selected_month,
        extract('year', ReplacementOrder.date) == selected_year,
        ReplacementOrder.amount_paid > 0
    ).order_by(ReplacementOrder.date.desc()).all()

    try:
        if not delivered_order_ids:
            top_products = []
        else:
            quantity_sum_expr = func.coalesce(func.sum(OrderItem.quantity), 0)
            profit_sum_expr = func.coalesce(func.sum((OrderItem.price - func.coalesce(OrderItem.purchase_price_snapshot, Product.purchase_price)) * OrderItem.quantity), 0)

            top_products_query = (
                db.session.query(
                    Product.id.label('product_id'),
                    Product.name.label('product_name'),
                    quantity_sum_expr.label('total_quantity'),
                    profit_sum_expr.label('total_profit')
                )
                .join(OrderItem, OrderItem.product_id == Product.id, isouter=True)
                .outerjoin(ProductVariant, ProductVariant.id == OrderItem.variant_id)
                .join(Order, Order.id == OrderItem.order_id)
                .filter(
                    Order.id.in_(delivered_order_ids),
                    Product.is_deleted == False,
                    or_(OrderItem.product_id == Product.id, ProductVariant.product_id == Product.id)
                )
                .group_by(Product.id, Product.name)
                .having(quantity_sum_expr > 0)
                .order_by(quantity_sum_expr.desc())
            )
            top_products = top_products_query.all()
    except Exception as e:
        current_app.logger.error(f"Error computing top products: {e}")
        top_products = []

    products_for_stock = Product.query.filter_by(is_deleted=False).order_by(Product.stock.desc()).all()

    damaged_logs_list = db.session.query(DamagedProductLog).filter(
        extract('month', DamagedProductLog.created_at) == selected_month,
        extract('year', DamagedProductLog.created_at) == selected_year
    ).order_by(DamagedProductLog.created_at.desc()).all()

    replacement_orders_loss_list = []
    if delivered_replacement_ids:
        replacement_orders_loss_list = ReplacementOrder.query.filter(
            ReplacementOrder.id.in_(delivered_replacement_ids),
            or_(
                func.coalesce(ReplacementOrder.damaged_products_loss, 0) > 0,
                func.coalesce(ReplacementOrder.delivery_fees_loss, 0) > 0
            )
        ).options(
            selectinload(ReplacementOrder.items).joinedload(ReplacementOrderItem._product)
        ).order_by(ReplacementOrder.date.desc()).all()

    return dict(
        total_products=total_products,
        total_stock_value=total_stock_value,
        total_pending_orders_value=total_pending_orders_value,
        delivery_orders_pending_value=delivery_orders_pending_value,
        replacement_delivery_orders_pending_value=replacement_delivery_orders_pending_value,
        pending_delivery_products_rows=pending_delivery_products_rows,
        pending_replacement_new_products_rows=pending_replacement_new_products_rows,
        delivery_orders=delivery_orders,
        replacement_delivery_orders=replacement_delivery_orders,
        net_profit=net_profit,
        monthly_employee_salaries=monthly_employee_salaries,
        employee_debt_list=employee_debt_list,
        employees_salary_details=employees_salary_details,
        monthly_bonuses=monthly_bonuses,
        monthly_deductions=monthly_deductions,
        monthly_advances=monthly_advances,
        total_sales_orders=total_sales_orders,
        total_sales_replacements=total_sales_replacements,
        total_sales_value=total_sales_value,
        total_cost_of_goods_sold=total_cost_of_goods_sold,
        monthly_cogs=total_cost_of_goods_sold,
        delivered_orders_list=delivered_orders_list,
        delivered_rate_orders_list=delivered_rate_orders_list,
        delivered_rate_replacements_list=delivered_rate_replacements_list,
        delivered_replacements_list=delivered_replacements_list,
        total_losses=total_losses,
        damaged_stock_loss=damaged_stock_loss,
        delivery_fees_loss=delivery_fees_loss,
        total_fixed_assets_value=total_fixed_assets_value,
        all_fixed_assets_list=all_fixed_assets_list,
        total_debt=total_debt,
        all_debts_details=all_debts_details,
        monthly_inventory_purchases=monthly_inventory_purchases,
        monthly_invoices_list=monthly_invoices_list,
        monthly_fixed_assets_expenses=monthly_fixed_assets_expenses,
        monthly_fixed_assets_list=monthly_fixed_assets_list,
        monthly_operational_expenses=monthly_operational_expenses,
        monthly_operational_expenses_list=monthly_operational_expenses_list,
        selected_month=selected_month,
        selected_year=selected_year,
        delivered_percentage=delivered_percentage,
        delivered_orders_count=delivered_orders_count,
        total_orders_count=total_orders_count,
        monthly_orders_created_count=monthly_orders_created_count,
        monthly_orders_new_count=monthly_orders_new_count,
        monthly_orders_delivery_count=monthly_orders_delivery_count,
        monthly_orders_delivered_count=monthly_orders_delivered_count,
        monthly_orders_other_count=monthly_orders_other_count,
        monthly_replacements_created_count=monthly_replacements_created_count,
        monthly_replacements_new_count=monthly_replacements_new_count,
        monthly_replacements_delivery_count=monthly_replacements_delivery_count,
        monthly_replacements_delivered_count=monthly_replacements_delivered_count,
        monthly_replacements_other_count=monthly_replacements_other_count,
        top_products=top_products,
        monthly_orders_other_list=monthly_orders_other_list,
        monthly_replacements_other_list=monthly_replacements_other_list,
        products_for_stock=products_for_stock,
        damaged_logs_list=damaged_logs_list,
        replacement_orders_loss_list=replacement_orders_loss_list,
        return_orders_count=return_orders_count,
        return_order_refund_total=return_order_refund_total,
        total_amount_paid=total_amount_paid,
        total_amount_paid_orders=total_amount_paid_orders,
        total_amount_paid_replacements=total_amount_paid_replacements,
        orders_with_payments=orders_with_payments,
        replacement_orders_with_payments=replacement_orders_with_payments,
        monthly_delivered_value=monthly_delivered_value,
        delivered_in_month_orders=delivered_in_month_orders,
        delivered_at_map=delivered_at_map,
        longest_status_info=longest_status_info,
        current_year=datetime.now().year,
    )


@main.route('/statistics')
@login_required
@permission_required('can_view_statistics')
def statistics():
    selected_month = request.args.get('month', datetime.now().month, type=int)
    selected_year = request.args.get('year', datetime.now().year, type=int)
    ctx = _build_stats_context(selected_month, selected_year)
    from app.permissions import has_permission as _hp
    def _sp(p): return _hp(p)
    ctx.update({
        'can_view_stats_stock': _sp('can_view_stats_stock'),
        'can_view_stats_fixed_assets': _sp('can_view_stats_fixed_assets'),
        'can_view_stats_total_debt': _sp('can_view_stats_total_debt'),
        'can_view_stats_capital_growth': _sp('can_view_stats_capital_growth'),
        'can_view_stats_daily': _sp('can_view_stats_daily'),
        'can_view_stats_pending_orders': _sp('can_view_stats_pending_orders'),
        'can_view_stats_net_profit': _sp('can_view_stats_net_profit'),
        'can_view_stats_losses': _sp('can_view_stats_losses'),
        'can_view_stats_delivery_rate': _sp('can_view_stats_delivery_rate'),
        'can_view_stats_sales': _sp('can_view_stats_sales'),
        'can_view_stats_amount_paid': _sp('can_view_stats_amount_paid'),
        'can_view_stats_monthly_delivered': _sp('can_view_stats_monthly_delivered'),
        'can_view_stats_monthly_invoices': _sp('can_view_stats_monthly_invoices'),
        'can_view_stats_fixed_assets_expenses': _sp('can_view_stats_fixed_assets_expenses'),
        'can_view_stats_employee_salaries': _sp('can_view_stats_employee_salaries'),
        'can_view_stats_employee_debt': _sp('can_view_stats_employee_debt'),
        'can_view_stats_operational_expenses': _sp('can_view_stats_operational_expenses'),
        'can_view_sold_products_by_quantity': _sp('can_view_sold_products_by_quantity'),
    })
    return render_template('statistics.html', **ctx)

def calculate_current_capital():
    fixed_assets_value = sum(
        expense.amount for expense in Expense.query.filter(
            Expense.category == 'أصول ثابتة'
        ).all()
    )
    
    stock_value = 0
    for product in Product.query.filter_by(is_deleted=False).all():
        stock_value += product.purchase_price * product.stock
    
    delivery_orders = Order.query.filter(Order.status == 'خرج للتوصيل').all()
    delivery_orders_remaining = sum(
        (o.remaining_amount or 0)
        for o in delivery_orders
    )
    
    replacement_delivery_orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'خرج للتوصيل'
    ).all()
    replacement_delivery_remaining = sum(
        (o.total_amount or 0) + (o.cod_fee_applied or 0) - (o.amount_paid or 0) - (o.customer_refund_amount or 0)
        for o in replacement_delivery_orders
    )
    pending_orders_value = delivery_orders_remaining + replacement_delivery_remaining

    supplier_returns_rows = db.session.query(
        Invoice.supplier_id,
        func.coalesce(func.sum(func.coalesce(SupplierReturn.total_amount, 0)), 0).label('total_returns')
    ).join(
        Invoice,
        SupplierReturn.invoice_id == Invoice.id
    ).group_by(Invoice.supplier_id).all()
    supplier_returns_map = {
        row.supplier_id: (row.total_returns or 0)
        for row in supplier_returns_rows
    }

    total_supplier_debt = 0
    for supplier in Supplier.query.all():
        invoices = Invoice.query.filter_by(supplier_id=supplier.id).all()
        manual_debts = SupplierDebt.query.filter_by(supplier_id=supplier.id, is_payment=False).all()
        supplier_debt = sum((invoice.total_amount or 0) - (invoice.paid_amount or 0) for invoice in invoices)
        supplier_debt += sum((debt.amount or 0) - (debt.paid_amount or 0) for debt in manual_debts)
        supplier_debt -= supplier_returns_map.get(supplier.id, 0)
        if supplier_debt > 0:
            total_supplier_debt += supplier_debt

    total_party_debt = 0
    for party in Party.query.all():
        party_debt = (party.total_payable or 0) - (party.total_receivable or 0)
        if party_debt > 0:
            total_party_debt += party_debt

    total_debt = total_supplier_debt + total_party_debt
    
    total_capital = fixed_assets_value + stock_value + pending_orders_value
    net_capital_after_debt = total_capital - total_debt
    
    return {
        'fixed_assets_value': fixed_assets_value,
        'stock_value': stock_value,
        'pending_orders_value': pending_orders_value,
        'total_capital': total_capital,
        'total_debt': total_debt,
        'net_capital_after_debt': net_capital_after_debt
    }

def auto_save_previous_month_snapshot():
    pass

@main.route('/api/save_capital_snapshot', methods=['POST'])
@login_required
@permission_required('can_view_statistics')
def save_capital_snapshot():
    try:
        data = request.get_json()
        year = data.get('year', datetime.now().year)
        month = data.get('month', datetime.now().month)
        
        capital_data = calculate_current_capital()
        
        snapshot = CapitalSnapshot.get_or_create_snapshot(year, month)
        snapshot.fixed_assets_value = capital_data['fixed_assets_value']
        snapshot.pending_orders_value = capital_data['pending_orders_value']
        snapshot.stock_value = capital_data['stock_value']
        snapshot.total_capital = capital_data['total_capital']
        snapshot.total_debt = capital_data['total_debt']
        snapshot.net_capital_after_debt = capital_data['net_capital_after_debt']
        snapshot.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم حفظ snapshot لشهر {month}/{year}',
            'data': {
                'year': year,
                'month': month,
                'total_capital': snapshot.total_capital
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/api/save_custom_capital_snapshot', methods=['POST'])
@login_required
@permission_required('can_view_statistics')
def save_custom_capital_snapshot():
    try:
        data = request.get_json()
        current_app.logger.info(f"📥 Received data: {data}")
        
        total_capital = data.get('total_capital')
        current_app.logger.info(f"💰 Total capital value: {total_capital}, type: {type(total_capital)}")
        
        if total_capital is None:
            current_app.logger.error("❌ Total capital is None")
            return jsonify({
                'success': False, 
                'message': 'لم يتم إرسال قيمة رأس المال'
            }), 400
            
        try:
            total_capital = float(total_capital)
        except (ValueError, TypeError) as e:
            current_app.logger.error(f"❌ Cannot convert to float: {e}")
            return jsonify({
                'success': False, 
                'message': 'قيمة رأس المال غير صحيحة'
            }), 400
            
        if total_capital <= 0:
            current_app.logger.error(f"❌ Total capital <= 0: {total_capital}")
            return jsonify({
                'success': False, 
                'message': 'يرجى إدخال قيمة أكبر من صفر'
            }), 400
        
        current_date = datetime.now()
        if current_date.month == 1:
            snap_month = 12
            snap_year = current_date.year - 1
        else:
            snap_month = current_date.month - 1
            snap_year = current_date.year
        
        current_app.logger.info(f"📅 Saving snapshot for: {snap_month}/{snap_year}")
        
        current_capital_data = calculate_current_capital()
        
        snapshot = CapitalSnapshot.get_or_create_snapshot(snap_year, snap_month)
        snapshot.total_capital = current_capital_data['total_capital']
        snapshot.fixed_assets_value = current_capital_data['fixed_assets_value']
        snapshot.stock_value = current_capital_data['stock_value']
        snapshot.pending_orders_value = current_capital_data['pending_orders_value']
        snapshot.total_debt = current_capital_data['total_debt']
        snapshot.net_capital_after_debt = current_capital_data['net_capital_after_debt']
        snapshot.updated_at = datetime.utcnow()
        
        current_app.logger.info(f"💾 About to commit snapshot: {snapshot}")
        current_app.logger.info(f"   Fixed Assets: {snapshot.fixed_assets_value}")
        current_app.logger.info(f"   Stock: {snapshot.stock_value}")
        current_app.logger.info(f"   Pending Orders: {snapshot.pending_orders_value}")
        current_app.logger.info(f"   Debt: {snapshot.total_debt}")
        current_app.logger.info(f"   Net After Debt: {snapshot.net_capital_after_debt}")
        db.session.commit()
        current_app.logger.info(f"✅ Snapshot saved successfully with ID: {snapshot.id}")
        
        # ── حفظ سجل في تاريخ نمو رأس المال ──
        last_history = CapitalGrowthHistory.query.order_by(CapitalGrowthHistory.saved_at.desc()).first()
        previous_capital = None
        if last_history:
            previous_capital = (
                last_history.net_capital_after_debt
                if last_history.net_capital_after_debt is not None
                else last_history.total_capital
            )
        growth_rate = None
        if previous_capital and previous_capital > 0:
            growth_rate = round((current_capital_data['net_capital_after_debt'] - previous_capital) / previous_capital * 100, 2)
        history_entry = CapitalGrowthHistory(
            saved_at=datetime(snap_year, snap_month, 1, 12, 0, 0),
            total_capital=current_capital_data['total_capital'],
            previous_capital=previous_capital,
            growth_rate=growth_rate,
            fixed_assets_value=current_capital_data['fixed_assets_value'],
            stock_value=current_capital_data['stock_value'],
            pending_orders_value=current_capital_data['pending_orders_value'],
            total_debt=current_capital_data['total_debt'],
            net_capital_after_debt=current_capital_data['net_capital_after_debt']
        )
        db.session.add(history_entry)
        db.session.commit()
        current_app.logger.info(f"✅ CapitalGrowthHistory entry saved: capital={history_entry.total_capital}, growth_rate={history_entry.growth_rate}")
        
        saved_snapshot = CapitalSnapshot.query.filter_by(year=snap_year, month=snap_month).first()
        if saved_snapshot:
            current_app.logger.info(f"✅ Verification: Found snapshot ID={saved_snapshot.id}, capital={saved_snapshot.total_capital}")
        else:
            current_app.logger.error(f"❌ Verification failed: No snapshot found for {snap_month}/{snap_year}")
        
        return jsonify({
            'success': True,
            'message': f'تم حفظ رأس المال للشهر السابق ({snap_month}/{snap_year}) بقيمة {current_capital_data["total_capital"]:.2f} ج.م',
            'snapshot_id': snapshot.id
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌❌ Error in save_custom_capital_snapshot: {str(e)}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False, 
            'message': f'حدث خطأ: {str(e)}'
        }), 500

@main.route('/api/capital_growth_rate')
@login_required
@permission_required('can_view_statistics')
def capital_growth_rate():
    try:
        current_capital = calculate_current_capital()
        current_date = datetime.now()
        
        last_snapshot = CapitalSnapshot.query.filter(
            or_(
                CapitalSnapshot.year < current_date.year,
                and_(
                    CapitalSnapshot.year == current_date.year,
                    CapitalSnapshot.month < current_date.month
                )
            )
        ).order_by(
            CapitalSnapshot.year.desc(),
            CapitalSnapshot.month.desc()
        ).first()
        
        if not last_snapshot:
            return jsonify({
                'success': True,
                'has_previous_data': False,
                'current_capital': current_capital['total_capital'],
                'current_breakdown': {
                    'fixed_assets': current_capital['fixed_assets_value'],
                    'stock': current_capital['stock_value'],
                    'pending_orders': current_capital['pending_orders_value'],
                    'total_debt': current_capital['total_debt'],
                    'net_after_debt': current_capital['net_capital_after_debt']
                },
                'total_debt': current_capital['total_debt'],
                'net_capital_after_debt': current_capital['net_capital_after_debt']
            })
        
        last_month_capital = last_snapshot.total_capital
        last_month_net_capital = (
            last_snapshot.net_capital_after_debt
            if last_snapshot.net_capital_after_debt is not None
            else (last_snapshot.total_capital - (last_snapshot.total_debt or 0))
        )
        capital_change = current_capital['net_capital_after_debt'] - last_month_net_capital
        growth_rate = (capital_change / last_month_net_capital * 100) if last_month_net_capital > 0 else 0
        
        return jsonify({
            'success': True,
            'has_previous_data': True,
            'current_capital': current_capital['total_capital'],
            'current_breakdown': {
                'fixed_assets': current_capital['fixed_assets_value'],
                'stock': current_capital['stock_value'],
                'pending_orders': current_capital['pending_orders_value'],
                'total_debt': current_capital['total_debt'],
                'net_after_debt': current_capital['net_capital_after_debt']
            },
            'last_month_capital': last_month_capital,
            'last_month_net_capital_after_debt': last_month_net_capital,
            'last_month_date': {
                'year': last_snapshot.year,
                'month': last_snapshot.month
            },
            'last_month_breakdown': {
                'fixed_assets': last_snapshot.fixed_assets_value,
                'stock': last_snapshot.stock_value,
                'pending_orders': last_snapshot.pending_orders_value,
                'total_debt': last_snapshot.total_debt,
                'net_after_debt': last_snapshot.net_capital_after_debt
            },
            'capital_change': capital_change,
            'growth_rate': round(growth_rate, 2),
            'total_debt': current_capital['total_debt'],
            'net_capital_after_debt': current_capital['net_capital_after_debt']
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in capital_growth_rate: {str(e)}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/api/capital_growth_history')
@login_required
@permission_required('can_view_statistics')
def capital_growth_history_api():
    try:
        offset = request.args.get('offset', 0, type=int)
        limit  = request.args.get('limit',  12, type=int)
        limit  = min(limit, 100)

        items, total = CapitalGrowthHistory.get_history(offset=offset, limit=limit)

        return jsonify({
            'success'  : True,
            'total'    : total,
            'offset'   : offset,
            'limit'    : limit,
            'has_more' : (offset + limit) < total,
            'items'    : [{
                'id'                   : item.id,
                'saved_at'             : item.saved_at.strftime('%Y-%m-%d %H:%M'),
                'period'               : item.saved_at.strftime('%Y-%m'),
                'total_capital'        : item.total_capital,
                'previous_capital'     : item.previous_capital,
                'growth_rate'          : item.growth_rate,
                'fixed_assets_value'   : item.fixed_assets_value,
                'stock_value'          : item.stock_value,
                'pending_orders_value' : item.pending_orders_value,
                'total_debt'           : item.total_debt,
                'net_capital_after_debt': item.net_capital_after_debt,
            } for item in items]
        })
    except Exception as e:
        current_app.logger.error(f"Error in capital_growth_history_api: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/damaged-products')
@login_required
@permission_required('can_view_damaged_products')
def damaged_products():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    query = DamagedProductLog.query.order_by(DamagedProductLog.created_at.desc())
    if month:
        query = query.filter(extract('month', DamagedProductLog.created_at) == month)
    if year:
        query = query.filter(extract('year', DamagedProductLog.created_at) == year)
    logs = query.all()

    total_loss = sum(l.total_loss for l in logs)
    return render_template('damaged_products.html', logs=logs, total_loss=total_loss, month=month, year=year)

@main.route('/damaged-products/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_products')
def add_damaged_product():
    products = Product.query.options(joinedload(Product.variants)).filter(Product.is_deleted == False, Product.is_bundle == False).all()

    def serialize_product(product):
        return {
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'purchase_price': product.purchase_price,
            'wholesale_price': product.wholesale_price,
            'stock': product.stock,
            'variants': [
                {
                    'id': v.id,
                    'group_name': v.group_name,
                    'variant_name': v.variant_name,
                    'price': v.price
                } for v in product.variants
            ]
        }

    products_data = [serialize_product(p) for p in products]

    if request.method == 'POST':
        try:
            product_id = int(request.form.get('product_id', '0'))
            quantity = int(request.form.get('quantity', '1'))
            notes = request.form.get('notes', '').strip() or None

            if quantity <= 0:
                flash('الكمية غير صحيحة', 'error')
                return redirect(url_for('main.add_damaged_product'))

            product = _get_locked_product(product_id)
            if not product:
                flash('المنتج غير موجود', 'error')
                return redirect(url_for('main.add_damaged_product'))

            if product.stock < quantity:
                flash(f'المخزون غير كافٍ للمنتج {product.name}. المتوفر: {product.stock}', 'error')
                return redirect(url_for('main.add_damaged_product'))

            if _block_if_deleted(product, 'هذا السجل', 'deduct'):
                return redirect(url_for('main.add_damaged_product'))

            product.stock -= quantity
            _log_stock_delta("damaged_product deduct", {product.id: -quantity})

            snapshot = float(product.purchase_price or 0)
            total_loss = snapshot * quantity
            log = DamagedProductLog(
                product_id=product.id,
                quantity=quantity,
                purchase_price_snapshot=snapshot,
                total_loss=total_loss,
                notes=notes,
                created_by=session.get('employee_id')
            )
            db.session.add(log)

            log_activity(
                action='create',
                entity_type='damaged_product',
                entity_id=log.id,
                entity_name=product.name,
                details=f'إضافة منتج تالف: {product.name} كمية {quantity} خسارة {total_loss:.2f}'
            )

            db.session.commit()
            flash('تم تسجيل المنتج التالف وتحديث المخزون', 'success')
            return redirect(url_for('main.damaged_products'))
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة المنتج التالف: {e}', 'error')
            return redirect(url_for('main.add_damaged_product'))

    return render_template('add_damaged_product.html', products=products_data)

@main.route('/damaged-products/<int:log_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_products')
def delete_damaged_product(log_id):
    log = DamagedProductLog.query.get(log_id)
    if not log:
        flash('السجل غير موجود', 'error')
        return redirect(url_for('main.damaged_products'))
    try:
        product_name = log.product.name if log.product else f'ID {log.product_id}'
        if log.product:
            product = _get_locked_product(log.product_id)
            if product:
                product.stock = (product.stock or 0) + (log.quantity or 0)
        db.session.delete(log)
        db.session.commit()
        log_activity(
            action='delete',
            entity_type='damaged_product',
            entity_id=log_id,
            entity_name=product_name,
            details='حذف سجل المنتج التالف وتمت إعادة الكمية للمخزون'
        )
        flash('تم حذف السجل', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'تعذر حذف السجل: {e}', 'error')
    return redirect(url_for('main.damaged_products'))

@main.route('/suppliers')
@login_required
@permission_required('can_view_suppliers')
def suppliers():
    search_query = request.args.get('search', '')
    if search_query:
        suppliers_list = Supplier.query.filter(
            Supplier.phone.contains(search_query)).all()
    else:
        suppliers_list = Supplier.query.all()
    
    for supplier in suppliers_list:
        invoices = Invoice.query.filter_by(supplier_id=supplier.id).all()
        manual_debts = SupplierDebt.query.filter_by(supplier_id=supplier.id, is_payment=False).all()
        total_debt = sum((invoice.total_amount or 0) - (invoice.paid_amount or 0) for invoice in invoices)
        total_debt += sum((debt.amount or 0) - (debt.paid_amount or 0) for debt in manual_debts)
        supplier.actual_total_debt = total_debt
    
    return render_template(
        'suppliers.html',
        suppliers=suppliers_list,
        search_query=search_query)

@main.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_suppliers')
def add_supplier():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        address = request.form['address']
        email = request.form.get('email', '')
        try:
            initial_debt = float(request.form.get('total_debt', 0))
        except ValueError:
            initial_debt = 0.0

        new_supplier = Supplier(
            name=name,
            phone=phone,
            address=address,
            email=email,
            total_debt=0
        )
        db.session.add(new_supplier)
        db.session.commit()

        if initial_debt and initial_debt > 0:
            try:
                new_debt = SupplierDebt(
                    supplier_id=new_supplier.id,
                    amount=initial_debt,
                    paid_amount=0,
                    notes='رصيد افتتاحي',
                    created_by=session.get('employee_id')
                )
                db.session.add(new_debt)
                db.session.commit()
                log_activity(
                    action='إضافة مديونية',
                    entity_type='supplier_debt',
                    entity_id=new_debt.id,
                    entity_name=f"مديونية للمورد {new_supplier.name}",
                    details=f"إضافة رصيد افتتاحي بقيمة {initial_debt:.2f} ج.م"
                )
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error creating opening supplier debt: {e}")
                flash('تم إنشاء المورد لكن حدث خطأ أثناء تسجيل الرصيد الافتتاحي', 'warning')
        flash("تم إضافة المورد بنجاح")
        return redirect(url_for('main.suppliers'))
    return render_template('add_supplier.html')

@main.route('/suppliers/<int:supplier_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_suppliers')
def edit_supplier(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    if request.method == 'POST':
        supplier.name = request.form['name']
        supplier.phone = request.form['phone']
        supplier.address = request.form['address']
        supplier.email = request.form.get('email', '')

        db.session.commit()
        flash("تم تعديل بيانات المورد بنجاح")
        return redirect(url_for('main.supplier_profile', supplier_id=supplier.id))
    return render_template('edit_supplier.html', supplier=supplier)

@main.route('/suppliers/<int:supplier_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_suppliers')
def delete_supplier(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    try:
        db.session.delete(supplier)
        db.session.commit()
        flash("تم حذف المورد بنجاح", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف المورد: {e}", "danger")
    return redirect(url_for('main.suppliers'))

@main.route('/suppliers/<int:supplier_id>/profile')
@login_required
@permission_required('can_view_suppliers')
def supplier_profile(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    
    invoices = Invoice.query.filter_by(supplier_id=supplier.id).order_by(Invoice.date.desc()).all()
    all_manual = SupplierDebt.query.filter_by(supplier_id=supplier.id).order_by(SupplierDebt.date.desc()).all()

    manual_debts = [d for d in all_manual if not d.is_payment]
    payment_logs = [d for d in all_manual if d.is_payment]

    total_debt = sum((invoice.total_amount or 0) - (invoice.paid_amount or 0) for invoice in invoices)
    total_debt += sum((debt.amount or 0) - (debt.paid_amount or 0) for debt in manual_debts)
    supplier.actual_total_debt = total_debt

    supplier.invoices = invoices
    supplier.debts = manual_debts
    supplier.payment_logs = payment_logs
    
    return render_template('supplier_profile.html', supplier=supplier)

@main.route('/suppliers/<int:supplier_id>/invoices')
@login_required
@permission_required('can_view_supplier_invoices')
def supplier_invoices(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    invoices = Invoice.query.filter_by(supplier_id=supplier_id).order_by(Invoice.date.desc()).all()
    
    total_debt = sum(invoice.total_amount - invoice.paid_amount for invoice in invoices)
    supplier.actual_total_debt = total_debt
    
    return render_template('supplier_invoices.html', supplier=supplier, invoices=invoices)

@main.route('/suppliers/<int:supplier_id>/account-history')
@login_required
@permission_required('can_view_supplier_history')
def supplier_account_history(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    invoices = Invoice.query.filter_by(supplier_id=supplier_id).order_by(Invoice.date.desc()).all()
    debts = SupplierDebt.query.filter_by(supplier_id=supplier_id).order_by(SupplierDebt.date.desc()).all()
    
    total_debt = sum((invoice.total_amount or 0) - (invoice.paid_amount or 0) for invoice in invoices)
    total_debt += sum((debt.amount or 0) - (debt.paid_amount or 0) for debt in debts)
    supplier.actual_total_debt = total_debt
    
    return render_template('supplier_account_history.html', supplier=supplier, invoices=invoices, debts=debts)

@main.route('/suppliers/<int:supplier_id>/pay-debt', methods=['POST'])
@login_required
@permission_required('can_pay_supplier_debt')
def pay_supplier_debt(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    
    # Get unified debt items (invoices and manual debts)
    debt_items = request.form.getlist('debt_items[]') or []
    debt_item_types = request.form.getlist('debt_item_types[]') or []
    total_payment_amount_raw = request.form.get('total_payment_amount', '0')
    payment_notes = request.form.get('payment_notes', '').strip()
    
    try:
        total_payment_amount = float(total_payment_amount_raw or 0)
    except ValueError:
        total_payment_amount = 0

    # Handle unified debt items with automatic distribution
    if debt_items and total_payment_amount > 0:
        try:
            # Verify that debt_items and debt_item_types have the same length
            if len(debt_items) != len(debt_item_types):
                flash('بيانات غير صحيحة', 'error')
                return redirect(url_for('main.supplier_account_history', supplier_id=supplier_id))
            
            # Collect all items (invoices and debts) with their metadata
            items_with_metadata = []
            
            for item_id, item_type in zip(debt_items, debt_item_types):
                try:
                    item_id = int(item_id)
                except ValueError:
                    continue
                
                if item_type == 'invoice':
                    invoice = Invoice.query.get(item_id)
                    if invoice and invoice.supplier_id == supplier_id:
                        remaining = (invoice.total_amount or 0) - (invoice.paid_amount or 0)
                        if remaining > 0:
                            items_with_metadata.append({
                                'type': 'invoice',
                                'object': invoice,
                                'remaining': remaining,
                                'created_date': invoice.date or datetime(1900, 1, 1)
                            })
                elif item_type == 'debt':
                    debt = SupplierDebt.query.get(item_id)
                    if debt and debt.supplier_id == supplier_id:
                        remaining = (debt.amount or 0) - (debt.paid_amount or 0)
                        if remaining > 0:
                            items_with_metadata.append({
                                'type': 'debt',
                                'object': debt,
                                'remaining': remaining,
                                'created_date': debt.created_at or datetime(1900, 1, 1)
                            })
            
            if not items_with_metadata:
                flash('لم يتم العثور على مستحقات صحيحة', 'error')
                return redirect(url_for('main.supplier_account_history', supplier_id=supplier_id))
            
            # Sort items by creation date (oldest first)
            items_sorted = sorted(items_with_metadata, key=lambda x: x['created_date'])
            
            # Distribute payment automatically from oldest to newest
            remaining_payment = total_payment_amount
            total_paid = 0.0
            payment_details = []
            
            for item_meta in items_sorted:
                if remaining_payment <= 0:
                    break
                
                item_remaining = item_meta['remaining']
                pay_amount = min(remaining_payment, item_remaining)
                
                if pay_amount > 0:
                    if item_meta['type'] == 'invoice':
                        invoice = item_meta['object']
                        invoice.paid_amount = (invoice.paid_amount or 0) + pay_amount
                        payment_details.append(f"فاتورة {invoice.invoice_number}: {pay_amount:.2f} ج.م")
                        
                        log_activity(
                            action='دفع مديونية',
                            entity_type='invoice',
                            entity_id=invoice.id,
                            entity_name=f"فاتورة رقم {invoice.invoice_number}",
                            details=f"دفع {pay_amount:.2f} ج.م للمورد {supplier.name}"
                        )
                    elif item_meta['type'] == 'debt':
                        debt = item_meta['object']
                        debt.paid_amount = (debt.paid_amount or 0) + pay_amount
                        debt_label = debt.notes or 'مديونية يدويّة'
                        payment_details.append(f"مديونية ({debt_label}): {pay_amount:.2f} ج.م")
                        
                        log_activity(
                            action='دفع مديونية',
                            entity_type='supplier_debt',
                            entity_id=debt.id,
                            entity_name=f"مديونية يدويّة للمورد {supplier.name}",
                            details=f"دفع {pay_amount:.2f} ج.م من المديونية"
                        )
                    
                    remaining_payment -= pay_amount
                    total_paid += pay_amount
            
            # Create payment log
            if total_paid > 0:
                payment_log = SupplierDebt(
                    supplier_id=supplier_id,
                    amount=0,
                    paid_amount=total_paid,
                    notes=(payment_notes if payment_notes else f"دفع متعدد: {total_paid:.2f} ج.م\n" + '\n'.join(payment_details)),
                    is_payment=True,
                    created_by=session.get('employee_id')
                )
                db.session.add(payment_log)
            
            db.session.commit()
            flash(f'تم دفع إجمالي {total_paid:.2f} ج.م بنجاح', 'success')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error paying debt items: {e}")
            flash('حدث خطأ أثناء عملية الدفع', 'error')
        return redirect(url_for('main.supplier_profile', supplier_id=supplier_id))

    flash('بيانات غير صحيحة', 'error')
    return redirect(url_for('main.supplier_account_history', supplier_id=supplier_id))

@main.route('/suppliers/<int:supplier_id>/add-debt', methods=['POST'])
@login_required
@permission_required('can_add_supplier_debt')
def add_supplier_debt(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    debt_amount = float(request.form.get('debt_amount', 0))
    notes = request.form.get('notes', '').strip()
    
    if debt_amount <= 0:
        flash('قيمة المديونية غير صحيحة', 'error')
        return redirect(url_for('main.supplier_account_history', supplier_id=supplier_id))
    
    try:
        new_debt = SupplierDebt(
            supplier_id=supplier_id,
            amount=debt_amount,
            paid_amount=0,
            notes=notes if notes else None,
            created_by=session.get('employee_id')
        )
        
        db.session.add(new_debt)
        db.session.commit()
        
        flash(f'تم إضافة مديونية بقيمة {debt_amount:.2f} ج.م بنجاح', 'success')
        
        log_activity(
            action='إضافة مديونية',
            entity_type='supplier_debt',
            entity_id=new_debt.id,
            entity_name=f"مديونية للمورد {supplier.name}",
            details=f"إضافة مديونية بقيمة {debt_amount:.2f} ج.م"
        )
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء إضافة المديونية', 'error')
        current_app.logger.error(f"Error adding supplier debt: {e}")
    
    return redirect(url_for('main.supplier_account_history', supplier_id=supplier_id))

@main.route('/suppliers/<int:supplier_id>/settle-debt', methods=['POST'])
@login_required
@permission_required('can_pay_supplier_debt')
def settle_supplier_debt(supplier_id):
    supplier = Supplier.query.get_or_404(supplier_id)
    amount = request.form.get('amount', type=float)
    method = (request.form.get('method') or '').strip()
    note = (request.form.get('note') or '').strip()
    
    if not amount or amount <= 0:
        flash('يرجى إدخال مبلغ صالح للسداد', 'error')
        return redirect(url_for('main.supplier_profile', supplier_id=supplier_id))
    
    try:
        payment_log = SupplierDebt(
            supplier_id=supplier_id,
            amount=0,
            paid_amount=amount,
            notes=f"سداد ({method}) - {note}" if note else f"سداد ({method})",
            is_payment=True,
            created_by=session.get('employee_id')
        )
        db.session.add(payment_log)
        db.session.commit()
        
        flash(f'تم تسجيل السداد بقيمة {amount:.2f} ج.م بنجاح', 'success')
        
        log_activity(
            action='سداد مديونية',
            entity_type='supplier_debt',
            entity_id=payment_log.id,
            entity_name=f"سداد للمورد {supplier.name}",
            details=f"سداد بقيمة {amount:.2f} ج.م - {method}"
        )
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تسجيل السداد', 'error')
        current_app.logger.error(f"Error settling supplier debt: {e}")
    
    return redirect(url_for('main.supplier_profile', supplier_id=supplier_id))

@main.route('/invoices')
@login_required
@permission_required('can_view_invoices')
def invoices():
    search_query = request.args.get('search', '').strip()
    
    query = Invoice.query
    
    if search_query:
        query = query.filter(Invoice.invoice_number.ilike(f'%{search_query}%'))
    
    invoices_list = query.order_by(Invoice.date.desc()).all()
    return render_template('invoices.html', invoices=invoices_list, search_query=search_query)

@main.route('/invoices/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_invoices')
@log_operation('create', 'invoice')
def add_invoice():
    if request.method == 'POST':
        submission_id = request.form.get('submission_id', '').strip()
        if not submission_id:
            flash('خطأ في إرسال النموذج. يرجى تحديث الصفحة والمحاولة مرة أخرى.', 'error')
            return redirect(url_for('main.add_invoice'))
        existing = Invoice.query.filter_by(submission_id=submission_id).first()
        if existing:
            flash('تم إضافة هذه الفاتورة مسبقًا.', 'info')
            return redirect(url_for('main.view_invoice', invoice_id=existing.id))

        supplier_phone = request.form.get('supplier_phone', '').strip()
        supplier_address = request.form.get('supplier_address', '').strip()
        supplier_id_form = request.form.get('supplier_id', '').strip()
        supplier_name_form = request.form.get('supplier_name', '').strip()

        supplier = None
        if supplier_id_form:
            supplier = Supplier.query.get(int(supplier_id_form))
            if not supplier:
                flash('المورد المحدد غير موجود', 'error')
                return redirect(url_for('main.add_invoice'))
        else:
            if supplier_name_form:
                supplier = Supplier(
                    name=supplier_name_form,
                    phone=supplier_phone or '',
                    address=supplier_address or None
                )
                db.session.add(supplier)
                db.session.flush()
            else:
                flash('يرجى اختيار المورد بالاسم أو إدخال اسم مورد جديد', 'error')
                return redirect(url_for('main.add_invoice'))

        supplier_id = supplier.id

        invoice_number = request.form.get('invoice_number', '').strip()
        if not invoice_number:
            def generate_invoice_number():
                base = datetime.utcnow().strftime('INV-%Y%m%d-')
                return base + uuid4().hex[:6].upper()
            invoice_number = generate_invoice_number()
            attempts = 0
            while Invoice.query.filter_by(invoice_number=invoice_number).first() and attempts < 5:
                invoice_number = generate_invoice_number()
                attempts += 1
        
        try:
            date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        except ValueError:
            flash('تاريخ الفاتورة غير صحيح', 'error')
            return redirect(url_for('main.add_invoice'))
            
        notes = request.form.get('notes', '')
        try:
            paid_amount = float(request.form.get('paid_amount', 0))
            if paid_amount < 0:
                flash('المبلغ المدفوع يجب أن يكون موجب', 'error')
                return redirect(url_for('main.add_invoice'))
        except ValueError:
            flash('المبلغ المدفوع غير صحيح', 'error')
            return redirect(url_for('main.add_invoice'))

        invoice = Invoice(
            supplier_id=supplier_id,
            invoice_number=invoice_number,
            submission_id=submission_id,
            date=date,
            notes=notes,
            paid_amount=paid_amount,
            total_amount=0
        )
        db.session.add(invoice)
        db.session.flush()

        products_data = request.form.to_dict(flat=False)
        product_indices = _extract_invoice_product_indices(products_data, 'product_name')
        total_amount = 0
        product_count = 0
        
        inv_increase_log = {}
        parsed_rows = []
        for row_number, index in enumerate(product_indices, start=1):
            product_name = products_data[f"products[{index}][product_name]"][0].strip()
            if not product_name:
                db.session.rollback()
                flash(f'اسم المنتج رقم {row_number} مطلوب', 'error')
                return redirect(url_for('main.add_invoice'))
                
            product_id_raw = products_data.get(f"products[{index}][product_id]", [None])[0]
            product_id = None
            if product_id_raw and str(product_id_raw).strip() != '':
                try:
                    product_id = int(product_id_raw)
                except (TypeError, ValueError):
                    db.session.rollback()
                    flash(f'معرف المنتج رقم {row_number} غير صحيح', 'error')
                    return redirect(url_for('main.add_invoice'))
            
            try:
                quantity = int(products_data[f"products[{index}][quantity]"][0])
                if quantity <= 0:
                    raise ValueError("الكمية يجب أن تكون أكبر من صفر")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'كمية المنتج رقم {row_number} غير صحيحة', 'error')
                return redirect(url_for('main.add_invoice'))
                
            try:
                purchase_price = float(products_data[f"products[{index}][purchase_price]"][0])
                if purchase_price < 0:
                    raise ValueError("سعر الشراء يجب أن يكون موجب")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'سعر شراء المنتج رقم {row_number} غير صحيح', 'error')
                return redirect(url_for('main.add_invoice'))
                
            try:
                selling_price = float(products_data[f"products[{index}][selling_price]"][0])
                if selling_price < 0:
                    raise ValueError("سعر البيع يجب أن يكون موجب")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'سعر بيع المنتج رقم {row_number} غير صحيح', 'error')
                return redirect(url_for('main.add_invoice'))

            parsed_rows.append({
                'row_number': row_number,
                'index': index,
                'product_name': product_name,
                'product_id': product_id,
                'quantity': quantity,
                'purchase_price': purchase_price,
                'selling_price': selling_price,
            })

        merged_rows = []
        existing_products_map = {}
        for row in parsed_rows:
            if row['product_id'] is None:
                row['line_count'] = 1
                merged_rows.append(row)
                continue

            pid = row['product_id']
            if pid in existing_products_map:
                existing_products_map[pid]['quantity'] += row['quantity']
                existing_products_map[pid]['line_count'] += 1
            else:
                row['line_count'] = 1
                existing_products_map[pid] = row
                merged_rows.append(row)

        merge_notices = []
        for row in merged_rows:
            index = row['index']
            product_name = row['product_name']
            product_id = row['product_id']
            quantity = row['quantity']
            purchase_price = row['purchase_price']
            selling_price = row['selling_price']

            if product_id is not None:
                product = _get_locked_product(product_id)
                if product:
                    if _block_if_deleted(product, 'هذه الفاتورة', 'deduct'):
                        db.session.rollback()
                        flash(f'المنتج "{product.name}" محذوف ولا يمكن إضافته إلى فاتورة', 'error')
                        return redirect(url_for('main.add_invoice'))
                    if row.get('line_count', 1) > 1:
                        merge_notices.append(
                            _format_invoice_merge_message(product.name, row['line_count'], quantity)
                        )

                    current_stock = int(product.stock or 0)
                    current_purchase_price = float(product.purchase_price or 0)
                    new_total_stock = current_stock + quantity
                    if new_total_stock > 0:
                        weighted_avg_price = (
                            (current_stock * current_purchase_price) + (quantity * purchase_price)
                        ) / new_total_stock
                        product.purchase_price = round(weighted_avg_price, 2)
                    product.stock = new_total_stock
                    inv_increase_log[product.id] = inv_increase_log.get(product.id, 0) + quantity
                    if (product.wholesale_price or 0) == 0 or purchase_price < (product.wholesale_price or 0):
                        product.wholesale_price = purchase_price
                    if selling_price != product.price:
                        product.price = selling_price
                else:
                    db.session.rollback()
                    flash(f'المنتج رقم {row["row_number"]} غير موجود', 'error')
                    return redirect(url_for('main.add_invoice'))
            else:
                product = None

            if not product:
                has_types = any(key.startswith(f"products[{index}][types]") for key in products_data.keys())
                
                product = Product(
                    name=product_name,
                    price=selling_price,
                    wholesale_price=purchase_price,
                    purchase_price=purchase_price,
                    stock=quantity,
                    has_size=False,
                    has_color=False,
                    has_style=False
                )
                db.session.add(product)
                db.session.flush()
                inv_increase_log[product.id] = inv_increase_log.get(product.id, 0) + quantity

                if has_types:
                    type_keys = [key for key in products_data.keys() if key.startswith(f"products[{index}][types]")]
                    type_ids = set()
                    
                    for key in type_keys:
                        parts = key.split('[')
                        if len(parts) >= 4:
                            type_id = parts[3].rstrip(']')
                            type_ids.add(type_id)
                    
                    for type_id in type_ids:
                        type_name_key = f"products[{index}][types][{type_id}][type_name]"
                        
                        if type_name_key in products_data:
                            type_name = products_data[type_name_key][0]
                            
                            if type_name == 'size':
                                product.has_size = True
                            elif type_name == 'color':
                                product.has_color = True
                            elif type_name == 'style':
                                product.has_style = True
                            
                            option_keys = [key for key in products_data.keys() 
                                         if key.startswith(f"products[{index}][types][{type_id}][options]")]
                            
                            for option_key in option_keys:
                                if '[name]' in option_key:
                                    option_name = products_data[option_key][0]
                                    option_price_key = option_key.replace('[name]', '[price]')
                                    option_price = float(products_data[option_price_key][0]) if option_price_key in products_data else 0
                                    
                                    if option_name.strip():
                                        variant = ProductVariant(
                                            product_id=product.id,
                                            group_name=type_name,
                                            variant_name=option_name.strip(),
                                            price=selling_price + option_price
                                        )
                                        db.session.add(variant)

            invoice_item = InvoiceItem(
                invoice_id=invoice.id,
                product_id=product.id,
                quantity=quantity,
                purchase_price=purchase_price,
                selling_price=selling_price
            )
            db.session.add(invoice_item)

            total_amount += purchase_price * quantity
            product_count += 1

        if product_count == 0:
            db.session.rollback()
            flash('يجب إضافة منتج واحد على الأقل', 'error')
            return redirect(url_for('main.add_invoice'))

        invoice.total_amount = total_amount

        supplier = Supplier.query.get(supplier_id)
        if supplier:
            supplier.total_debt += (total_amount - paid_amount)

        try:
            db.session.commit()
            for notice in merge_notices:
                flash(notice, 'info')
            _log_stock_delta("supplier_invoice increase", inv_increase_log)
            return redirect(url_for('main.invoices'))
        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ أثناء إضافة الفاتورة: {str(e)}", 'error')
            return redirect(url_for('main.add_invoice'))

    submission_id = str(uuid4())
    return render_template('add_invoice.html', submission_id=submission_id)

@main.route('/invoices/<int:invoice_id>')
@login_required
@permission_required('can_view_invoice_details')
def view_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    calculated_total = sum(item.purchase_price * item.quantity for item in invoice.items)
    if abs(invoice.total_amount - calculated_total) > 0.01:
        invoice.total_amount = calculated_total
        db.session.commit()
    
    return render_template('view_invoice.html', invoice=invoice)

@main.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_invoices')
def delete_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)

    has_returns = SupplierReturn.query.filter_by(invoice_id=invoice.id).first() is not None
    if has_returns:
        flash('لا يمكن حذف الفاتورة لأن عليها مرتجع مورد. يرجى تسوية أو حذف المرتجع أولاً.', 'error')
        return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

    del_increase_revert_log = {}
    for item in invoice.items:
        product = _get_locked_product(item.product_id)
        if product:
            product.stock = (product.stock or 0) - item.quantity
            del_increase_revert_log[product.id] = del_increase_revert_log.get(product.id, 0) - item.quantity

    supplier = invoice.supplier
    if supplier:
        remaining_amount = (invoice.total_amount or 0) - (invoice.paid_amount or 0)
        supplier.total_debt = (supplier.total_debt or 0) - remaining_amount

    db.session.delete(invoice)
    db.session.commit()
    flash("تم حذف الفاتورة بنجاح")
    _log_stock_delta("supplier_invoice delete restock", del_increase_revert_log)
    return redirect(url_for('main.invoices'))

@main.route('/invoices/<int:invoice_id>/return', methods=['POST'])
@login_required
@permission_required('can_add_invoices')
def create_supplier_return(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)

    returned_quantity_rows = db.session.query(
        SupplierReturnItem.invoice_item_id,
        func.sum(SupplierReturnItem.quantity)
    ).join(
        SupplierReturn,
        SupplierReturn.id == SupplierReturnItem.return_id
    ).filter(
        SupplierReturn.invoice_id == invoice_id
    ).group_by(
        SupplierReturnItem.invoice_item_id
    ).all()
    returned_quantity_by_item = {
        item_id: int(total_qty or 0)
        for item_id, total_qty in returned_quantity_rows
    }
    
    return_record = SupplierReturn(
        invoice_id=invoice_id,
        notes=request.form.get('notes', ''),
        created_by=session.get('employee_id')
    )
    
    total_amount = 0
    has_items = False
    
    ret_decrease_log = {}
    for item in invoice.items:
        quantity_key = f'quantity_{item.id}'
        try:
            quantity = int(request.form.get(quantity_key, 0))
        except (TypeError, ValueError):
            flash(f'قيمة كمية المرتجع للمنتج {item.product.name} غير صحيحة', 'error')
            return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

        if quantity < 0:
            flash(f'كمية المرتجع للمنتج {item.product.name} لا يمكن أن تكون سالبة', 'error')
            return redirect(url_for('main.view_invoice', invoice_id=invoice_id))
        
        if quantity > 0:
            already_returned_qty = returned_quantity_by_item.get(item.id, 0)
            available_qty = item.quantity - already_returned_qty

            if quantity > available_qty:
                flash(
                    f'كمية المرتجع لـ {item.product.name} تتجاوز المتاح. المتاح حالياً {available_qty} من أصل {item.quantity}',
                    'error'
                )
                return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

            returned_quantity_by_item[item.id] = already_returned_qty + quantity
            
            has_items = True
            total = quantity * item.purchase_price
            
            return_item = SupplierReturnItem(
                invoice_item_id=item.id,
                product_id=item.product_id,
                quantity=quantity,
                price=item.purchase_price,
                total=total
            )
            return_record.items.append(return_item)
            total_amount += total
            
            product = _get_locked_product(item.product_id)
            if not product:
                db.session.rollback()
                flash(f'المنتج المرتجع {item.product.name} غير موجود', 'error')
                return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

            if product.stock < quantity:
                db.session.rollback()
                flash(f'المخزون الحالي للمنتج {product.name} غير كافٍ للمرتجع. المتوفر حالياً {product.stock}', 'error')
                return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

            if _block_if_deleted(product, 'هذا المرتجع', 'deduct'):
                db.session.rollback()
                return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

            if product:
                product.stock -= quantity
                ret_decrease_log[product.id] = ret_decrease_log.get(product.id, 0) - quantity
    
    if not has_items:
        flash('الرجاء إدخال كمية مرتجع لمنتج واحد على الأقل', 'warning')
        return redirect(url_for('main.view_invoice', invoice_id=invoice_id))
    
    return_record.total_amount = total_amount
    
    supplier = invoice.supplier
    if supplier:
        supplier.total_debt -= total_amount
    
    try:
        db.session.add(return_record)
        db.session.commit()
        _log_stock_delta("supplier_return decrease", ret_decrease_log)
        
        log_activity(
            action='إضافة مرتجع',
            entity_type='supplier_return',
            entity_id=return_record.id,
            entity_name=f"مرتجع فاتورة {invoice.invoice_number}",
            details=f"مرتجع بقيمة {total_amount:.2f} جنيه - عدد المنتجات: {len(return_record.items)}"
        )
        
        flash(f'تم تسجيل المرتجع بنجاح - المبلغ: {total_amount:.2f} جنيه', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating supplier return: {e}")
        flash('حدث خطأ أثناء تسجيل المرتجع', 'error')
    
    return redirect(url_for('main.view_invoice', invoice_id=invoice_id))

@main.route('/invoices/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('can_edit_invoices')
@log_operation('update', 'invoice')
def edit_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        old_total = invoice.total_amount
        old_paid = invoice.paid_amount
        old_supplier_id = invoice.supplier_id
        
        invoice.invoice_number = request.form.get('invoice_number', '').strip()
        
        try:
            invoice.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        except ValueError:
            flash('تاريخ الفاتورة غير صحيح', 'error')
            return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
        
        invoice.notes = request.form.get('notes', '')
        
        try:
            new_paid_amount = float(request.form.get('paid_amount', 0))
            if new_paid_amount < 0:
                flash('المبلغ المدفوع يجب أن يكون موجب', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
        except ValueError:
            flash('المبلغ المدفوع غير صحيح', 'error')
            return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
        
        supplier_id_form = request.form.get('supplier_id', '').strip()
        if supplier_id_form and int(supplier_id_form) != old_supplier_id:
            new_supplier = Supplier.query.get(int(supplier_id_form))
            if not new_supplier:
                flash('المورد المحدد غير موجود', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
            
            old_supplier = Supplier.query.get(old_supplier_id)
            if old_supplier:
                old_remaining = old_total - old_paid
                old_supplier.total_debt = (old_supplier.total_debt or 0) - old_remaining
            
            invoice.supplier_id = int(supplier_id_form)
        
        old_items = {}
        for item in invoice.items:
            old_items[item.product_id] = old_items.get(item.product_id, 0) + item.quantity
        
        InvoiceItem.query.filter_by(invoice_id=invoice.id).delete()
        
        products_data = request.form.to_dict(flat=False)
        product_indices = _extract_invoice_product_indices(products_data, 'product_id')
        total_amount = 0
        product_count = 0
        new_items = {}
        parsed_rows = []
        
        for row_number, index in enumerate(product_indices, start=1):
            product_id_raw = products_data[f"products[{index}][product_id]"][0]
            product_name = products_data.get(f"products[{index}][product_name]", [''])[0].strip()
            
            if not product_id_raw or str(product_id_raw).strip() == '':
                if product_name:
                    db.session.rollback()
                    flash(f'يرجى اختيار منتج صحيح من البحث للصف رقم {row_number}', 'error')
                    return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
                continue

            try:
                product_id = int(product_id_raw)
            except (TypeError, ValueError):
                db.session.rollback()
                flash(f'معرف المنتج رقم {row_number} غير صحيح', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
            
            try:
                quantity = int(products_data[f"products[{index}][quantity]"][0])
                if quantity <= 0:
                    raise ValueError("الكمية يجب أن تكون أكبر من صفر")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'كمية المنتج رقم {row_number} غير صحيحة', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
            
            try:
                purchase_price = float(products_data[f"products[{index}][purchase_price]"][0])
                if purchase_price < 0:
                    raise ValueError("سعر الشراء يجب أن يكون موجب")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'سعر شراء المنتج رقم {row_number} غير صحيح', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
            
            try:
                selling_price = float(products_data[f"products[{index}][selling_price]"][0])
                if selling_price < 0:
                    raise ValueError("سعر البيع يجب أن يكون موجب")
            except (ValueError, IndexError):
                db.session.rollback()
                flash(f'سعر بيع المنتج رقم {row_number} غير صحيح', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))

            parsed_rows.append({
                'row_number': row_number,
                'product_id': product_id,
                'quantity': quantity,
                'purchase_price': purchase_price,
                'selling_price': selling_price,
            })

        merged_rows = {}
        for row in parsed_rows:
            pid = row['product_id']
            if pid in merged_rows:
                merged_rows[pid]['quantity'] += row['quantity']
                merged_rows[pid]['line_count'] += 1
            else:
                row['line_count'] = 1
                merged_rows[pid] = row

        merge_notices = []
        for row in merged_rows.values():
            product = _get_locked_product(row['product_id'])
            if not product:
                db.session.rollback()
                flash(f'المنتج رقم {row["row_number"]} غير موجود', 'error')
                return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))

            quantity = row['quantity']
            purchase_price = row['purchase_price']
            selling_price = row['selling_price']
            pid = row['product_id']

            if row.get('line_count', 1) > 1:
                merge_notices.append(
                    _format_invoice_merge_message(product.name, row['line_count'], quantity)
                )
            
            new_items[pid] = quantity
            
            current_stock = int(product.stock or 0)
            current_purchase_price = float(product.purchase_price or 0)
            if current_stock > 0:
                weighted_avg_price = (
                    ((current_stock - quantity) * current_purchase_price) + (quantity * purchase_price)
                ) / current_stock
                product.purchase_price = round(weighted_avg_price, 2)
            
            if (product.wholesale_price or 0) == 0 or purchase_price < (product.wholesale_price or 0):
                product.wholesale_price = purchase_price
            
            if selling_price != product.price:
                product.price = selling_price
            
            invoice_item = InvoiceItem(
                invoice_id=invoice.id,
                product_id=product.id,
                quantity=quantity,
                purchase_price=purchase_price,
                selling_price=selling_price
            )
            db.session.add(invoice_item)
            
            total_amount += purchase_price * quantity
            product_count += 1
        
        if product_count == 0:
            db.session.rollback()
            flash('يجب إضافة منتج واحد على الأقل', 'error')
            return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
        
        all_product_ids = set(old_items.keys()) | set(new_items.keys())
        locked_products = _get_locked_products(all_product_ids)
        for pid in all_product_ids:
            product = locked_products.get(pid)
            if product:
                old_qty = old_items.get(pid, 0)
                new_qty = new_items.get(pid, 0)
                qty_diff = new_qty - old_qty
                if qty_diff < 0 and _block_if_deleted(product, 'هذه الفاتورة', 'deduct'):
                    db.session.rollback()
                    flash(f'لا يمكن تعديل الفاتورة: تحتوي على منتج محذوف ("{product.name}")', 'error')
                    return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
                product.stock = (product.stock or 0) + qty_diff
        
        invoice.total_amount = total_amount
        invoice.paid_amount = new_paid_amount
        
        supplier = Supplier.query.get(invoice.supplier_id)
        if supplier:
            new_remaining = total_amount - new_paid_amount
            old_remaining = old_total - old_paid
            
            if invoice.supplier_id == old_supplier_id:
                supplier.total_debt = (supplier.total_debt or 0) - old_remaining + new_remaining
            else:
                supplier.total_debt = (supplier.total_debt or 0) + new_remaining
        
        try:
            db.session.commit()
            for notice in merge_notices:
                flash(notice, 'info')
            flash('تم تعديل الفاتورة بنجاح')
            return redirect(url_for('main.view_invoice', invoice_id=invoice.id))
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء تعديل الفاتورة: {str(e)}', 'error')
            return redirect(url_for('main.edit_invoice', invoice_id=invoice_id))
    
    suppliers = Supplier.query.all()
    return render_template('edit_invoice.html', invoice=invoice, suppliers=suppliers)

@main.route('/my-account')
@login_required
def my_account():
    employee = Employee.query.get(session['employee_id'])

    current_month = request.args.get('month', datetime.now().month, type=int)
    current_year = request.args.get('year', datetime.now().year, type=int)

    salary_summary = employee.get_salary_summary(current_month, current_year)

    monthly_transactions = employee.get_monthly_transactions(
        current_month, current_year)

    start_date = datetime(current_year, current_month, 1)
    end_date = datetime(current_year + 1, 1, 1) if current_month == 12 else datetime(current_year, current_month + 1, 1)
    days_in_month = (end_date - start_date).days
    if employee.requires_attendance:
        accrued_salary = Employee._calc_accrued_salary_attendance(
            salary_summary['net_salary'] or 0, employee.id, current_month, current_year, days_in_month)
    else:
        accrued_salary = Employee._calc_accrued_salary(salary_summary['net_salary'] or 0, current_month, current_year, days_in_month)

    recent_orders = Order.query.filter(
        Order.employee_id == employee.id,
        Order.date >= start_date,
        Order.date < end_date
    ).order_by(Order.date.desc()).all()

    total_orders = len(recent_orders)

    # Use history-based query: orders that reached 'وصل' this month,
    # even if their current status is 'استبدال'.
    delivered_this_month = employee._query_period_orders(start_date, end_date)
    delivered_orders_count = salary_summary['delivered_orders_count']
    total_sales = salary_summary['delivered_sales']
    total_commission = salary_summary['sales_commission']

    return render_template('my_account.html',
                           employee=employee,
                           transactions=monthly_transactions,
                           base_salary=salary_summary['base_salary'],
                           sales_commission=salary_summary['sales_commission'],
                           total_salary=salary_summary['total_salary'],
                           total_deductions=salary_summary['total_deductions'],
                           total_bonuses=salary_summary['total_bonuses'],
                           total_advances=salary_summary['total_advances'],
                           net_salary=salary_summary['net_salary'],
                           accrued_salary=accrued_salary,
                           current_month=current_month,
                           current_year=current_year,
                           recent_orders=recent_orders,
                           delivered_orders_this_month=delivered_this_month,
                           total_orders=total_orders,
                           delivered_orders_count=delivered_orders_count,
                           total_sales=total_sales,
                           total_commission=total_commission,
                           today=date.today().isoformat())

@main.route('/employees/<int:employee_id>/activity-log')
@login_required
@permission_required('can_view_employee_activity')
def employee_activity_log(employee_id):
    employee = Employee.query.get_or_404(employee_id)

    current_employee = Employee.query.get(session['employee_id'])
    if not current_employee.is_admin and current_employee.id != employee.id:
        flash('ليس لديك صلاحية لعرض سجل نشاط هذا الموظف')
        return redirect(url_for('main.employees'))

    selected_date = request.args.get('date', '')
    action_type = request.args.get('action_type', '')

    query = EmployeeActivityLog.query.filter_by(employee_id=employee.id)

    if selected_date:
        query = query.filter(
            func.date(
                EmployeeActivityLog.created_at) == selected_date)

    if action_type:
        query = query.filter(EmployeeActivityLog.action == action_type)

    activity_logs = query.order_by(
        EmployeeActivityLog.created_at.desc()).limit(100).all()

    return render_template('employee_activity_log.html',
                           employee=employee,
                           activity_logs=activity_logs,
                           selected_date=selected_date,
                           is_admin=current_employee.is_admin)

@main.route('/activity-log')
@login_required
@permission_required('can_view_activity_log')
def activity_log():
    current_employee = Employee.query.get(session['employee_id'])
    if not current_employee.is_admin:
        flash('ليس لديك صلاحية لعرض سجل العمليات الشامل')
        return redirect(url_for('main.dashboard'))

    period = request.args.get('period', 'month')
    employee_filter = request.args.get('employee_id', '')
    action_filter = request.args.get('action', '')
    entity_filter = request.args.get('entity_type', '')

    today = date.today()
    if period == 'day':
        start_date = today
        end_date = today
    elif period == 'month':
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = None
        end_date = None

    query = EmployeeActivityLog.query.join(Employee)

    if start_date and end_date:
        query = query.filter(
            func.date(EmployeeActivityLog.created_at) >= start_date,
            func.date(EmployeeActivityLog.created_at) <= end_date
        )

    if employee_filter:
        query = query.filter(
            EmployeeActivityLog.employee_id == int(employee_filter))

    if action_filter:
        query = query.filter(EmployeeActivityLog.action == action_filter)

    if entity_filter:
        query = query.filter(EmployeeActivityLog.entity_type == entity_filter)

    activity_logs = query.order_by(
        EmployeeActivityLog.created_at.desc()).limit(500).all()

    total_activities = EmployeeActivityLog.query.count()
    today_activities = EmployeeActivityLog.query.filter(
        func.date(EmployeeActivityLog.created_at) == today
    ).count()
    month_activities = EmployeeActivityLog.query.filter(
        func.date(EmployeeActivityLog.created_at) >= start_date if start_date else True,
        func.date(EmployeeActivityLog.created_at) <= end_date if end_date else True
    ).count()

    employees = Employee.query.all()
    actions = db.session.query(EmployeeActivityLog.action).distinct().all()
    entity_types = db.session.query(
        EmployeeActivityLog.entity_type).distinct().all()

    return render_template('activity_log.html',
                           activity_logs=activity_logs,
                           period=period,
                           employee_filter=employee_filter,
                           action_filter=action_filter,
                           entity_filter=entity_filter,
                           total_activities=total_activities,
                           today_activities=today_activities,
                           month_activities=month_activities,
                           employees=employees,
                           actions=[a[0] for a in actions],
                           entity_types=[e[0] for e in entity_types])

@main.route('/products/restore/<int:product_id>', methods=['POST'])
@login_required
@permission_required('can_restore_products')
def restore_product(product_id):
    product = Product.query.get_or_404(product_id)
    next_url = request.form.get('next') or request.args.get('next')
    if product.is_deleted:
        product.is_deleted = False
        db.session.commit()

        log_activity(
            'restore',
            'product',
            product_id,
            product.name,
            f"استرجاع منتج: {product.name}"
        )

        flash(f"تم استرجاع المنتج {product.name} بنجاح", "success")
    else:
        flash("المنتج غير محذوف", "warning")
    if next_url:
        try:
            return redirect(next_url)
        except Exception:
            pass
    return redirect(url_for('main.products'))

@main.route('/expenses')
@login_required
@permission_required('can_view_expenses')
def expenses():
    current_month = date.today().replace(day=1)
    next_month = current_month.replace(
        month=current_month.month +
        1) if current_month.month < 12 else current_month.replace(
        year=current_month.year +
        1,
        month=1)

    current_month_expenses = Expense.query.filter(
        Expense.date >= current_month,
        Expense.date < next_month
    ).all()

    operational_expenses = sum(
        exp.amount for exp in current_month_expenses if exp.category == 'مصاريف تشغيلية')
    fixed_assets = sum(
        exp.amount for exp in current_month_expenses if exp.category == 'أصول ثابتة')

    return render_template('expenses.html',
                           operational_expenses=operational_expenses,
                           fixed_assets=fixed_assets,
                           today=date.today().strftime('%Y-%m-%d'))

@main.route('/expenses/operational')
@login_required
@permission_required('can_view_operational_expenses')
def operational_expenses():
    current_date = date.today()
    
    selected_year = request.args.get('year', current_date.year)
    selected_month = request.args.get('month', current_date.month)
    
    if selected_year:
        selected_year = int(selected_year)
    if selected_month:
        selected_month = int(selected_month)
    
    query = Expense.query.filter(Expense.category == 'مصاريف تشغيلية')

    if selected_year:
        query = query.filter(extract('year', Expense.date) == selected_year)
    
    if selected_month:
        query = query.filter(extract('month', Expense.date) == selected_month)

    expenses = query.order_by(Expense.date.desc()).all()

    total_expenses = sum(exp.amount for exp in expenses)
    expenses_count = len(expenses)

    return render_template('operational_expenses.html',
                           expenses=expenses,
                           total_expenses=total_expenses,
                           expenses_count=expenses_count,
                           selected_year=selected_year,
                           selected_month=selected_month,
                           today=date.today().strftime('%Y-%m-%d'))

@main.route('/expenses/fixed-assets')
@login_required
@permission_required('can_view_fixed_assets')
def fixed_assets_expenses():
    current_date = date.today()
    
    selected_year = request.args.get('year', current_date.year)
    selected_month = request.args.get('month', current_date.month)
    
    if selected_year:
        selected_year = int(selected_year)
    if selected_month:
        selected_month = int(selected_month)
    
    query = Expense.query.filter(Expense.category == 'أصول ثابتة')

    if selected_year:
        query = query.filter(extract('year', Expense.date) == selected_year)
    
    if selected_month:
        query = query.filter(extract('month', Expense.date) == selected_month)

    expenses = query.order_by(Expense.date.desc()).all()

    total_expenses = sum(exp.amount for exp in expenses)
    expenses_count = len(expenses)

    return render_template('fixed_assets_expenses.html',
                           expenses=expenses,
                           total_expenses=total_expenses,
                           expenses_count=expenses_count,
                           selected_year=selected_year,
                           selected_month=selected_month,
                           today=date.today().strftime('%Y-%m-%d'))

@main.route('/expenses/add', methods=['GET', 'POST'])
@login_required
@permission_required('can_add_expenses')
def add_expense():
    if request.method == 'GET':
        return render_template(
            'add_expense.html',
            today=date.today().strftime('%Y-%m-%d'))

    if request.method == 'POST':
        try:
            category = request.form['category']
            original_amount = float(request.form['amount'])

            if category == 'مصاريف تشغيلية':
                expense = Expense(
                    title=request.form['title'],
                    amount=original_amount,
                    category=category,
                    date=datetime.strptime(
                        request.form['date'],
                        '%Y-%m-%d').date(),
                    description=request.form.get(
                        'description',
                        ''),
                    receipt_number=request.form.get(
                        'receipt_number',
                        ''),
                    supplier=request.form.get(
                        'supplier',
                        ''),
                    payment_method='دفعة واحدة',
                    created_by=session.get('employee_id'))
                db.session.add(expense)
                db.session.commit()
                flash('تم إضافة المصروف التشغيلي بنجاح', 'success')

            elif category == 'أصول ثابتة':
                distribution_type = request.form.get(
                    'distribution_type', 'current_month')

                if distribution_type == 'current_month':
                    expense = Expense(
                        title=request.form['title'],
                        amount=original_amount,
                        category=category,
                        date=datetime.strptime(
                            request.form['date'],
                            '%Y-%m-%d').date(),
                        description=request.form.get(
                            'description',
                            ''),
                        receipt_number=request.form.get(
                            'receipt_number',
                            ''),
                        supplier=request.form.get(
                            'supplier',
                            ''),
                        payment_method='الشهر الحالي فقط',
                        created_by=session.get('employee_id'))
                    db.session.add(expense)
                    db.session.commit()
                    flash('تم إضافة الأصل الثابت للشهر الحالي بنجاح', 'success')

                elif distribution_type == 'selected_months':
                    selected_years = request.form.getlist('year[]')
                    selected_months = request.form.getlist('month[]')

                    if selected_years and selected_months and len(
                            selected_years) == len(selected_months):
                        divided_amount = original_amount / len(selected_years)
                        month_names = [
                            'يناير',
                            'فبراير',
                            'مارس',
                            'أبريل',
                            'مايو',
                            'يونيو',
                            'يوليو',
                            'أغسطس',
                            'سبتمبر',
                            'أكتوبر',
                            'نوفمبر',
                            'ديسمبر']

                        for i in range(len(selected_years)):
                            year = int(selected_years[i])
                            month = int(selected_months[i])
                            expense_date = datetime(year, month, 1).date()

                            expense = Expense(
                                title=request.form['title'],
                                amount=divided_amount,
                                category=category,
                                date=expense_date,
                                description=request.form.get(
                                    'description',
                                    ''),
                                receipt_number=request.form.get(
                                    'receipt_number',
                                    ''),
                                supplier=request.form.get(
                                    'supplier',
                                    ''),
                                payment_method=f'مقسوم على {len(selected_years)} شهور',
                                created_by=session.get('employee_id'))
                            db.session.add(expense)

                        db.session.commit()

                        selections = []
                        for i in range(len(selected_years)):
                            month_name = month_names[int(
                                selected_months[i]) - 1]
                            year = selected_years[i]
                            selections.append(f"{month_name} {year}")

                        flash(
                            f'تم إضافة الأصل الثابت مقسماً على الشهور: {", ".join(selections)} بنجاح', 'success')
                    else:
                        flash('يرجى اختيار شهور محددة', 'error')
                        return redirect(url_for('main.add_expense'))

            return redirect(url_for('main.expenses'))

        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء إضافة المصروف: {str(e)}', 'error')
            return redirect(url_for('main.add_expense'))

@main.route('/expenses/<int:expense_id>')
@login_required
@permission_required('can_view_expenses')
def view_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    return render_template('expense_detail.html', expense=expense)

@main.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required
@permission_required('can_delete_expenses')
def delete_expense(expense_id):
    try:
        expense = Expense.query.get_or_404(expense_id)

        if expense.category == 'أصول ثابتة' and expense.payment_method and 'مقسوم على' in expense.payment_method:
            related_expenses = Expense.query.filter(
                Expense.title == expense.title,
                Expense.category == expense.category,
                Expense.payment_method == expense.payment_method
            ).all()

            for related_expense in related_expenses:
                db.session.delete(related_expense)
        else:
            db.session.delete(expense)

        db.session.commit()

        details = f"حذف مصروف: {expense.title} - المبلغ: {expense.amount}"
        log_activity('delete', 'expense', expense_id, expense.title, details)

        return jsonify({'success': True, 'message': 'تم حذف المصروف بنجاح'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False,
                        'message': f'حدث خطأ أثناء حذف المصروف: {str(e)}'})

@main.route('/orders/import-old', methods=['POST'])
@login_required
@permission_required('can_add_orders')
def import_old_orders():
    
    if 'file' not in request.files:
        flash('لم يتم اختيار ملف', 'error')
        return redirect(url_for('main.orders'))
    
    file = request.files['file']
    if file.filename == '':
        flash('لم يتم اختيار ملف', 'error')
        return redirect(url_for('main.orders'))
    
    if not file.filename.lower().endswith('.csv'):
        flash('يجب اختيار ملف CSV فقط', 'error')
        return redirect(url_for('main.orders'))
    
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    
    max_size = 10 * 1024 * 1024
    if file_size > max_size:
        flash(f'حجم الملف كبير جداً ({file_size / (1024*1024):.1f} ميجابايت). الحد الأقصى 10 ميجابايت', 'error')
        return redirect(url_for('main.orders'))
    
    try:
        import csv
        import io
        import signal
        
        def timeout_handler(signum, frame):
            raise TimeoutError("انتهت مهلة معالجة الملف")
        
        try:
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(300)
        except (AttributeError, OSError):
            pass
        
        flash('بدء قراءة محتوى الملف...', 'info')
        
        file_content = file.stream.read()
        
        if not file_content:
            flash('الملف فارغ', 'error')
            return redirect(url_for('main.orders'))
        
        try:
            file_content = file_content.decode("UTF8")
        except UnicodeDecodeError:
            try:
                file_content = file_content.decode("UTF-8-SIG")
            except UnicodeDecodeError:
                try:
                    file_content = file_content.decode("cp1256")
                except UnicodeDecodeError:
                    flash('خطأ في ترميز الملف. يرجى التأكد من أن الملف محفوظ بترميز UTF-8', 'error')
                    return redirect(url_for('main.orders'))
        
        flash(f'تم قراءة {len(file_content)} حرف من الملف', 'info')
        
        stream = io.StringIO(file_content, newline=None)
        csv_reader = csv.reader(stream)
        
        all_rows = list(csv_reader)
        if not all_rows:
            flash('الملف فارغ', 'error')
            return redirect(url_for('main.orders'))
        
        header = all_rows[0]
        data_rows = all_rows[1:] if len(header) > 0 and any('اسم' in str(col) or 'phone' in str(col).lower() for col in header) else all_rows
        
        flash(f'تم العثور على {len(data_rows)} صف للمعالجة', 'info')
        
        orders_created = 0
        customers_created = 0
        errors = []
        
        for row_num, row in enumerate(data_rows, start=2):
            try:
                if len(row) < 3:
                    errors.append(f"الصف {row_num}: عدد الأعمدة غير كافي ({len(row)} أعمدة)")
                    continue
                
                customer_name = row[0].strip() if len(row) > 0 else ""
                customer_phone = row[1].strip() if len(row) > 1 else ""
                order_details = row[2].strip() if len(row) > 2 else ""
                
                if customer_phone and customer_phone != "00000000000":
                    phone_clean = ''.join(filter(str.isdigit, customer_phone))
                    if phone_clean.startswith('20') and len(phone_clean) > 10:
                        phone_clean = phone_clean[2:]
                    
                    if len(phone_clean) >= 11:
                        customer_phone = phone_clean[:11]
                    elif len(phone_clean) >= 10:
                        customer_phone = phone_clean[:10]
                    elif len(phone_clean) >= 7:
                        customer_phone = phone_clean
                    else:
                        customer_phone = "00000000000"
                
                amount_paid_str = row[3].strip() if len(row) > 3 else "0"
                remaining_amount_str = row[4].strip() if len(row) > 4 else "0"
                
                def clean_amount(amount_str):
                    if not amount_str:
                        return 0
                    cleaned = amount_str.replace('EGP', '').replace(',', '').replace('٬', '').replace('٫', '.').replace(' ', '').strip()
                    try:
                        return float(cleaned) if cleaned else 0
                    except ValueError:
                        import re
                        cleaned = re.sub(r'[^\d.]', '', cleaned)
                        try:
                            return float(cleaned) if cleaned else 0
                        except ValueError:
                            return 0
                
                amount_paid = clean_amount(amount_paid_str)
                remaining_amount = clean_amount(remaining_amount_str)
                
                address_details = ""
                governorate = ""
                tracking_number = ""
                order_date_str = ""
                
                if len(row) >= 9:
                    address_details = row[5].strip() if len(row) > 5 else ""
                    governorate = row[6].strip() if len(row) > 6 else ""
                    tracking_number = row[7].strip() if len(row) > 7 else ""
                    order_date_str = row[8].strip() if len(row) > 8 else ""
                    
                    if not address_details or not address_details.strip():
                        address_details = "عنوان غير محدد"
                    
                    if not governorate or not governorate.strip():
                        governorate = "غير محدد"
                    
                    if not tracking_number or not tracking_number.strip():
                        tracking_number = "غير محدد"
                elif len(row) >= 8:
                    address_details = row[5].strip() if len(row) > 5 else ""
                    tracking_number = row[6].strip() if len(row) > 6 else ""
                    order_date_str = row[7].strip() if len(row) > 7 else ""
                    
                    if not address_details or not address_details.strip():
                        address_details = "عنوان غير محدد"
                    
                    if not tracking_number or not tracking_number.strip():
                        tracking_number = "غير محدد"
                        
                elif len(row) >= 7:
                    address_details = row[5].strip() if len(row) > 5 else ""
                    tracking_number = row[6].strip() if len(row) > 6 else ""
                    
                    if not address_details or not address_details.strip():
                        address_details = "عنوان غير محدد"
                    
                    if not tracking_number or not tracking_number.strip():
                        tracking_number = "غير محدد"
                        
                elif len(row) >= 6:
                    address_details = row[5].strip() if len(row) > 5 else ""
                    
                    if not address_details or not address_details.strip():
                        address_details = "عنوان غير محدد"
                
                if len(row) >= 6 and not governorate:
                    col5_value = row[5].strip() if len(row) > 5 else ""
                    col6_value = row[6].strip() if len(row) > 6 else ""
                    
                    governorates_check = [
                        'القاهرة', 'الجيزة', 'الإسكندرية', 'الشرقية', 'الغربية', 'المنوفية', 
                        'القليوبية', 'البحيرة', 'المنيا', 'أسيوط', 'سوهاج', 'قنا', 'الأقصر', 
                        'أسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط', 'بور سعيد', 
                        'الإسماعيلية', 'السويس', 'شمال سيناء', 'جنوب سيناء', 'البحر الأحمر', 
                        'مطروح', 'كفر الشيخ', 'الدقهلية', 'اسكندريه', 'اسيوط', 'سوهاج', 'قنا',
                        'الاقصر', 'اسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط',
                        'بورسعيد', 'الاسماعيليه', 'السويس', 'شمال سيناء', 'جنوب سيناء',
                        'البحر الاحمر', 'مطروح', 'كفر الشيخ', 'الدقهليه', 'الشرقيه', 'الغربيه',
                        'المنوفيه', 'القليوبيه', 'البحيره', 'المنيا', 'اسيوط', 'سوهاج', 'قنا',
                        'الاقصر', 'اسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط',
                        'بورسعيد', 'الاسماعيليه', 'السويس', 'شمال سيناء', 'جنوب سيناء',
                        'البحر الاحمر', 'مطروح', 'كفر الشيخ', 'الدقهليه', 'الشرقيه', 'الغربيه',
                        'المنوفيه', 'القليوبيه', 'البحيره', 'المنيا', 'اسيوط', 'سوهاج', 'قنا',
                        'الاقصر', 'اسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط',
                        'بورسعيد', 'الاسماعيليه', 'السويس', 'شمال سيناء', 'جنوب سيناء',
                        'البحر الاحمر', 'مطروح', 'كفر الشيخ', 'الدقهليه'
                    ]
                    
                    if any(gov.lower() in col5_value.lower() for gov in governorates_check):
                        governorate = col5_value
                        address_details = col6_value
                        if len(row) > 7:
                            tracking_number = row[7].strip()
                        if len(row) > 8:
                            order_date_str = row[8].strip()
                        
                        if not address_details or not address_details.strip():
                            address_details = "عنوان غير محدد"
                        
                        if not governorate or not governorate.strip():
                            governorate = "غير محدد"
                        
                        if not tracking_number or not tracking_number.strip():
                            tracking_number = "غير محدد"
                
                if not governorate and address_details:
                    governorates = [
                        'القاهرة', 'الجيزة', 'الإسكندرية', 'الشرقية', 'الغربية', 'المنوفية', 
                        'القليوبية', 'البحيرة', 'المنيا', 'أسيوط', 'سوهاج', 'قنا', 'الأقصر', 
                        'أسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط', 'بور سعيد', 
                        'الإسماعيلية', 'السويس', 'شمال سيناء', 'جنوب سيناء', 'البحر الأحمر', 
                        'مطروح', 'كفر الشيخ', 'الدقهلية', 'اسكندريه', 'اسيوط', 'سوهاج', 'قنا',
                        'الاقصر', 'اسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط',
                        'بورسعيد', 'الاسماعيليه', 'السويس', 'شمال سيناء', 'جنوب سيناء',
                        'البحر الاحمر', 'مطروح', 'كفر الشيخ', 'الدقهليه', 'الشرقيه', 'الغربيه',
                        'المنوفيه', 'القليوبيه', 'البحيره', 'المنيا', 'اسيوط', 'سوهاج', 'قنا',
                        'الاقصر', 'اسوان', 'بني سويف', 'الفيوم', 'الوادي الجديد', 'دمياط',
                        'بورسعيد', 'الاسماعيليه', 'السويس', 'شمال سيناء', 'جنوب سيناء',
                        'البحر الاحمر', 'مطروح', 'كفر الشيخ', 'الدقهليه'
                    ]
                    
                    for gov in governorates:
                        if gov.lower() in address_details.lower():
                            governorate = gov
                            break
                    
                    if not governorate and len(row) > 6:
                        col6_value = row[6].strip()
                        for gov in governorates:
                            if gov.lower() in col6_value.lower():
                                governorate = gov
                                break
                    
                    if not governorate:
                        governorate = 'غير محدد'
                
                order_date = datetime.now()
                if order_date_str:
                    try:
                        if '/' in order_date_str:
                            parts = order_date_str.split('/')
                            if len(parts) == 3:
                                if len(parts[0]) == 4:
                                    order_date = datetime.strptime(order_date_str, '%Y/%m/%d')
                                else:
                                    order_date = datetime.strptime(order_date_str, '%d/%m/%Y')
                        elif '-' in order_date_str:
                            order_date = datetime.strptime(order_date_str, '%Y-%m-%d')
                        else:
                            if len(order_date_str) == 8:
                                order_date = datetime.strptime(order_date_str, '%Y%m%d')
                            else:
                                order_date = datetime.strptime(order_date_str, '%Y/%m/%d')
                    except ValueError:
                        order_date = datetime.now()
                
                if not customer_name or not customer_name.strip():
                    customer_name = "غير محدد"
                
                if not order_details or not order_details.strip():
                    order_details = "طلب بدون تفاصيل"
                
                if not customer_phone or not customer_phone.strip():
                    customer_phone = "00000000000"
                
                customer = Customer.query.filter_by(phone=customer_phone).first()
                if not customer:
                    customer = Customer(
                        name=customer_name,
                        phone=customer_phone,
                        governorate=governorate,
                        address_details=address_details
                    )
                    db.session.add(customer)
                    db.session.flush()
                    customers_created += 1
                else:
                    if customer.name != customer_name or customer.governorate != governorate or customer.address_details != address_details:
                        customer.name = customer_name
                        customer.governorate = governorate
                        customer.address_details = address_details
                
                total_amount = amount_paid + remaining_amount
                
                if tracking_number and len(tracking_number) > 100:
                    tracking_number = tracking_number[:100]
                
                order = Order(
                    customer_id=customer.id,
                    date=order_date,
                    status='وصل',
                    delivery_fees=0,
                    total_amount=total_amount,
                    amount_paid=amount_paid,
                    notes=order_details,
                    tracking_number=tracking_number
                )
                
                db.session.add(order)
                orders_created += 1
                
                if orders_created % 100 == 0:
                    db.session.commit()
                    flash(f'تم حفظ {orders_created} طلب حتى الآن...', 'info')
                
            except Exception as e:
                error_details = f"الصف {row_num}: خطأ في معالجة البيانات - {str(e)}"
                errors.append(error_details)
                continue
        
        try:
            db.session.commit()
            flash('تم حفظ جميع البيانات بنجاح', 'success')
        except Exception as commit_error:
            db.session.rollback()
            flash(f'خطأ في حفظ البيانات: {str(commit_error)}', 'error')
            return redirect(url_for('main.orders'))
        
        success_msg = f"✅ تم استيراد {orders_created} طلب بنجاح من {len(data_rows)} صف"
        if customers_created > 0:
            success_msg += f"، تم إنشاء {customers_created} عميل جديد"
        
        flash(success_msg, 'success')
        
        if errors:
            error_msg = f"حدثت {len(errors)} أخطاء: "
            if len(errors) <= 3:
                error_msg += "; ".join(errors)
            else:
                error_msg += "; ".join(errors[:3]) + f" و {len(errors) - 3} أخطاء أخرى"
            flash(error_msg, 'warning')
        
        try:
            signal.alarm(0)
        except (AttributeError, OSError):
            pass
            
        return redirect(url_for('main.orders'))
        
    except Exception as e:
        try:
            signal.alarm(0)
        except (AttributeError, OSError):
            pass
        
        flash(f'خطأ في قراءة الملف: {str(e)}', 'error')
        flash('يرجى التحقق من تنسيق الملف والتأكد من أنه ملف CSV صحيح', 'warning')
        return redirect(url_for('main.orders'))

@main.route('/api/daily_statistics')
@login_required
@permission_required('can_view_statistics')
def daily_statistics():
    try:
        date_param = request.args.get('date')
        if date_param:
            try:
                selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                selected_date = date.today()
        else:
            selected_date = date.today()
        
        from app.models import OrderStatusHistory, ReplacementOrderStatusHistory
        from collections import defaultdict

        created_orders_today = Order.query.filter(
            func.date(Order.date) == selected_date
        ).order_by(Order.id.desc()).all()

        order_history_rows = db.session.query(
            OrderStatusHistory.order_id,
            OrderStatusHistory.status,
        ).filter(
            func.date(OrderStatusHistory.timestamp) == selected_date
        ).all()

        orders_by_status = defaultdict(set)
        for _oid, _st in order_history_rows:
            orders_by_status[_st].add(_oid)

        delivery_order_ids = list(orders_by_status.get('خرج للتوصيل', set()))
        delivered_order_ids = list(orders_by_status.get('وصل', set()))
        new_order_ids = list(orders_by_status.get('جديد', set()))

        new_orders_today       = Order.query.filter(Order.id.in_(new_order_ids)).all()       if new_order_ids       else []
        delivery_orders_today  = Order.query.filter(Order.id.in_(delivery_order_ids)).all()  if delivery_order_ids  else []
        delivered_orders_today = Order.query.filter(Order.id.in_(delivered_order_ids)).all() if delivered_order_ids else []

        new_orders_count       = len(new_orders_today)
        new_orders_value       = sum((o.total_amount or 0) for o in new_orders_today)
        new_orders_fees        = sum((o.delivery_fees or 0) for o in new_orders_today)

        delivery_orders_count  = len(delivery_orders_today)
        delivery_orders_value  = sum((o.total_amount or 0) for o in delivery_orders_today)
        delivery_orders_fees   = sum((o.delivery_fees or 0) for o in delivery_orders_today)

        delivered_orders_count = len(delivered_orders_today)
        delivered_orders_value = sum((o.total_amount or 0) for o in delivered_orders_today)
        delivered_orders_fees  = sum((o.delivery_fees or 0) for o in delivered_orders_today)

        replacement_history_rows = db.session.query(
            ReplacementOrderStatusHistory.replacement_order_id,
            ReplacementOrderStatusHistory.status,
        ).filter(
            func.date(ReplacementOrderStatusHistory.timestamp) == selected_date
        ).all()

        replacements_by_status = defaultdict(set)
        for _rid, _st in replacement_history_rows:
            replacements_by_status[_st].add(_rid)

        r_delivery_ids  = list(replacements_by_status.get('خرج للتوصيل', set()))
        r_delivered_ids = list(replacements_by_status.get('وصل', set()))
        r_new_ids       = list(replacements_by_status.get('جديد', set()))

        replacement_created_orders    = ReplacementOrder.query.filter(
            func.date(ReplacementOrder.date) == selected_date
        ).order_by(ReplacementOrder.id.desc()).all()
        replacement_new_orders       = ReplacementOrder.query.filter(ReplacementOrder.id.in_(r_new_ids)).all() if r_new_ids else []
        replacement_delivery_orders  = ReplacementOrder.query.filter(ReplacementOrder.id.in_(r_delivery_ids)).all()  if r_delivery_ids  else []
        replacement_delivered_orders = ReplacementOrder.query.filter(ReplacementOrder.id.in_(r_delivered_ids)).all() if r_delivered_ids else []

        replacement_new_count              = len(replacement_new_orders)
        replacement_new_value              = sum((ro.total_amount or 0) for ro in replacement_new_orders)
        replacement_new_fees_company       = sum((ro.delivery_fees or 0) for ro in replacement_new_orders)
        replacement_new_fees_customer      = sum((ro.delivery_fees_customer or 0) for ro in replacement_new_orders)

        replacement_delivery_count         = len(replacement_delivery_orders)
        replacement_delivery_value         = sum((ro.total_amount or 0) for ro in replacement_delivery_orders)
        replacement_delivery_fees_company  = sum((ro.delivery_fees or 0) for ro in replacement_delivery_orders)
        replacement_delivery_fees_customer = sum((ro.delivery_fees_customer or 0) for ro in replacement_delivery_orders)

        replacement_delivered_count         = len(replacement_delivered_orders)
        replacement_delivered_value         = sum((ro.total_amount or 0) for ro in replacement_delivered_orders)
        replacement_delivered_fees_company  = sum((ro.delivery_fees or 0) for ro in replacement_delivered_orders)
        replacement_delivered_fees_customer = sum((ro.delivery_fees_customer or 0) for ro in replacement_delivered_orders)

        # ── الإجمالي اليومي بحسب الطلبات الفريدة فقط ───────────────────────
        order_activity_ids = {o.id for o in created_orders_today}
        order_activity_ids.update(oid for oid, _ in order_history_rows)

        replacement_activity_ids = {ro.id for ro in replacement_created_orders}
        replacement_activity_ids.update(rid for rid, _ in replacement_history_rows)

        unique_orders_today = Order.query.filter(Order.id.in_(order_activity_ids)).all() if order_activity_ids else []
        unique_replacements_today = ReplacementOrder.query.filter(ReplacementOrder.id.in_(replacement_activity_ids)).all() if replacement_activity_ids else []

        total_count = len(order_activity_ids) + len(replacement_activity_ids)
        total_value = round(
            sum((o.total_amount or 0) for o in unique_orders_today)
            + sum((ro.total_amount or 0) for ro in unique_replacements_today),
            2,
        )
        total_fees = round(
            sum((o.delivery_fees or 0) for o in unique_orders_today)
            + sum((ro.delivery_fees or 0) + (ro.delivery_fees_customer or 0) for ro in unique_replacements_today),
            2,
        )
        
        orders_paid_today = unique_orders_today
        orders_paid_amount = sum((order.amount_paid or 0) for order in orders_paid_today)
        
        replacement_orders_paid_today = unique_replacements_today
        replacement_paid_amount = sum((ro.amount_paid or 0) for ro in replacement_orders_paid_today)
        
        total_paid_today = orders_paid_amount + replacement_paid_amount
        
        invoices = Invoice.query.filter(func.date(Invoice.date) <= selected_date).all()
        manual_debts = SupplierDebt.query.filter(
            func.date(SupplierDebt.date) <= selected_date,
            SupplierDebt.is_payment == False
        ).all()
        supplier_returns_total = db.session.query(
            func.coalesce(func.sum(func.coalesce(SupplierReturn.total_amount, 0)), 0)
        ).filter(
            func.date(SupplierReturn.return_date) <= selected_date
        ).scalar() or 0
        total_supplier_debt = sum((inv.total_amount or 0) - (inv.paid_amount or 0) for inv in invoices)
        total_supplier_debt += sum((d.amount or 0) - (d.paid_amount or 0) for d in manual_debts)
        total_supplier_debt -= supplier_returns_total
        total_supplier_debt = max(total_supplier_debt, 0)
        
        return jsonify({
            'success': True,
            'date': selected_date.strftime('%Y-%m-%d'),
            'new_orders': {
                'count': new_orders_count,
                'value': round(new_orders_value, 2),
                'fees': round(new_orders_fees, 2)
            },
            'created_orders': {
                'count': new_orders_count,
                'value': round(new_orders_value, 2),
                'fees': round(new_orders_fees, 2)
            },
            'delivery_orders': {
                'count': delivery_orders_count,
                'value': round(delivery_orders_value, 2),
                'fees': round(delivery_orders_fees, 2)
            },
            'delivered_orders': {
                'count': delivered_orders_count,
                'value': round(delivered_orders_value, 2),
                'fees': round(delivered_orders_fees, 2)
            },
            'replacement_new_orders': {
                'count': replacement_new_count,
                'value': round(replacement_new_value, 2),
                'fees_company': round(replacement_new_fees_company, 2),
                'fees_customer': round(replacement_new_fees_customer, 2)
            },
            'replacement_created_orders': {
                'count': replacement_new_count,
                'value': round(replacement_new_value, 2),
                'fees_company': round(replacement_new_fees_company, 2),
                'fees_customer': round(replacement_new_fees_customer, 2)
            },
            'replacement_delivery_orders': {
                'count': replacement_delivery_count,
                'value': round(replacement_delivery_value, 2),
                'fees_company': round(replacement_delivery_fees_company, 2),
                'fees_customer': round(replacement_delivery_fees_customer, 2)
            },
            'replacement_delivered_orders': {
                'count': replacement_delivered_count,
                'value': round(replacement_delivered_value, 2),
                'fees_company': round(replacement_delivered_fees_company, 2),
                'fees_customer': round(replacement_delivered_fees_customer, 2)
            },
            'total': {
                'count': total_count,
                'value': round(total_value, 2),
                'fees': round(total_fees, 2)
            },
            'total_paid_today': round(total_paid_today, 2),
            'total_supplier_debt': round(total_supplier_debt, 2)
        })
    except Exception as e:
        current_app.logger.error(f"Error in daily_statistics: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@main.route('/api/daily_fees_breakdown')
@login_required
@permission_required('can_view_statistics')
def daily_fees_breakdown():
    try:
        date_param = request.args.get('date')
        if date_param:
            try:
                selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                selected_date = date.today()
        else:
            selected_date = date.today()

        from app.models import OrderStatusHistory, ReplacementOrderStatusHistory
        from collections import defaultdict

        # --- طلبات عادية خرجت للتوصيل اليوم ---
        order_delivery_rows = db.session.query(
            OrderStatusHistory.order_id,
            Order
        ).join(
            Order,
            Order.id == OrderStatusHistory.order_id
        ).options(
            selectinload(Order.customer)
        ).filter(
            func.date(OrderStatusHistory.timestamp) == selected_date,
            OrderStatusHistory.status == 'خرج للتوصيل'
        ).order_by(
            OrderStatusHistory.timestamp.asc(),
            OrderStatusHistory.id.asc()
        ).all()

        # تجميع الطلبات العادية – المحافظات بنفس قيمة الرسوم في صف واحد
        orders_by_gov = defaultdict(lambda: {'count': 0, 'fee_rate': 0.0})
        seen_order_ids = set()
        for _, o in order_delivery_rows:
            if o.id in seen_order_ids:
                continue
            seen_order_ids.add(o.id)
            if o.is_nearest_post_branch:
                gov = 'لأقرب فرع بريد'
            else:
                gov = (o.customer.governorate or 'غير محدد') if o.customer else 'غير محدد'
            orders_by_gov[gov]['count'] += 1
            if orders_by_gov[gov]['fee_rate'] == 0.0:
                orders_by_gov[gov]['fee_rate'] = float(o.delivery_fees or 0)

        # تجميع المحافظات التي لها نفس قيمة الرسوم
        fee_groups = defaultdict(lambda: {'governorates': [], 'count': 0})
        for gov, v in orders_by_gov.items():
            fee_key = round(v['fee_rate'], 2)
            fee_groups[fee_key]['governorates'].append(gov)
            fee_groups[fee_key]['count'] += v['count']

        orders_breakdown = [
            {'governorates': sorted(v['governorates']), 'count': v['count'], 'fees': fee_key}
            for fee_key, v in sorted(fee_groups.items(), key=lambda x: -x[0])
        ]

        # --- طلبات استبدال خرجت للتوصيل اليوم ---
        replacements_breakdown = []
        rep_delivery_rows = db.session.query(
            ReplacementOrderStatusHistory.replacement_order_id,
            ReplacementOrder
        ).join(
            ReplacementOrder,
            ReplacementOrder.id == ReplacementOrderStatusHistory.replacement_order_id
        ).options(
            selectinload(ReplacementOrder.customer)
        ).filter(
            func.date(ReplacementOrderStatusHistory.timestamp) == selected_date,
            ReplacementOrderStatusHistory.status == 'خرج للتوصيل'
        ).order_by(
            ReplacementOrderStatusHistory.timestamp.asc(),
            ReplacementOrderStatusHistory.id.asc()
        ).all()

        seen_rep_ids = set()
        for _, ro in rep_delivery_rows:
            if ro.id in seen_rep_ids:
                continue
            seen_rep_ids.add(ro.id)
            gov = ro.alternative_governorate
            if not gov and ro.customer:
                gov = ro.customer.governorate
            gov = gov or 'غير محدد'
            replacements_breakdown.append({
                'id': ro.id,
                'governorate': gov,
                'fees_company': round(float(ro.delivery_fees or 0), 2),
                'fees_customer': round(float(ro.delivery_fees_customer or 0), 2),
                'total_fees': round(float(ro.delivery_fees or 0) + float(ro.delivery_fees_customer or 0), 2)
            })

        return jsonify({
            'success': True,
            'date': selected_date.strftime('%Y-%m-%d'),
            'orders_breakdown': orders_breakdown,
            'replacements_breakdown': replacements_breakdown
        })
    except Exception as e:
        current_app.logger.error(f"Error in daily_fees_breakdown: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/daily_statistics_details')
@login_required
@permission_required('can_view_statistics')
def daily_statistics_details():
    try:
        date_param = request.args.get('date')
        order_type = request.args.get('type')
        
        if date_param:
            try:
                selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                selected_date = date.today()
        else:
            selected_date = date.today()
        
        from app.models import OrderStatusHistory, ReplacementOrderStatusHistory
        from collections import defaultdict

        orders = []

        def _orders_by_status_rows(status_val):
            """كل سجلات الحالة للطلبات داخل اليوم المختار."""
            return (
                db.session.query(OrderStatusHistory, Order)
                .join(Order, Order.id == OrderStatusHistory.order_id)
                .filter(
                    func.date(OrderStatusHistory.timestamp) == selected_date,
                    OrderStatusHistory.status == status_val,
                )
                .order_by(OrderStatusHistory.timestamp.asc(), OrderStatusHistory.id.asc())
                .all()
            )

        def _replacements_by_status_rows(status_val):
            """كل سجلات الحالة لطلبات الاستبدال داخل اليوم المختار."""
            return (
                db.session.query(ReplacementOrderStatusHistory, ReplacementOrder)
                .join(ReplacementOrder, ReplacementOrder.id == ReplacementOrderStatusHistory.replacement_order_id)
                .filter(
                    func.date(ReplacementOrderStatusHistory.timestamp) == selected_date,
                    ReplacementOrderStatusHistory.status == status_val,
                )
                .order_by(ReplacementOrderStatusHistory.timestamp.asc(), ReplacementOrderStatusHistory.id.asc())
                .all()
            )

        if order_type == 'new':
            orders_query = _orders_by_status_rows('جديد')

            seen_order_ids = set()
            for history, order in orders_query:
                if order.id in seen_order_ids:
                    continue
                seen_order_ids.add(order.id)
                total = (order.total_amount or 0)
                fees = (order.delivery_fees or 0)
                paid = (order.amount_paid or 0)
                remaining = total + fees - paid

                orders.append({
                    'id': order.id,
                    'customer_name': order.customer_name or 'غير معروف',
                    'total_amount': round(total, 2),
                    'delivery_fees': round(fees, 2),
                    'amount_paid': round(paid, 2),
                    'remaining': round(remaining, 2),
                    'date': order.date.strftime('%Y-%m-%d') if order.date else '',
                    'status_updated_at': history.timestamp.strftime('%Y-%m-%d %H:%M:%S') if history.timestamp else ''
                })

        elif order_type == 'delivery':
            orders_query = _orders_by_status_rows('خرج للتوصيل')

            seen_order_ids = set()
            for history, order in orders_query:
                if order.id in seen_order_ids:
                    continue
                seen_order_ids.add(order.id)
                total = (order.total_amount or 0)
                fees = (order.delivery_fees or 0)
                paid = (order.amount_paid or 0)
                remaining = total + fees - paid

                orders.append({
                    'id': order.id,
                    'customer_name': order.customer_name or 'غير معروف',
                    'total_amount': round(total, 2),
                    'delivery_fees': round(fees, 2),
                    'amount_paid': round(paid, 2),
                    'remaining': round(remaining, 2),
                    'date': order.date.strftime('%Y-%m-%d') if order.date else '',
                    'status_updated_at': history.timestamp.strftime('%Y-%m-%d %H:%M:%S') if history.timestamp else ''
                })

        elif order_type == 'delivered':
            orders_query = _orders_by_status_rows('وصل')

            seen_order_ids = set()
            for history, order in orders_query:
                if order.id in seen_order_ids:
                    continue
                seen_order_ids.add(order.id)
                total = (order.total_amount or 0)
                fees = (order.delivery_fees or 0)
                paid = (order.amount_paid or 0)
                remaining = total + fees - paid

                orders.append({
                    'id': order.id,
                    'customer_name': order.customer_name or 'غير معروف',
                    'total_amount': round(total, 2),
                    'delivery_fees': round(fees, 2),
                    'amount_paid': round(paid, 2),
                    'remaining': round(remaining, 2),
                    'date': order.date.strftime('%Y-%m-%d') if order.date else '',
                    'status_updated_at': history.timestamp.strftime('%Y-%m-%d %H:%M:%S') if history.timestamp else ''
                })

        elif order_type in ['replacement_new', 'replacement_delivery', 'replacement_delivered']:
            if order_type == 'replacement_new':
                replacement_orders_query = _replacements_by_status_rows('جديد')
            elif order_type == 'replacement_delivery':
                replacement_orders_query = _replacements_by_status_rows('خرج للتوصيل')
            elif order_type == 'replacement_delivered':
                replacement_orders_query = _replacements_by_status_rows('وصل')
            else:
                replacement_orders_query = []
            
            seen_rep_ids = set()
            for history, ro in replacement_orders_query:
                if ro.id in seen_rep_ids:
                    continue
                seen_rep_ids.add(ro.id)
                total = (ro.total_amount or 0)
                fees_company = (ro.delivery_fees or 0)
                fees_customer = (ro.delivery_fees_customer or 0)
                paid = (ro.amount_paid or 0)
                remaining = total + fees_customer - paid
                
                orders.append({
                    'id': ro.id,
                    'customer_name': ro.customer.name if ro.customer else 'غير معروف',
                    'total_amount': round(total, 2),
                    'delivery_fees': round(fees_customer, 2),
                    'delivery_fees_company': round(fees_company, 2),
                    'amount_paid': round(paid, 2),
                    'remaining': round(remaining, 2),
                    'date': ro.date.strftime('%Y-%m-%d') if ro.date else '',
                    'status_updated_at': history.timestamp.strftime('%Y-%m-%d %H:%M:%S') if history.timestamp else ''
                })
        
        return jsonify({
            'success': True,
            'orders': orders
        })
    except Exception as e:
        current_app.logger.error(f"Error in daily_statistics_details: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@main.route('/orders/<int:order_id>/status-history')
@login_required
@permission_required('can_view_orders')
def order_status_history(order_id):
    order = Order.query.options(
        joinedload(Order.registered_by),
        joinedload(Order.called_by),
        joinedload(Order.verified_by)
    ).get_or_404(order_id)

    status_timeline = order.get_status_timeline()
    edit_logs = order.edit_logs.order_by(OrderEditLog.timestamp).all()

    events = []
    for r in status_timeline:
        review_desc = _delivery_call_review_description_from_note(r.notes)
        if review_desc:
            events.append({
                'event_type': 'review',
                'timestamp': r.timestamp,
                'description': review_desc,
                'employee': r.changed_by,
                'notes': None
            })
            continue
        events.append({
            'event_type': 'status',
            'timestamp': r.timestamp,
            'description': r.status,
            'employee': r.changed_by,
            'notes': r.notes
        })
    for l in edit_logs:
        events.append({
            'event_type': 'edit',
            'timestamp': l.timestamp,
            'description': l.description,
            'employee': l.employee,
            'notes': None
        })

    review_items = [
        (order.registered, order.registered_at, order.registered_by, 'تم التسجيل'),
        (order.customer_called, order.customer_called_at, order.called_by, 'تم الاتصال بالعميل'),
        (order.customer_verified, order.customer_verified_at, order.verified_by, 'تم التحقق من العميل')
    ]
    for flag, ts, emp, desc in review_items:
        if flag:
            events.append({
                'event_type': 'review',
                'timestamp': ts,
                'description': desc,
                'employee': emp,
                'notes': None
            })

    events.sort(key=lambda x: x['timestamp'] or datetime.min, reverse=True)

    return render_template(
        'order_status_history.html',
        order=order,
        status_timeline=status_timeline,
        events=events
    )

@main.route('/api/orders/<int:order_id>/status-history')
@login_required
@permission_required('can_view_orders')
def api_order_status_history(order_id):
    try:
        order = Order.query.get_or_404(order_id)
        
        status_timeline = order.get_status_timeline()
        
        history_data = []
        for record in status_timeline:
            history_data.append({
                'id': record.id,
                'status': record.status,
                'timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S') if record.timestamp else None,
                'changed_by': record.changed_by.name if record.changed_by else None,
                'changed_by_id': record.changed_by_employee_id,
                'notes': record.notes
            })
        
        return jsonify({
            'success': True,
            'order_id': order.id,
            'current_status': order.status,
            'history': history_data
        })
    except Exception as e:
        current_app.logger.error(f"Error in api_order_status_history: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@main.route('/replacement-orders/<int:order_id>/status-history')
@login_required
@permission_required('can_view_replacements')
def replacement_order_status_history(order_id):
    order = ReplacementOrder.query.options(
        joinedload(ReplacementOrder.registered_by),
        joinedload(ReplacementOrder.called_by),
        joinedload(ReplacementOrder.verified_by)
    ).get_or_404(order_id)

    status_timeline = order.get_status_timeline()
    edit_logs = order.edit_logs.order_by(OrderEditLog.timestamp).all()

    events = []
    for r in status_timeline:
        review_desc = _delivery_call_review_description_from_note(r.notes)
        if review_desc:
            events.append({
                'event_type': 'review',
                'timestamp': r.timestamp,
                'description': review_desc,
                'employee': r.changed_by,
                'notes': None
            })
            continue
        events.append({
            'event_type': 'status',
            'timestamp': r.timestamp,
            'description': r.status,
            'employee': r.changed_by,
            'notes': r.notes
        })
    for l in edit_logs:
        events.append({
            'event_type': 'edit',
            'timestamp': l.timestamp,
            'description': l.description,
            'employee': l.employee,
            'notes': None
        })

    review_items = [
        (order.registered, order.registered_at, order.registered_by, 'تم التسجيل'),
        (order.customer_called, order.customer_called_at, order.called_by, 'تم الاتصال بالعميل'),
        (order.customer_verified, order.customer_verified_at, order.verified_by, 'تم التحقق من العميل')
    ]
    for flag, ts, emp, desc in review_items:
        if flag:
            events.append({
                'event_type': 'review',
                'timestamp': ts,
                'description': desc,
                'employee': emp,
                'notes': None
            })

    events.sort(key=lambda x: x['timestamp'] or datetime.min, reverse=True)

    return render_template(
        'replacement_order_status_history.html',
        order=order,
        status_timeline=status_timeline,
        events=events
    )

@main.route('/api/replacement-orders/<int:order_id>/status-history')
@login_required
@permission_required('can_view_replacements')
def api_replacement_order_status_history(order_id):
    try:
        order = ReplacementOrder.query.get_or_404(order_id)
        
        status_timeline = order.get_status_timeline()
        
        history_data = []
        for record in status_timeline:
            history_data.append({
                'id': record.id,
                'status': record.status,
                'timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S') if record.timestamp else None,
                'changed_by': record.changed_by.name if record.changed_by else None,
                'changed_by_id': record.changed_by_employee_id,
                'notes': record.notes
            })
        
        return jsonify({
            'success': True,
            'order_id': order.id,
            'current_status': order.status,
            'history': history_data
        })
    except Exception as e:
        current_app.logger.error(f"Error in api_replacement_order_status_history: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


STOCKTAKE_ITEMS_PER_PAGE = 50


@main.route('/products/stocktake')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_index():
    in_progress = StockTake.query.filter_by(status='in_progress').first()
    return render_template('stocktake_index.html', in_progress=in_progress)


@main.route('/products/stocktake/new')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_new():
    existing = StockTake.query.filter_by(status='in_progress').first()
    if existing:
        flash('يوجد جرد قيد التنفيذ بالفعل. تم تحويلك إليه.', 'info')
        return redirect(url_for('main.stocktake_new_page', id=existing.id))

    employee_id = session.get('employee_id')
    try:
        stocktake = StockTake(status='in_progress', started_by_id=employee_id)
        db.session.add(stocktake)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        existing = StockTake.query.filter_by(status='in_progress').first()
        if existing:
            flash('تم إنشاء جرد بواسطة موظف آخر في نفس اللحظة. تم تحويلك إليه.', 'info')
            return redirect(url_for('main.stocktake_new_page', id=existing.id))
        flash('حدث خطأ أثناء إنشاء الجرد.', 'danger')
        return redirect(url_for('main.stocktake_index'))

    return redirect(url_for('main.stocktake_new_page', id=stocktake.id))


@main.route('/products/stocktake/<int:id>/new')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_new_page(id):
    stocktake = StockTake.query.get_or_404(id)

    search = request.args.get('search', '').strip()

    products_query = Product.query.filter(
        Product.is_deleted == False,
        Product.is_bundle == False,
    )

    if search:
        products_query = products_query.filter(Product.name.ilike(f'%{search}%'))

    products = products_query.order_by(Product.name.asc()).all()

    existing_items = {
        item.product_id: item
        for item in StockTakeItem.query.filter_by(stocktake_id=id).all()
        if item.product_id is not None
    }

    return render_template(
        'stocktake_new.html',
        stocktake=stocktake,
        products=products,
        existing_items=existing_items,
        search=search,
    )


@main.route('/products/stocktake/<int:id>/item', methods=['POST'])
@login_required
@permission_required('can_manage_stocktake')
def stocktake_save_item(id):
    stocktake = StockTake.query.get_or_404(id)
    if stocktake.status != 'in_progress':
        return jsonify({'error': 'هذا الجرد لم يعد قيد التنفيذ'}), 409

    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'طلب غير صالح'}), 400

    product_id = data.get('product_id')
    counted_stock = data.get('counted_stock')

    if product_id is None:
        return jsonify({'error': 'معرف المنتج مطلوب'}), 400

    product = Product.query.get(product_id)
    if not product or product.is_deleted or product.is_bundle:
        return jsonify({'error': 'المنتج غير صالح أو محذوف'}), 400

    if counted_stock is not None:
        try:
            counted_stock = int(counted_stock)
        except (TypeError, ValueError):
            return jsonify({'error': 'قيمة الجرد يجب أن تكون رقمًا صحيحًا'}), 400
        if counted_stock < 0:
            return jsonify({'error': 'قيمة الجرد لا يمكن أن تكون سالبة'}), 400

    from sqlalchemy.dialects.postgresql import insert as pg_insert

    system_stock = product.stock

    stmt = pg_insert(StockTakeItem).values(
        stocktake_id=id,
        product_id=product_id,
        product_name_snapshot=product.name,
        system_stock=system_stock,
        counted_stock=counted_stock,
        diff=(counted_stock - system_stock) if counted_stock is not None else None,
        was_skipped=False,
    )
    stmt = stmt.on_conflict_do_update(
        constraint='uq_stocktake_product',
        set_={
            'system_stock': system_stock,
            'counted_stock': counted_stock,
            'diff': (counted_stock - system_stock) if counted_stock is not None else None,
            'was_skipped': False,
            'product_name_snapshot': product.name,
        },
    )
    db.session.execute(stmt)
    db.session.commit()

    return jsonify({'success': True}), 200


@main.route('/products/stocktake/<int:id>/finish', methods=['POST'])
@login_required
@permission_required('can_manage_stocktake')
def stocktake_finish(id):
    stocktake = StockTake.query.get_or_404(id)
    if stocktake.status != 'in_progress':
        flash('هذا الجرد لم يعد قيد التنفيذ.', 'warning')
        return redirect(url_for('main.stocktake_summary', id=id))

    active_products = Product.query.filter(
        Product.is_deleted == False,
        Product.is_bundle == False,
    ).all()
    existing_product_ids = {
        item.product_id
        for item in StockTakeItem.query.filter_by(stocktake_id=id).all()
        if item.product_id is not None
    }

    new_items = []
    for prod in active_products:
        if prod.id not in existing_product_ids:
            new_items.append(StockTakeItem(
                stocktake_id=id,
                product_id=prod.id,
                product_name_snapshot=prod.name,
                system_stock=prod.stock,
                counted_stock=prod.stock,
                diff=0,
                was_skipped=True,
            ))

    if new_items:
        db.session.bulk_save_objects(new_items)
        db.session.commit()

    return redirect(url_for('main.stocktake_summary', id=id))


@main.route('/products/stocktake/<int:id>/summary')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_summary(id):
    stocktake = StockTake.query.get_or_404(id)
    items = StockTakeItem.query.filter_by(stocktake_id=id).order_by(StockTakeItem.id).all()

    surplus = sum(1 for i in items if i.counted_stock is not None and i.diff is not None and i.diff > 0)
    shortage = sum(1 for i in items if i.counted_stock is not None and i.diff is not None and i.diff < 0)
    matched = sum(1 for i in items if i.counted_stock is not None and i.diff == 0)

    return render_template(
        'stocktake_summary.html',
        stocktake=stocktake,
        items=items,
        surplus=surplus,
        shortage=shortage,
        matched=matched,
    )


@main.route('/products/stocktake/<int:id>/apply', methods=['POST'])
@login_required
@permission_required('can_manage_stocktake')
def stocktake_apply(id):
    employee_id = session.get('employee_id')

    from sqlalchemy import text as sa_text
    result = db.session.execute(
        sa_text(
            "UPDATE stock_take SET status = 'applied', applied_by_id = :eid, applied_at = now() "
            "WHERE id = :id AND status = 'in_progress' RETURNING id"
        ),
        {'id': id, 'eid': employee_id},
    )
    row = result.fetchone()
    if not row:
        db.session.rollback()
        flash('تعذر تطبيق الجرد — ربما تم تطبيقه أو التخلص منه بالفعل.', 'warning')
        return redirect(url_for('main.stocktake_summary', id=id))

    stocktake = StockTake.query.get(id)

    items = StockTakeItem.query.filter(
        StockTakeItem.stocktake_id == id,
        StockTakeItem.product_id.isnot(None),
        StockTakeItem.counted_stock.isnot(None),
    ).all()

    for item in items:
        db.session.execute(
            sa_text("UPDATE product SET stock = :stock WHERE id = :pid"),
            {'stock': item.counted_stock, 'pid': item.product_id},
        )

    from app.pdf_utils import find_and_register_arabic_font, arabic_text
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm

    font_name = find_and_register_arabic_font()
    if not font_name:
        font_name = 'ArabicFont'

    buf = io.BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    p.setFont(font_name, 16)
    title = f'نتيجة جرد المخزون - {datetime.now().strftime("%d/%m/%Y")}'
    p.drawCentredString(width / 2.0, height - 2 * cm, arabic_text(title))

    surplus_count = sum(1 for i in items if i.diff is not None and i.diff > 0)
    shortage_count = sum(1 for i in items if i.diff is not None and i.diff < 0)
    matched_count = sum(1 for i in items if i.diff == 0)

    p.setFont(font_name, 11)
    y = height - 3.5 * cm
    summary_line = f'المنتجات الزائدة: {surplus_count}  |  المنتجات الناقصة: {shortage_count}  |  المتطابقة: {matched_count}'
    p.drawCentredString(width / 2.0, y, arabic_text(summary_line))

    y -= 1.5 * cm
    right_x = width - 1.5 * cm
    left_x = 1.5 * cm
    col_w = (right_x - left_x) / 5

    p.setFont(font_name, 10)
    headers = ['اسم المنتج', 'المخزون قبل', 'المخزون بعد', 'الفرق', 'الحالة']
    x_positions = [right_x - col_w, right_x - 2 * col_w, right_x - 3 * col_w, right_x - 4 * col_w, left_x]
    for i, h in enumerate(headers):
        p.drawString(x_positions[i], y, arabic_text(h))
    y -= 0.6 * cm

    p.setFont(font_name, 9)
    for item in items:
        if y < 2 * cm:
            p.showPage()
            y = height - 2 * cm
            p.setFont(font_name, 10)
            for i, h in enumerate(headers):
                p.drawString(x_positions[i], y, arabic_text(h))
            y -= 0.6 * cm
            p.setFont(font_name, 9)

        name = item.product_name_snapshot or ''
        sys_stock = str(item.system_stock or 0)
        cnt_stock = str(item.counted_stock or 0)
        diff_str = str(item.diff or 0)

        if item.diff is not None and item.diff > 0:
            status = 'فائض'
        elif item.diff is not None and item.diff < 0:
            status = 'ناقص'
        else:
            status = 'متطابق'

        p.drawString(x_positions[0], y, arabic_text(name[:60]))
        p.drawString(x_positions[1], y, sys_stock)
        p.drawString(x_positions[2], y, cnt_stock)
        p.drawString(x_positions[3], y, diff_str)
        p.drawString(x_positions[4], y, arabic_text(status))

        y -= 0.5 * cm

    p.save()
    buf.seek(0)

    instance_dir = os.path.join(current_app.instance_path, 'stocktakes')
    os.makedirs(instance_dir, exist_ok=True)
    pdf_filename = f'stocktake_{id}.pdf'
    pdf_path = os.path.join(instance_dir, pdf_filename)
    with open(pdf_path, 'wb') as f:
        f.write(buf.getvalue())

    stocktake.pdf_path = pdf_path
    db.session.commit()

    flash('تم تطبيق الجرد وتحديث المخزون بنجاح.', 'success')
    return redirect(url_for('main.stocktake_summary', id=id))


@main.route('/products/stocktake/<int:id>/discard', methods=['POST'])
@login_required
@permission_required('can_manage_stocktake')
def stocktake_discard(id):
    employee_id = session.get('employee_id')

    from sqlalchemy import text as sa_text
    result = db.session.execute(
        sa_text(
            "UPDATE stock_take SET status = 'discarded', discarded_by_id = :eid, discarded_at = now() "
            "WHERE id = :id AND status = 'in_progress' RETURNING id"
        ),
        {'id': id, 'eid': employee_id},
    )
    row = result.fetchone()
    if not row:
        db.session.rollback()
        flash('تعذر التخلص من الجرد — ربما تم تطبيقه أو التخلص منه بالفعل.', 'warning')
        return redirect(url_for('main.stocktake_summary', id=id))

    db.session.commit()

    flash('تم التخلص من الجرد.', 'info')
    return redirect(url_for('main.stocktake_summary', id=id))


@main.route('/products/stocktake/<int:id>/download')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_download(id):
    stocktake = StockTake.query.get_or_404(id)
    if not stocktake.pdf_path:
        flash('لم يتم تطبيق هذا الجرد بعد، لا يوجد ملف PDF.', 'warning')
        return redirect(url_for('main.stocktake_summary', id=id))
    return send_file(
        stocktake.pdf_path,
        as_attachment=True,
        download_name=f'stocktake_{id}.pdf',
    )


@main.route('/products/stocktake/list')
@login_required
@permission_required('can_manage_stocktake')
def stocktake_list():
    stocktakes = StockTake.query.order_by(StockTake.created_at.desc()).all()
    return render_template('stocktake_list.html', stocktakes=stocktakes)


@main.route('/api/orders/eligible_wassalha')
@login_required
def eligible_wassalha_orders_api():
    orders = Order.query.filter(
        Order.status == 'جديد',
        Order.registered == False,
        Order.is_nearest_post_branch == False,
        ~Order._customer_governorate.contains('فرع بريد'),
        Order.delivery_fees != 40,
    ).order_by(Order.id.desc()).options(
        selectinload(Order.items).selectinload(OrderItem._product).selectinload(Product.bundle_items).selectinload(BundleItem.product),
        selectinload(Order.items).selectinload(OrderItem.variant),
        selectinload(Order.items).selectinload(OrderItem.size_variant),
        selectinload(Order.items).selectinload(OrderItem.color_variant),
        selectinload(Order.items).selectinload(OrderItem.style_variant),
    ).all()
    return jsonify([{
        'id':             o.id,
        'customer_name':  o.customer_name,
        'customer_phone': o.customer_phone,
        'governorate':    o.customer_governorate or '',
        'date':           o.date.strftime('%Y-%m-%d') if o.date else '',
        'items': [{
            'product_name':   i.product.name if i.product else 'منتج محذوف',
            'quantity':       i.quantity,
            'variants':       [{'label': k, 'name': v.variant_name} for k, v in i.selected_variants.items()],
            'is_out_of_stock': bool(i.product and not i.product.is_bundle and (i.product.stock or 0) == 0),
            'is_bundle':      bool(i.product and i.product.is_bundle and i.product.bundle_items),
            'bundle_items': _build_bundle_items_json(i),
        } for i in o.items],
        'remaining':      o.remaining_amount or 0,
    } for o in orders])


def _build_bundle_items_json(item):
    if not item.product or not item.product.is_bundle or not item.product.bundle_items:
        return []
    bv_map = item.bundle_variants_map
    result = []
    for idx, bi in enumerate(item.product.bundle_items):
        bv = bv_map.get(idx, {})
        variants = [{'label': k, 'name': v.variant_name} for k, v in bv.get('variants', {}).items()]
        result.append({
            'product_name': bi.product.name if bi.product else 'منتج محذوف',
            'variants':     variants,
        })
    return result


@main.route('/api/replacement-orders/eligible_wassalha')
@login_required
def eligible_wassalha_replacement_orders_api():
    orders = ReplacementOrder.query.filter(
        ReplacementOrder.status == 'جديد',
        ReplacementOrder.registered == False,
        ReplacementOrder.delivery_fees != 40,
    ).order_by(ReplacementOrder.id.desc()).options(
        selectinload(ReplacementOrder.customer),
        selectinload(ReplacementOrder.items).selectinload(ReplacementOrderItem._product).selectinload(Product.bundle_items).selectinload(BundleItem.product),
        selectinload(ReplacementOrder.items).selectinload(ReplacementOrderItem.variant),
        selectinload(ReplacementOrder.items).selectinload(ReplacementOrderItem.size_variant),
        selectinload(ReplacementOrder.items).selectinload(ReplacementOrderItem.color_variant),
        selectinload(ReplacementOrder.items).selectinload(ReplacementOrderItem.style_variant),
    ).all()
    orders = [o for o in orders if not (o.customer and o.customer.governorate and 'فرع بريد' in o.customer.governorate)]
    return jsonify([{
        'id':             o.id,
        'customer_name':  o.customer_name,
        'customer_phone': o.customer_phone,
        'governorate':    o.customer_governorate or '',
        'date':           o.date.strftime('%Y-%m-%d') if o.date else '',
        'items': [{
            'product_name': i.product.name if i.product else 'منتج محذوف',
            'quantity':     i.quantity,
            'variants':     [{'label': k, 'name': v.variant_name} for k, v in i.selected_variants.items()],
            'is_deleted':     bool(i.product and i.product.is_deleted),
            'is_bundle':      bool(i.product and i.product.is_bundle and i.product.bundle_items),
            'has_product':    bool(i.product and i.product.name),
            'is_out_of_stock': bool(i.product and not i.product.is_bundle and (i.product.stock or 0) == 0),
            'bundle_items': _build_bundle_items_json(i),
        } for i in o.items if i.state in ('سليم', 'منتج جديد', 'جديد')],
        'remaining':      o.remaining_amount or 0,
    } for o in orders])


@main.route('/orders/export_wassalha', methods=['POST'])
def export_wassalha_orders():
    from datetime import datetime, date
    order_ids = request.form.getlist('order_ids')
    if not order_ids:
        flash('اختر طلب واحد على الأقل', 'warning')
        return redirect(url_for('main.orders_by_status', status='جديد'))
    orders = Order.query.filter(Order.id.in_(order_ids)).all()
    if not orders:
        flash('لا يوجد طلبات جديدة غير مسجلة في وصلها', 'warning')
        return redirect(url_for('main.orders_by_status', status='جديد'))
    buf = _generate_wassalha_xlsx(orders)
    employee_id = session.get('employee_id')
    now = datetime.utcnow()
    for order in orders:
        order.registered       = True
        order.registered_at    = now
        order.registered_by_id = employee_id
    db.session.commit()
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="wassalha_orders_{date.today()}.xlsx"'
    return resp


@main.route('/replacement-orders/export_wassalha', methods=['POST'])
def export_wassalha_replacement_orders():
    from datetime import datetime, date
    order_ids = request.form.getlist('order_ids')
    if not order_ids:
        flash('اختر طلب واحد على الأقل', 'warning')
        return redirect(url_for('main.replacement_orders_by_status', status='جديد'))
    orders = ReplacementOrder.query.filter(ReplacementOrder.id.in_(order_ids)).all()
    if not orders:
        flash('لا يوجد طلبات استبدال جديدة غير مسجلة في وصلها', 'warning')
        return redirect(url_for('main.replacement_orders_by_status', status='جديد'))
    buf = _generate_wassalha_xlsx(orders)
    employee_id = session.get('employee_id')
    now = datetime.utcnow()
    for order in orders:
        order.registered       = True
        order.registered_at    = now
        order.registered_by_id = employee_id
    db.session.commit()
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="wassalha_replacements_{date.today()}.xlsx"'
    return resp


@main.route('/auto-login')
def auto_login():
    try:
        from .models import Employee
        employee = Employee.query.filter_by(username='nebaxusbeta', is_active=True).first()
    except Exception:
        employee = None
    if not employee:
        flash('لم يتم العثور على حساب افتراضي')
        return redirect(url_for('main.login'))
    session['employee_id'] = employee.id
    session['employee_name'] = employee.name
    session['employee_username'] = employee.username
    session['is_admin'] = employee.is_admin
    return redirect(url_for('main.dashboard'))


@main.route('/api/trial-stats')
def trial_stats():
    from .trial import orders_this_month_count, is_read_only
    used = orders_this_month_count()
    return jsonify({
        'limit': 500,
        'used': used,
        'remaining': max(0, 500 - used),
        'read_only': is_read_only(),
        'pct': min(100, int(used / 500 * 100)),
    })


@main.route('/export', methods=['GET'])
@login_required
def export_data():
    import os
    from datetime import datetime
    from .trial import generate_nbx
    export_dir = os.environ.get('NEBAXUS_EXPORT_PATH') or os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'exports'
    )
    os.makedirs(export_dir, exist_ok=True)
    buf = generate_nbx()
    filename = f'nebaxus-export-{datetime.now():%Y-%m-%d}.nbx'
    filepath = os.path.join(export_dir, filename)
    with open(filepath, 'wb') as f:
        f.write(buf.getvalue())
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename,
    )


@main.route('/api/backup', methods=['POST'])
@login_required
def trigger_backup():
    import shutil
    import os
    from datetime import datetime
    db_dir = os.environ.get('NEBAXUS_INSTANCE_PATH') or os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'app', 'instance'
    )
    backup_dir = os.environ.get('NEBAXUS_BACKUP_PATH') or os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'backups'
    )
    os.makedirs(backup_dir, exist_ok=True)
    src = os.path.join(db_dir, 'dukana.db')
    if not os.path.exists(src):
        return jsonify({'success': False, 'error': 'لا يوجد ملف قاعدة بيانات للنسخ الاحتياطي'})
    date_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    dst = os.path.join(backup_dir, f'manual-{date_str}.db')
    shutil.copy2(src, dst)
    return jsonify({'success': True, 'path': dst})